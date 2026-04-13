"""Standalone engineering sidecar for merged V1 runs.

This tool is intentionally kept outside the V1 UI and outside the Step 2
default workflow. It operates on completed run directories and only performs
write/verify actions when the caller explicitly opts in via CLI flags.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook

from ..config import load_config
from ..data.points import CalibrationPoint, load_points_from_excel
from ..export import export_temperature_compensation_artifacts
from ..senco_format import format_senco_values
from ..v2.config import AppConfig, CoefficientsConfig
from ..v2.adapters.analyzer_coefficient_downloader import (
    CsvIoLogger,
    _resolve_gas_analyzer_class,
    download_coefficients_to_analyzers,
    load_download_plan,
    load_download_targets,
)
from ..v2.export.ratio_poly_report import (
    export_ratio_poly_report_from_summary_frame,
    load_summary_workbook_rows,
)
from . import verify_short_run


COEFF_DISPLAY_FORMAT = "0.00000E00"
DEFAULT_GAS_PPM = (0.0, 200.0, 400.0, 600.0, 800.0, 1000.0)
VERIFY_CO2_TEMP_C = 30.0
VERIFY_CO2_PPM = (200.0, 500.0, 800.0)
VERIFY_H2O_TARGETS = (
    (0.0, 0.0, 50.0),
    (20.0, 20.0, 70.0),
)


def _log(message: str) -> None:
    print(message, flush=True)


def _default_coefficients_payload() -> dict[str, Any]:
    return {
        "enabled": True,
        "auto_fit": True,
        "model": "ratio_poly_rt_p",
        "summary_columns": {
            "co2": {"target": "ppm_CO2_Tank", "ratio": "R_CO2", "temperature": "Temp", "pressure": "BAR"},
            "h2o": {"target": "ppm_H2O_Dew", "ratio": "R_H2O", "temperature": "Temp", "pressure": "BAR"},
        },
    }


def _load_coefficients_config_lazy(config_path: str | Path | None) -> CoefficientsConfig:
    # Keep merged-sidecar merge/report flows independent from SQLAlchemy-backed
    # postprocess modules so summary-only paths still run in lean environments.
    if not config_path:
        config = AppConfig.from_dict({"coefficients": _default_coefficients_payload()})
    else:
        config = AppConfig.from_json_file(str(Path(config_path)))
    coeff_cfg = config.coefficients
    coeff_cfg.enabled = True
    coeff_cfg.auto_fit = True
    return coeff_cfg


def _safe_float(value: Any) -> Optional[float]:
    if value in (None, "", "null", "None"):
        return None
    try:
        numeric = float(value)
    except Exception:
        return None
    if not math.isfinite(numeric):
        return None
    return numeric


def _safe_bool(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y", "on", "ok"}


def _phase_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"co2", "气路"}:
        return "co2"
    if text in {"h2o", "水路"}:
        return "h2o"
    return text


def _first_available(row: Mapping[str, Any], keys: Sequence[str]) -> Any:
    for key in keys:
        if key in row and row.get(key) not in (None, ""):
            return row.get(key)
    return None


def _first_valid_float(row: Mapping[str, Any], keys: Sequence[str]) -> Optional[float]:
    for key in keys:
        if key not in row:
            continue
        numeric = _safe_float(row.get(key))
        if numeric is not None:
            return numeric
    return None


def _summary_phase(row: Mapping[str, Any]) -> str:
    return _phase_key(_first_available(row, ("PhaseKey", "PointPhase", "流程阶段")))


def _point_identity_from_row(
    row: Mapping[str, Any],
) -> Tuple[str, Optional[float], Optional[float], Optional[float], Optional[float], Optional[float]]:
    phase = _summary_phase(row)
    chamber_temp = _first_valid_float(row, ("TempSet", "温箱目标温度C", "EnvTempC"))
    co2_ppm = _first_valid_float(row, ("ppm_CO2_Tank", "目标二氧化碳浓度ppm"))
    hgen_temp = _first_valid_float(row, ("HgenTempSet", "湿度发生器_目标温度(℃)", "湿度发生器目标温度C"))
    hgen_rh = _first_valid_float(row, ("HgenRhSet", "湿度发生器_目标湿度(%RH)", "湿度发生器目标湿度%"))
    pressure = _first_valid_float(row, ("PressureTarget", "目标压力hPa"))
    return (phase, chamber_temp, co2_ppm, hgen_temp, hgen_rh, pressure)


def _point_sort_key(row: Mapping[str, Any]) -> Tuple[float, int, float, float, float]:
    phase = _summary_phase(row)
    phase_order = 0 if phase == "co2" else 1
    chamber_temp = _first_valid_float(row, ("TempSet", "温箱目标温度C", "EnvTempC")) or 0.0
    co2_ppm = _first_valid_float(row, ("ppm_CO2_Tank", "目标二氧化碳浓度ppm")) or -1.0
    hgen_temp = _first_valid_float(row, ("HgenTempSet", "湿度发生器_目标温度(℃)", "湿度发生器目标温度C")) or -1.0
    pressure = _first_valid_float(row, ("PressureTarget", "目标压力hPa")) or -1.0
    return (chamber_temp, phase_order, co2_ppm if phase == "co2" else hgen_temp, pressure, 0.0)


def _load_csv_rows(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _latest_matching(run_dir: Path, pattern: str) -> Optional[Path]:
    matches = [path for path in run_dir.glob(pattern) if path.is_file()]
    if not matches:
        return None
    return max(matches, key=lambda item: item.stat().st_mtime)


def _resolve_summary_paths(run_dir: Path) -> List[Path]:
    gas = _latest_matching(run_dir, "分析仪汇总_气路_*.csv") or _latest_matching(run_dir, "分析仪汇总_气路_*.xlsx")
    water = _latest_matching(run_dir, "分析仪汇总_水路_*.csv") or _latest_matching(run_dir, "分析仪汇总_水路_*.xlsx")
    if gas and water:
        return [gas, water]
    combined = _latest_matching(run_dir, "分析仪汇总_*.csv") or _latest_matching(run_dir, "分析仪汇总_*.xlsx")
    if combined:
        return [combined]
    raise FileNotFoundError(f"未找到分析仪汇总文件: {run_dir}")


def _resolve_points_readable_path(run_dir: Path) -> Path:
    path = _latest_matching(run_dir, "points_readable_*.csv")
    if path is None:
        raise FileNotFoundError(f"未找到 points_readable_*.csv: {run_dir}")
    return path


def _resolve_temperature_observation_path(run_dir: Path) -> Optional[Path]:
    path = run_dir / "temperature_calibration_observations.csv"
    return path if path.exists() else None


def _write_csv(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    header: List[str] = []
    for row in rows:
        for key in row.keys():
            if key not in header:
                header.append(str(key))
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=header)
        writer.writeheader()
        for row in rows:
            writer.writerow(dict(row))


def _write_workbook(path: Path, sheets: Mapping[str, Sequence[Mapping[str, Any]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    try:
        for title, rows in sheets.items():
            ws = wb.create_sheet(title=title[:31])
            header: List[str] = []
            for row in rows:
                for key in row.keys():
                    if key not in header:
                        header.append(str(key))
            if header:
                ws.append(header)
                for row in rows:
                    ws.append([row.get(col) for col in header])
            else:
                ws.append(["说明"])
                ws.append(["无数据"])
        wb.save(path)
    finally:
        wb.close()


def _load_merge_inputs(run_dirs: Sequence[Path]) -> Dict[str, Any]:
    bundle: Dict[str, Any] = {
        "summary_rows_by_run": {},
        "point_rows_by_run": {},
        "temperature_rows_by_run": {},
    }
    for run_dir in run_dirs:
        summary_frame = load_summary_workbook_rows(_resolve_summary_paths(run_dir))
        bundle["summary_rows_by_run"][str(run_dir)] = summary_frame.to_dict(orient="records")
        bundle["point_rows_by_run"][str(run_dir)] = _load_csv_rows(_resolve_points_readable_path(run_dir))
        temp_path = _resolve_temperature_observation_path(run_dir)
        bundle["temperature_rows_by_run"][str(run_dir)] = _load_csv_rows(temp_path) if temp_path else []
    return bundle


def _merge_point_rows(
    run_dirs: Sequence[Path],
    point_rows_by_run: Mapping[str, Sequence[Mapping[str, Any]]],
    *,
    allowed_gas_ppm: Sequence[float],
) -> Tuple[
    List[Dict[str, Any]],
    Dict[Tuple[str, Optional[float], Optional[float], Optional[float], Optional[float], Optional[float]], Dict[str, Any]],
]:
    selected: Dict[Tuple[str, Optional[float], Optional[float], Optional[float], Optional[float], Optional[float]], Dict[str, Any]] = {}
    selected_source: Dict[Tuple[str, Optional[float], Optional[float], Optional[float], Optional[float], Optional[float]], Dict[str, Any]] = {}
    allowed_ppm_set = {float(item) for item in allowed_gas_ppm}
    for run_dir in run_dirs:
        run_key = str(run_dir)
        for row in point_rows_by_run.get(run_key, []):
            key = _point_identity_from_row(row)
            if key[0] == "co2":
                if key[2] is None or float(key[2]) not in allowed_ppm_set:
                    continue
            selected[key] = dict(row)
            selected_source[key] = {
                "source_run": run_key,
                "point_row": str(row.get("校准点行号") or row.get("PointRow") or "").strip(),
                "phase": key[0],
            }
    merged_rows = sorted(selected.values(), key=_point_sort_key)
    return merged_rows, selected_source


def _merge_summary_rows(
    selected_sources: Mapping[Tuple[str, Optional[float], Optional[float], Optional[float], Optional[float], Optional[float]], Mapping[str, Any]],
    summary_rows_by_run: Mapping[str, Sequence[Mapping[str, Any]]],
) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[str, Optional[float], Optional[float], Optional[float], Optional[float], Optional[float], str], Dict[str, Any]] = {}
    for key, selected_info in selected_sources.items():
        source_run = str(selected_info.get("source_run") or "")
        selected_point_row = str(selected_info.get("point_row") or "").strip()
        selected_phase = str(selected_info.get("phase") or "")
        for row in summary_rows_by_run.get(source_run, []):
            same_identity = _point_identity_from_row(row) == key
            same_point_row = selected_point_row and str(row.get("PointRow") or "").strip() == selected_point_row
            same_phase = _summary_phase(row) == selected_phase
            if not same_identity and not (same_point_row and same_phase):
                continue
            analyzer = str(row.get("Analyzer") or "").strip().upper()
            grouped[(key[0], key[1], key[2], key[3], key[4], key[5], analyzer)] = dict(row)
    rows = list(grouped.values())
    rows.sort(key=lambda row: (_point_sort_key(row), str(row.get("Analyzer") or "")))
    return rows


def _merge_temperature_rows(
    run_dirs: Sequence[Path],
    temperature_rows_by_run: Mapping[str, Sequence[Mapping[str, Any]]],
) -> List[Dict[str, Any]]:
    selected: Dict[Tuple[str, Optional[float]], Dict[str, Any]] = {}
    for run_dir in run_dirs:
        run_key = str(run_dir)
        for row in temperature_rows_by_run.get(run_key, []):
            analyzer = str(row.get("analyzer_id") or "").strip().upper()
            temp_set = _safe_float(row.get("temp_setpoint_c"))
            if analyzer and temp_set is not None:
                selected[(analyzer, temp_set)] = dict(row)
    rows = list(selected.values())
    rows.sort(key=lambda row: (str(row.get("analyzer_id") or ""), _safe_float(row.get("temp_setpoint_c")) or 0.0))
    return rows


def _count_summary_by_phase(rows: Sequence[Mapping[str, Any]]) -> Dict[str, int]:
    gas = sum(1 for row in rows if _summary_phase(row) == "co2")
    water = sum(1 for row in rows if _summary_phase(row) == "h2o")
    total = int(len(rows))
    return {"gas": gas, "water": water, "unknown": max(0, total - gas - water), "total": total}


def _split_summary_rows_by_phase(
    rows: Sequence[Mapping[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    gas_rows: List[Dict[str, Any]] = []
    water_rows: List[Dict[str, Any]] = []
    unknown_rows: List[Dict[str, Any]] = []
    for row in rows:
        payload = dict(row)
        phase = _summary_phase(payload)
        if phase == "co2":
            gas_rows.append(payload)
        elif phase == "h2o":
            water_rows.append(payload)
        else:
            unknown_rows.append(payload)
    return gas_rows, water_rows, unknown_rows


def _count_points_by_phase(rows: Sequence[Mapping[str, Any]]) -> Dict[str, int]:
    gas = sum(1 for row in rows if _summary_phase(row) == "co2")
    water = sum(1 for row in rows if _summary_phase(row) == "h2o")
    return {"gas": gas, "water": water, "total": gas + water}


def _build_merge_manifest(
    run_dirs: Sequence[Path],
    merged_points: Sequence[Mapping[str, Any]],
    merged_summary: Sequence[Mapping[str, Any]],
    merged_temperature_rows: Sequence[Mapping[str, Any]],
) -> Dict[str, Any]:
    return {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "run_dirs": [str(item) for item in run_dirs],
        "gas_ppm_filter": list(DEFAULT_GAS_PPM),
        "point_counts": _count_points_by_phase(merged_points),
        "summary_counts": _count_summary_by_phase(merged_summary),
        "temperature_observation_count": int(len(merged_temperature_rows)),
        "dedupe_rule": "完整点位键去重；后提供的 run_dir 覆盖先前 run_dir；气路和水路不互相覆盖",
        "verify_plan": {
            "co2_temp_c": VERIFY_CO2_TEMP_C,
            "co2_ppm": list(VERIFY_CO2_PPM),
            "h2o_targets": [
                {"temp_c": temp_c, "hgen_temp_c": hgen_temp, "hgen_rh_pct": rh}
                for temp_c, hgen_temp, rh in VERIFY_H2O_TARGETS
            ],
        },
    }


def _load_temperature_result_rows(result_csv: Path) -> List[Dict[str, Any]]:
    return _load_csv_rows(result_csv)


def _normalize_analyzer_name(value: Any) -> str:
    return str(value or "").strip().upper()


def _build_temperature_apply_plan(rows: Sequence[Mapping[str, Any]]) -> Dict[str, Dict[str, Dict[str, float]]]:
    plan: Dict[str, Dict[str, Dict[str, float]]] = {}
    for row in rows:
        analyzer_id = _normalize_analyzer_name(row.get("analyzer_id"))
        fit_type = str(row.get("fit_type") or "").strip().lower()
        if not analyzer_id or fit_type not in {"cell", "shell"}:
            continue
        if str(row.get("availability") or "").strip().lower() != "available":
            continue
        if not _safe_bool(row.get("fit_ok")):
            continue
        try:
            coeffs = {
                "A": float(row.get("A", 0.0)),
                "B": float(row.get("B", 0.0)),
                "C": float(row.get("C", 0.0)),
                "D": float(row.get("D", 0.0)),
            }
        except Exception:
            continue
        plan.setdefault(analyzer_id, {})[fit_type] = coeffs
    return plan


def _parse_senco_command(command: str) -> Optional[Tuple[str, int, List[float]]]:
    text = str(command or "").strip()
    if not text:
        return None
    parts = [part.strip() for part in text.split(",")]
    if len(parts) < 4:
        return None
    head = parts[0].upper()
    if not head.startswith("SENCO"):
        return None
    try:
        index = int(head.replace("SENCO", ""))
    except Exception:
        return None
    coeffs: List[float] = []
    for part in parts[3:]:
        try:
            coeffs.append(float(part))
        except Exception:
            return None
    return "FFF", index, coeffs


def _build_expected_gas_groups(download_plan_rows: Sequence[Mapping[str, Any]]) -> Dict[str, Dict[int, List[float]]]:
    expected: Dict[str, Dict[int, List[float]]] = defaultdict(dict)
    for row in download_plan_rows:
        analyzer = _normalize_analyzer_name(row.get("Analyzer"))
        for key in ("PrimaryCommand", "SecondaryCommand"):
            parsed = _parse_senco_command(str(row.get(key) or ""))
            if not analyzer or parsed is None:
                continue
            _, index, coeffs = parsed
            expected[analyzer][index] = coeffs
    return dict(expected)


def _read_coefficient_groups_for_targets(
    *,
    config_path: str | Path,
    groups: Sequence[int],
    output_dir: Path,
) -> List[Dict[str, Any]]:
    io_logger = CsvIoLogger(output_dir / "coefficient_readback_io.csv")
    rows: List[Dict[str, Any]] = []
    targets = load_download_targets(config_path)
    gas_class = _resolve_gas_analyzer_class()
    try:
        for target in targets:
            analyzer = gas_class(
                port=target.port,
                baudrate=target.baudrate,
                timeout=target.timeout,
                device_id=target.device_id,
                io_logger=io_logger,
            )
            try:
                analyzer.open()
                for group in groups:
                    error = ""
                    parsed: Dict[str, float] = {}
                    try:
                        parsed = analyzer.read_coefficient_group(int(group))
                    except Exception as exc:
                        error = str(exc)
                    rows.append(
                        {
                            "Analyzer": target.analyzer,
                            "Port": target.port,
                            "DeviceId": target.device_id,
                            "Group": int(group),
                            "Readback": json.dumps(parsed, ensure_ascii=False, sort_keys=True) if parsed else "",
                            "ReadbackError": error,
                        }
                    )
            finally:
                try:
                    analyzer.close()
                except Exception:
                    pass
    finally:
        io_logger.close()
    return rows


def _compare_readback(
    before_rows: Sequence[Mapping[str, Any]],
    write_rows: Sequence[Mapping[str, Any]],
    after_rows: Sequence[Mapping[str, Any]],
    expected_by_analyzer_group: Mapping[str, Mapping[int, Sequence[float]]],
) -> List[Dict[str, Any]]:
    before_map = {
        (_normalize_analyzer_name(row.get("Analyzer")), int(row.get("Group", 0))): dict(row)
        for row in before_rows
    }
    write_map = {
        (_normalize_analyzer_name(row.get("Analyzer")), int(row.get("Group", 0))): dict(row)
        for row in write_rows
    }
    after_map = {
        (_normalize_analyzer_name(row.get("Analyzer")), int(row.get("Group", 0))): dict(row)
        for row in after_rows
    }
    out: List[Dict[str, Any]] = []
    keys = sorted(set(before_map.keys()) | set(write_map.keys()) | set(after_map.keys()))
    for analyzer, group in keys:
        expected = list(expected_by_analyzer_group.get(analyzer, {}).get(group, []))
        before = before_map.get((analyzer, group), {})
        write = write_map.get((analyzer, group), {})
        after = after_map.get((analyzer, group), {})
        try:
            readback = json.loads(str(after.get("Readback") or "") or "{}")
        except Exception:
            readback = {}
        readback_values = [float(readback.get(f"C{i}")) for i in range(len(expected))] if expected else []
        match = bool(expected) and len(readback_values) == len(expected) and all(
            abs(readback_values[i] - float(expected[i])) <= 1e-9 for i in range(len(expected))
        )
        out.append(
            {
                "Analyzer": analyzer,
                "Group": group,
                "BeforeReadback": before.get("Readback", ""),
                "ExpectedWrite": json.dumps(expected, ensure_ascii=False) if expected else "",
                "WriteAttempted": write.get("Attempted", ""),
                "WriteOk": write.get("WriteOk", ""),
                "WriteError": write.get("Error", ""),
                "AfterReadback": after.get("Readback", ""),
                "ReadbackError": after.get("ReadbackError", ""),
                "ReadbackMatch": match,
            }
        )
    return out


def _write_temperature_compensation(
    *,
    config_path: str | Path,
    temperature_rows: Sequence[Mapping[str, Any]],
    output_dir: Path,
) -> Dict[str, Any]:
    plan = _build_temperature_apply_plan(temperature_rows)
    before = _read_coefficient_groups_for_targets(config_path=config_path, groups=(7, 8), output_dir=output_dir / "before")
    io_logger = CsvIoLogger(output_dir / "temperature_write_io.csv")
    results: List[Dict[str, Any]] = []
    targets = load_download_targets(config_path)
    gas_class = _resolve_gas_analyzer_class()
    try:
        for target in targets:
            fit_map = plan.get(target.analyzer, {})
            analyzer = gas_class(
                port=target.port,
                baudrate=target.baudrate,
                timeout=target.timeout,
                device_id=target.device_id,
                io_logger=io_logger,
            )
            try:
                analyzer.open()
                pending_groups = [
                    (fit_type, group)
                    for fit_type, group in (("cell", 7), ("shell", 8))
                    if fit_map.get(fit_type)
                ]
                mode_switch_attempted = False
                entered_mode2 = False
                mode_enter_error = ""
                if pending_groups:
                    mode_switch_attempted = True
                    try:
                        entered_mode2 = bool(analyzer.set_mode(2))
                    except Exception as exc:
                        mode_enter_error = str(exc)
                    else:
                        if not entered_mode2:
                            mode_enter_error = "MODE_2_ACK_FAILED"
                for fit_type, group in (("cell", 7), ("shell", 8)):
                    coeffs = fit_map.get(fit_type)
                    row: Dict[str, Any] = {
                        "Analyzer": target.analyzer,
                        "Port": target.port,
                        "DeviceId": target.device_id,
                        "FitType": fit_type,
                        "Group": group,
                        "Attempted": bool(coeffs),
                        "WriteOk": False,
                        "Error": "",
                    }
                    if coeffs:
                        if mode_enter_error:
                            row["Error"] = mode_enter_error
                        else:
                            try:
                                acked = analyzer.set_senco(group, coeffs["A"], coeffs["B"], coeffs["C"], coeffs["D"])
                                row["WriteOk"] = bool(acked)
                                row["CommandString"] = (
                                    f"SENCO{group},YGAS,FFF,"
                                    + ",".join(
                                        format_senco_values((coeffs["A"], coeffs["B"], coeffs["C"], coeffs["D"]))
                                    )
                                )
                            except Exception as exc:
                                row["Error"] = str(exc)
                    results.append(row)
                if mode_switch_attempted:
                    try:
                        mode_exit_ok = bool(analyzer.set_mode(1))
                    except Exception as exc:
                        mode_exit_error = str(exc)
                    else:
                        mode_exit_error = "" if mode_exit_ok else "MODE_1_ACK_FAILED"
                    if mode_exit_error:
                        for row in results:
                            if row.get("Analyzer") == target.analyzer and row.get("Attempted") and not row.get("Error"):
                                row["Error"] = mode_exit_error
            finally:
                try:
                    analyzer.close()
                except Exception:
                    pass
    finally:
        io_logger.close()
    after = _read_coefficient_groups_for_targets(config_path=config_path, groups=(7, 8), output_dir=output_dir / "after")
    return {
        "before_rows": before,
        "write_rows": results,
        "after_rows": after,
    }


def _build_verify_points_workbook(source_points: Path, output_path: Path) -> Dict[str, Any]:
    points = load_points_from_excel(source_points)
    selected_indexes: List[int] = []
    for point in points:
        if point.co2_ppm is not None:
            if float(point.temp_chamber_c) == VERIFY_CO2_TEMP_C and float(point.co2_ppm) in set(VERIFY_CO2_PPM):
                selected_indexes.append(int(point.index))
        elif point.hgen_temp_c is not None and point.hgen_rh_pct is not None:
            identity = (float(point.temp_chamber_c), float(point.hgen_temp_c), float(point.hgen_rh_pct))
            if identity in VERIFY_H2O_TARGETS:
                selected_indexes.append(int(point.index))
    selected_indexes = sorted(set(selected_indexes))
    source_wb = load_workbook(source_points)
    target_wb = Workbook()
    try:
        source_ws = source_wb.active
        target_ws = target_wb.active
        target_ws.title = source_ws.title
        next_row = 1
        for row_idx in [1, 2] + selected_indexes:
            for col_idx in range(1, source_ws.max_column + 1):
                target_ws.cell(row=next_row, column=col_idx, value=source_ws.cell(row=row_idx, column=col_idx).value)
            next_row += 1
        output_path.parent.mkdir(parents=True, exist_ok=True)
        target_wb.save(output_path)
    finally:
        source_wb.close()
        target_wb.close()
    return {
        "path": output_path,
        "selected_excel_rows": selected_indexes,
        "point_count": len(selected_indexes),
    }


def _build_verify_point_rows_from_workbook(points_path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    points = load_points_from_excel(points_path)
    for point in points:
        rows.append(
            {
                "流程阶段": "h2o" if point.is_h2o_point else "co2",
                "温箱目标温度C": point.temp_chamber_c,
                "目标二氧化碳浓度ppm": point.co2_ppm,
                "湿度发生器目标温度C": point.hgen_temp_c,
                "湿度发生器目标湿度%": point.hgen_rh_pct,
                "目标压力hPa": point.target_pressure_hpa,
            }
        )
    return rows


def _build_verify_subset_summary(
    merged_summary_rows: Sequence[Mapping[str, Any]],
    verify_point_rows: Sequence[Mapping[str, Any]],
) -> List[Dict[str, Any]]:
    selected_keys = {_point_identity_from_row(row) for row in verify_point_rows}
    rows = [dict(row) for row in merged_summary_rows if _point_identity_from_row(row) in selected_keys]
    rows.sort(key=lambda row: (_point_sort_key(row), str(row.get("Analyzer") or "")))
    return rows


def _rmse(values: Sequence[float]) -> Optional[float]:
    if not values:
        return None
    return math.sqrt(sum(v * v for v in values) / float(len(values)))


def _mean(values: Sequence[float]) -> Optional[float]:
    if not values:
        return None
    return sum(values) / float(len(values))


def _build_verify_effect_rows(
    *,
    before_rows: Sequence[Mapping[str, Any]],
    after_rows: Sequence[Mapping[str, Any]],
) -> List[Dict[str, Any]]:
    def _collect(rows: Sequence[Mapping[str, Any]], gas: str) -> Dict[str, Dict[str, Optional[float]]]:
        grouped: Dict[str, List[float]] = defaultdict(list)
        for row in rows:
            if _phase_key(row.get("PointPhase")) != gas:
                continue
            analyzer = _normalize_analyzer_name(row.get("Analyzer"))
            if gas == "co2":
                target = _safe_float(row.get("ppm_CO2_Tank"))
                measured = _safe_float(row.get("ppm_CO2"))
            else:
                target = _safe_float(row.get("ppm_H2O_Dew"))
                measured = _safe_float(row.get("ppm_H2O"))
            if analyzer and target is not None and measured is not None:
                grouped[analyzer].append(measured - target)
        out: Dict[str, Dict[str, Optional[float]]] = {}
        for analyzer in sorted(grouped.keys()):
            errs = grouped.get(analyzer, [])
            out[analyzer] = {
                "rmse": _rmse(errs),
                "bias": _mean(errs),
            }
        return out

    rows: List[Dict[str, Any]] = []
    for gas in ("co2", "h2o"):
        before = _collect(before_rows, gas)
        after = _collect(after_rows, gas)
        analyzers = sorted(set(before.keys()) | set(after.keys()))
        for analyzer in analyzers:
            before_rmse = before.get(analyzer, {}).get("rmse")
            after_rmse = after.get(analyzer, {}).get("rmse")
            before_bias = before.get(analyzer, {}).get("bias")
            after_bias = after.get(analyzer, {}).get("bias")
            verdict = "缺数据"
            if before_rmse is not None and after_rmse is not None:
                if after_rmse < before_rmse and (
                    before_bias is None
                    or after_bias is None
                    or abs(after_bias) <= abs(before_bias)
                ):
                    verdict = "改善"
                elif after_rmse <= before_rmse * 1.05:
                    verdict = "基本持平"
                else:
                    verdict = "变差"
            rows.append(
                {
                    "分析仪": analyzer,
                    "气体": gas.upper(),
                    "写前RMSE": before_rmse,
                    "写前Bias": before_bias,
                    "写后RMSE": after_rmse,
                    "写后Bias": after_bias,
                    "判定": verdict,
                }
            )
    return rows


def _build_temperature_effect_rows(verify_temperature_rows: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[str, str], List[float]] = defaultdict(list)
    for row in verify_temperature_rows:
        analyzer = _normalize_analyzer_name(row.get("analyzer_id"))
        ref = _safe_float(row.get("ref_temp_c"))
        cell = _safe_float(row.get("analyzer_cell_temp_raw_c") or row.get("cell_temp_raw_c"))
        shell = _safe_float(row.get("analyzer_shell_temp_raw_c") or row.get("shell_temp_raw_c"))
        if analyzer and ref is not None and cell is not None:
            grouped[(analyzer, "cell")].append(cell - ref)
        if analyzer and ref is not None and shell is not None:
            grouped[(analyzer, "shell")].append(shell - ref)
    rows: List[Dict[str, Any]] = []
    for (analyzer, fit_type), errors in sorted(grouped.items()):
        rmse = _rmse(errors)
        bias = _mean(errors)
        max_abs = max(abs(err) for err in errors) if errors else None
        verdict = "失败"
        if max_abs is not None:
            if max_abs <= 0.1:
                verdict = "通过"
            elif max_abs <= 0.2:
                verdict = "关注"
        rows.append(
            {
                "分析仪": analyzer,
                "温补通道": "cell(SENCO7)" if fit_type == "cell" else "shell(SENCO8)",
                "RMSE(°C)": rmse,
                "Bias(°C)": bias,
                "最大绝对误差(°C)": max_abs,
                "判定": verdict,
            }
        )
    return rows


def _build_final_conclusion_rows(
    verify_effect_rows: Sequence[Mapping[str, Any]],
    temperature_effect_rows: Sequence[Mapping[str, Any]],
) -> List[Dict[str, Any]]:
    gas_map: Dict[str, List[str]] = defaultdict(list)
    for row in verify_effect_rows:
        gas_map[_normalize_analyzer_name(row.get("分析仪"))].append(str(row.get("判定") or ""))
    temp_map: Dict[str, List[str]] = defaultdict(list)
    for row in temperature_effect_rows:
        temp_map[_normalize_analyzer_name(row.get("分析仪"))].append(str(row.get("判定") or ""))
    analyzers = sorted(set(gas_map.keys()) | set(temp_map.keys()))
    rows: List[Dict[str, Any]] = []
    for analyzer in analyzers:
        gas_verdicts = gas_map.get(analyzer, [])
        temp_verdicts = temp_map.get(analyzer, [])
        overall = "需人工复核"
        if gas_verdicts and all(item in {"改善", "基本持平"} for item in gas_verdicts) and temp_verdicts and all(
            item in {"通过", "关注"} for item in temp_verdicts
        ):
            overall = "建议保留当前系数"
        elif any(item == "变差" for item in gas_verdicts) or any(item == "失败" for item in temp_verdicts):
            overall = "建议复核或回退"
        rows.append(
            {
                "分析仪": analyzer,
                "气体验证判定": " / ".join(gas_verdicts) if gas_verdicts else "未执行",
                "温度验证判定": " / ".join(temp_verdicts) if temp_verdicts else "未执行",
                "最终结论": overall,
            }
        )
    return rows


def _build_overview_rows(
    manifest: Mapping[str, Any],
    *,
    gas_report_path: Optional[Path],
    temperature_dir: Path,
    verify_dir: Optional[Path],
    temperature_write_compare: Sequence[Mapping[str, Any]],
    gas_write_compare: Sequence[Mapping[str, Any]],
) -> List[Dict[str, Any]]:
    return [
        {"项目": "生成时间", "值": manifest.get("created_at", "")},
        {"项目": "合并运行目录", "值": " | ".join(manifest.get("run_dirs", []))},
        {"项目": "气路点数", "值": manifest.get("point_counts", {}).get("gas", "")},
        {"项目": "水路点数", "值": manifest.get("point_counts", {}).get("water", "")},
        {"项目": "总点数", "值": manifest.get("point_counts", {}).get("total", "")},
        {"项目": "气路汇总行数", "值": manifest.get("summary_counts", {}).get("gas", "")},
        {"项目": "水路汇总行数", "值": manifest.get("summary_counts", {}).get("water", "")},
        {"项目": "总汇总行数", "值": manifest.get("summary_counts", {}).get("total", "")},
        {"项目": "气体系数报表", "值": str(gas_report_path) if gas_report_path else "未生成"},
        {"项目": "温度补偿目录", "值": str(temperature_dir)},
        {"项目": "验证目录", "值": str(verify_dir) if verify_dir else "未执行"},
        {"项目": "温补写后读回匹配数", "值": sum(1 for row in temperature_write_compare if row.get("ReadbackMatch"))},
        {"项目": "气体写后读回匹配数", "值": sum(1 for row in gas_write_compare if row.get("ReadbackMatch"))},
    ]


def _format_coefficient_sheet(ws, coefficient_columns: Sequence[str]) -> None:
    header = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
    for col_name in coefficient_columns:
        if col_name not in header:
            continue
        idx = header.index(col_name) + 1
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=idx).number_format = COEFF_DISPLAY_FORMAT


def _write_chinese_summary_workbook(
    *,
    output_path: Path,
    overview_rows: Sequence[Mapping[str, Any]],
    merge_manifest: Mapping[str, Any],
    merged_point_rows: Sequence[Mapping[str, Any]],
    merged_summary_rows: Sequence[Mapping[str, Any]],
    temperature_coeff_rows: Sequence[Mapping[str, Any]],
    temperature_write_compare: Sequence[Mapping[str, Any]],
    gas_fit_summary_rows: Sequence[Mapping[str, Any]],
    gas_coefficient_rows: Sequence[Mapping[str, Any]],
    gas_download_rows: Sequence[Mapping[str, Any]],
    gas_write_compare: Sequence[Mapping[str, Any]],
    verify_effect_rows: Sequence[Mapping[str, Any]],
    temperature_effect_rows: Sequence[Mapping[str, Any]],
    final_rows: Sequence[Mapping[str, Any]],
) -> None:
    sheets: Dict[str, Sequence[Mapping[str, Any]]] = {
        "总览": overview_rows,
        "合并核对": [
            {"项目": "气路点位", "数量": merge_manifest.get("point_counts", {}).get("gas", "")},
            {"项目": "水路点位", "数量": merge_manifest.get("point_counts", {}).get("water", "")},
            {"项目": "总点位", "数量": merge_manifest.get("point_counts", {}).get("total", "")},
            {"项目": "气路汇总行", "数量": merge_manifest.get("summary_counts", {}).get("gas", "")},
            {"项目": "水路汇总行", "数量": merge_manifest.get("summary_counts", {}).get("water", "")},
            {"项目": "总汇总行", "数量": merge_manifest.get("summary_counts", {}).get("total", "")},
        ],
        "温度校准系数": temperature_coeff_rows,
        "温度写后读回": temperature_write_compare,
        "气体系数拟合汇总": gas_fit_summary_rows,
        "气体分析仪校准系数": gas_coefficient_rows,
        "气体下发计划": gas_download_rows,
        "气体写后读回": gas_write_compare,
        "30C与两组水路验证": verify_effect_rows,
        "温度校准验证": temperature_effect_rows,
        "按分析仪最终结论": final_rows,
        "合并点位明细": merged_point_rows,
        "合并分析仪汇总": merged_summary_rows,
    }
    _write_workbook(output_path, sheets)
    wb = load_workbook(output_path)
    try:
        if "温度校准系数" in wb.sheetnames:
            _format_coefficient_sheet(wb["温度校准系数"], ("A", "B", "C", "D"))
        if "气体分析仪校准系数" in wb.sheetnames:
            _format_coefficient_sheet(
                wb["气体分析仪校准系数"],
                ("Constant", "R", "R2", "R3", "T_k", "T_k2", "R*T_k", "P", "R*T_k*P"),
            )
        wb.save(output_path)
    finally:
        wb.close()


def _parse_args(argv: Optional[Iterable[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Standalone engineering sidecar for merging completed V1 runs and optional post-process steps."
    )
    parser.add_argument("--config", default="configs/default_config.json", help="Config path.")
    parser.add_argument(
        "--run-dir",
        dest="run_dirs",
        action="append",
        required=True,
        help="Completed V1 run directory. Later entries override earlier duplicates.",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Output directory. Defaults to logs/merged_sidecar_YYYYMMDD_HHMMSS.",
    )
    parser.add_argument("--write-temperature", action="store_true", help="Write SENCO7/SENCO8 after fitting.")
    parser.add_argument("--write-gas", action="store_true", help="Write SENCO1~4 after fitting.")
    parser.add_argument("--run-verify", action="store_true", help="Run the 30C + two-water-point verification subset.")
    parser.add_argument("--verify-run-id", default=None, help="Optional verification run id.")
    return parser.parse_args(list(argv) if argv is not None else None)


def main(argv: Optional[Iterable[str]] = None) -> int:
    args = _parse_args(argv)
    cfg = load_config(args.config)
    run_dirs = [Path(item).resolve() for item in args.run_dirs]
    output_dir = (
        Path(args.output_dir).resolve()
        if args.output_dir
        else Path(cfg.get("paths", {}).get("output_dir") or ".").resolve()
        / f"merged_sidecar_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    _log(f"旁路总控输出目录: {output_dir}")
    _log(f"合并顺序（后者覆盖前者）: {', '.join(str(item) for item in run_dirs)}")

    merge_inputs = _load_merge_inputs(run_dirs)
    merged_point_rows, selected_sources = _merge_point_rows(
        run_dirs,
        merge_inputs["point_rows_by_run"],
        allowed_gas_ppm=DEFAULT_GAS_PPM,
    )
    merged_summary_rows = _merge_summary_rows(selected_sources, merge_inputs["summary_rows_by_run"])
    merged_temperature_rows = _merge_temperature_rows(run_dirs, merge_inputs["temperature_rows_by_run"])

    manifest = _build_merge_manifest(run_dirs, merged_point_rows, merged_summary_rows, merged_temperature_rows)
    gas_summary_rows, water_summary_rows, unknown_summary_rows = _split_summary_rows_by_phase(merged_summary_rows)
    (output_dir / "merge_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    _write_csv(output_dir / "merged_point_selection.csv", merged_point_rows)
    _write_csv(output_dir / "merged_analyzer_summary.csv", merged_summary_rows)
    _write_csv(output_dir / "merged_analyzer_summary_gas.csv", gas_summary_rows)
    _write_csv(output_dir / "merged_analyzer_summary_h2o.csv", water_summary_rows)
    if unknown_summary_rows:
        _write_csv(output_dir / "merged_analyzer_summary_unknown_phase.csv", unknown_summary_rows)
    _write_csv(output_dir / "merged_temperature_calibration_observations.csv", merged_temperature_rows)

    _log(
        "合并核对: "
        f"点位 气={manifest['point_counts']['gas']} 水={manifest['point_counts']['water']} 总={manifest['point_counts']['total']} | "
        f"汇总 气={manifest['summary_counts']['gas']} 水={manifest['summary_counts']['water']} 总={manifest['summary_counts']['total']}"
    )
    if manifest["summary_counts"].get("unknown", 0):
        _log(f"汇总相位警告: unknown={manifest['summary_counts']['unknown']}")

    temperature_dir = output_dir / "temperature_compensation"
    temperature_dir.mkdir(parents=True, exist_ok=True)
    temperature_bundle = export_temperature_compensation_artifacts(
        temperature_dir,
        merged_temperature_rows,
        polynomial_order=3,
        export_commands=True,
    )
    temperature_coeff_rows = temperature_bundle["results"]
    temperature_write_compare: List[Dict[str, Any]] = []
    if args.write_temperature:
        _log("开始写入温度校准系数（SENCO7/8）...")
        temp_write = _write_temperature_compensation(
            config_path=args.config,
            temperature_rows=temperature_coeff_rows,
            output_dir=output_dir / "temperature_write_readback",
        )
        temperature_expected = {
            analyzer: {
                7 if fit_type == "cell" else 8: [coeffs["A"], coeffs["B"], coeffs["C"], coeffs["D"]]
                for fit_type, coeffs in fit_map.items()
            }
            for analyzer, fit_map in _build_temperature_apply_plan(temperature_coeff_rows).items()
        }
        temperature_write_compare = _compare_readback(
            temp_write["before_rows"],
            temp_write["write_rows"],
            temp_write["after_rows"],
            temperature_expected,
        )
        _write_csv(output_dir / "temperature_write_readback" / "temperature_write_match_check.csv", temperature_write_compare)

    coeff_cfg = _load_coefficients_config_lazy(args.config)
    merged_summary_frame = pd.DataFrame(merged_summary_rows)
    gas_report_path = export_ratio_poly_report_from_summary_frame(
        merged_summary_frame,
        out_dir=output_dir / "gas_coefficients",
        coeff_cfg=coeff_cfg,
    )
    gas_fit_summary_rows: List[Dict[str, Any]] = []
    gas_coefficient_rows: List[Dict[str, Any]] = []
    gas_download_rows: List[Dict[str, Any]] = []
    gas_write_compare: List[Dict[str, Any]] = []
    if gas_report_path is not None:
        gas_book = pd.ExcelFile(gas_report_path)
        gas_fit_summary_rows = (
            pd.read_excel(gas_report_path, sheet_name="汇总").to_dict(orient="records")
            if "汇总" in gas_book.sheet_names
            else []
        )
        gas_coefficient_rows = (
            pd.read_excel(gas_report_path, sheet_name="简化系数").to_dict(orient="records")
            if "简化系数" in gas_book.sheet_names
            else list(gas_fit_summary_rows)
        )
        gas_download_rows = load_download_plan(gas_report_path)
        _write_csv(output_dir / "gas_coefficients" / "download_plan.csv", gas_download_rows)
        if args.write_gas:
            _log("开始写入气体分析仪校准系数（SENCO1~4）...")
            before_rows = _read_coefficient_groups_for_targets(
                config_path=args.config,
                groups=(1, 2, 3, 4),
                output_dir=output_dir / "post_gas_write_readback" / "before",
            )
            download_coefficients_to_analyzers(
                report_path=gas_report_path,
                config_path=args.config,
                output_dir=output_dir / "post_gas_write_readback" / "download",
            )
            after_rows = _read_coefficient_groups_for_targets(
                config_path=args.config,
                groups=(1, 2, 3, 4),
                output_dir=output_dir / "post_gas_write_readback" / "after",
            )
            expected = _build_expected_gas_groups(gas_download_rows)
            synthetic_write_rows = []
            for analyzer, groups in expected.items():
                for group in groups.keys():
                    synthetic_write_rows.append(
                        {
                            "Analyzer": analyzer,
                            "Group": group,
                            "Attempted": True,
                            "WriteOk": True,
                            "Error": "",
                        }
                    )
            gas_write_compare = _compare_readback(before_rows, synthetic_write_rows, after_rows, expected)
            _write_csv(output_dir / "post_gas_write_readback" / "gas_write_match_check.csv", gas_write_compare)

    verify_effect_rows: List[Dict[str, Any]] = []
    temperature_effect_rows: List[Dict[str, Any]] = []
    verify_dir: Optional[Path] = None
    if args.run_verify:
        verify_root = output_dir / "verify30"
        verify_root.mkdir(parents=True, exist_ok=True)
        verify_points = _build_verify_points_workbook(
            Path(cfg.get("paths", {}).get("points_excel") or "points.xlsx"),
            output_dir / "verify_points_30C_两组水路.xlsx",
        )
        _log(
            "开始执行验证子集: "
            f"CO2@30℃ -> {list(VERIFY_CO2_PPM)} ppm; "
            f"H2O -> {VERIFY_H2O_TARGETS}. "
            f"点数={verify_points['point_count']}"
        )
        verify_args = [
            "--config",
            str(Path(args.config).resolve()),
            "--temp",
            str(VERIFY_CO2_TEMP_C),
            "--points-excel",
            str(verify_points["path"]),
            "--output-dir",
            str(verify_root),
        ]
        if args.verify_run_id:
            verify_args.extend(["--run-id", args.verify_run_id])
        rc = verify_short_run.main(verify_args)
        if rc != 0:
            _log(f"验证运行返回非零状态: {rc}")
        if args.verify_run_id and (verify_root / args.verify_run_id).is_dir():
            verify_dir = verify_root / args.verify_run_id
        else:
            candidates = [path for path in verify_root.glob("verify_short_*") if path.is_dir()]
            if candidates:
                verify_dir = max(candidates, key=lambda path: path.stat().st_mtime)
        if verify_dir is not None:
            verify_summary_rows = load_summary_workbook_rows(_resolve_summary_paths(verify_dir)).to_dict(orient="records")
            before_subset = _build_verify_subset_summary(
                merged_summary_rows,
                _build_verify_point_rows_from_workbook(Path(verify_points["path"])),
            )
            verify_effect_rows = _build_verify_effect_rows(before_rows=before_subset, after_rows=verify_summary_rows)
            verify_temp_path = verify_dir / "temperature_calibration_observations.csv"
            if verify_temp_path.exists():
                temperature_effect_rows = _build_temperature_effect_rows(_load_csv_rows(verify_temp_path))

    final_rows = _build_final_conclusion_rows(verify_effect_rows, temperature_effect_rows)
    overview_rows = _build_overview_rows(
        manifest,
        gas_report_path=gas_report_path,
        temperature_dir=temperature_dir,
        verify_dir=verify_dir,
        temperature_write_compare=temperature_write_compare,
        gas_write_compare=gas_write_compare,
    )
    summary_path = output_dir / f"校准汇总与验证结论_{datetime.now().strftime('%Y%m%d')}.xlsx"
    _write_chinese_summary_workbook(
        output_path=summary_path,
        overview_rows=overview_rows,
        merge_manifest=manifest,
        merged_point_rows=merged_point_rows,
        merged_summary_rows=merged_summary_rows,
        temperature_coeff_rows=temperature_coeff_rows,
        temperature_write_compare=temperature_write_compare,
        gas_fit_summary_rows=gas_fit_summary_rows,
        gas_coefficient_rows=gas_coefficient_rows,
        gas_download_rows=gas_download_rows,
        gas_write_compare=gas_write_compare,
        verify_effect_rows=verify_effect_rows,
        temperature_effect_rows=temperature_effect_rows,
        final_rows=final_rows,
    )
    _log(f"旁路总控完成，中文总结文件: {summary_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
