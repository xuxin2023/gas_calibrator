"""V1 corrected-delivery/report helper kept on the V1 runtime side.

This entrypoint is reused by the V1 runner and the guarded online acceptance
tool, so it should not depend on V2 runtime or offline bridge modules.
"""

from __future__ import annotations

import csv
import inspect
import json
import math
import re
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook

from ..config import (
    V1_CO2_ONLY_H2O_NOT_SUPPORTED_MESSAGE,
    load_config,
    require_v1_h2o_zero_span_supported,
    v1_h2o_zero_span_capability,
)
from ..coefficients.write_readiness import (
    build_write_readiness_decision,
    summarize_device_write_verify,
    summarize_fit_quality,
    summarize_runtime_parity,
)
from ..devices.gas_analyzer import GasAnalyzer
from ..export import build_corrected_water_points_report
from ..h2o_summary_selection import normalize_h2o_summary_selection
from ..senco_format import format_senco_values, senco_readback_matches
from ._no500_filter import filter_no_500_frame

_PRESSURE_WRITE_MIN_GAUGE_CONTROLLER_OVERLAP = 5
_PRESSURE_WRITE_MAX_GAUGE_CONTROLLER_MEAN_ABS_HPA = 3.0
_PRESSURE_WRITE_MAX_GAUGE_CONTROLLER_MAX_ABS_HPA = 8.0
_STARTUP_PRESSURE_WRITE_MIN_SAMPLES = 3
_STARTUP_PRESSURE_WRITE_MAX_REFERENCE_SPAN_HPA = 2.0
_READBACK_SOURCE_EXPLICIT_C0 = "parsed_from_explicit_c0_line"
_READBACK_SOURCE_AMBIGUOUS = "parsed_from_ambiguous_line"
_READBACK_SOURCE_NONE = "no_valid_coefficient_line"


class _TranscriptIoLogger:
    def __init__(self) -> None:
        self.rows: list[dict[str, Any]] = []

    def log_io(
        self,
        *,
        port: str,
        device: str,
        direction: str,
        command: Any = None,
        response: Any = None,
        error: Any = None,
        duration_ms: Any = None,
        **_kwargs: Any,
    ) -> None:
        self.rows.append(
            {
                "ts": datetime.now().isoformat(timespec="milliseconds"),
                "port": str(port or ""),
                "device": str(device or ""),
                "direction": str(direction or ""),
                "duration_ms": "" if duration_ms in (None, "") else str(duration_ms),
                "command": "" if command is None else str(command),
                "response": "" if response is None else str(response),
                "error": "" if error is None else str(error),
            }
        )


def _safe_float(value: Any) -> Optional[float]:
    if value in (None, "", "null", "None"):
        return None
    try:
        numeric = float(value)
    except Exception:
        return None
    if not math.isfinite(numeric):
        return None
    return numeric


def _safe_bool(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on", "ok"}


def _normalize_device_id(value: Any) -> str:
    text = str(value or "").strip().upper()
    if not text:
        return ""
    if text.isdigit():
        return f"{int(text):03d}"
    return text


def _normalize_analyzer(value: Any, *, index: Optional[int] = None) -> str:
    text = str(value or "").strip().upper()
    if text:
        return text
    if index is not None:
        return f"GA{int(index):02d}"
    return ""


def _latest_matching(run_dir: Path, pattern: str) -> Optional[Path]:
    matches = [path for path in run_dir.glob(pattern) if path.is_file()]
    if not matches:
        return None
    return max(matches, key=lambda item: item.stat().st_mtime)


def _resolve_summary_paths(run_dir: Path) -> List[Path]:
    gas = _latest_matching(run_dir, "分析仪汇总_气路_*.xlsx") or _latest_matching(run_dir, "分析仪汇总_气路_*.csv")
    water = _latest_matching(run_dir, "分析仪汇总_水路_*.xlsx") or _latest_matching(run_dir, "分析仪汇总_水路_*.csv")
    if gas and water:
        return [gas, water]
    raise FileNotFoundError(f"未找到气路/水路汇总文件: {run_dir}")


def _resolve_samples_path(run_dir: Path) -> Path:
    path = _latest_matching(run_dir, "samples_*.csv") or (run_dir / "samples.csv")
    if path is None or not path.exists():
        raise FileNotFoundError(f"未找到 samples csv: {run_dir}")
    return path


def _write_csv(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    header: List[str] = []
    for row in rows:
        for key in row.keys():
            if key not in header:
                header.append(str(key))
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=header)
        writer.writeheader()
        for row in rows:
            writer.writerow(dict(row))


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")


def _truth_source_label(source: Any) -> str:
    text = str(source or "").strip()
    if text == _READBACK_SOURCE_EXPLICIT_C0:
        return "explicit_c0"
    if text == _READBACK_SOURCE_AMBIGUOUS:
        return "ambiguous"
    return "none"


def _resolve_io_logger(ga: Any) -> Any:
    ser = getattr(ga, "ser", None)
    logger = getattr(ser, "io_logger", None)
    if logger is not None:
        return logger
    return getattr(ga, "io_logger", None)


def _io_row_count(ga: Any) -> int:
    logger = _resolve_io_logger(ga)
    rows = getattr(logger, "rows", None)
    if isinstance(rows, list):
        return len(rows)
    return 0


def _capture_io_action(
    ga: Any,
    start_index: int,
    *,
    phase: str,
    action: str,
    ok: bool,
    error: str = "",
    metadata: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    logger = _resolve_io_logger(ga)
    rows = getattr(logger, "rows", None)
    sliced_rows = [dict(row) for row in list(rows or [])[max(0, int(start_index)) :]]
    payload: Dict[str, Any] = {
        "kind": "io_action",
        "phase": str(phase or ""),
        "action": str(action or ""),
        "ok": bool(ok),
        "error": str(error or ""),
        "rows": sliced_rows,
    }
    if metadata:
        payload.update(dict(metadata))
    return payload


def _empty_action_capture(action: str = "") -> Dict[str, Any]:
    return {
        "kind": "io_action",
        "phase": "",
        "action": str(action or ""),
        "ok": False,
        "error": "",
        "rows": [],
    }


def _capture_getco_event(
    *,
    phase: str,
    action: str,
    group: int,
    capture: Mapping[str, Any] | None,
) -> Dict[str, Any]:
    payload = _empty_readback_capture()
    payload.update(dict(capture or {}))
    source = str(payload.get("source") or _READBACK_SOURCE_NONE)
    error_text = str(payload.get("error") or "")
    if source == _READBACK_SOURCE_AMBIGUOUS and not error_text:
        error_text = f"READBACK_SOURCE_UNTRUSTED:{source}"
    if source == _READBACK_SOURCE_NONE and not error_text:
        error_text = "NO_VALID_COEFFICIENT_LINE"
    return {
        "kind": "getco",
        "phase": str(phase or ""),
        "action": str(action or ""),
        "group": int(group),
        "ok": source == _READBACK_SOURCE_EXPLICIT_C0,
        "error": "" if source == _READBACK_SOURCE_EXPLICIT_C0 else error_text,
        "truth_source": _truth_source_label(source),
        "source": source,
        "command": str(payload.get("command") or ""),
        "target_id": str(payload.get("target_id") or ""),
        "source_line": str(payload.get("source_line") or ""),
        "source_line_has_explicit_c0": bool(payload.get("source_line_has_explicit_c0", False)),
        "raw_transcript_lines": list(payload.get("raw_transcript_lines") or []),
        "attempt_transcripts": list(payload.get("attempt_transcripts") or []),
    }


def _write_writeback_raw_transcript_log(path: Path, sessions: Sequence[Mapping[str, Any]]) -> None:
    lines: list[str] = []
    for index, session in enumerate(sessions, start=1):
        lines.append(
            "# session "
            + " ".join(
                [
                    f"index={int(index)}",
                    f"analyzer={session.get('Analyzer') or ''}",
                    f"port={session.get('Port') or ''}",
                    f"target={session.get('TargetDeviceId') or ''}",
                    f"live={session.get('LiveDeviceId') or ''}",
                    f"status={session.get('Status') or ''}",
                ]
            ).strip()
        )
        fatal_error = str(session.get("FatalError") or "").strip()
        if fatal_error:
            lines.append(f"fatal_error: {fatal_error}")
        for event in list(session.get("Events") or []):
            lines.append(
                "## event "
                + " ".join(
                    [
                        f"phase={event.get('phase') or ''}",
                        f"action={event.get('action') or ''}",
                        f"ok={bool(event.get('ok', False))}",
                        f"group={event.get('group') if event.get('group') not in (None, '') else ''}",
                        f"truth_source={event.get('truth_source') or ''}",
                    ]
                ).strip()
            )
            error_text = str(event.get("error") or "").strip()
            if error_text:
                lines.append(f"error: {error_text}")
            for key in ("command", "target_id", "source_line", "requested_mode", "value", "coefficients"):
                if key not in event:
                    continue
                value = event.get(key)
                if value in (None, "", [], {}):
                    continue
                lines.append(f"{key}: {json.dumps(value, ensure_ascii=False, default=str)}")
            if event.get("source_line_has_explicit_c0") is not None:
                lines.append(f"source_line_has_explicit_c0: {bool(event.get('source_line_has_explicit_c0', False))}")
            for row in list(event.get("rows") or []):
                lines.append("io_row: " + json.dumps(row, ensure_ascii=False, default=str))
            for raw_line in list(event.get("raw_transcript_lines") or []):
                lines.append(f"raw_line: {raw_line}")
            attempt_transcripts = list(event.get("attempt_transcripts") or [])
            if attempt_transcripts:
                lines.append("attempt_transcripts: " + json.dumps(attempt_transcripts, ensure_ascii=False, default=str))
        io_rows = list(session.get("IoRows") or [])
        if io_rows:
            lines.append("## io_rows_all")
            for row in io_rows:
                lines.append("io_row_all: " + json.dumps(row, ensure_ascii=False, default=str))
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def _build_writeback_truth_groups(detail_rows: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for row in detail_rows:
        try:
            group = int(row.get("Group"))
        except Exception:
            continue
        source = str(row.get("ReadbackSource") or _READBACK_SOURCE_NONE)
        truth_source = _truth_source_label(source)
        verified = bool(_safe_bool(row.get("ReadbackOk"))) and truth_source == "explicit_c0"
        rows.append(
            {
                "Analyzer": _normalize_analyzer(row.get("Analyzer")),
                "Port": str(row.get("Port") or ""),
                "TargetDeviceId": _normalize_device_id(row.get("TargetDeviceId")),
                "LiveDeviceId": _normalize_device_id(row.get("LiveDeviceId")),
                "Group": int(group),
                "TruthSource": truth_source,
                "Verified": bool(verified),
                "ReadbackSource": source,
                "ReadbackSourceLine": str(row.get("ReadbackSourceLine") or ""),
                "ReadbackSourceHasExplicitC0": bool(row.get("ReadbackSourceHasExplicitC0", False)),
                "ReadbackCommand": str(row.get("ReadbackCommand") or ""),
                "ReadbackTargetId": str(row.get("ReadbackTargetId") or ""),
                "WriteStatus": str(row.get("WriteStatus") or ""),
                "VerifyStatus": str(row.get("VerifyStatus") or ""),
                "RollbackStatus": str(row.get("RollbackStatus") or ""),
                "Error": str(row.get("Error") or ""),
            }
        )
    return rows


def _build_writeback_truth_summary(
    *,
    summary_rows: Sequence[Mapping[str, Any]],
    truth_group_rows: Sequence[Mapping[str, Any]],
    sessions: Sequence[Mapping[str, Any]],
) -> Dict[str, Any]:
    truth_rows = list(truth_group_rows)
    source_counts = Counter(str(row.get("TruthSource") or "none") for row in truth_rows)
    verified_count = sum(1 for row in truth_rows if bool(row.get("Verified")))
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "explicit_c0_required": True,
        "group_count": len(truth_rows),
        "verified_group_count": int(verified_count),
        "unverified_group_count": int(len(truth_rows) - verified_count),
        "all_groups_verified": bool(truth_rows) and verified_count == len(truth_rows),
        "truth_source_counts": {
            "explicit_c0": int(source_counts.get("explicit_c0", 0)),
            "ambiguous": int(source_counts.get("ambiguous", 0)),
            "none": int(source_counts.get("none", 0)),
        },
        "summary_rows": [dict(row) for row in summary_rows],
        "sessions": [
            {
                "Analyzer": _normalize_analyzer(session.get("Analyzer")),
                "Port": str(session.get("Port") or ""),
                "TargetDeviceId": _normalize_device_id(session.get("TargetDeviceId")),
                "LiveDeviceId": _normalize_device_id(session.get("LiveDeviceId")),
                "Status": str(session.get("Status") or ""),
                "FatalError": str(session.get("FatalError") or ""),
                "EventCount": len(list(session.get("Events") or [])),
                "IoRowCount": len(list(session.get("IoRows") or [])),
            }
            for session in sessions
        ],
    }


def _annotate_rows_with_actual_device_ids(
    rows: Sequence[Mapping[str, Any]],
    actual_device_ids: Mapping[str, str],
    *,
    analyzer_key: str = "Analyzer",
) -> List[Dict[str, Any]]:
    normalized_ids = {
        _normalize_analyzer(key): _normalize_device_id(value)
        for key, value in dict(actual_device_ids or {}).items()
        if _normalize_analyzer(key) and _normalize_device_id(value)
    }
    annotated: List[Dict[str, Any]] = []
    for row in rows:
        payload = dict(row)
        resolved_analyzer_key = analyzer_key if analyzer_key in payload else ""
        if not resolved_analyzer_key:
            for key in payload.keys():
                header = str(key or "").strip()
                if header.lower() == "analyzer" or "分析" in header:
                    resolved_analyzer_key = str(key)
                    break
        if not resolved_analyzer_key:
            for key in payload.keys():
                analyzer_text = _normalize_analyzer(payload.get(key))
                if analyzer_text.startswith("GA"):
                    resolved_analyzer_key = str(key)
                    break
        analyzer = _normalize_analyzer(payload.get(resolved_analyzer_key))
        payload.setdefault("ActualDeviceId", normalized_ids.get(analyzer, ""))
        annotated.append(payload)
    return annotated


def _annotate_workbook_with_actual_device_ids(workbook_path: Path, actual_device_ids: Mapping[str, str]) -> None:
    normalized_ids = {
        _normalize_analyzer(key): _normalize_device_id(value)
        for key, value in dict(actual_device_ids or {}).items()
        if _normalize_analyzer(key) and _normalize_device_id(value)
    }
    if not normalized_ids or not workbook_path.exists():
        return

    analyzer_value_re = re.compile(r"^GA\d{2}$", re.IGNORECASE)
    wb = load_workbook(workbook_path)
    try:
        for ws in wb.worksheets:
            if ws.max_row < 2 or ws.max_column < 1:
                continue
            header = [str(ws.cell(1, idx).value or "").strip() for idx in range(1, ws.max_column + 1)]
            if "ActualDeviceId" in header:
                continue

            analyzer_col: Optional[int] = None
            for idx, value in enumerate(header, start=1):
                normalized = str(value or "").strip().lower()
                if normalized == "analyzer" or "分析" in str(value or ""):
                    analyzer_col = idx
                    break

            if analyzer_col is None:
                candidates = []
                for row_idx in range(2, min(ws.max_row, 16) + 1):
                    cell_value = str(ws.cell(row_idx, 1).value or "").strip()
                    if cell_value:
                        candidates.append(cell_value)
                if candidates and all(analyzer_value_re.match(value) for value in candidates):
                    analyzer_col = 1

            if analyzer_col is None:
                continue

            ws.insert_cols(analyzer_col + 1, amount=1)
            ws.cell(1, analyzer_col + 1).value = "ActualDeviceId"
            for row_idx in range(2, ws.max_row + 1):
                analyzer = _normalize_analyzer(ws.cell(row_idx, analyzer_col).value)
                ws.cell(row_idx, analyzer_col + 1).value = normalized_ids.get(analyzer, "")
        wb.save(workbook_path)
    finally:
        wb.close()


def _append_dataframe_sheet(workbook_path: Path, sheet_name: str, frame: pd.DataFrame) -> None:
    wb = load_workbook(workbook_path)
    try:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name[:31])
        if frame.empty:
            ws.append(["说明"])
            ws.append(["无数据"])
        else:
            ws.append(list(frame.columns))
            for row in frame.itertuples(index=False, name=None):
                ws.append(list(row))
        wb.save(workbook_path)
    finally:
        wb.close()


def _load_temperature_gate_hits(run_dir: Path) -> List[Dict[str, Any]]:
    observations_path = run_dir / "temperature_calibration_observations.csv"
    if not observations_path.exists():
        return []

    rows: List[Dict[str, Any]] = []
    with observations_path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            cell_valid = _safe_bool(row.get("valid_for_cell_fit"))
            shell_valid = _safe_bool(row.get("valid_for_shell_fit"))
            cell_reason = str(row.get("cell_fit_gate_reason") or "").strip()
            shell_reason = str(row.get("shell_fit_gate_reason") or "").strip()
            if cell_valid and shell_valid and not cell_reason and not shell_reason:
                continue
            rows.append(
                {
                    "analyzer_id": _normalize_analyzer(row.get("analyzer_id")),
                    "analyzer_device_id": _normalize_device_id(row.get("analyzer_device_id")),
                    "snapshot_time": row.get("snapshot_time") or row.get("timestamp"),
                    "route_type": row.get("route_type"),
                    "ref_temp_c": _safe_float(row.get("ref_temp_c")),
                    "cell_temp_raw_c": _safe_float(row.get("cell_temp_raw_c")),
                    "shell_temp_raw_c": _safe_float(row.get("shell_temp_raw_c")),
                    "cell_temp_span_c": _safe_float(row.get("cell_temp_span_c")),
                    "shell_temp_span_c": _safe_float(row.get("shell_temp_span_c")),
                    "valid_for_cell_fit": cell_valid,
                    "valid_for_shell_fit": shell_valid,
                    "cell_fit_gate_reason": cell_reason,
                    "shell_fit_gate_reason": shell_reason,
                }
            )
    return rows


def _filter_no_500_summary_paths(run_dir: Path, output_dir: Path) -> tuple[List[Path], List[Dict[str, Any]]]:
    filtered_paths: List[Path] = []
    stats_rows: List[Dict[str, Any]] = []
    for source_path in _resolve_summary_paths(run_dir):
        if source_path.suffix.lower() == ".csv":
            frame = pd.read_csv(source_path, encoding="utf-8-sig")
            filtered_frame, stats = filter_no_500_frame(frame)
            out_path = output_dir / f"{source_path.stem}_no_500hpa.csv"
            filtered_frame.to_csv(out_path, index=False, encoding="utf-8-sig")
        else:
            workbook = pd.read_excel(source_path, sheet_name=None)
            sheets: Dict[str, pd.DataFrame] = {}
            original_rows = 0
            removed_rows = 0
            kept_rows = 0
            for ws_name, frame in workbook.items():
                filtered_frame, one_stats = filter_no_500_frame(frame)
                sheets[str(ws_name)] = filtered_frame
                original_rows += int(one_stats["original_rows"])
                removed_rows += int(one_stats["removed_rows"])
                kept_rows += int(one_stats["kept_rows"])
            out_path = output_dir / f"{source_path.stem}_no_500hpa.xlsx"
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                for ws_name, frame in sheets.items():
                    frame.to_excel(writer, sheet_name=str(ws_name)[:31], index=False)
            stats = {"original_rows": original_rows, "removed_rows": removed_rows, "kept_rows": kept_rows}
        filtered_paths.append(out_path)
        stats_rows.append({"source": str(source_path), "filtered": str(out_path), **stats})
    return filtered_paths, stats_rows


def extract_run_device_ids(run_dir: str | Path) -> Dict[str, str]:
    run_dir = Path(run_dir)
    frame = pd.read_csv(_resolve_samples_path(run_dir), encoding="utf-8-sig")
    mapping: Dict[str, str] = {}
    for index in range(1, 9):
        analyzer = f"GA{index:02d}"
        values: List[str] = []
        for column in (
            f"气体分析仪{index}_设备ID",
            f"gas_analyzer_{index}_device_id",
            f"ga{index:02d}_device_id",
        ):
            if column not in frame.columns:
                continue
            for value in frame[column].tolist():
                normalized = _normalize_device_id(value)
                if normalized:
                    values.append(normalized)
        if values:
            mapping[analyzer] = Counter(values).most_common(1)[0][0]
    return mapping


def _coeff_value(row: Mapping[str, Any], index: int) -> float:
    numeric = _safe_float(row.get(f"a{index}"))
    return 0.0 if numeric is None else float(numeric)


def _row_identity(row: Mapping[str, Any]) -> tuple[str, str]:
    analyzer = _normalize_analyzer(row.get("Analyzer") or row.get("分析仪") or row.get("鍒嗘瀽浠?"))
    gas = str(row.get("Gas") or row.get("气体") or row.get("姘斾綋") or "").strip().upper()
    return analyzer, gas


def _summary_row_by_identity(summary_frame: pd.DataFrame) -> Dict[tuple[str, str], Dict[str, Any]]:
    mapping: Dict[tuple[str, str], Dict[str, Any]] = {}
    if summary_frame.empty:
        return mapping
    for row in summary_frame.to_dict(orient="records"):
        identity = _row_identity(row)
        if identity[0] and identity[1]:
            mapping[identity] = dict(row)
    return mapping


def _frame_row_by_identity(frame: pd.DataFrame) -> Dict[tuple[str, str], Dict[str, Any]]:
    mapping: Dict[tuple[str, str], Dict[str, Any]] = {}
    if frame.empty:
        return mapping
    for row in frame.to_dict(orient="records"):
        identity = _row_identity(row)
        if identity[0] and identity[1]:
            mapping[identity] = dict(row)
    return mapping


def _corrected_delivery_guard_cfg(corrected_cfg: Mapping[str, Any] | None) -> Dict[str, Any]:
    cfg = dict(corrected_cfg or {})
    return {
        "enabled": bool(cfg.get("simplification_guard_enabled", True)),
        "rmse_ratio_threshold": float(cfg.get("simplification_guard_rmse_ratio_threshold", 1.2) or 1.2),
        "prediction_diff_threshold_by_gas": dict(
            cfg.get(
                "simplification_guard_max_prediction_diff_threshold_by_gas",
                {"CO2": 5.0, "H2O": 0.5},
            )
            or {}
        ),
    }


def _select_download_source_rows(
    simplified_frame: pd.DataFrame,
    *,
    original_frame: pd.DataFrame | None = None,
    summary_frame: pd.DataFrame | None = None,
    corrected_cfg: Mapping[str, Any] | None = None,
) -> List[Dict[str, Any]]:
    if simplified_frame.empty:
        return []

    guard_cfg = _corrected_delivery_guard_cfg(corrected_cfg)
    summary_by_identity = _summary_row_by_identity(summary_frame if summary_frame is not None else pd.DataFrame())
    original_by_identity = _frame_row_by_identity(original_frame if original_frame is not None else pd.DataFrame())
    selected_rows: List[Dict[str, Any]] = []
    for simplified_row in simplified_frame.to_dict(orient="records"):
        identity = _row_identity(simplified_row)
        analyzer, gas = identity
        payload = dict(simplified_row)
        payload["coefficient_source"] = "simplified"
        payload["fallback_reason"] = ""
        summary_row = summary_by_identity.get(identity, {})
        if guard_cfg["enabled"]:
            original_row = original_by_identity.get(identity)
            simplified_rmse = _safe_float(summary_row.get("simplified_rmse"))
            original_rmse = _safe_float(summary_row.get("original_rmse"))
            prediction_diff = _safe_float(summary_row.get("max_prediction_diff_between_original_and_simplified"))
            reasons: List[str] = []
            if (
                simplified_rmse is not None
                and original_rmse is not None
                and original_rmse > 0.0
                and simplified_rmse > original_rmse * float(guard_cfg["rmse_ratio_threshold"])
            ):
                reasons.append("simplified_rmse_ratio_exceeded")
            diff_threshold = _safe_float(guard_cfg["prediction_diff_threshold_by_gas"].get(gas))
            if prediction_diff is not None and diff_threshold is not None and prediction_diff > diff_threshold:
                reasons.append("simplified_prediction_diff_exceeded")
            if reasons and original_row:
                payload = dict(original_row)
                payload["coefficient_source"] = "original_fallback"
                payload["fallback_reason"] = ";".join(reasons)
        if analyzer and gas:
            selected_rows.append(payload)
    return selected_rows


def build_corrected_download_plan_rows(
    simplified_frame: pd.DataFrame,
    *,
    actual_device_ids: Optional[Mapping[str, str]] = None,
    original_frame: pd.DataFrame | None = None,
    summary_frame: pd.DataFrame | None = None,
    corrected_cfg: Mapping[str, Any] | None = None,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    normalized_ids = {
        _normalize_analyzer(key): _normalize_device_id(value)
        for key, value in dict(actual_device_ids or {}).items()
        if _normalize_analyzer(key) and _normalize_device_id(value)
    }
    if simplified_frame.empty:
        return rows
    for row in _select_download_source_rows(
        simplified_frame,
        original_frame=original_frame,
        summary_frame=summary_frame,
        corrected_cfg=corrected_cfg,
    ):
        analyzer, gas = _row_identity(row)
        if not analyzer or gas not in {"CO2", "H2O"}:
            continue
        primary_group = 1 if gas == "CO2" else 2
        secondary_group = 3 if gas == "CO2" else 4
        primary_values = [_coeff_value(row, idx) for idx in range(4)] + [0.0, 0.0]
        secondary_values = [_coeff_value(row, idx) for idx in range(4, 9)] + [0.0]
        payload: Dict[str, Any] = {
            "Analyzer": analyzer,
            "ActualDeviceId": normalized_ids.get(analyzer, ""),
            "Gas": gas,
            "ModeEnterCommand": "MODE,YGAS,FFF,2",
            "ModeExitCommand": "MODE,YGAS,FFF,1",
            "PrimarySENCO": str(primary_group),
            "PrimaryValues": ",".join(format_senco_values(primary_values)),
            "PrimaryCommand": f"SENCO{primary_group},YGAS,FFF," + ",".join(format_senco_values(primary_values)),
            "SecondarySENCO": str(secondary_group),
            "SecondaryValues": ",".join(format_senco_values(secondary_values)),
            "SecondaryCommand": f"SENCO{secondary_group},YGAS,FFF," + ",".join(format_senco_values(secondary_values)),
            "CoefficientSource": str(row.get("coefficient_source") or "simplified"),
            "FallbackReason": str(row.get("fallback_reason") or ""),
        }
        for idx, value in enumerate(primary_values):
            payload[f"PrimaryC{idx}"] = format_senco_values([value])[0]
        for idx, value in enumerate(secondary_values):
            payload[f"SecondaryC{idx}"] = format_senco_values([value])[0]
        for idx in range(9):
            payload[f"a{idx}"] = _coeff_value(row, idx)
        rows.append(payload)
    rows.sort(key=lambda item: (str(item.get("Analyzer") or ""), str(item.get("Gas") or "")))
    return rows


def _temperature_coefficients_path(run_dir: Path) -> Optional[Path]:
    direct = run_dir / "temperature_compensation_coefficients.csv"
    if direct.exists():
        return direct
    return _latest_matching(run_dir, "temperature_compensation_coefficients*.csv")


def load_temperature_coefficient_rows(run_dir: str | Path) -> List[Dict[str, Any]]:
    run_dir = Path(run_dir)
    path = _temperature_coefficients_path(run_dir)
    if path is None:
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _latest_startup_pressure_calibration_summary(run_dir: Path) -> Optional[Path]:
    summaries = [
        path
        for path in run_dir.glob("startup_pressure_sensor_calibration_*/summary.csv")
        if path.is_file()
    ]
    if not summaries:
        return None
    return max(summaries, key=lambda item: item.stat().st_mtime)


def load_startup_pressure_calibration_rows(run_dir: str | Path) -> List[Dict[str, Any]]:
    run_dir = Path(run_dir)
    summary_path = _latest_startup_pressure_calibration_summary(run_dir)
    if summary_path is None:
        return []
    with summary_path.open("r", encoding="utf-8-sig", newline="") as handle:
        source_rows = list(csv.DictReader(handle))
    detail_path = summary_path.with_name("detail.csv")
    detail_rows: List[Dict[str, Any]] = []
    if detail_path.exists():
        with detail_path.open("r", encoding="utf-8-sig", newline="") as handle:
            detail_rows = list(csv.DictReader(handle))

    rows: List[Dict[str, Any]] = []
    for row in source_rows:
        analyzer = _normalize_analyzer(row.get("Analyzer"))
        device_id = _normalize_device_id(row.get("DeviceId"))
        offset = _safe_float(row.get("OffsetA_kPa"))
        if not analyzer or offset is None:
            continue
        summary_status = str(row.get("Status") or "").strip().lower()
        summary_samples = int(_safe_float(row.get("Samples")) or 0)
        pressure_write_recommended = True
        pressure_write_reason = ""
        reference_span_hpa: float | str = ""
        detail_samples = 0
        pressure_gauge_samples = 0
        analyzer_detail_rows = [item for item in detail_rows if _normalize_analyzer(item.get("Analyzer")) == analyzer]
        if analyzer_detail_rows:
            detail_samples = len(analyzer_detail_rows)
            pressure_gauge_samples = sum(1 for item in analyzer_detail_rows if str(item.get("ReferenceSource") or "").strip() == "pressure_gauge")
            reference_values = [_safe_float(item.get("ReferenceHpa")) for item in analyzer_detail_rows]
            reference_values = [float(value) for value in reference_values if value is not None]
            if reference_values:
                reference_span_hpa = float(max(reference_values) - min(reference_values))
        if summary_status and summary_status != "ok":
            pressure_write_recommended = False
            pressure_write_reason = f"startup_pressure_summary_status_{summary_status}"
        elif summary_samples < _STARTUP_PRESSURE_WRITE_MIN_SAMPLES:
            pressure_write_recommended = False
            pressure_write_reason = "startup_pressure_insufficient_samples"
        elif analyzer_detail_rows and pressure_gauge_samples != detail_samples:
            pressure_write_recommended = False
            pressure_write_reason = "startup_pressure_reference_not_pressure_gauge"
        elif isinstance(reference_span_hpa, float) and reference_span_hpa > _STARTUP_PRESSURE_WRITE_MAX_REFERENCE_SPAN_HPA:
            pressure_write_recommended = False
            pressure_write_reason = "startup_pressure_reference_unstable"
        rows.append(
            {
                "Analyzer": analyzer,
                "DeviceId": device_id,
                "ReferenceSource": "startup_pressure_sensor_calibration",
                "Samples": summary_samples,
                "OffsetA_kPa": float(offset),
                "ResidualMeanAbs_kPa": "",
                "ResidualMaxAbs_kPa": "",
                "StartupDetailSamples": detail_samples,
                "StartupPressureGaugeSamples": pressure_gauge_samples,
                "StartupReferenceSpanHpa": reference_span_hpa,
                "PressureWriteRecommended": pressure_write_recommended,
                "PressureWriteReason": pressure_write_reason,
                "WriteApplied": row.get("WriteApplied", ""),
                "ReadbackOk": row.get("ReadbackOk", ""),
                "Status": row.get("Status", ""),
                "Error": row.get("Error", ""),
                "Command": "SENCO9,YGAS,FFF," + ",".join(format_senco_values((float(offset), 1.0, 0.0, 0.0))),
                "SourceSummary": str(summary_path),
            }
        )
    rows.sort(key=lambda item: str(item.get("Analyzer") or ""))
    return rows


def compute_pressure_offset_rows(
    run_dir: str | Path,
    *,
    fallback_to_controller: bool = False,
) -> List[Dict[str, Any]]:
    run_dir = Path(run_dir)
    frame = pd.read_csv(_resolve_samples_path(run_dir), encoding="utf-8-sig")
    pressure_mode_col = "压力执行模式" if "压力执行模式" in frame.columns else "PressureMode"
    gauge_candidates = [
        "数字压力计压力hPa",
        "pressure_gauge_hpa",
        "PressureGaugeHpa",
    ]
    controller_candidates = ["压力控制器压力hPa", "pressure_controller_hpa", "PressureControllerHpa"]
    gauge_col = next((column for column in gauge_candidates if column in frame.columns), None)
    controller_col = next((column for column in controller_candidates if column in frame.columns), None)
    ref_candidates = list(gauge_candidates)
    if fallback_to_controller:
        ref_candidates.extend(["压力控制器压力hPa", "pressure_controller_hpa", "PressureControllerHpa"])
    ref_col = next((column for column in ref_candidates if column in frame.columns), None)
    if ref_col is None:
        return []

    ambient_mask = (
        frame[pressure_mode_col].astype(str).str.strip().str.lower().eq("ambient_open")
        if pressure_mode_col in frame.columns
        else pd.Series([True] * len(frame))
    )
    rows: List[Dict[str, Any]] = []
    for index in range(1, 9):
        analyzer = f"GA{index:02d}"
        device_id_col = f"气体分析仪{index}_设备ID"
        pressure_col = f"气体分析仪{index}_分析仪压力kPa"
        usable_col = f"气体分析仪{index}_分析仪可用帧"
        if device_id_col not in frame.columns or pressure_col not in frame.columns:
            continue
        subset = frame.loc[ambient_mask].copy()
        if usable_col in subset.columns:
            subset = subset[subset[usable_col].map(_safe_bool)]
        subset["ref_hpa"] = pd.to_numeric(subset[ref_col], errors="coerce")
        subset["analyzer_kpa"] = pd.to_numeric(subset[pressure_col], errors="coerce")
        subset = subset.dropna(subset=["ref_hpa", "analyzer_kpa"])
        if subset.empty:
            continue
        device_ids = [_normalize_device_id(value) for value in subset[device_id_col].tolist() if _normalize_device_id(value)]
        if not device_ids:
            continue
        device_id = Counter(device_ids).most_common(1)[0][0]
        residuals = (subset["ref_hpa"] / 10.0) - subset["analyzer_kpa"]
        offset = float(residuals.mean())
        corrected = subset["analyzer_kpa"] + offset
        abs_error_kpa = ((subset["ref_hpa"] / 10.0) - corrected).abs()
        overlap_samples = 0
        overlap_mean_abs_hpa: float | str = ""
        overlap_max_abs_hpa: float | str = ""
        pressure_write_recommended = True
        pressure_write_reason = ""
        if ref_col == controller_col:
            pressure_write_recommended = False
            pressure_write_reason = "controller_reference_fallback_used"
        elif gauge_col and controller_col:
            overlap = subset.copy()
            overlap["gauge_hpa"] = pd.to_numeric(overlap[gauge_col], errors="coerce")
            overlap["controller_hpa"] = pd.to_numeric(overlap[controller_col], errors="coerce")
            overlap = overlap.dropna(subset=["gauge_hpa", "controller_hpa"])
            overlap_samples = int(len(overlap))
            if overlap_samples < _PRESSURE_WRITE_MIN_GAUGE_CONTROLLER_OVERLAP:
                pressure_write_recommended = False
                pressure_write_reason = "insufficient_gauge_controller_overlap"
            else:
                gauge_controller_diff_hpa = (overlap["gauge_hpa"] - overlap["controller_hpa"]).abs()
                overlap_mean_abs_hpa = float(gauge_controller_diff_hpa.mean())
                overlap_max_abs_hpa = float(gauge_controller_diff_hpa.max())
                if (
                    float(overlap_mean_abs_hpa) > _PRESSURE_WRITE_MAX_GAUGE_CONTROLLER_MEAN_ABS_HPA
                    or float(overlap_max_abs_hpa) > _PRESSURE_WRITE_MAX_GAUGE_CONTROLLER_MAX_ABS_HPA
                ):
                    pressure_write_recommended = False
                    pressure_write_reason = "ambient_open_backpressure_too_large"
        else:
            pressure_write_recommended = False
            pressure_write_reason = "insufficient_gauge_controller_overlap"
        rows.append(
            {
                "Analyzer": analyzer,
                "DeviceId": device_id,
                "ReferenceSource": ref_col,
                "Samples": int(len(subset)),
                "OffsetA_kPa": offset,
                "ResidualMeanAbs_kPa": float(abs_error_kpa.mean()),
                "ResidualMaxAbs_kPa": float(abs_error_kpa.max()),
                "GaugeControllerOverlapSamples": overlap_samples,
                "GaugeControllerMeanAbsDiff_hPa": overlap_mean_abs_hpa,
                "GaugeControllerMaxAbsDiff_hPa": overlap_max_abs_hpa,
                "PressureWriteRecommended": pressure_write_recommended,
                "PressureWriteReason": pressure_write_reason,
                "Command": "SENCO9,YGAS,FFF," + ",".join(format_senco_values((offset, 1.0, 0.0, 0.0))),
            }
        )
    rows.sort(key=lambda item: str(item.get("Analyzer") or ""))
    return rows


def _load_targets_from_cfg(cfg: Mapping[str, Any]) -> List[Dict[str, Any]]:
    devices_cfg = dict(cfg.get("devices", {}) or {})
    gas_cfg = devices_cfg.get("gas_analyzers")
    if not isinstance(gas_cfg, list) or not gas_cfg:
        single = devices_cfg.get("gas_analyzer")
        gas_cfg = [single] if isinstance(single, Mapping) else []
    defaults = dict(devices_cfg.get("gas_analyzer", {}) or {})
    targets: List[Dict[str, Any]] = []
    for idx, item in enumerate(gas_cfg, start=1):
        if not isinstance(item, Mapping) or not item.get("enabled", True):
            continue
        port = str(item.get("port") or "").strip()
        if not port:
            continue
        targets.append(
            {
                "Analyzer": _normalize_analyzer(item.get("name"), index=idx),
                "Port": port,
                "Baudrate": int(item.get("baud", item.get("baudrate", defaults.get("baud", 115200))) or 115200),
                "Timeout": float(item.get("timeout", defaults.get("timeout", 0.6)) or 0.6),
                "ConfiguredDeviceId": _normalize_device_id(item.get("device_id") or defaults.get("device_id") or "000"),
                "ActiveSend": bool(item.get("active_send", defaults.get("active_send", True))),
                "FtdHz": int(item.get("ftd_hz", defaults.get("ftd_hz", 1)) or 1),
                "AverageFilter": int(item.get("average_filter", defaults.get("average_filter", 49)) or 49),
            }
        )
    return targets


def scan_live_targets(cfg: Mapping[str, Any], output_dir: str | Path) -> List[Dict[str, Any]]:
    output_dir = Path(output_dir)
    rows: List[Dict[str, Any]] = []
    for target in _load_targets_from_cfg(cfg):
        live_id = ""
        raw_line = ""
        error = ""
        ga = GasAnalyzer(
            target["Port"],
            target["Baudrate"],
            timeout=float(target["Timeout"]),
            device_id=target["ConfiguredDeviceId"] or "000",
        )
        try:
            ga.open()
            try:
                ga.set_comm_way_with_ack(False, require_ack=False)
            except Exception:
                pass
            for _ in range(4):
                raw_line = str(
                    ga.read_latest_data(
                        prefer_stream=True,
                        drain_s=0.2,
                        read_timeout_s=0.05,
                        allow_passive_fallback=True,
                    )
                    or ""
                )
                parsed = ga.parse_line(raw_line)
                live_id = _normalize_device_id((parsed or {}).get("id"))
                if live_id:
                    break
                time.sleep(0.1)
        except Exception as exc:
            error = str(exc)
        finally:
            try:
                ga.close()
            except Exception:
                pass
        rows.append({**target, "LiveDeviceId": live_id, "Raw": raw_line, "Error": error})
    _write_csv(output_dir / "scan.csv", rows)
    return rows


def _parse_senco_command(command: str) -> tuple[int, List[float]]:
    parts = [part.strip() for part in str(command or "").split(",") if str(part).strip()]
    if len(parts) < 4 or not parts[0].upper().startswith("SENCO"):
        raise ValueError(f"invalid SENCO command: {command}")
    return int(parts[0].upper().replace("SENCO", "")), [float(part) for part in parts[3:]]


def _empty_readback_capture() -> Dict[str, Any]:
    return {
        "source": _READBACK_SOURCE_NONE,
        "coefficients": {},
        "source_line": "",
        "source_line_has_explicit_c0": False,
        "raw_transcript_lines": [],
        "attempt_transcripts": [],
        "command": "",
        "target_id": "",
        "error": "",
    }


def _read_group_capture(ga: GasAnalyzer, group: int) -> Dict[str, Any]:
    reader = getattr(ga, "read_coefficient_group_capture", None)
    if callable(reader):
        capture = dict(reader(int(group)))
        payload = _empty_readback_capture()
        payload.update(capture)
        payload["coefficients"] = dict(payload.get("coefficients") or {})
        payload["raw_transcript_lines"] = list(payload.get("raw_transcript_lines") or [])
        payload["attempt_transcripts"] = list(payload.get("attempt_transcripts") or [])
        return payload

    parsed = ga.read_coefficient_group(int(group))
    payload = _empty_readback_capture()
    payload.update(
        {
            "source": _READBACK_SOURCE_EXPLICIT_C0,
            "coefficients": dict(parsed or {}),
            "source_line": "",
            "source_line_has_explicit_c0": True,
        }
    )
    return payload


def _read_group_as_list(
    ga: GasAnalyzer,
    group: int,
    expected_len: int,
) -> tuple[List[float], Optional[str], Dict[str, Any]]:
    capture = _read_group_capture(ga, int(group))
    source = str(capture.get("source") or _READBACK_SOURCE_NONE)
    parsed = dict(capture.get("coefficients") or {})
    if source != _READBACK_SOURCE_EXPLICIT_C0:
        if source == _READBACK_SOURCE_AMBIGUOUS:
            error_text = f"READBACK_SOURCE_UNTRUSTED:{source}"
        else:
            error_text = str(capture.get("error") or "NO_VALID_COEFFICIENT_LINE")
        return [], error_text, capture
    if not isinstance(parsed, Mapping) or not parsed:
        return [], "READBACK_EMPTY", capture
    values: List[float] = []
    for idx in range(expected_len):
        key = f"C{idx}"
        if key not in parsed:
            return [], f"READBACK_PARSE_MISSING:{key}", capture
        try:
            values.append(float(parsed.get(key)))
        except Exception as exc:
            return [], f"READBACK_PARSE_ERROR:{key}:{exc}", capture
    return values, None, capture


def _read_group_with_retry(
    ga: GasAnalyzer,
    group: int,
    expected_len: int,
    *,
    attempts: int = 3,
    retry_delay_s: float = 0.15,
) -> tuple[List[float], Optional[str], Dict[str, Any]]:
    last_values: List[float] = []
    last_error = ""
    last_capture = _empty_readback_capture()
    total_attempts = max(1, int(attempts))
    for idx in range(total_attempts):
        try:
            values, read_error, capture = _read_group_as_list(ga, int(group), expected_len)
        except Exception as exc:
            values = []
            read_error = str(exc)
            capture = _empty_readback_capture()
        last_capture = dict(capture or {})
        if values:
            return values, None, capture
        last_values = []
        last_error = str(read_error or "READBACK_MISSING")
        if idx + 1 < total_attempts and retry_delay_s > 0:
            time.sleep(max(0.01, float(retry_delay_s)))
    return last_values, last_error or "READBACK_MISSING", last_capture


def _read_group_with_match_retry(
    ga: GasAnalyzer,
    group: int,
    coeffs: Sequence[float],
    *,
    attempts: int = 3,
    retry_delay_s: float = 0.15,
) -> tuple[List[float], Optional[str], Dict[str, Any]]:
    expected = [float(value) for value in coeffs]
    expected_len = len(expected)
    last_values: List[float] = []
    last_error = ""
    last_capture = _empty_readback_capture()
    total_attempts = max(1, int(attempts))
    for idx in range(total_attempts):
        values, read_error, capture = _read_group_with_retry(
            ga,
            int(group),
            expected_len,
            attempts=1,
            retry_delay_s=retry_delay_s,
        )
        last_capture = dict(capture or {})
        if values:
            last_values = values
            last_error = ""
            if senco_readback_matches(expected, values):
                return values, None, last_capture
        else:
            last_error = str(read_error or "READBACK_MISSING")
        if idx + 1 < total_attempts and retry_delay_s > 0:
            time.sleep(max(0.01, float(retry_delay_s)))
    if last_values:
        return last_values, last_error or "READBACK_MISMATCH", last_capture
    return [], last_error or "READBACK_MISSING", last_capture


def _normalize_mode_value(value: Any) -> Any:
    if value in (None, "", "null", "None"):
        return "UNKNOWN"
    try:
        return int(value)
    except Exception:
        return str(value)


def _best_effort_mode_snapshot(ga: GasAnalyzer) -> Any:
    reader = getattr(ga, "read_current_mode_snapshot", None)
    if callable(reader):
        try:
            snapshot = reader()
        except Exception:
            snapshot = None
        if isinstance(snapshot, dict):
            return _normalize_mode_value(snapshot.get("mode"))
    return "UNKNOWN"


def _mode_matches_expected(mode_value: Any, expected_mode: int) -> bool:
    normalized = _normalize_mode_value(mode_value)
    if normalized == "UNKNOWN":
        return False
    try:
        return int(normalized) == int(expected_mode)
    except Exception:
        return False


def _aggregate_detail_status(
    detail_rows: Sequence[Mapping[str, Any]],
    key: str,
    *,
    failure_reason: str = "",
    default: str,
) -> str:
    statuses = [str(row.get(key) or "").strip().lower() for row in detail_rows if str(row.get(key) or "").strip()]
    if not statuses:
        return "failed" if failure_reason else default
    if any(status == "failed" for status in statuses):
        return "failed"
    if any(status == "success" for status in statuses):
        if all(status in {"success", "not_needed", "not_requested"} for status in statuses):
            return "success"
        return "partial"
    if any(status == "skipped" for status in statuses):
        return "failed" if failure_reason and key != "rollback_status" else "skipped"
    return statuses[-1] if statuses else default


def _capture_detail_fields(prefix: str, capture: Mapping[str, Any] | None) -> Dict[str, Any]:
    payload = dict(capture or {})
    return {
        f"{prefix}Source": str(payload.get("source") or _READBACK_SOURCE_NONE),
        f"{prefix}SourceLine": str(payload.get("source_line") or ""),
        f"{prefix}SourceHasExplicitC0": bool(payload.get("source_line_has_explicit_c0", False)),
        f"{prefix}Command": str(payload.get("command") or ""),
        f"{prefix}TargetId": str(payload.get("target_id") or ""),
        f"{prefix}Transcript": json.dumps(list(payload.get("raw_transcript_lines") or []), ensure_ascii=False),
        f"{prefix}AttemptTranscripts": json.dumps(list(payload.get("attempt_transcripts") or []), ensure_ascii=False),
    }


def _invoke_set_senco(ga: GasAnalyzer, group: int, coeffs: Sequence[float]) -> bool:
    setter = getattr(ga, "set_senco")
    try:
        params = list(inspect.signature(setter).parameters.values())
    except (TypeError, ValueError):
        params = []
    if any(param.kind == inspect.Parameter.VAR_POSITIONAL for param in params):
        return bool(setter(int(group), *coeffs))
    return bool(setter(int(group), coeffs))


def _set_mode_with_required_ack(ga: GasAnalyzer, mode: int) -> bool:
    setter = getattr(ga, "set_mode_with_ack", None)
    if callable(setter):
        return bool(setter(int(mode), require_ack=True))
    fallback = getattr(ga, "set_mode", None)
    if callable(fallback):
        return bool(fallback(int(mode)))
    raise AttributeError("gas analyzer does not provide set_mode or set_mode_with_ack")


def write_senco_groups_with_full_verification(
    ga: GasAnalyzer,
    *,
    expected_groups: Mapping[int, Sequence[float]],
    restore_mode: int = 1,
    readback_attempts: int = 3,
    retry_delay_s: float = 0.15,
    compare_atol: float = 1e-9,
) -> Dict[str, Any]:
    normalized_groups = {
        int(group): [float(value) for value in coeffs]
        for group, coeffs in dict(expected_groups or {}).items()
    }
    mode_before = _best_effort_mode_snapshot(ga)
    mode_after: Any = mode_before
    detail_rows: List[Dict[str, Any]] = []
    detail_by_group: Dict[int, Dict[str, Any]] = {}
    before_snapshots: Dict[int, List[float]] = {}
    failure_reason = ""
    unsafe = False
    entered_mode = False
    calibration_mode_requested = False
    written_groups: List[int] = []
    mode_exit_attempted = False
    mode_exit_confirmed = False
    rollback_attempted = False
    rollback_confirmed = False
    session_events: List[Dict[str, Any]] = []
    mode_enter_capture = _empty_action_capture("set_mode_2")
    mode_exit_capture = _empty_action_capture("set_mode_restore")

    for group, coeffs in sorted(normalized_groups.items()):
        detail = {
            "group": int(group),
            "coeff_before": [],
            "coeff_target": [float(value) for value in coeffs],
            "coeff_readback": [],
            "coeff_rollback_target": [],
            "coeff_rollback_readback": [],
            "coeff_before_capture": _empty_readback_capture(),
            "coeff_readback_capture": _empty_readback_capture(),
            "coeff_rollback_capture": _empty_readback_capture(),
            "mode_before": mode_before,
            "mode_after": "UNKNOWN",
            "write_status": "pending",
            "verify_status": "pending",
            "rollback_status": "not_needed",
            "failure_reason": "",
            "compare_tolerance": float(compare_atol),
            "mode_exit_attempted": False,
            "mode_exit_confirmed": False,
            "rollback_attempted": False,
            "rollback_confirmed": False,
            "write_capture": _empty_action_capture(f"set_senco_group_{int(group)}"),
            "rollback_write_capture": _empty_action_capture(f"rollback_senco_group_{int(group)}"),
        }
        detail_rows.append(detail)
        detail_by_group[int(group)] = detail

        before_values, before_error, before_capture = _read_group_with_retry(
            ga,
            int(group),
            len(coeffs),
            attempts=readback_attempts,
            retry_delay_s=retry_delay_s,
        )
        detail["coeff_before_capture"] = dict(before_capture or {})
        session_events.append(
            _capture_getco_event(
                phase="prewrite",
                action=f"getco_before_group_{int(group)}",
                group=int(group),
                capture=before_capture,
            )
        )
        if before_error:
            detail["write_status"] = "skipped"
            detail["verify_status"] = "skipped"
            detail["failure_reason"] = f"PREWRITE_SNAPSHOT_FAILED:{before_error}"
            failure_reason = f"group={int(group)} {detail['failure_reason']}"
            break
        detail["coeff_before"] = list(before_values)
        before_snapshots[int(group)] = list(before_values)

    if not failure_reason and normalized_groups:
        try:
            calibration_mode_requested = True
            mode_enter_start = _io_row_count(ga)
            try:
                mode_enter_result = _set_mode_with_required_ack(ga, 2)
            except Exception as exc:
                mode_enter_capture = _capture_io_action(
                    ga,
                    mode_enter_start,
                    phase="writeback",
                    action="set_mode_2",
                    ok=False,
                    error=str(exc),
                    metadata={"requested_mode": 2},
                )
                session_events.append(dict(mode_enter_capture))
                raise
            mode_enter_capture = _capture_io_action(
                ga,
                mode_enter_start,
                phase="writeback",
                action="set_mode_2",
                ok=bool(mode_enter_result),
                metadata={"requested_mode": 2, "value": bool(mode_enter_result)},
            )
            session_events.append(dict(mode_enter_capture))
            if not mode_enter_result:
                raise RuntimeError("MODE=2 not acknowledged before SENCO write")
            entered_mode = True

            for group, coeffs in sorted(normalized_groups.items()):
                detail = detail_by_group[int(group)]
                try:
                    write_start = _io_row_count(ga)
                    try:
                        write_result = _invoke_set_senco(ga, int(group), coeffs)
                    except Exception as exc:
                        write_capture = _capture_io_action(
                            ga,
                            write_start,
                            phase="writeback",
                            action=f"set_senco_group_{int(group)}",
                            ok=False,
                            error=str(exc),
                            metadata={"group": int(group), "coefficients": list(coeffs)},
                        )
                        detail["write_capture"] = dict(write_capture)
                        session_events.append(dict(write_capture))
                        raise
                    write_capture = _capture_io_action(
                        ga,
                        write_start,
                        phase="writeback",
                        action=f"set_senco_group_{int(group)}",
                        ok=bool(write_result),
                        metadata={"group": int(group), "coefficients": list(coeffs), "value": bool(write_result)},
                    )
                    detail["write_capture"] = dict(write_capture)
                    session_events.append(dict(write_capture))
                    if not write_result:
                        raise RuntimeError("WRITE_ACK_FAILED")
                    detail["write_status"] = "success"
                    if int(group) not in written_groups:
                        written_groups.append(int(group))
                    values, readback_error, readback_capture = _read_group_with_match_retry(
                        ga,
                        int(group),
                        coeffs,
                        attempts=readback_attempts,
                        retry_delay_s=retry_delay_s,
                    )
                    detail["coeff_readback"] = list(values)
                    detail["coeff_readback_capture"] = dict(readback_capture or {})
                    session_events.append(
                        _capture_getco_event(
                            phase="verify",
                            action=f"getco_after_write_group_{int(group)}",
                            group=int(group),
                            capture=readback_capture,
                        )
                    )
                    if senco_readback_matches(coeffs, values, atol=compare_atol):
                        detail["verify_status"] = "success"
                    else:
                        detail["verify_status"] = "failed"
                        detail["failure_reason"] = str(
                            readback_error or f"READBACK_MISMATCH atol={float(compare_atol):g}"
                        )
                        failure_reason = f"group={int(group)} {detail['failure_reason']}"
                        break
                except Exception as exc:
                    detail["write_status"] = "failed"
                    detail["verify_status"] = "failed"
                    detail["failure_reason"] = str(exc)
                    if int(group) not in written_groups:
                        written_groups.append(int(group))
                    failure_reason = f"group={int(group)} {detail['failure_reason']}"
                    break
        except Exception as exc:
            failure_reason = str(exc)

    if failure_reason and written_groups:
        rollback_attempted = True
        rollback_failures: List[str] = []
        for group in written_groups:
            detail = detail_by_group.get(int(group))
            if detail is None:
                continue
            detail["rollback_attempted"] = True
            rollback_target = list(before_snapshots.get(int(group), []))
            if not rollback_target:
                detail["rollback_status"] = "failed"
                detail["rollback_confirmed"] = False
                detail["failure_reason"] = (
                    f"{detail['failure_reason']}; ROLLBACK_TARGET_MISSING".strip("; ")
                )
                rollback_failures.append(f"group={int(group)} ROLLBACK_TARGET_MISSING")
                continue
            detail["coeff_rollback_target"] = list(rollback_target)
            try:
                rollback_write_start = _io_row_count(ga)
                try:
                    rollback_write_result = _invoke_set_senco(ga, int(group), rollback_target)
                except Exception as exc:
                    rollback_write_capture = _capture_io_action(
                        ga,
                        rollback_write_start,
                        phase="rollback",
                        action=f"rollback_senco_group_{int(group)}",
                        ok=False,
                        error=str(exc),
                        metadata={"group": int(group), "coefficients": list(rollback_target)},
                    )
                    detail["rollback_write_capture"] = dict(rollback_write_capture)
                    session_events.append(dict(rollback_write_capture))
                    raise
                rollback_write_capture = _capture_io_action(
                    ga,
                    rollback_write_start,
                    phase="rollback",
                    action=f"rollback_senco_group_{int(group)}",
                    ok=bool(rollback_write_result),
                    metadata={"group": int(group), "coefficients": list(rollback_target), "value": bool(rollback_write_result)},
                )
                detail["rollback_write_capture"] = dict(rollback_write_capture)
                session_events.append(dict(rollback_write_capture))
                if not rollback_write_result:
                    raise RuntimeError("ROLLBACK_WRITE_ACK_FAILED")
                rollback_values, rollback_error, rollback_capture = _read_group_with_match_retry(
                    ga,
                    int(group),
                    rollback_target,
                    attempts=readback_attempts,
                    retry_delay_s=retry_delay_s,
                )
                detail["coeff_rollback_readback"] = list(rollback_values)
                detail["coeff_rollback_capture"] = dict(rollback_capture or {})
                session_events.append(
                    _capture_getco_event(
                        phase="rollback",
                        action=f"getco_after_rollback_group_{int(group)}",
                        group=int(group),
                        capture=rollback_capture,
                    )
                )
                if senco_readback_matches(rollback_target, rollback_values, atol=compare_atol):
                    detail["rollback_status"] = "success"
                    detail["rollback_confirmed"] = True
                else:
                    detail["rollback_status"] = "failed"
                    detail["rollback_confirmed"] = False
                    error_text = str(
                        rollback_error or f"ROLLBACK_READBACK_MISMATCH atol={float(compare_atol):g}"
                    )
                    detail["failure_reason"] = f"{detail['failure_reason']}; {error_text}".strip("; ")
                    rollback_failures.append(f"group={int(group)} {error_text}")
            except Exception as exc:
                detail["rollback_status"] = "failed"
                detail["rollback_confirmed"] = False
                error_text = str(exc)
                detail["failure_reason"] = f"{detail['failure_reason']}; {error_text}".strip("; ")
                rollback_failures.append(f"group={int(group)} {error_text}")
        if rollback_failures:
            unsafe = True
            failure_reason = f"{failure_reason}; rollback_failed={' | '.join(rollback_failures)}"
        rollback_confirmed = rollback_attempted and not rollback_failures

    for detail in detail_rows:
        if detail["write_status"] == "pending":
            detail["write_status"] = "skipped" if failure_reason else "not_requested"
        if detail["verify_status"] == "pending":
            detail["verify_status"] = "skipped" if failure_reason else "not_requested"

    restore_error = ""
    if calibration_mode_requested or entered_mode or written_groups or failure_reason:
        mode_exit_attempted = True
        mode_exit_start = _io_row_count(ga)
        try:
            mode_exit_result = _set_mode_with_required_ack(ga, int(restore_mode))
        except Exception as exc:
            restore_error = str(exc)
            mode_exit_capture = _capture_io_action(
                ga,
                mode_exit_start,
                phase="restore",
                action=f"set_mode_{int(restore_mode)}",
                ok=False,
                error=restore_error,
                metadata={"requested_mode": int(restore_mode)},
            )
        else:
            mode_exit_capture = _capture_io_action(
                ga,
                mode_exit_start,
                phase="restore",
                action=f"set_mode_{int(restore_mode)}",
                ok=bool(mode_exit_result),
                metadata={"requested_mode": int(restore_mode), "value": bool(mode_exit_result)},
            )
            if not mode_exit_result:
                restore_error = f"MODE={int(restore_mode)} not acknowledged during restore"
        session_events.append(dict(mode_exit_capture))
    mode_after = _best_effort_mode_snapshot(ga)
    if mode_exit_attempted:
        mode_exit_confirmed = not restore_error and _mode_matches_expected(mode_after, int(restore_mode))
        if restore_error:
            unsafe = True
            failure_reason = "; ".join(item for item in [failure_reason, restore_error] if item)
        if not mode_exit_confirmed:
            unsafe = True
            failure_reason = "; ".join(
                item
                for item in [
                    failure_reason,
                    f"mode_exit_unconfirmed expected={int(restore_mode)} observed={mode_after}",
                ]
                if item
            )

    for detail in detail_rows:
        detail["mode_after"] = mode_after
        detail["mode_exit_attempted"] = mode_exit_attempted
        detail["mode_exit_confirmed"] = mode_exit_confirmed
        if detail["rollback_status"] == "not_needed" and not detail["rollback_attempted"]:
            detail["rollback_confirmed"] = False

    ok = bool(normalized_groups) and not failure_reason and not unsafe
    if not normalized_groups:
        ok = True
    return {
        "ok": ok,
        "unsafe": bool(unsafe),
        "mode_before": mode_before,
        "mode_after": mode_after,
        "mode_exit_attempted": bool(mode_exit_attempted),
        "mode_exit_confirmed": bool(mode_exit_confirmed),
        "rollback_attempted": bool(rollback_attempted),
        "rollback_confirmed": bool(rollback_confirmed),
        "write_status": _aggregate_detail_status(
            detail_rows,
            "write_status",
            failure_reason=failure_reason,
            default="not_requested",
        ),
        "verify_status": _aggregate_detail_status(
            detail_rows,
            "verify_status",
            failure_reason=failure_reason,
            default="not_requested",
        ),
        "rollback_status": _aggregate_detail_status(
            detail_rows,
            "rollback_status",
            failure_reason=failure_reason,
            default="not_needed",
        ),
        "failure_reason": failure_reason,
        "compare_tolerance": float(compare_atol),
        "mode_enter_capture": mode_enter_capture,
        "mode_exit_capture": mode_exit_capture,
        "session_events": session_events,
        "detail_rows": detail_rows,
    }


def _restore_stream_settings(
    ga: GasAnalyzer,
    target_cfg: Mapping[str, Any],
    *,
    return_captures: bool = False,
) -> List[Dict[str, Any]] | None:
    captures: List[Dict[str, Any]] = []
    steps = (
        ("restore_mode_2", lambda: ga.set_mode_with_ack(2, require_ack=False), {"requested_mode": 2}),
        (
            "restore_ftd",
            lambda: ga.set_active_freq_with_ack(int(target_cfg.get("FtdHz", 1) or 1), require_ack=False),
            {"value": int(target_cfg.get("FtdHz", 1) or 1)},
        ),
        (
            "restore_average_filter",
            lambda: ga.set_average_filter_with_ack(int(target_cfg.get("AverageFilter", 49) or 49), require_ack=False),
            {"value": int(target_cfg.get("AverageFilter", 49) or 49)},
        ),
        (
            "restore_comm_way",
            lambda: ga.set_comm_way_with_ack(bool(target_cfg.get("ActiveSend", True)), require_ack=False),
            {"value": bool(target_cfg.get("ActiveSend", True))},
        ),
    )
    for action, fn, metadata in steps:
        start_index = _io_row_count(ga)
        try:
            result = fn()
        except Exception as exc:
            capture = _capture_io_action(
                ga,
                start_index,
                phase="restore",
                action=action,
                ok=False,
                error=str(exc),
                metadata=metadata,
            )
            captures.append(capture)
            raise
        captures.append(
            _capture_io_action(
                ga,
                start_index,
                phase="restore",
                action=action,
                ok=True,
                metadata={**metadata, "result": result},
            )
        )
    if return_captures:
        return captures
    return None


def write_coefficients_to_live_devices(
    *,
    cfg: Mapping[str, Any],
    output_dir: str | Path,
    download_plan_rows: Sequence[Mapping[str, Any]],
    temperature_rows: Sequence[Mapping[str, Any]],
    pressure_rows: Sequence[Mapping[str, Any]],
    actual_device_ids: Mapping[str, str],
    write_pressure_rows: bool = True,
) -> Dict[str, Any]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    scan_rows = scan_live_targets(cfg, output_dir)
    live_by_id = {str(row.get("LiveDeviceId") or ""): row for row in scan_rows if str(row.get("LiveDeviceId") or "")}

    temp_by_analyzer: Dict[str, Dict[int, List[float]]] = {}
    for row in temperature_rows:
        analyzer = _normalize_analyzer(row.get("analyzer_id"))
        channel = str(row.get("senco_channel") or "").strip().upper()
        if analyzer and channel in {"SENCO7", "SENCO8"}:
            group = 7 if channel == "SENCO7" else 8
            temp_by_analyzer.setdefault(analyzer, {})[group] = [
                float(row.get("A") or 0.0),
                float(row.get("B") or 0.0),
                float(row.get("C") or 0.0),
                float(row.get("D") or 0.0),
            ]

    gas_by_analyzer: Dict[str, Dict[int, List[float]]] = {}
    for row in download_plan_rows:
        analyzer = _normalize_analyzer(row.get("Analyzer"))
        for key in ("PrimaryCommand", "SecondaryCommand"):
            command = str(row.get(key) or "").strip()
            if not command:
                continue
            group, coeffs = _parse_senco_command(command)
            gas_by_analyzer.setdefault(analyzer, {})[group] = coeffs

    pressure_by_analyzer: Dict[str, Dict[int, List[float]]] = {}
    skipped_pressure_by_analyzer: Dict[str, str] = {}
    if write_pressure_rows:
        for row in pressure_rows:
            analyzer = _normalize_analyzer(row.get("Analyzer"))
            offset = _safe_float(row.get("OffsetA_kPa"))
            if not analyzer or offset is None:
                continue
            recommended = row.get("PressureWriteRecommended")
            if recommended not in (None, "") and not _safe_bool(recommended):
                skipped_pressure_by_analyzer[analyzer] = str(row.get("PressureWriteReason") or "pressure_write_not_recommended")
                continue
            pressure_by_analyzer.setdefault(analyzer, {})[9] = [float(offset), 1.0, 0.0, 0.0]

    summary_rows: List[Dict[str, Any]] = []
    detail_rows: List[Dict[str, Any]] = []
    writeback_sessions: List[Dict[str, Any]] = []
    analyzers = sorted(set(actual_device_ids) | set(temp_by_analyzer) | set(gas_by_analyzer) | set(pressure_by_analyzer) | set(skipped_pressure_by_analyzer))
    for analyzer in analyzers:
        target_device_id = _normalize_device_id(actual_device_ids.get(analyzer))
        live_target = live_by_id.get(target_device_id)
        if analyzer in skipped_pressure_by_analyzer:
            detail_rows.append(
                {
                    "Analyzer": analyzer,
                    "Port": live_target["Port"] if live_target is not None else "",
                    "TargetDeviceId": target_device_id,
                    "LiveDeviceId": live_target.get("LiveDeviceId") if live_target is not None else "",
                    "Group": 9,
                    "Expected": "",
                    "Readback": "",
                    "ReadbackOk": False,
                    "ReadbackVerified": False,
                    "ReadbackTruthSource": "none",
                    "Error": f"SKIPPED_PRESSURE_WRITE:{skipped_pressure_by_analyzer[analyzer]}",
                }
            )
        expected_groups: Dict[int, List[float]] = {}
        expected_groups.update(temp_by_analyzer.get(analyzer, {}))
        expected_groups.update(gas_by_analyzer.get(analyzer, {}))
        expected_groups.update(pressure_by_analyzer.get(analyzer, {}))
        if not target_device_id or live_target is None:
            summary_rows.append({"Analyzer": analyzer, "TargetDeviceId": target_device_id, "ExpectedGroups": len(expected_groups), "MatchedGroups": 0, "Status": "missing_live_device"})
            continue
        matched = 0
        status = "ok"
        io_logger = _TranscriptIoLogger()
        session_events: List[Dict[str, Any]] = []
        session_fatal_error = ""
        ga = GasAnalyzer(
            live_target["Port"],
            live_target["Baudrate"],
            timeout=float(live_target["Timeout"]),
            device_id=target_device_id,
            io_logger=io_logger,
        )
        try:
            ga.open()
            prepare_start = _io_row_count(ga)
            try:
                prepare_result = ga.set_comm_way_with_ack(False, require_ack=False)
            except Exception as exc:
                session_events.append(
                    _capture_io_action(
                        ga,
                        prepare_start,
                        phase="prepare",
                        action="set_comm_way_false_noack",
                        ok=False,
                        error=str(exc),
                    )
                )
                raise
            session_events.append(
                _capture_io_action(
                    ga,
                    prepare_start,
                    phase="prepare",
                    action="set_comm_way_false_noack",
                    ok=True,
                    metadata={"value": bool(prepare_result)},
                )
            )
            verify_result = write_senco_groups_with_full_verification(
                ga,
                expected_groups=expected_groups,
            )
            session_events.extend(list(verify_result.get("session_events") or []))
            for detail in list(verify_result.get("detail_rows") or []):
                group = int(detail.get("group"))
                coeffs = expected_groups.get(group, [])
                readback_values = list(detail.get("coeff_readback") or [])
                readback_ok = str(detail.get("verify_status") or "").strip().lower() == "success"
                readback_source = str((detail.get("coeff_readback_capture") or {}).get("source") or _READBACK_SOURCE_NONE)
                readback_verified = bool(readback_ok) and _truth_source_label(readback_source) == "explicit_c0"
                if readback_ok:
                    matched += 1
                else:
                    status = "error" if bool(verify_result.get("unsafe")) else "partial"
                detail_rows.append(
                    {
                        "Analyzer": analyzer,
                        "Port": live_target["Port"],
                        "TargetDeviceId": target_device_id,
                        "LiveDeviceId": live_target.get("LiveDeviceId"),
                        "Group": group,
                        "Expected": json.dumps([float(value) for value in coeffs], ensure_ascii=False),
                        "Readback": json.dumps(readback_values, ensure_ascii=False),
                        "ReadbackOk": readback_ok,
                        "ReadbackVerified": bool(readback_verified),
                        "ReadbackTruthSource": _truth_source_label(readback_source),
                        "Error": str(detail.get("failure_reason") or ""),
                        "CoeffBefore": json.dumps(list(detail.get("coeff_before") or []), ensure_ascii=False),
                        "CoeffRollbackTarget": json.dumps(list(detail.get("coeff_rollback_target") or []), ensure_ascii=False),
                        "CoeffRollbackReadback": json.dumps(list(detail.get("coeff_rollback_readback") or []), ensure_ascii=False),
                        "ModeBefore": detail.get("mode_before"),
                        "ModeAfter": detail.get("mode_after"),
                        "ModeExitAttempted": bool(verify_result.get("mode_exit_attempted", False)),
                        "ModeExitConfirmed": bool(verify_result.get("mode_exit_confirmed", False)),
                        "RollbackAttempted": bool(
                            detail.get("rollback_attempted", verify_result.get("rollback_attempted", False))
                        ),
                        "RollbackConfirmed": bool(
                            detail.get("rollback_confirmed", verify_result.get("rollback_confirmed", False))
                        ),
                        "WriteStatus": detail.get("write_status"),
                        "VerifyStatus": detail.get("verify_status"),
                        "RollbackStatus": detail.get("rollback_status"),
                        "OverallWriteStatus": verify_result.get("write_status"),
                        "OverallVerifyStatus": verify_result.get("verify_status"),
                        "OverallRollbackStatus": verify_result.get("rollback_status"),
                        "CompareTolerance": detail.get("compare_tolerance"),
                        **_capture_detail_fields("CoeffBefore", detail.get("coeff_before_capture")),
                        **_capture_detail_fields("Readback", detail.get("coeff_readback_capture")),
                        **_capture_detail_fields("RollbackReadback", detail.get("coeff_rollback_capture")),
                    }
                )
            if not verify_result.get("ok"):
                status = "error" if bool(verify_result.get("unsafe")) else "partial"
            restore_captures = _restore_stream_settings(ga, live_target, return_captures=True) or []
            session_events.extend(list(restore_captures))
        except Exception as exc:
            status = "error"
            session_fatal_error = str(exc)
            detail_rows.append(
                {
                    "Analyzer": analyzer,
                    "Port": live_target["Port"],
                    "TargetDeviceId": target_device_id,
                    "LiveDeviceId": live_target.get("LiveDeviceId"),
                    "Group": "session",
                    "Expected": "",
                    "Readback": "",
                    "ReadbackOk": False,
                    "ReadbackVerified": False,
                    "ReadbackTruthSource": "none",
                    "Error": str(exc),
                }
            )
        finally:
            try:
                ga.close()
            except Exception:
                pass
        writeback_sessions.append(
            {
                "Analyzer": analyzer,
                "Port": live_target["Port"],
                "TargetDeviceId": target_device_id,
                "LiveDeviceId": live_target.get("LiveDeviceId"),
                "Status": status if matched < len(expected_groups) else "ok",
                "FatalError": session_fatal_error,
                "Events": session_events,
                "IoRows": list(io_logger.rows),
            }
        )
        summary_rows.append(
            {
                "Analyzer": analyzer,
                "Port": live_target["Port"],
                "TargetDeviceId": target_device_id,
                "LiveDeviceId": live_target.get("LiveDeviceId"),
                "ExpectedGroups": len(expected_groups),
                "MatchedGroups": matched,
                "Status": status if matched < len(expected_groups) else "ok",
            }
        )
    _write_csv(output_dir / "summary.csv", summary_rows)
    _write_csv(output_dir / "detail.csv", detail_rows)
    truth_group_rows = _build_writeback_truth_groups(detail_rows)
    truth_summary = _build_writeback_truth_summary(
        summary_rows=summary_rows,
        truth_group_rows=truth_group_rows,
        sessions=writeback_sessions,
    )
    transcript_path = output_dir / "writeback_raw_transcript.log"
    truth_summary_path = output_dir / "writeback_truth_summary.json"
    truth_groups_path = output_dir / "writeback_truth_groups.csv"
    _write_writeback_raw_transcript_log(transcript_path, writeback_sessions)
    _write_json(truth_summary_path, truth_summary)
    _write_csv(truth_groups_path, truth_group_rows)
    return {
        "scan_rows": scan_rows,
        "summary_rows": summary_rows,
        "detail_rows": detail_rows,
        "writeback_sessions": writeback_sessions,
        "writeback_raw_transcript_path": str(transcript_path),
        "writeback_truth_summary_path": str(truth_summary_path),
        "writeback_truth_groups_path": str(truth_groups_path),
    }


def _writeback_summary_all_ok(write_result: Mapping[str, Any]) -> bool:
    summary_rows = list(write_result.get("summary_rows") or [])
    if not summary_rows:
        return False
    return all(str(row.get("Status") or "").strip().lower() == "ok" for row in summary_rows)


def _resolve_postrun_override_path(cfg: Mapping[str, Any], raw_path: Any) -> Optional[str]:
    text = str(raw_path or "").strip()
    if not text:
        return None
    candidate = Path(text)
    if candidate.is_absolute():
        return str(candidate.resolve())
    base_dir = Path(str(cfg.get("_base_dir") or Path.cwd()))
    return str((base_dir / candidate).resolve())


def _load_runtime_snapshot_cfg(run_dir: Path) -> Dict[str, Any]:
    snapshot_path = run_dir / "runtime_config_snapshot.json"
    if not snapshot_path.exists():
        return {}
    try:
        return load_config(str(snapshot_path))
    except Exception:
        try:
            return json.loads(snapshot_path.read_text(encoding="utf-8-sig"))
        except Exception:
            return {}


def _resolve_corrected_selected_pressure_points(
    runtime_cfg: Mapping[str, Any] | None,
    coeff_cfg: Mapping[str, Any] | None,
) -> tuple[Any, Any, str]:
    workflow_cfg = dict(runtime_cfg.get("workflow") or {}) if isinstance(runtime_cfg, Mapping) else {}
    original_points = workflow_cfg.get("selected_pressure_points")
    coeff_payload = dict(coeff_cfg or {})
    has_override = (
        "postrun_selected_pressure_points_override" in coeff_payload
        or "selected_pressure_points_override" in coeff_payload
    )
    override_points = coeff_payload.get("postrun_selected_pressure_points_override")
    if override_points is None and "selected_pressure_points_override" in coeff_payload:
        override_points = coeff_payload.get("selected_pressure_points_override")
    if has_override:
        return original_points, override_points, "postrun_override"
    return original_points, original_points, "runtime_snapshot"


def _resolve_corrected_temperature_keys(coeff_cfg: Mapping[str, Any] | None) -> Dict[str, str]:
    summary_columns = dict((coeff_cfg or {}).get("summary_columns") or {})
    resolved: Dict[str, str] = {}
    for gas in ("co2", "h2o"):
        gas_cfg = summary_columns.get(gas) or summary_columns.get(gas.upper()) or {}
        if isinstance(gas_cfg, Mapping):
            temperature_key = str(gas_cfg.get("temperature") or "").strip()
            if temperature_key:
                resolved[gas] = temperature_key
    return resolved


def _build_fit_quality_summary(summary: pd.DataFrame) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for row in summary.to_dict(orient="records"):
        analyzer, gas = _row_identity(row)
        if not analyzer or not gas:
            continue
        rows.append(
            {
                "Analyzer": analyzer,
                "ActualDeviceId": row.get("ActualDeviceId", ""),
                "Gas": gas,
                "TemperatureColumnUsed": row.get("temperature_column_used") or row.get("温度源列", ""),
                "ModelFeaturePolicy": row.get("model_feature_policy") or row.get("模型特征策略", ""),
                "FitInputQuality": row.get("fit_input_quality") or row.get("拟合输入质量", ""),
                "FitInputWarning": row.get("fit_input_warning") or row.get("拟合输入告警", ""),
                "DeliveryRecommendation": row.get("delivery_recommendation_label") or row.get("下发建议", ""),
                "DeliveryRecommendationCode": row.get("delivery_recommendation") or "",
                "OriginalRmse": row.get("original_rmse"),
                "SimplifiedRmse": row.get("simplified_rmse"),
                "MaxPredictionDiff": row.get("max_prediction_diff_between_original_and_simplified"),
                "OverallSuggestion": row.get("综合建议", ""),
            }
        )
    return rows


def _build_coefficient_source_summary(download_plan_rows: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for row in download_plan_rows:
        rows.append(
            {
                "Analyzer": _normalize_analyzer(row.get("Analyzer")),
                "ActualDeviceId": _normalize_device_id(row.get("ActualDeviceId")),
                "Gas": str(row.get("Gas") or "").strip().upper(),
                "CoefficientSource": str(row.get("CoefficientSource") or "simplified"),
                "FallbackReason": str(row.get("FallbackReason") or ""),
            }
        )
    return rows


def _build_device_write_verify_summary(write_result: Mapping[str, Any]) -> List[Dict[str, Any]]:
    summary_rows = list(write_result.get("summary_rows") or [])
    detail_rows = list(write_result.get("detail_rows") or [])
    if not summary_rows:
        return [{"Status": "not_requested", "MatchedGroups": 0, "ExpectedGroups": 0, "FailureReasons": ""}]

    rows: List[Dict[str, Any]] = []
    for row in summary_rows:
        analyzer = _normalize_analyzer(row.get("Analyzer"))
        detail_subset = [item for item in detail_rows if _normalize_analyzer(item.get("Analyzer")) == analyzer]
        failure_reasons = sorted({str(item.get("Error") or "").strip() for item in detail_subset if str(item.get("Error") or "").strip()})
        failed_groups = [str(item.get("Group")) for item in detail_subset if not _safe_bool(item.get("ReadbackOk"))]
        rows.append(
            {
                "Analyzer": analyzer,
                "TargetDeviceId": _normalize_device_id(row.get("TargetDeviceId")),
                "LiveDeviceId": _normalize_device_id(row.get("LiveDeviceId")),
                "Status": str(row.get("Status") or "").strip(),
                "MatchedGroups": int(_safe_float(row.get("MatchedGroups")) or 0),
                "ExpectedGroups": int(_safe_float(row.get("ExpectedGroups")) or 0),
                "FailedGroups": ",".join(failed_groups),
                "FailureReasons": ";".join(failure_reasons),
            }
        )
    return rows


def _load_runtime_parity_summary(runtime_parity_summary_path: Any) -> Dict[str, Any]:
    text = str(runtime_parity_summary_path or "").strip()
    if not text:
        return {}
    path = Path(text)
    if not path.exists():
        return {
            "parity_verdict": "not_audited",
            "runtime_parity_quality": "not_audited",
            "runtime_parity_summary_path": str(path),
            "parity_note": "runtime parity summary path does not exist",
        }
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        return {}
    if str(payload.get("probe_type") or "").strip().lower() == "baseline_ygas_stream":
        stream_formats = sorted(
            {
                str(item or "").strip().lower()
                for item in list(payload.get("stream_formats_seen") or [])
                if str(item or "").strip()
            }
        )
        legacy_stream_only = bool(payload.get("legacy_stream_only")) or stream_formats == ["legacy"]
        payload.setdefault("legacy_stream_only", legacy_stream_only)
        if legacy_stream_only:
            payload.setdefault("parity_verdict", "parity_inconclusive_missing_runtime_inputs")
            payload.setdefault("runtime_parity_quality", "parity_inconclusive_missing_runtime_inputs")
            payload.setdefault("final_write_ready", False)
            payload.setdefault("readiness_code", "legacy_stream_insufficient_for_runtime_parity")
            payload.setdefault("readiness_reason", "legacy_stream_insufficient_for_runtime_parity")
        elif not str(payload.get("parity_verdict") or "").strip():
            payload.setdefault("parity_verdict", "not_audited")
            payload.setdefault("runtime_parity_quality", "not_audited")
            payload.setdefault("final_write_ready", False)
    payload.setdefault("runtime_parity_summary_path", str(path))
    return payload


def _build_runtime_parity_summary(parity_payload: Mapping[str, Any] | None) -> List[Dict[str, Any]]:
    payload = dict(parity_payload or {})
    parity = summarize_runtime_parity(payload)
    if not payload:
        return [
            {
                "RuntimeParityQuality": "not_audited",
                "ParityVerdict": "not_audited",
                "LegacyStreamOnly": False,
                "VisibleRuntimeInputsAvailable": "",
                "VisibleRuntimeInputsMissing": "",
                "BestCandidate": "",
                "CandidateCountTested": 0,
                "ParitySummaryPath": "",
                "ParityNote": "runtime parity audit not provided",
            }
        ]
    candidate_rows = list(payload.get("candidate_rows") or [])
    tested_count = sum(1 for row in candidate_rows if str(row.get("candidate_status") or "").strip() == "tested")
    best_candidate = dict(payload.get("best_candidate") or {})
    return [
        {
            "RuntimeParityQuality": parity["quality"],
            "ParityVerdict": parity["verdict"],
            "LegacyStreamOnly": bool(parity["legacy_stream_only"]),
            "VisibleRuntimeInputsAvailable": ";".join(payload.get("visible_runtime_inputs_available") or []),
            "VisibleRuntimeInputsMissing": ";".join(payload.get("visible_runtime_inputs_missing") or []),
            "BestCandidate": best_candidate.get("candidate_name") or "",
            "CandidateCountTested": tested_count,
            "ParitySummaryPath": str(payload.get("runtime_parity_summary_path") or payload.get("output_summary_path") or ""),
            "ParityNote": str(payload.get("parity_note") or payload.get("conclusion_hint") or payload.get("readiness_reason") or ""),
        }
    ]


def _build_write_readiness_summary(
    *,
    fit_quality_summary: Sequence[Mapping[str, Any]],
    coefficient_source_summary: Sequence[Mapping[str, Any]],
    device_write_verify_summary: Sequence[Mapping[str, Any]],
    runtime_parity_payload: Mapping[str, Any] | None,
) -> List[Dict[str, Any]]:
    fit_gate = summarize_fit_quality(fit_quality_summary)
    writeback_quality = summarize_device_write_verify(device_write_verify_summary)
    parity_gate = summarize_runtime_parity(runtime_parity_payload)
    coefficient_sources = sorted(
        {
            str(row.get("CoefficientSource") or "").strip()
            for row in coefficient_source_summary
            if str(row.get("CoefficientSource") or "").strip()
        }
    )
    readiness = build_write_readiness_decision(
        fit_quality=fit_gate["quality"],
        delivery_recommendation=fit_gate["delivery_recommendation"],
        coefficient_source=";".join(coefficient_sources),
        writeback_status=writeback_quality,
        runtime_parity_verdict=parity_gate["verdict"],
        legacy_stream_only=bool(parity_gate["legacy_stream_only"]),
    )
    return [
        {
            "corrected_fit_quality": fit_gate["quality"],
            "delivery_recommendation": fit_gate["delivery_recommendation"],
            "device_write_verify_quality": writeback_quality,
            "runtime_parity_quality": parity_gate["quality"],
            "runtime_parity_verdict": parity_gate["verdict"],
            "legacy_stream_only": bool(parity_gate["legacy_stream_only"]),
            "coefficient_source": ";".join(coefficient_sources),
            "final_write_ready": bool(readiness["final_write_ready"]),
            "readiness_code": readiness["readiness_code"],
            "readiness_reason": readiness["readiness_reason"],
            "readiness_summary": readiness["readiness_summary"],
        }
    ]


def _write_corrected_summary_markdown(
    output_dir: Path,
    *,
    run_dir: Path,
    pressure_mode: str,
    capability: Mapping[str, Any],
    filter_stats: Sequence[Mapping[str, Any]],
    pressure_points_summary: Mapping[str, Any],
    fit_quality_summary: Sequence[Mapping[str, Any]],
    coefficient_source_summary: Sequence[Mapping[str, Any]],
    device_write_verify_summary: Sequence[Mapping[str, Any]],
    runtime_parity_summary: Sequence[Mapping[str, Any]],
    write_readiness_summary: Sequence[Mapping[str, Any]],
    run_structure_hints: Sequence[Mapping[str, Any]],
) -> None:
    readiness_row = dict(write_readiness_summary[0]) if write_readiness_summary else {}
    summary_lines = [
        "# corrected-entry no-500 summary",
        "",
        f"- run_dir: {run_dir}",
        f"- output_dir: {output_dir}",
        f"- pressure_row_source: {pressure_mode}",
        f"- h2o_zero_span_status: {capability.get('status')}",
        f"- h2o_zero_span_note: {capability.get('note')}",
        f"- original_pressure_points: {pressure_points_summary.get('original')}",
        f"- corrected_pressure_points: {pressure_points_summary.get('effective')}",
        f"- pressure_points_source: {pressure_points_summary.get('source')}",
        f"- corrected_fit_quality: {readiness_row.get('corrected_fit_quality', 'unknown')}",
        f"- device_write_verify_quality: {readiness_row.get('device_write_verify_quality', 'not_requested')}",
        f"- runtime_parity_quality: {readiness_row.get('runtime_parity_quality', 'not_audited')}",
        f"- final_write_ready: {readiness_row.get('final_write_ready', False)}",
        f"- readiness_code: {readiness_row.get('readiness_code', 'unknown')}",
        f"- readiness_reason: {readiness_row.get('readiness_reason', '')}",
        "",
        "## filter summary",
    ]
    for row in filter_stats:
        summary_lines.append(
            f"- {Path(str(row.get('source') or '')).name}: original={row.get('original_rows')} removed={row.get('removed_rows')} kept={row.get('kept_rows')}"
        )

    summary_lines.extend(["", "## fit_quality_summary"])
    if fit_quality_summary:
        for row in fit_quality_summary:
            summary_lines.append(
                f"- {row.get('Analyzer')}/{row.get('Gas')}: input={row.get('FitInputQuality')} warning={row.get('FitInputWarning')} temp_col={row.get('TemperatureColumnUsed')} model_policy={row.get('ModelFeaturePolicy')} suggestion={row.get('OverallSuggestion')}"
            )
    else:
        summary_lines.append("- none")

    summary_lines.extend(["", "## coefficient_source_summary"])
    if coefficient_source_summary:
        for row in coefficient_source_summary:
            summary_lines.append(
                f"- {row.get('Analyzer')}/{row.get('Gas')}: source={row.get('CoefficientSource')} fallback_reason={row.get('FallbackReason')}"
            )
    else:
        summary_lines.append("- none")

    summary_lines.extend(["", "## runtime_parity_summary"])
    if runtime_parity_summary:
        for row in runtime_parity_summary:
            summary_lines.append(
                f"- verdict={row.get('ParityVerdict')} quality={row.get('RuntimeParityQuality')} legacy_only={row.get('LegacyStreamOnly')} tested={row.get('CandidateCountTested')} best={row.get('BestCandidate', '')}"
            )
    else:
        summary_lines.append("- none")

    summary_lines.extend(["", "## device_write_verify_summary"])
    if device_write_verify_summary:
        for row in device_write_verify_summary:
            summary_lines.append(
                f"- {row.get('Analyzer', '') or 'n/a'}: status={row.get('Status')} matched={row.get('MatchedGroups')} expected={row.get('ExpectedGroups')} reasons={row.get('FailureReasons', '')}"
            )
    else:
        summary_lines.append("- none")

    summary_lines.extend(["", "## write_readiness_summary"])
    if write_readiness_summary:
        for row in write_readiness_summary:
            summary_lines.append(
                f"- corrected_fit_quality={row.get('corrected_fit_quality')} device_write_verify_quality={row.get('device_write_verify_quality')} runtime_parity_quality={row.get('runtime_parity_quality')} final_write_ready={row.get('final_write_ready')} code={row.get('readiness_code')}"
            )
    else:
        summary_lines.append("- none")

    if run_structure_hints:
        summary_lines.extend(["", "## run structure hints"])
        for row in run_structure_hints:
            summary_lines.append(f"- [{row.get('Status')}] {row.get('Summary')}: {row.get('Detail')}")

    (output_dir / "summary.md").write_text("\n".join(summary_lines) + "\n", encoding="utf-8")


def _match_anchor_temp_groups(
    values: Sequence[float],
    expected: Sequence[float],
    *,
    tolerance_c: float,
) -> tuple[List[float], List[float]]:
    matched: List[float] = []
    missing: List[float] = []
    for target in expected:
        if any(abs(float(value) - float(target)) <= tolerance_c for value in values):
            matched.append(float(target))
        else:
            missing.append(float(target))
    return matched, missing


def _build_run_structure_hints(
    *,
    run_dir: Path,
    coeff_cfg: Mapping[str, Any] | None,
    runtime_cfg: Mapping[str, Any] | None,
    h2o_selected: pd.DataFrame,
    h2o_anchor_gate_hits: pd.DataFrame,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = [
        {
            "CheckCode": "single_run_scope",
            "Status": "info",
            "Summary": "当前自动后处理只覆盖本轮 run",
            "Detail": "如果一次校准分成多轮完成，建议改用 merged calibration sidecar，并按 ActualDeviceId 合并后再计算最终系数。",
        }
    ]

    workflow_cfg = dict(runtime_cfg.get("workflow") or {}) if isinstance(runtime_cfg, Mapping) else {}
    pressure_points = [str(item).strip().lower() for item in list(workflow_cfg.get("selected_pressure_points") or []) if str(item).strip()]
    if pressure_points:
        non_ambient = [item for item in pressure_points if item != "ambient"]
        if non_ambient:
            rows.append(
                {
                    "CheckCode": "pressure_structure",
                    "Status": "ok",
                    "Summary": "当前 run 包含非 ambient 压力工况",
                    "Detail": "这更接近推荐运行结构，有利于后处理横向对比压力相关误差。",
                }
            )
        else:
            rows.append(
                {
                    "CheckCode": "pressure_structure",
                    "Status": "warn",
                    "Summary": "当前 run 仅包含 ambient 压力工况",
                    "Detail": "不修改本轮结果，但若想更接近 2026-04-03 的约束力，建议后续至少保留 1 个 sealed 压力点。",
                }
            )
    else:
        rows.append(
            {
                "CheckCode": "pressure_structure",
                "Status": "info",
                "Summary": "未显式记录 selected_pressure_points",
                "Detail": "自动后处理无法仅从配置判断本轮是否保留了 sealed 压力工况。",
            }
        )

    selection_cfg = normalize_h2o_summary_selection(
        (dict(coeff_cfg or {}).get("h2o_summary_selection") if isinstance(coeff_cfg, Mapping) else None)
    )
    expected_anchor_temps = [
        float(value)
        for value in list(selection_cfg.get("include_co2_zero_ppm_temp_groups_c") or [])
        if selection_cfg.get("include_co2_zero_ppm_rows", True)
    ]
    if expected_anchor_temps:
        anchor_frame = h2o_selected.copy()
        if not anchor_frame.empty:
            if "SelectionOrigin" in anchor_frame.columns:
                anchor_frame = anchor_frame[
                    anchor_frame["SelectionOrigin"].astype(str).str.strip().eq("co2_zero_ppm_anchor")
                ].copy()
            elif "PointPhase" in anchor_frame.columns:
                anchor_frame = anchor_frame[
                    anchor_frame["PointPhase"].astype(str).str.lower().isin({"气路", "co2"})
                ].copy()
                if "ppm_CO2_Tank" in anchor_frame.columns:
                    anchor_frame["ppm_CO2_Tank"] = pd.to_numeric(anchor_frame["ppm_CO2_Tank"], errors="coerce")
                    anchor_frame = anchor_frame[anchor_frame["ppm_CO2_Tank"].sub(0.0).abs().le(0.5)].copy()
        actual_anchor_temps: List[float] = []
        if not anchor_frame.empty:
            source = "EnvTempC" if "EnvTempC" in anchor_frame.columns else "Temp"
            actual_anchor_temps = [
                float(value)
                for value in pd.to_numeric(anchor_frame[source], errors="coerce").dropna().tolist()
            ]
        matched_temps, missing_temps = _match_anchor_temp_groups(
            actual_anchor_temps,
            expected_anchor_temps,
            tolerance_c=float(selection_cfg.get("temp_tolerance_c", 0.6) or 0.6),
        )
        if missing_temps:
            rows.append(
                {
                    "CheckCode": "h2o_anchor_coverage",
                    "Status": "warn",
                    "Summary": "H2O 0ppm 气路锚点覆盖不完整",
                    "Detail": f"期望温组 {expected_anchor_temps}°C，当前仅匹配到 {matched_temps}°C，缺少 {missing_temps}°C。",
                }
            )
        else:
            rows.append(
                {
                    "CheckCode": "h2o_anchor_coverage",
                    "Status": "ok",
                    "Summary": "H2O 0ppm 气路锚点覆盖完整",
                    "Detail": f"已覆盖 {matched_temps}°C 这些 H2O 零点锚点温组。",
                }
            )

    gate_rows = h2o_anchor_gate_hits.copy()
    if gate_rows.empty:
        rows.append(
            {
                "CheckCode": "h2o_anchor_quality_gate",
                "Status": "ok",
                "Summary": "本轮没有命中 H2O 锚点质量门禁",
                "Detail": "用于 H2O 拟合的 0ppm 气路锚点未出现 ppm_H2O_Dew 超限剔除。",
            }
        )
    else:
        if "PointRow" in gate_rows.columns:
            gate_rows["PointRow"] = pd.to_numeric(gate_rows["PointRow"], errors="coerce")
        hit_rows = [
            str(int(value))
            for value in gate_rows.get("PointRow", pd.Series(dtype=float)).dropna().tolist()
        ]
        rows.append(
            {
                "CheckCode": "h2o_anchor_quality_gate",
                "Status": "warn",
                "Summary": f"本轮有 {len(gate_rows)} 个 H2O 气路锚点被质量门禁剔除",
                "Detail": "命中 PointRow="
                + ", ".join(hit_rows[:8])
                + (" ..." if len(hit_rows) > 8 else "")
                + "。这不会改变原始点位记录，但会从 H2O 拟合中排除这些锚点。",
            }
        )
    return rows


def build_corrected_delivery(
    *,
    run_dir: str | Path,
    output_dir: str | Path,
    coeff_cfg: Optional[Mapping[str, Any]] = None,
    fallback_pressure_to_controller: bool = False,
    pressure_row_source: str = "startup_calibration",
) -> Dict[str, Any]:
    run_dir = Path(run_dir).resolve()
    output_dir = Path(output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    filtered_paths, filter_stats = _filter_no_500_summary_paths(run_dir, output_dir)
    runtime_cfg = _load_runtime_snapshot_cfg(run_dir)
    report_coeff_cfg = dict(coeff_cfg or {})
    original_pressure_points, corrected_pressure_points, pressure_points_source = _resolve_corrected_selected_pressure_points(
        runtime_cfg,
        report_coeff_cfg,
    )
    if corrected_pressure_points is not None:
        report_coeff_cfg["selected_pressure_points"] = corrected_pressure_points
    report_coeff_cfg["original_selected_pressure_points"] = original_pressure_points
    report_coeff_cfg["selected_pressure_points_source"] = pressure_points_source
    gas_temperature_keys = _resolve_corrected_temperature_keys(report_coeff_cfg)
    report_path = output_dir / "calibration_coefficients.xlsx"
    report = build_corrected_water_points_report(
        filtered_paths,
        output_path=report_path,
        coeff_cfg=report_coeff_cfg,
        gas_temperature_keys=gas_temperature_keys,
    )
    simplified = pd.DataFrame(report.get("simplified", pd.DataFrame())).copy()
    summary = pd.DataFrame(report.get("summary", pd.DataFrame())).copy()
    original = pd.DataFrame(report.get("original", pd.DataFrame())).copy()
    points = pd.DataFrame(report.get("points", pd.DataFrame())).copy()
    ranges = pd.DataFrame(report.get("ranges", pd.DataFrame())).copy()
    topn = pd.DataFrame(report.get("topn", pd.DataFrame())).copy()
    h2o_selected = pd.DataFrame(report.get("h2o_selected_rows", pd.DataFrame())).copy()
    h2o_anchor_gate_hits = pd.DataFrame(report.get("h2o_anchor_gate_hits", pd.DataFrame())).copy()
    actual_device_ids = extract_run_device_ids(run_dir)
    summary = pd.DataFrame(_annotate_rows_with_actual_device_ids(summary.to_dict(orient="records"), actual_device_ids))
    simplified = pd.DataFrame(_annotate_rows_with_actual_device_ids(simplified.to_dict(orient="records"), actual_device_ids))
    original = pd.DataFrame(_annotate_rows_with_actual_device_ids(original.to_dict(orient="records"), actual_device_ids))
    points = pd.DataFrame(_annotate_rows_with_actual_device_ids(points.to_dict(orient="records"), actual_device_ids))
    ranges = pd.DataFrame(_annotate_rows_with_actual_device_ids(ranges.to_dict(orient="records"), actual_device_ids))
    topn = pd.DataFrame(_annotate_rows_with_actual_device_ids(topn.to_dict(orient="records"), actual_device_ids))
    h2o_selected = pd.DataFrame(_annotate_rows_with_actual_device_ids(h2o_selected.to_dict(orient="records"), actual_device_ids))
    h2o_anchor_gate_hits = pd.DataFrame(
        _annotate_rows_with_actual_device_ids(h2o_anchor_gate_hits.to_dict(orient="records"), actual_device_ids)
    )
    temperature_gate_hits = pd.DataFrame(
        _annotate_rows_with_actual_device_ids(
            _load_temperature_gate_hits(run_dir),
            actual_device_ids,
            analyzer_key="analyzer_id",
        )
    )

    analyzer_summary_rows = []
    analyzer_values = summary.get("分析仪")
    analyzers = set(actual_device_ids)
    if isinstance(analyzer_values, pd.Series):
        analyzers.update(str(item) for item in analyzer_values.dropna().unique())
    for analyzer in sorted(analyzers):
        analyzer_summary_rows.append({"Analyzer": analyzer, "ActualDeviceId": actual_device_ids.get(analyzer, "")})

    download_plan_rows = build_corrected_download_plan_rows(
        simplified,
        actual_device_ids=actual_device_ids,
        original_frame=original,
        summary_frame=summary,
        corrected_cfg=report_coeff_cfg,
    )
    temperature_rows = load_temperature_coefficient_rows(run_dir)
    pressure_mode = str(pressure_row_source or "startup_calibration").strip().lower()
    if pressure_mode == "current_ambient":
        pressure_rows = compute_pressure_offset_rows(run_dir, fallback_to_controller=fallback_pressure_to_controller)
    elif pressure_mode == "none":
        pressure_rows = []
    else:
        pressure_rows = load_startup_pressure_calibration_rows(run_dir)
        if not pressure_rows and fallback_pressure_to_controller:
            pressure_rows = compute_pressure_offset_rows(run_dir, fallback_to_controller=True)

    run_structure_hint_cfg = (
        dict(runtime_cfg.get("workflow", {}).get("postrun_corrected_delivery", {}).get("run_structure_hints", {}))
        if isinstance(runtime_cfg, Mapping)
        else {}
    )
    run_structure_hints = pd.DataFrame(
        _build_run_structure_hints(
            run_dir=run_dir,
            coeff_cfg=coeff_cfg,
            runtime_cfg=runtime_cfg,
            h2o_selected=h2o_selected,
            h2o_anchor_gate_hits=h2o_anchor_gate_hits,
        )
        if bool(run_structure_hint_cfg.get("enabled", True))
        else []
    )

    _append_dataframe_sheet(report_path, "download_plan", pd.DataFrame(download_plan_rows))
    _append_dataframe_sheet(report_path, "分析仪汇总", pd.DataFrame(analyzer_summary_rows))
    _append_dataframe_sheet(report_path, "temperature_plan", pd.DataFrame(temperature_rows))
    _append_dataframe_sheet(report_path, "pressure_plan", pd.DataFrame(pressure_rows))
    fit_quality_summary = _build_fit_quality_summary(summary)
    coefficient_source_summary = _build_coefficient_source_summary(download_plan_rows)
    runtime_parity_payload: Dict[str, Any] = {}
    runtime_parity_summary = _build_runtime_parity_summary(runtime_parity_payload)
    write_readiness_summary = _build_write_readiness_summary(
        fit_quality_summary=fit_quality_summary,
        coefficient_source_summary=coefficient_source_summary,
        device_write_verify_summary=_build_device_write_verify_summary({}),
        runtime_parity_payload=runtime_parity_payload,
    )
    _append_dataframe_sheet(report_path, "fit_quality_summary", pd.DataFrame(fit_quality_summary))
    _append_dataframe_sheet(report_path, "coefficient_source_summary", pd.DataFrame(coefficient_source_summary))
    _append_dataframe_sheet(report_path, "runtime_parity_summary", pd.DataFrame(runtime_parity_summary))
    _append_dataframe_sheet(report_path, "write_readiness_summary", pd.DataFrame(write_readiness_summary))
    if not h2o_selected.empty:
        _append_dataframe_sheet(report_path, "H2O锚点入选", h2o_selected)
    if not h2o_anchor_gate_hits.empty:
        _append_dataframe_sheet(report_path, "H2O锚点门禁", h2o_anchor_gate_hits)
    if not temperature_gate_hits.empty:
        _append_dataframe_sheet(report_path, "温补异常快照", temperature_gate_hits)
    if not run_structure_hints.empty:
        _append_dataframe_sheet(report_path, "推荐运行结构提示", run_structure_hints)
    _annotate_workbook_with_actual_device_ids(report_path, actual_device_ids)

    _write_csv(output_dir / "download_plan_no_500.csv", download_plan_rows)
    _write_csv(output_dir / "fit_summary_no_500.csv", summary.to_dict(orient="records"))
    _write_csv(output_dir / "simplified_coefficients_no_500.csv", simplified.to_dict(orient="records"))
    _write_csv(output_dir / "original_coefficients_no_500.csv", original.to_dict(orient="records"))
    _write_csv(output_dir / "points_with_actual_ids_no_500.csv", points.to_dict(orient="records"))
    _write_csv(output_dir / "range_analysis_with_actual_ids_no_500.csv", ranges.to_dict(orient="records"))
    _write_csv(output_dir / "topn_with_actual_ids_no_500.csv", topn.to_dict(orient="records"))
    _write_csv(output_dir / "h2o_selected_rows_with_actual_ids.csv", h2o_selected.to_dict(orient="records"))
    _write_csv(output_dir / "h2o_anchor_gate_hits.csv", h2o_anchor_gate_hits.to_dict(orient="records"))
    _write_csv(output_dir / "temperature_fit_gate_hits.csv", temperature_gate_hits.to_dict(orient="records"))
    _write_csv(output_dir / "temperature_coefficients_target.csv", temperature_rows)
    _write_csv(output_dir / "pressure_offset_current_ambient_summary.csv", pressure_rows)
    _write_csv(output_dir / "run_structure_hints.csv", run_structure_hints.to_dict(orient="records"))
    _write_csv(output_dir / "filter_summary.csv", filter_stats)
    _write_csv(output_dir / "fit_quality_summary.csv", fit_quality_summary)
    _write_csv(output_dir / "coefficient_source_summary.csv", coefficient_source_summary)
    _write_csv(output_dir / "runtime_parity_summary.csv", runtime_parity_summary)
    _write_csv(output_dir / "write_readiness_summary.csv", write_readiness_summary)
    capability = v1_h2o_zero_span_capability(coeff_cfg if isinstance(coeff_cfg, Mapping) else {})
    pressure_points_summary = {
        "original": original_pressure_points,
        "effective": corrected_pressure_points,
        "source": pressure_points_source,
    }
    _write_corrected_summary_markdown(
        output_dir,
        run_dir=run_dir,
        pressure_mode=pressure_mode,
        capability=capability,
        filter_stats=filter_stats,
        pressure_points_summary=pressure_points_summary,
        fit_quality_summary=fit_quality_summary,
        coefficient_source_summary=coefficient_source_summary,
        device_write_verify_summary=_build_device_write_verify_summary({}),
        runtime_parity_summary=runtime_parity_summary,
        write_readiness_summary=write_readiness_summary,
        run_structure_hints=run_structure_hints.to_dict(orient="records"),
    )
    return {
        "report_path": str(report_path),
        "output_dir": str(output_dir),
        "filtered_summary_paths": [str(path) for path in filtered_paths],
        "filter_stats": filter_stats,
        "actual_device_ids": actual_device_ids,
        "download_plan_rows": download_plan_rows,
        "temperature_rows": temperature_rows,
        "pressure_rows": pressure_rows,
        "pressure_row_source": pressure_mode,
        "pressure_points_summary": pressure_points_summary,
        "fit_quality_summary": fit_quality_summary,
        "corrected_fit_quality": write_readiness_summary[0]["corrected_fit_quality"],
        "coefficient_source_summary": coefficient_source_summary,
        "runtime_parity_summary": runtime_parity_summary,
        "runtime_parity_payload": runtime_parity_payload,
        "runtime_parity_quality": write_readiness_summary[0]["runtime_parity_quality"],
        "device_write_verify_quality": write_readiness_summary[0]["device_write_verify_quality"],
        "final_write_ready": write_readiness_summary[0]["final_write_ready"],
        "readiness_code": write_readiness_summary[0]["readiness_code"],
        "readiness_reason": write_readiness_summary[0]["readiness_reason"],
        "write_readiness_summary": write_readiness_summary,
        "h2o_selected_rows": h2o_selected.to_dict(orient="records"),
        "h2o_anchor_gate_hits": h2o_anchor_gate_hits.to_dict(orient="records"),
        "temperature_gate_hits": temperature_gate_hits.to_dict(orient="records"),
        "run_structure_hints": run_structure_hints.to_dict(orient="records"),
    }


def run_from_cli(
    *,
    run_dir: str,
    config_path: Optional[str] = None,
    output_dir: Optional[str] = None,
    write_devices: bool = False,
    verify_report: bool = False,
    verification_template: Optional[str] = None,
    fallback_pressure_to_controller: bool = False,
    pressure_row_source: str = "startup_calibration",
    write_pressure_coefficients: bool = False,
    verify_short_run_cfg: Optional[Mapping[str, Any]] = None,
    runtime_parity_summary_path: Optional[str] = None,
) -> Dict[str, Any]:
    run_dir_path = Path(run_dir).resolve()
    target_dir = Path(output_dir).resolve() if output_dir else run_dir_path / f"corrected_autodelivery_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    target_dir.mkdir(parents=True, exist_ok=True)
    cfg_path = Path(config_path).resolve() if config_path else (run_dir_path / "runtime_config_snapshot.json")
    cfg = load_config(str(cfg_path))
    coeff_cfg = cfg.get("coefficients", {}) if isinstance(cfg.get("coefficients", {}), dict) else {}
    capability = v1_h2o_zero_span_capability(coeff_cfg)
    require_v1_h2o_zero_span_supported(coeff_cfg, context="run_v1_corrected_autodelivery")
    delivery = build_corrected_delivery(
        run_dir=run_dir_path,
        output_dir=target_dir,
        coeff_cfg=coeff_cfg,
        fallback_pressure_to_controller=fallback_pressure_to_controller,
        pressure_row_source=pressure_row_source,
    )
    write_result: Dict[str, Any] = {}
    if write_devices:
        write_result = write_coefficients_to_live_devices(
            cfg=cfg,
            output_dir=target_dir / "device_write",
            download_plan_rows=delivery["download_plan_rows"],
            temperature_rows=delivery["temperature_rows"],
            pressure_rows=delivery["pressure_rows"],
            actual_device_ids=delivery["actual_device_ids"],
            write_pressure_rows=write_pressure_coefficients,
        )

    verify_outputs: Dict[str, Any] = {}
    if verify_report and verification_template:
        from . import validate_verification_doc

        verification_targets = []
        for row in list(write_result.get("scan_rows") or []):
            live_id = _normalize_device_id(row.get("LiveDeviceId"))
            analyzer = _normalize_analyzer(row.get("Analyzer"))
            target_id = _normalize_device_id(delivery["actual_device_ids"].get(analyzer))
            if not live_id or live_id != target_id:
                continue
            verification_targets.append(
                {
                    "name": analyzer.lower(),
                    "port": row.get("Port"),
                    "baud": int(row.get("Baudrate") or 115200),
                    "device_id": live_id,
                    "enabled": True,
                }
            )
        targets_json_path = target_dir / "verification_targets.json"
        targets_json_path.write_text(json.dumps({"devices": {"gas_analyzers": verification_targets}}, ensure_ascii=False, indent=2), encoding="utf-8")
        verify_dir = target_dir / "verification_report"
        verify_dir.mkdir(parents=True, exist_ok=True)
        validate_verification_doc.run_from_cli(
            config=str(cfg_path),
            template=str(Path(verification_template).resolve()),
            targets_json=str(targets_json_path),
            output_dir=str(verify_dir),
        )
        verify_outputs = {"targets_json": str(targets_json_path), "output_dir": str(verify_dir)}

    short_verify_outputs: Dict[str, Any] = {}
    short_verify_cfg = dict(verify_short_run_cfg or {})
    if bool(short_verify_cfg.get("enabled", False)):
        if not write_devices:
            short_verify_outputs = {"skipped": True, "reason": "write_devices_disabled"}
        elif not _writeback_summary_all_ok(write_result):
            short_verify_outputs = {"skipped": True, "reason": "writeback_incomplete"}
        else:
            from . import verify_short_run

            points_excel_override = _resolve_postrun_override_path(cfg, short_verify_cfg.get("points_excel"))
            short_verify_outputs = verify_short_run.run_short_verification(
                config_path=str(cfg_path),
                temp_c=float(short_verify_cfg.get("temp_c", 20.0) or 20.0),
                skip_co2_ppm=[
                    int(item)
                    for item in list(short_verify_cfg.get("skip_co2_ppm") or [])
                    if str(item).strip()
                ],
                enable_connect_check=bool(short_verify_cfg.get("enable_connect_check", False)),
                run_id=f"postrun_verify_short_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                points_excel_override=points_excel_override,
                output_dir_override=str(target_dir / "short_verify"),
                actual_device_ids=delivery["actual_device_ids"],
            )

    device_write_verify_summary = _build_device_write_verify_summary(write_result)
    runtime_parity_payload = _load_runtime_parity_summary(runtime_parity_summary_path)
    runtime_parity_summary = _build_runtime_parity_summary(runtime_parity_payload)
    write_readiness_summary = _build_write_readiness_summary(
        fit_quality_summary=list(delivery.get("fit_quality_summary") or []),
        coefficient_source_summary=list(delivery.get("coefficient_source_summary") or []),
        device_write_verify_summary=device_write_verify_summary,
        runtime_parity_payload=runtime_parity_payload,
    )
    _append_dataframe_sheet(Path(delivery["report_path"]), "device_write_verify_summary", pd.DataFrame(device_write_verify_summary))
    _append_dataframe_sheet(Path(delivery["report_path"]), "runtime_parity_summary", pd.DataFrame(runtime_parity_summary))
    _append_dataframe_sheet(Path(delivery["report_path"]), "write_readiness_summary", pd.DataFrame(write_readiness_summary))
    _write_csv(target_dir / "runtime_parity_summary.csv", runtime_parity_summary)
    _write_csv(target_dir / "write_readiness_summary.csv", write_readiness_summary)
    _write_corrected_summary_markdown(
        target_dir,
        run_dir=run_dir_path,
        pressure_mode=str(delivery.get("pressure_row_source") or pressure_row_source),
        capability=capability,
        filter_stats=list(delivery.get("filter_stats") or []),
        pressure_points_summary=dict(delivery.get("pressure_points_summary") or {}),
        fit_quality_summary=list(delivery.get("fit_quality_summary") or []),
        coefficient_source_summary=list(delivery.get("coefficient_source_summary") or []),
        device_write_verify_summary=device_write_verify_summary,
        runtime_parity_summary=runtime_parity_summary,
        write_readiness_summary=write_readiness_summary,
        run_structure_hints=list(delivery.get("run_structure_hints") or []),
    )

    payload = {
        **delivery,
        "h2o_zero_span_capability": {
            **capability,
            "note": capability.get("note") or V1_CO2_ONLY_H2O_NOT_SUPPORTED_MESSAGE,
        },
        "corrected_fit_quality": write_readiness_summary[0]["corrected_fit_quality"],
        "device_write_verify_quality": write_readiness_summary[0]["device_write_verify_quality"],
        "runtime_parity_quality": write_readiness_summary[0]["runtime_parity_quality"],
        "runtime_parity_summary": runtime_parity_summary,
        "runtime_parity_payload": runtime_parity_payload,
        "final_write_ready": write_readiness_summary[0]["final_write_ready"],
        "readiness_code": write_readiness_summary[0]["readiness_code"],
        "readiness_reason": write_readiness_summary[0]["readiness_reason"],
        "write_readiness_summary": write_readiness_summary,
        "device_write_verify_summary": device_write_verify_summary,
        "write_result": write_result,
        "verify_outputs": verify_outputs,
        "short_verify_outputs": short_verify_outputs,
    }
    (target_dir / "autodelivery_summary.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    return payload
