from __future__ import annotations

from dataclasses import dataclass
import json
import math
import re
from pathlib import Path
from statistics import mean, stdev
from typing import Any, Dict, Iterable, Optional, Sequence

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd

from ...senco_format import format_senco_value
from ...coefficients.fit_ratio_poly import fit_ratio_poly_rt_p
from ...coefficients.model_metrics import compute_metrics
from ...coefficients.prediction_analysis import analyze_by_range
from ..config.models import CoefficientSummaryColumnConfig, CoefficientsConfig, H2OSummarySelectionConfig
from ..core.models import SamplingResult

_TEMP_RE = re.compile(r"([-+]?\d+(?:\.\d+)?)(?:°|℃)?C?")
_REPORT_SCOPE_TEXT = "按水路纠正规则"
_TOPN_LIMIT = 10
_PREFIT_QUALITY_SHEET = "拟合前基础质量检查"
_QUALITY_DETAIL_SHEET = "数据质量分析_分通道"
_QUALITY_NOTES_SHEET = "数据质量分析_说明"
_QUALITY_WORKBOOK_SUMMARY_SHEET = "汇总"
_QUALITY_WORKBOOK_DETAIL_SHEET = "分通道明细"
_QUALITY_WORKBOOK_NOTES_SHEET = "说明"
_PREFIT_QUALITY_COLUMNS = [
    "设备",
    "ID",
    "RISE ppm",
    "Bias ppm",
    "R²质量",
    "结论",
]
_QUALITY_DETAIL_COLUMNS = [
    "设备",
    "气路RISE ppm",
    "气路Bias ppm",
    "气路R²",
    "水路RISE ppm(等效)",
    "水路Bias ppm(等效)",
    "水路R²",
    "RISE ppm",
    "Bias ppm",
    "R²质量",
]
_PREFIT_QUALITY_FALLBACK = "数据缺失"
_QUALITY_ATTENTION = "可校准（关注）"
_QUALITY_WATER_EQ_SCALE = 40.0
_ANALYZER_SHEET_ID_RE = re.compile(r"(\d+)_ID([0-9A-Za-z_-]+)$")
_COEFF_DISPLAY_NUMBER_FORMAT = "0.00000E00"
_SUMMARY_FILL_MAP = {
    "建议采用": PatternFill(fill_type="solid", fgColor="C6EFCE"),
    "视业务确认": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "暂不采用": PatternFill(fill_type="solid", fgColor="F4CCCC"),
}


@dataclass(frozen=True)
class RatioPolyFitRecord:
    analyzer: str
    gas: str
    column_cfg: CoefficientSummaryColumnConfig
    selected_frame: pd.DataFrame
    result: Any
    residuals: pd.DataFrame
    metrics_original: Dict[str, float]
    metrics_simplified: Dict[str, float]
    effect: str
    effect_note: str
    suggestion: str
    suggestion_note: str


@dataclass(frozen=True)
class QualityAnalysisBundle:
    summary: pd.DataFrame
    detail: pd.DataFrame
    notes: pd.DataFrame


def _safe_float(value: Any) -> Optional[float]:
    try:
        numeric = float(value)
    except Exception:
        return None
    if math.isfinite(numeric):
        return numeric
    return None


def _mean(values: Iterable[Any]) -> Optional[float]:
    numeric = [value for value in (_safe_float(item) for item in values) if value is not None]
    if not numeric:
        return None
    return round(mean(numeric), 6)


def _std(values: Iterable[Any]) -> Optional[float]:
    numeric = [value for value in (_safe_float(item) for item in values) if value is not None]
    if len(numeric) < 2:
        return None
    return round(stdev(numeric), 6)


def _phase_display(route: str) -> str:
    return "水路" if str(route or "").strip().lower() == "h2o" else "气路"


def _phase_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"气路", "co2"}:
        return "co2"
    if text in {"水路", "h2o"}:
        return "h2o"
    return text


def _dewpoint_saturation_pressure_hpa(dewpoint_c: float) -> float:
    temp_c = float(dewpoint_c)
    if temp_c >= 0.0:
        return 6.1121 * math.exp((18.678 - temp_c / 234.5) * (temp_c / (257.14 + temp_c)))
    return 6.1115 * math.exp((23.036 - temp_c / 333.7) * (temp_c / (279.82 + temp_c)))


def _dewpoint_to_h2o_mmol_per_mol(dewpoint_c: Any, pressure_hpa: Any) -> Optional[float]:
    dewpoint = _safe_float(dewpoint_c)
    pressure = _safe_float(pressure_hpa)
    if dewpoint is None or pressure is None or pressure <= 0:
        return None
    vapor_pressure = _dewpoint_saturation_pressure_hpa(dewpoint)
    return round(1000.0 * vapor_pressure / pressure, 6)


def _build_point_tag(point_index: int, route: str) -> str:
    return f"{str(route or '').strip().lower()}_{int(point_index)}"


def _build_point_title(sample: SamplingResult) -> str:
    point = sample.point
    phase = _phase_display(point.route)
    temp_text = f"{float(point.temperature_c):g}°C"
    if str(point.route).strip().lower() == "h2o":
        humidity = "" if point.humidity_pct is None else f" RH={float(point.humidity_pct):g}%"
        pressure = "" if point.pressure_hpa is None else f" P={float(point.pressure_hpa):g}hPa"
        return f"{phase} {temp_text}{humidity}{pressure}".strip()
    co2 = "" if point.co2_ppm is None else f" CO2={float(point.co2_ppm):g}ppm"
    pressure = "" if point.pressure_hpa is None else f" P={float(point.pressure_hpa):g}hPa"
    return f"{phase} {temp_text}{co2}{pressure}".strip()


def _env_temp_from_row(row: Dict[str, Any]) -> Optional[float]:
    env_temp = _safe_float(row.get("EnvTempC"))
    if env_temp is not None:
        return env_temp
    for key in ("PointTitle", "PointTag"):
        match = _TEMP_RE.search(str(row.get(key) or ""))
        if match:
            return _safe_float(match.group(1))
    return None


def _range_bins(gas: str) -> list[float]:
    return [0.0, 200.0, 400.0, 800.0, 1200.0, 2000.0] if gas == "co2" else [0.0, 5.0, 10.0, 20.0, 40.0]


def _sample_is_usable(sample: SamplingResult) -> bool:
    if getattr(sample, "frame_usable", True) is False:
        return False
    return any(
        value is not None
        for value in (
            sample.co2_ratio_f,
            sample.co2_ratio_raw,
            sample.h2o_ratio_f,
            sample.h2o_ratio_raw,
            sample.co2_ppm,
            sample.h2o_mmol,
        )
    )


def _sample_alignment_key(sample: SamplingResult) -> str:
    sample_index = int(getattr(sample, "sample_index", 0) or 0)
    if sample_index > 0:
        return f"idx:{sample_index}"
    return f"ts:{sample.timestamp.isoformat()}"


def _dedupe_reference_samples(samples: Sequence[SamplingResult]) -> list[SamplingResult]:
    selected: dict[str, SamplingResult] = {}
    for sample in sorted(
        samples,
        key=lambda item: (
            int(getattr(item, "sample_index", 0) or 0),
            item.timestamp.isoformat(),
            str(item.analyzer_id or ""),
        ),
    ):
        key = _sample_alignment_key(sample)
        current = selected.get(key)
        if current is None:
            selected[key] = sample
            continue
        current_score = int(current.thermometer_temp_c is not None) + int(current.pressure_gauge_hpa is not None)
        sample_score = int(sample.thermometer_temp_c is not None) + int(sample.pressure_gauge_hpa is not None)
        if sample_score > current_score:
            selected[key] = sample
    return list(selected.values())


def _point_integrity_text(*, expected_count: int, present: set[str], usable: set[str]) -> str:
    if expected_count <= 0:
        return "鏃犲垎鏋愪华"
    if not present:
        return "鏃犲抚"
    missing = expected_count - len(present)
    unusable = len(present - usable)
    if len(usable) == expected_count:
        return "瀹屾暣"
    if len(usable) == 0 and len(present) > 0:
        return "浠呭紓甯稿抚"
    if missing > 0 and unusable > 0:
        return "閮ㄥ垎缂哄け涓斿惈寮傚父甯?"
    if missing > 0:
        return "閮ㄥ垎缂哄け"
    if unusable > 0:
        return "鍚紓甯稿抚"
    return "閮ㄥ垎鍙敤"


def _format_senco_float(value: Any) -> str:
    numeric = _safe_float(value)
    if numeric is None:
        numeric = 0.0
    return format_senco_value(numeric)


def _build_senco_payloads(coefficients: Dict[str, Any], gas: str) -> list[tuple[int, list[str]]]:
    gas_key = str(gas or "").strip().lower()
    if gas_key == "co2":
        primary_index, secondary_index = 1, 3
    elif gas_key == "h2o":
        primary_index, secondary_index = 2, 4
    else:
        raise ValueError(f"Unsupported gas for SENCO mapping: {gas}")

    primary_values = [
        _format_senco_float(coefficients.get("a0")),
        _format_senco_float(coefficients.get("a1")),
        _format_senco_float(coefficients.get("a2")),
        _format_senco_float(coefficients.get("a3")),
        _format_senco_float(0.0),
        _format_senco_float(0.0),
    ]
    secondary_values = [
        _format_senco_float(coefficients.get("a4")),
        _format_senco_float(coefficients.get("a5")),
        _format_senco_float(coefficients.get("a6")),
        _format_senco_float(coefficients.get("a7")),
        _format_senco_float(coefficients.get("a8")),
        _format_senco_float(0.0),
    ]
    return [
        (primary_index, primary_values),
        (secondary_index, secondary_values),
    ]


def _build_download_plan_rows(records: Sequence[RatioPolyFitRecord]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        payloads = _build_senco_payloads(record.result.simplified_coefficients, record.gas)
        row: dict[str, Any] = {
            "Analyzer": record.analyzer,
            "Gas": record.gas.upper(),
            "ModeEnterCommand": "MODE,YGAS,FFF,2",
            "ModeExitCommand": "MODE,YGAS,FFF,1",
        }
        for slot, (senco_index, values) in enumerate(payloads, start=1):
            prefix = "Primary" if slot == 1 else "Secondary"
            row[f"{prefix}SENCO"] = senco_index
            row[f"{prefix}Values"] = ",".join(values)
            row[f"{prefix}Command"] = f"SENCO{senco_index},YGAS,FFF,{','.join(values)}"
            for value_idx, value in enumerate(values):
                row[f"{prefix}C{value_idx}"] = value
        for coeff_name, coeff_value in sorted(record.result.simplified_coefficients.items()):
            row[coeff_name] = coeff_value
        rows.append(row)
    return rows


def _normalize_device_id(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text.zfill(3) if text.isdigit() else text


def _load_analyzer_ids_from_sheet_workbooks(summary_frame: pd.DataFrame) -> dict[str, str]:
    mapping: dict[str, str] = {}
    if "SourcePath" not in summary_frame.columns:
        return mapping

    parent_dirs = sorted({str(Path(value).resolve().parent) for value in summary_frame["SourcePath"].dropna() if str(value).strip()})
    for raw_dir in parent_dirs:
        directory = Path(raw_dir)
        for pattern in ("co2_analyzer_sheets_*.xlsx", "h2o_analyzer_sheets_*.xlsx"):
            for workbook_path in sorted(directory.glob(pattern)):
                try:
                    workbook = load_workbook(workbook_path, read_only=True)
                except Exception:
                    continue
                try:
                    for sheet_name in workbook.sheetnames:
                        match = _ANALYZER_SHEET_ID_RE.search(str(sheet_name))
                        if not match:
                            continue
                        analyzer = f"GA{int(match.group(1)):02d}"
                        device_id = _normalize_device_id(match.group(2))
                        if analyzer and device_id:
                            mapping.setdefault(analyzer, device_id)
                finally:
                    workbook.close()
    return mapping


def _load_runtime_analyzer_ids(summary_frame: pd.DataFrame) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for column in ("DeviceId", "AnalyzerId", "ID"):
        if column not in summary_frame.columns:
            continue
        for row in summary_frame.loc[:, ["Analyzer", column]].dropna().to_dict(orient="records"):
            analyzer = str(row.get("Analyzer") or "").strip().upper()
            device_id = _normalize_device_id(row.get(column))
            if analyzer and device_id:
                mapping.setdefault(analyzer, device_id)
        if mapping:
            return mapping

    sheet_mapping = _load_analyzer_ids_from_sheet_workbooks(summary_frame)
    if sheet_mapping:
        return sheet_mapping

    if "SourcePath" not in summary_frame.columns:
        return mapping

    source_paths = sorted({str(value).strip() for value in summary_frame["SourcePath"].dropna() if str(value).strip()})
    for raw_path in source_paths:
        snapshot_path = Path(raw_path).resolve().parent / "runtime_config_snapshot.json"
        if not snapshot_path.exists():
            continue
        try:
            payload = json.loads(snapshot_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        analyzers = (((payload or {}).get("devices") or {}).get("gas_analyzers") or [])
        for item in analyzers:
            if not isinstance(item, dict):
                continue
            analyzer = str(item.get("name") or "").strip().upper()
            device_id = _normalize_device_id(item.get("device_id"))
            if analyzer and device_id:
                mapping.setdefault(analyzer, device_id)
    return mapping


def _prefit_quality_conclusion(rise_ppm: Any, bias_ppm: Any, r2_value: Any) -> str:
    rise = _safe_float(rise_ppm)
    bias = _safe_float(bias_ppm)
    quality = _safe_float(r2_value)
    if rise is None or bias is None or quality is None:
        return _PREFIT_QUALITY_FALLBACK
    abs_bias = abs(bias)
    if quality < 0.70 or rise >= 180.0:
        return "建议剔除"
    if quality < 0.90 and rise < 100.0 and abs_bias < 80.0:
        return "边界可用"
    if quality >= 0.95 and rise <= 80.0 and abs_bias <= 50.0:
        return "可校准"
    return _QUALITY_ATTENTION


def _build_quality_design_matrix(route_frame: pd.DataFrame) -> np.ndarray:
    primary = route_frame["PrimaryQuality"].to_numpy(float)
    pressure = route_frame["PressureQuality"].to_numpy(float)
    temperature = route_frame["TempQuality"].to_numpy(float)
    auxiliary = route_frame["AuxQuality"].to_numpy(float)
    return np.column_stack(
        [
            np.ones(len(route_frame)),
            primary,
            pressure,
            temperature,
            auxiliary,
            primary * primary,
            pressure * pressure,
            temperature * temperature,
            auxiliary * auxiliary,
            primary * pressure,
            primary * temperature,
            primary * auxiliary,
            pressure * temperature,
            pressure * auxiliary,
            temperature * auxiliary,
        ]
    )


def _prepare_quality_route_frame(
    summary_frame: pd.DataFrame,
    *,
    phase_key: str,
    target_column: str,
    primary_column: str,
    auxiliary_column: str,
) -> pd.DataFrame:
    if summary_frame.empty:
        return pd.DataFrame()

    working = summary_frame.copy()
    working["Analyzer"] = working.get("Analyzer", "").map(lambda value: str(value or "").strip().upper())
    if "PhaseKey" not in working.columns:
        working["PhaseKey"] = working.get("PointPhase", "").map(_phase_key)
    else:
        working["PhaseKey"] = working["PhaseKey"].map(_phase_key)
    working["PointTagNorm"] = working.get("PointTag", "").map(lambda value: str(value or "").strip())
    working["TargetQuality"] = pd.to_numeric(working.get(target_column), errors="coerce")
    working["PrimaryQuality"] = pd.to_numeric(working.get(primary_column), errors="coerce")
    working["AuxQuality"] = pd.to_numeric(working.get(auxiliary_column), errors="coerce")
    pressure_primary = (
        pd.to_numeric(working.get("P"), errors="coerce")
        if "P" in working.columns
        else pd.Series(np.nan, index=working.index, dtype=float)
    )
    pressure_fallback = (
        pd.to_numeric(working.get("BAR"), errors="coerce")
        if "BAR" in working.columns
        else pd.Series(np.nan, index=working.index, dtype=float)
    )
    working["PressureQuality"] = pressure_primary.fillna(pressure_fallback)
    temp_primary = (
        pd.to_numeric(working.get("Temp"), errors="coerce")
        if "Temp" in working.columns
        else pd.Series(np.nan, index=working.index, dtype=float)
    )
    temp_fallback = (
        pd.to_numeric(working.get("T1"), errors="coerce")
        if "T1" in working.columns
        else pd.Series(np.nan, index=working.index, dtype=float)
    )
    working["TempQuality"] = temp_primary.fillna(temp_fallback)

    working = working[
        (working["PhaseKey"] == phase_key)
        & working["Analyzer"].ne("")
        & working["PointTagNorm"].ne("")
        & working["TargetQuality"].notna()
        & working["PrimaryQuality"].notna()
        & working["AuxQuality"].notna()
        & working["PressureQuality"].notna()
        & working["TempQuality"].notna()
    ].copy()
    if working.empty:
        return pd.DataFrame()

    expected_analyzers = int(working["Analyzer"].nunique())
    common_counts = working.groupby("PointTagNorm")["Analyzer"].nunique()
    common_tags = common_counts[common_counts == expected_analyzers].index
    working = working[working["PointTagNorm"].isin(common_tags)].copy()
    return working


def _quality_score(prediction: np.ndarray, reference: np.ndarray) -> float:
    pred = np.asarray(prediction, dtype=float)
    ref = np.asarray(reference, dtype=float)
    if pred.size < 2 or ref.size < 2:
        return 1.0
    pred_std = float(np.std(pred))
    ref_std = float(np.std(ref))
    if pred_std <= 1e-12 or ref_std <= 1e-12:
        return 1.0 if np.allclose(pred, ref, atol=1e-9, rtol=1e-6) else 0.0
    corr = float(np.corrcoef(pred, ref)[0, 1])
    if not math.isfinite(corr):
        return 0.0
    return max(0.0, min(corr, 1.0))


def _build_route_quality_metrics(
    route_frame: pd.DataFrame,
    *,
    scale: float,
    rise_label: str,
    bias_label: str,
    quality_label: str,
) -> pd.DataFrame:
    if route_frame.empty:
        return pd.DataFrame(columns=["设备", rise_label, bias_label, quality_label])

    design = _build_quality_design_matrix(route_frame)
    coefficients, *_ = np.linalg.lstsq(design, route_frame["TargetQuality"].to_numpy(float), rcond=None)
    working = route_frame.copy()
    working["ConvertedQuality"] = design @ coefficients
    working["ReferenceQuality"] = working.groupby("PointTagNorm")["ConvertedQuality"].transform("median")
    working["ErrorQuality"] = (working["ConvertedQuality"] - working["ReferenceQuality"]) * float(scale)

    rows: list[dict[str, Any]] = []
    for analyzer, analyzer_frame in working.groupby("Analyzer", sort=True):
        prediction = analyzer_frame["ConvertedQuality"].to_numpy(float) * float(scale)
        reference = analyzer_frame["ReferenceQuality"].to_numpy(float) * float(scale)
        rise = int(round(float(np.sqrt(np.mean(np.square(analyzer_frame["ErrorQuality"].to_numpy(float)))))))
        bias = int(round(float(analyzer_frame["ErrorQuality"].mean())))
        rows.append(
            {
                "设备": analyzer,
                rise_label: rise,
                bias_label: bias,
                quality_label: round(_quality_score(prediction, reference), 2),
            }
        )

    return pd.DataFrame(rows)


def _combine_quality_metrics(gas_row: Optional[dict[str, Any]], water_row: Optional[dict[str, Any]]) -> tuple[Optional[int], Optional[int], Optional[float]]:
    if not gas_row or not water_row:
        return None, None, None

    gas_rise = _safe_float(gas_row.get("气路RISE ppm"))
    gas_bias = _safe_float(gas_row.get("气路Bias ppm"))
    gas_quality = _safe_float(gas_row.get("气路R²"))
    water_rise = _safe_float(water_row.get("水路RISE ppm(等效)"))
    water_bias = _safe_float(water_row.get("水路Bias ppm(等效)"))
    water_quality = _safe_float(water_row.get("水路R²"))
    if None in {gas_rise, gas_bias, gas_quality, water_rise, water_bias, water_quality}:
        return None, None, None

    combined_rise = int(round(math.sqrt((gas_rise**2 + water_rise**2) / 2.0)))
    bias_sum = float(gas_bias + water_bias)
    if abs(bias_sum) <= 1e-12:
        combined_bias = 0
    else:
        combined_bias = int(round(math.copysign((abs(gas_bias) + abs(water_bias)) / 2.0, bias_sum)))
    combined_quality = round(float((gas_quality + water_quality) / 2.0), 2)
    return combined_rise, combined_bias, combined_quality


def _quality_focus_reason(detail_row: dict[str, Any], conclusion: str) -> str:
    gas_rise = abs(_safe_float(detail_row.get("气路RISE ppm")) or 0.0)
    gas_bias = abs(_safe_float(detail_row.get("气路Bias ppm")) or 0.0)
    gas_quality = _safe_float(detail_row.get("气路R²")) or 0.0
    water_rise = abs(_safe_float(detail_row.get("水路RISE ppm(等效)")) or 0.0)
    water_bias = abs(_safe_float(detail_row.get("水路Bias ppm(等效)")) or 0.0)
    water_quality = _safe_float(detail_row.get("水路R²")) or 0.0
    gas_score = gas_rise + gas_bias + (1.0 - gas_quality) * 100.0
    water_score = water_rise + water_bias + (1.0 - water_quality) * 100.0
    route_name = "气路" if gas_score >= water_score else "水路"
    return f"{route_name}{'异常' if conclusion == '建议剔除' else '偏移较大'}"


def _build_quality_notes_frame(
    summary_frame: pd.DataFrame,
    *,
    summary: pd.DataFrame,
    detail: pd.DataFrame,
) -> pd.DataFrame:
    sources = _collect_source_names(summary_frame)
    lines: list[str | None] = [
        "新一轮 8 台设备数据质量分析说明",
        None,
        "本次使用文件：",
    ]
    if sources:
        lines.extend([f"{idx}) {name}" for idx, name in enumerate(sources, start=1)])
    else:
        lines.append("1) 本次运行结果直接汇总生成")
    lines.extend(
        [
            "",
            "计算口径（这次和前两版不同）：",
            "• 按 GA01~GA08 各自 sheet 逐台读取，不再把气路/水路直接混在一起取均值。",
            "• 气路只用 CO2 相关响应（R_CO2、P、Temp、R_H2O）建公共转换模型。",
            "• 水路只用 H2O 相关响应（R_H2O、P、Temp、R_CO2）建公共转换模型。",
            "• 在每个共同点位上，对 8 台设备取中位数作为本轮参考，再计算每台相对中位数的偏差。",
            "• 主表中的 RISE ppm / Bias ppm 是综合指标：",
            "  - 气路直接按 ppm 统计；",
            f"  - 水路按固定 {_QUALITY_WATER_EQ_SCALE:g} 倍折算到 0–1000 ppm 等效尺度，再与气路合并；",
            "  - 合并时，RISE 用 RMS，Bias 用带方向的绝对均值，避免气路/水路互相抵消。",
            "",
            "建议优先关注：",
        ]
    )

    focus = detail.merge(summary.loc[:, ["设备", "结论"]], on="设备", how="left")
    flagged = focus[
        focus["结论"].isin(["建议剔除", _QUALITY_ATTENTION])
        & (
            (focus["气路RISE ppm"].fillna(0).abs() >= 100)
            | (focus["水路RISE ppm(等效)"].fillna(0).abs() >= 100)
            | (focus["气路Bias ppm"].fillna(0).abs() >= 80)
            | (focus["水路Bias ppm(等效)"].fillna(0).abs() >= 80)
        )
    ].copy()
    if flagged.empty:
        lines.append("• 本轮未发现需要额外关注的设备")
    else:
        for row in flagged.sort_values("设备").head(3).to_dict(orient="records"):
            lines.append(f"• {row['设备']}：{_quality_focus_reason(row, str(row.get('结论') or ''))}")
    return pd.DataFrame({"说明": lines})


def build_quality_analysis_bundle(summary_frame: pd.DataFrame) -> QualityAnalysisBundle:
    empty_summary = pd.DataFrame(columns=_PREFIT_QUALITY_COLUMNS)
    empty_detail = pd.DataFrame(columns=_QUALITY_DETAIL_COLUMNS)
    if summary_frame.empty:
        return QualityAnalysisBundle(summary=empty_summary, detail=empty_detail, notes=pd.DataFrame({"说明": []}))

    gas_frame = _prepare_quality_route_frame(
        summary_frame,
        phase_key="co2",
        target_column="ppm_CO2_Tank",
        primary_column="R_CO2",
        auxiliary_column="R_H2O",
    )
    water_frame = _prepare_quality_route_frame(
        summary_frame,
        phase_key="h2o",
        target_column="ppm_H2O_Dew",
        primary_column="R_H2O",
        auxiliary_column="R_CO2",
    )

    gas_metrics = _build_route_quality_metrics(
        gas_frame,
        scale=1.0,
        rise_label="气路RISE ppm",
        bias_label="气路Bias ppm",
        quality_label="气路R²",
    )
    water_metrics = _build_route_quality_metrics(
        water_frame,
        scale=_QUALITY_WATER_EQ_SCALE,
        rise_label="水路RISE ppm(等效)",
        bias_label="水路Bias ppm(等效)",
        quality_label="水路R²",
    )
    gas_lookup = {row["设备"]: row for row in gas_metrics.to_dict(orient="records")}
    water_lookup = {row["设备"]: row for row in water_metrics.to_dict(orient="records")}

    device_ids = _load_runtime_analyzer_ids(summary_frame)
    analyzers = sorted(
        {
            str(value or "").strip().upper()
            for value in pd.concat(
                [summary_frame.get("Analyzer", pd.Series(dtype=str)), gas_metrics.get("设备", pd.Series(dtype=str)), water_metrics.get("设备", pd.Series(dtype=str))],
                ignore_index=True,
            )
            if str(value or "").strip()
        }
    )

    summary_rows: list[dict[str, Any]] = []
    detail_rows: list[dict[str, Any]] = []
    for analyzer in analyzers:
        gas_row = dict(gas_lookup.get(analyzer, {}))
        water_row = dict(water_lookup.get(analyzer, {}))
        combined_rise, combined_bias, combined_quality = _combine_quality_metrics(gas_row or None, water_row or None)
        detail_row = {
            "设备": analyzer,
            "气路RISE ppm": gas_row.get("气路RISE ppm"),
            "气路Bias ppm": gas_row.get("气路Bias ppm"),
            "气路R²": gas_row.get("气路R²"),
            "水路RISE ppm(等效)": water_row.get("水路RISE ppm(等效)"),
            "水路Bias ppm(等效)": water_row.get("水路Bias ppm(等效)"),
            "水路R²": water_row.get("水路R²"),
            "RISE ppm": combined_rise,
            "Bias ppm": combined_bias,
            "R²质量": combined_quality,
        }
        detail_rows.append(detail_row)
        summary_rows.append(
            {
                "设备": analyzer,
                "ID": _normalize_device_id(device_ids.get(analyzer, "")),
                "RISE ppm": combined_rise,
                "Bias ppm": combined_bias,
                "R²质量": combined_quality,
                "结论": _prefit_quality_conclusion(combined_rise, combined_bias, combined_quality),
            }
        )

    summary = pd.DataFrame(summary_rows, columns=_PREFIT_QUALITY_COLUMNS)
    detail = pd.DataFrame(detail_rows, columns=_QUALITY_DETAIL_COLUMNS)
    notes = _build_quality_notes_frame(summary_frame, summary=summary, detail=detail)
    return QualityAnalysisBundle(summary=summary, detail=detail, notes=notes)


def _relative_error(errors: pd.Series, truth: pd.Series) -> pd.Series:
    truth_abs = truth.abs()
    rel = pd.Series([math.nan] * len(truth), index=truth.index, dtype=float)
    mask = truth_abs > 1e-12
    rel.loc[mask] = errors.loc[mask] / truth.loc[mask] * 100.0
    return rel


def _score_summary(r2: float, rmse_pct: float) -> tuple[str, str, str, str]:
    if rmse_pct <= 1.0 and r2 >= 0.95:
        return ("拟合可用", "误差水平稳定", "建议采用", "绿色")
    if rmse_pct <= 3.0 and r2 >= 0.90:
        return ("谨慎采用", "误差可接受", "视业务确认", "黄色")
    return ("暂不建议", "误差偏大", "暂不采用", "红色")


def _collect_source_names(summary_frame: pd.DataFrame) -> list[str]:
    if "SourcePath" not in summary_frame.columns:
        return []
    source_series = summary_frame["SourcePath"].dropna().astype(str).map(str.strip)
    return sorted({Path(value).name for value in source_series if value})


def _build_notes_frame(summary_frame: pd.DataFrame, *, temperature_key: str, pressure_key: str) -> pd.DataFrame:
    sources = _collect_source_names(summary_frame)
    if sources:
        source_note = f"{len(sources)} 份分析仪汇总合并，覆盖 {', '.join(sources)}"
    else:
        source_note = "本次运行结果直接汇总生成"

    rows = [
        {"说明项": "数据来源", "说明内容": source_note},
        {"说明项": "统计口径", "说明内容": "CO2 只使用 PointPhase=气路；H2O 使用全部水路点 + 气路 -20/-10/0°C 全部点 + 10°C 仅 0ppm"},
        {"说明项": "温度列", "说明内容": f"拟合温度列使用 {temperature_key}"},
        {"说明项": "压力列", "说明内容": f"拟合压力列使用 {pressure_key}"},
        {"说明项": "数据质量分析", "说明内容": f"设备质量分析按气路/水路分别建立公共二次转换模型，水路再按固定 {_QUALITY_WATER_EQ_SCALE:g} 倍折算到 0–1000 ppm 等效尺度后与气路合并"},
        {"说明项": "修正原因", "说明内容": "按确认后的固定规则补入 H2O 所需干气点"},
        {"说明项": "算法", "说明内容": "ratio_poly_rt_p 全量拟合 + 系数回代验证 + 逐点对账分析"},
        {"说明项": "颜色含义", "说明内容": "绿色=建议采用；黄色=视业务确认；红色=暂不采用"},
    ]
    return pd.DataFrame(rows)


def _apply_summary_fill(output_path: Path) -> None:
    workbook = load_workbook(output_path)
    try:
        sheet = workbook["汇总"]
        header_cells = {str(cell.value or "").strip(): cell.column for cell in sheet[1]}
        suggestion_col = header_cells.get("综合建议")
        if suggestion_col is None:
            workbook.save(output_path)
            return
        for row_idx in range(2, sheet.max_row + 1):
            suggestion = str(sheet.cell(row_idx, suggestion_col).value or "").strip()
            fill = _SUMMARY_FILL_MAP.get(suggestion)
            if fill is None:
                continue
            for col_idx in range(1, sheet.max_column + 1):
                sheet.cell(row_idx, col_idx).fill = fill
        metadata_headers = {"分析仪", "气体", "数据范围"}
        for sheet_name in ("简化系数", "原始系数"):
            coeff_sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
            if coeff_sheet is None:
                continue
            coefficient_columns = [
                index
                for index, cell in enumerate(coeff_sheet[1], start=1)
                if str(cell.value or "").strip()
                and str(cell.value or "").strip() not in metadata_headers
                and not str(cell.value or "").strip().endswith("_term")
            ]
            for row_idx in range(2, coeff_sheet.max_row + 1):
                for col_idx in coefficient_columns:
                    coeff_sheet.cell(row_idx, col_idx).number_format = _COEFF_DISPLAY_NUMBER_FORMAT
        workbook.save(output_path)
    finally:
        workbook.close()


def build_analyzer_summary_frame(
    results: Sequence[SamplingResult],
    *,
    expected_analyzers: Optional[Sequence[str]] = None,
    reference_on_aligned_rows: bool = True,
) -> pd.DataFrame:
    grouped: Dict[tuple[str, int, str], list[SamplingResult]] = {}
    point_groups: Dict[tuple[int, str], list[SamplingResult]] = {}
    for result in results:
        analyzer_id = str(result.analyzer_id).strip().upper()
        route = str(result.point.route).strip().lower()
        key = (analyzer_id, int(result.point.index), route)
        grouped.setdefault(key, []).append(result)
        point_groups.setdefault((int(result.point.index), route), []).append(result)

    expected_ids = sorted(
        {
            str(item or "").strip().upper()
            for item in list(expected_analyzers or [])
            if str(item or "").strip()
        }
    )
    if not expected_ids:
        expected_ids = sorted({str(result.analyzer_id).strip().upper() for result in results if str(result.analyzer_id).strip()})

    rows: list[dict[str, Any]] = []
    for num, ((analyzer_id, point_index, route), samples) in enumerate(sorted(grouped.items()), start=1):
        first = samples[0]
        point = first.point
        point_title = _build_point_title(first)
        point_tag = _build_point_tag(point_index, route)
        point_samples = point_groups.get((point_index, route), [])
        point_by_analyzer: Dict[str, list[SamplingResult]] = {}
        for point_sample in point_samples:
            point_by_analyzer.setdefault(str(point_sample.analyzer_id).strip().upper(), []).append(point_sample)
        point_present = {name for name, values in point_by_analyzer.items() if values}
        point_usable = {
            name
            for name, values in point_by_analyzer.items()
            if any(_sample_is_usable(sample) for sample in values)
        }
        point_expected = expected_ids or sorted(point_present)
        usable_alignment_keys = {_sample_alignment_key(sample) for sample in samples if _sample_is_usable(sample)}
        if reference_on_aligned_rows and usable_alignment_keys:
            reference_pool = [sample for sample in point_samples if _sample_alignment_key(sample) in usable_alignment_keys]
        else:
            reference_pool = list(point_samples)
        reference_samples = _dedupe_reference_samples(reference_pool or point_samples or samples)
        usable_samples = [sample for sample in samples if _sample_is_usable(sample)]
        aligned_reference_samples = usable_samples if usable_samples else list(samples)
        thermometer_mean = _mean(sample.thermometer_temp_c for sample in aligned_reference_samples)
        chamber_reference_mean = _mean(sample.temperature_c for sample in aligned_reference_samples)
        analyzer_temp_mean = _mean(sample.analyzer_chamber_temp_c for sample in samples)
        reference_temp = thermometer_mean if thermometer_mean is not None else (
            chamber_reference_mean if chamber_reference_mean is not None else analyzer_temp_mean
        )
        reference_pressure_hpa = _mean(sample.pressure_gauge_hpa for sample in reference_samples)
        if reference_pressure_hpa is None:
            reference_pressure_hpa = _mean(sample.pressure_hpa for sample in reference_samples)
        analyzer_pressure_kpa = _mean(
            sample.analyzer_pressure_kpa
            if sample.analyzer_pressure_kpa is not None
            else (
                sample.pressure_gauge_hpa / 10.0
                if sample.pressure_gauge_hpa is not None
                else (sample.pressure_hpa / 10.0 if sample.pressure_hpa is not None else None)
            )
            for sample in reference_samples
        )
        dew_pressure_hpa = reference_pressure_hpa
        if dew_pressure_hpa is None:
            dew_pressure_hpa = _mean(
                sample.pressure_hpa
                if sample.pressure_hpa is not None
                else (sample.analyzer_pressure_kpa * 10.0 if sample.analyzer_pressure_kpa is not None else None)
                for sample in reference_samples
            )
        missing_analyzers = sorted(set(point_expected) - point_present)
        unusable_analyzers = sorted(point_present - point_usable)
        expected_count = len(point_expected)
        row = {
            "Analyzer": analyzer_id,
            "NUM": num,
            "PointRow": point_index,
            "PointPhase": _phase_display(route),
            "PointTag": point_tag,
            "PointTitle": point_title,
            "TempSet": point.temperature_c,
            "HgenTempSet": point.temperature_c if route == "h2o" else None,
            "HgenRhSet": point.humidity_pct if route == "h2o" else None,
            "Temp": reference_temp,
            "Dew": _mean(sample.dew_point_c for sample in reference_samples),
            "P": reference_pressure_hpa,
            "ppm_CO2_Tank": point.co2_ppm,
            "PressureTarget": point.pressure_hpa,
            "AnalyzerCoverage": f"{len(point_usable)}/{expected_count}" if expected_count else "0/0",
            "UsableAnalyzers": len(point_usable),
            "ExpectedAnalyzers": expected_count,
            "PointIntegrity": "完整" if usable_samples else "缺失",
            "MissingAnalyzers": ",".join(missing_analyzers),
            "UnusableAnalyzers": ",".join(unusable_analyzers),
            "PointIntegrity": _point_integrity_text(expected_count=expected_count, present=point_present, usable=point_usable),
            "ValidFrames": len(usable_samples),
            "TotalFrames": len(samples),
            "FrameStatus": "全部可用" if len(usable_samples) == len(samples) else ("部分可用" if usable_samples else "无可用帧"),
            "ppm_H2O_Dew": _dewpoint_to_h2o_mmol_per_mol(_mean(sample.dew_point_c for sample in reference_samples), dew_pressure_hpa),
            "ppm_CO2": _mean(sample.co2_ppm for sample in usable_samples),
            "ppm_H2O": _mean(sample.h2o_mmol for sample in usable_samples),
            "R_CO2": _mean((sample.co2_ratio_f if sample.co2_ratio_f is not None else sample.co2_ratio_raw) for sample in usable_samples),
            "R_CO2_dev": _std((sample.co2_ratio_f if sample.co2_ratio_f is not None else sample.co2_ratio_raw) for sample in usable_samples),
            "R_H2O": _mean((sample.h2o_ratio_f if sample.h2o_ratio_f is not None else sample.h2o_ratio_raw) for sample in usable_samples),
            "R_H2O_dev": _std((sample.h2o_ratio_f if sample.h2o_ratio_f is not None else sample.h2o_ratio_raw) for sample in usable_samples),
            "Raw_REF": _mean(sample.ref_signal for sample in usable_samples),
            "Raw_CO2": _mean(sample.co2_signal for sample in usable_samples),
            "Raw_H2O": _mean(sample.h2o_signal for sample in usable_samples),
            "T1": analyzer_temp_mean if analyzer_temp_mean is not None else reference_temp,
            "T2": _mean(sample.case_temp_c for sample in usable_samples),
            "BAR": analyzer_pressure_kpa,
            "P_fit": analyzer_pressure_kpa,
            "EnvTempC": reference_temp if reference_temp is not None else point.temperature_c,
            "PhaseKey": _phase_key(route),
            "AnalyzerMeanMode": "primary_or_first_usable",
            "ReferenceAlignedRows": bool(reference_on_aligned_rows),
            "ReferencePressureHpa": reference_pressure_hpa,
            "ReferenceThermometerTempC": thermometer_mean,
        }
        rows.append(row)

    return pd.DataFrame(rows)


def load_summary_workbook_rows(paths: Sequence[str | Path]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for raw_path in paths:
        path = Path(raw_path)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            frame = pd.read_csv(path)
            for row in frame.to_dict(orient="records"):
                payload = dict(row)
                payload["Analyzer"] = str(payload.get("Analyzer") or "").strip().upper()
                payload["SourcePath"] = str(path)
                payload["EnvTempC"] = _env_temp_from_row(payload)
                payload["PhaseKey"] = _phase_key(payload.get("PointPhase"))
                rows.append(payload)
            continue

        workbook = pd.read_excel(path, sheet_name=None)
        for analyzer, frame in workbook.items():
            for row in frame.to_dict(orient="records"):
                payload = dict(row)
                payload["Analyzer"] = str(payload.get("Analyzer") or analyzer or "").strip().upper()
                payload["SourcePath"] = str(path)
                payload["EnvTempC"] = _env_temp_from_row(payload)
                payload["PhaseKey"] = _phase_key(payload.get("PointPhase"))
                rows.append(payload)
    return pd.DataFrame(rows)


def select_corrected_fit_rows(
    frame: pd.DataFrame,
    *,
    gas: str,
    selection: H2OSummarySelectionConfig,
    temperature_key: str,
) -> pd.DataFrame:
    working = frame.copy()
    working["PhaseKey"] = working.get("PhaseKey", working.get("PointPhase", "")).map(_phase_key)
    working["FitTemp"] = pd.to_numeric(working.get(temperature_key), errors="coerce")
    working["EnvTempC"] = working.apply(lambda row: _env_temp_from_row(dict(row)), axis=1)
    working["ppm_CO2_Tank_num"] = pd.to_numeric(working.get("ppm_CO2_Tank"), errors="coerce")

    gas_key = str(gas or "").strip().lower()
    if gas_key == "co2":
        return working[working["PhaseKey"] == "co2"].copy()
    if gas_key != "h2o":
        raise ValueError(f"Unsupported gas: {gas}")

    phase_mask = working["PhaseKey"] == "h2o" if selection.include_h2o_phase else False
    co2_mask = working["PhaseKey"] == "co2"

    co2_temp_mask = False
    for target in selection.include_co2_temp_groups_c:
        co2_temp_mask = co2_temp_mask | (
            co2_mask & (working["EnvTempC"] - float(target)).abs().le(float(selection.temp_tolerance_c))
        )

    zero_temp_mask = False
    if selection.include_co2_zero_ppm_rows:
        zero_ppm_mask = working["ppm_CO2_Tank_num"].sub(float(selection.co2_zero_ppm_target)).abs().le(
            float(selection.co2_zero_ppm_tolerance)
        )
        for target in selection.include_co2_zero_ppm_temp_groups_c:
            zero_temp_mask = zero_temp_mask | (
                co2_mask
                & zero_ppm_mask
                & (working["EnvTempC"] - float(target)).abs().le(float(selection.temp_tolerance_c))
            )

    return working[phase_mask | co2_temp_mask | zero_temp_mask].copy()


def _available_fit_gases(frame: pd.DataFrame) -> tuple[str, ...]:
    if frame.empty:
        return ()

    if "PhaseKey" in frame.columns:
        phase_series = frame["PhaseKey"]
    elif "PointPhase" in frame.columns:
        phase_series = frame["PointPhase"]
    else:
        return ()

    phases = {
        _phase_key(value)
        for value in phase_series.dropna().tolist()
        if str(value or "").strip()
    }
    ordered = []
    for gas in ("co2", "h2o"):
        if gas in phases:
            ordered.append(gas)
    return tuple(ordered)


def build_ratio_poly_fit_records(
    summary_frame: pd.DataFrame,
    *,
    coeff_cfg: CoefficientsConfig,
) -> list[RatioPolyFitRecord]:
    if summary_frame.empty:
        return []

    temperature_key = str(coeff_cfg.report_temperature_key or "Temp")
    pressure_key = str(coeff_cfg.report_pressure_key or "P_fit")
    records: list[RatioPolyFitRecord] = []

    for analyzer in sorted(str(value) for value in summary_frame["Analyzer"].dropna().unique()):
        analyzer_frame = summary_frame[summary_frame["Analyzer"] == analyzer].copy()
        for gas in _available_fit_gases(analyzer_frame):
            selected = select_corrected_fit_rows(
                analyzer_frame,
                gas=gas,
                selection=coeff_cfg.h2o_summary_selection,
                temperature_key=temperature_key,
            )
            if selected.empty:
                continue

            column_cfg = coeff_cfg.summary_columns[gas]
            fit_frame = selected.copy()
            if pressure_key == "P_fit" and "P_fit" not in fit_frame.columns and "BAR" in fit_frame.columns:
                fit_frame["P_fit"] = pd.to_numeric(fit_frame["BAR"], errors="coerce")
            fit_frame["FitTemp"] = pd.to_numeric(fit_frame.get(temperature_key), errors="coerce")

            try:
                result = fit_ratio_poly_rt_p(
                    fit_frame.to_dict(orient="records"),
                    gas=gas,
                    target_key=column_cfg.target,
                    ratio_keys=(column_cfg.ratio,),
                    temp_keys=("FitTemp",),
                    pressure_keys=(pressure_key,),
                    pressure_scale=float(column_cfg.pressure_scale),
                    ratio_degree=int(coeff_cfg.ratio_degree),
                    temperature_offset_c=float(coeff_cfg.temperature_offset_c),
                    add_intercept=bool(coeff_cfg.add_intercept),
                    simplify_coefficients=bool(coeff_cfg.simplify_coefficients),
                    simplification_method=str(coeff_cfg.simplification_method),
                    target_digits=int(coeff_cfg.target_digits),
                    min_samples=0,
                    train_ratio=0.7,
                    val_ratio=0.15,
                    random_seed=42,
                    shuffle_dataset=True,
                    evaluation_bins=_range_bins(gas),
                )
            except Exception:
                continue

            residuals = pd.DataFrame(result.residuals).copy()
            residuals["index"] = range(len(residuals))
            residuals["Y_true"] = pd.to_numeric(residuals["target"], errors="coerce")
            residuals["Y_pred_orig"] = pd.to_numeric(residuals["prediction_original"], errors="coerce")
            residuals["Y_pred_simple"] = pd.to_numeric(residuals["prediction_simplified"], errors="coerce")
            residuals["error_orig"] = pd.to_numeric(residuals["error_original"], errors="coerce")
            residuals["error_simple"] = pd.to_numeric(residuals["error_simplified"], errors="coerce")
            residuals["rel_error_orig_pct"] = _relative_error(residuals["error_orig"], residuals["Y_true"])
            residuals["rel_error_simple_pct"] = _relative_error(residuals["error_simple"], residuals["Y_true"])
            residuals["pred_diff"] = residuals["Y_pred_simple"] - residuals["Y_pred_orig"]
            residuals["abs_error_orig"] = residuals["error_orig"].abs()
            residuals["abs_error_simple"] = residuals["error_simple"].abs()
            residuals["abs_pred_diff"] = residuals["pred_diff"].abs()

            metrics_orig = compute_metrics(
                residuals["Y_true"].to_numpy(float),
                residuals["Y_pred_orig"].to_numpy(float),
            )
            metrics_simple = compute_metrics(
                residuals["Y_true"].to_numpy(float),
                residuals["Y_pred_simple"].to_numpy(float),
            )
            rmse_change = float(metrics_simple["RMSE"] - metrics_orig["RMSE"])
            rmse_pct = 0.0 if metrics_orig["RMSE"] == 0.0 else abs(rmse_change / metrics_orig["RMSE"] * 100.0)
            effect, effect_note, suggestion, suggestion_note = _score_summary(
                float(metrics_simple["R2"]),
                rmse_pct,
            )
            records.append(
                RatioPolyFitRecord(
                    analyzer=analyzer,
                    gas=gas,
                    column_cfg=column_cfg,
                    selected_frame=selected,
                    result=result,
                    residuals=residuals,
                    metrics_original={key: float(value) for key, value in metrics_orig.items()},
                    metrics_simplified={key: float(value) for key, value in metrics_simple.items()},
                    effect=effect,
                    effect_note=effect_note,
                    suggestion=suggestion,
                    suggestion_note=suggestion_note,
                )
            )
    return records


def _build_summary_row(
    record: RatioPolyFitRecord,
    *,
    coeff_cfg: CoefficientsConfig,
    temperature_key: str,
    pressure_key: str,
) -> dict[str, Any]:
    rmse_change = float(record.metrics_simplified["RMSE"] - record.metrics_original["RMSE"])
    rmse_pct = 0.0 if record.metrics_original["RMSE"] == 0.0 else rmse_change / record.metrics_original["RMSE"] * 100.0
    residuals = record.residuals

    return {
        "分析仪": record.analyzer,
        "气体": record.gas.upper(),
        "数据范围": _REPORT_SCOPE_TEXT,
        "总样本数": int(len(record.selected_frame)),
        "参与拟合样本数": int(record.result.n),
        "目标列": record.column_cfg.target,
        "比值列(R)": record.column_cfg.ratio,
        "温度列(T)": temperature_key,
        "压力列(P)": pressure_key,
        "原始方程RMSE": float(record.metrics_original["RMSE"]),
        "原始方程R2": float(record.metrics_original["R2"]),
        "原始方程Bias": float(record.metrics_original["Bias"]),
        "原始方程MaxError": float(record.metrics_original["MaxError"]),
        "简化方程RMSE": float(record.metrics_simplified["RMSE"]),
        "简化方程R2": float(record.metrics_simplified["R2"]),
        "简化方程Bias": float(record.metrics_simplified["Bias"]),
        "简化方程MaxError": float(record.metrics_simplified["MaxError"]),
        "RMSE变化量": rmse_change,
        "RMSE相对变化(%)": rmse_pct,
        "简化方程MAE": float((record.result.stats or {}).get("mae_simplified", residuals["abs_error_simple"].mean())),
        "简化方程最大绝对误差": float(residuals["abs_error_simple"].max()),
        "原始与简化预测最大差值": float(residuals["abs_pred_diff"].max()),
        "原始与简化预测平均差值": float(residuals["pred_diff"].mean()),
        "拟合效果评价": record.effect,
        "拟合效果摘要": record.effect_note,
        "综合建议": record.suggestion,
        "建议说明": record.suggestion_note,
        "模型": record.result.model,
        "系数简化方法": str(coeff_cfg.simplification_method),
        "系数有效数字": int(coeff_cfg.target_digits),
        "R多项式阶数": int(coeff_cfg.ratio_degree),
    }


def _build_coefficient_row(record: RatioPolyFitRecord, *, simplified: bool) -> dict[str, Any]:
    row: dict[str, Any] = {
        "分析仪": record.analyzer,
        "气体": record.gas.upper(),
        "数据范围": _REPORT_SCOPE_TEXT,
    }
    coefficients = record.result.simplified_coefficients if simplified else record.result.original_coefficients
    for name in record.result.feature_names:
        row[name] = coefficients.get(name)
        row[f"{name}_term"] = record.result.feature_terms.get(name, "")
    return row


def _build_point_table(
    record: RatioPolyFitRecord,
    *,
    temperature_key: str,
    pressure_key: str,
) -> pd.DataFrame:
    residuals = record.residuals
    return pd.DataFrame(
        {
            "分析仪": record.analyzer,
            "气体": record.gas.upper(),
            "数据范围": _REPORT_SCOPE_TEXT,
            "index": residuals["index"],
            "点位行号": residuals["PointRow"],
            "点位相位": residuals["PointPhase"],
            "点位标签": residuals["PointTag"],
            "点位标题": residuals["PointTitle"],
            "Y_true": residuals["Y_true"],
            "Y_pred_orig": residuals["Y_pred_orig"],
            "error_orig": residuals["error_orig"],
            "rel_error_orig_pct": residuals["rel_error_orig_pct"],
            "Y_pred_simple": residuals["Y_pred_simple"],
            "error_simple": residuals["error_simple"],
            "rel_error_simple_pct": residuals["rel_error_simple_pct"],
            "pred_diff": residuals["pred_diff"],
            "abs_error_orig": residuals["abs_error_orig"],
            "abs_error_simple": residuals["abs_error_simple"],
            "abs_pred_diff": residuals["abs_pred_diff"],
            "R": pd.to_numeric(residuals["R"], errors="coerce"),
            temperature_key: pd.to_numeric(residuals["T_c"], errors="coerce"),
            pressure_key: pd.to_numeric(residuals["P"], errors="coerce"),
        }
    )


def _build_range_table(record: RatioPolyFitRecord) -> pd.DataFrame:
    range_table = analyze_by_range(
        record.residuals["Y_true"].to_numpy(float),
        record.residuals["error_orig"].to_numpy(float),
        record.residuals["error_simple"].to_numpy(float),
        bins=_range_bins(record.gas),
    )
    range_table.insert(0, "数据范围", _REPORT_SCOPE_TEXT)
    range_table.insert(0, "气体", record.gas.upper())
    range_table.insert(0, "分析仪", record.analyzer)
    return range_table


def _build_topn_tables(record: RatioPolyFitRecord) -> list[pd.DataFrame]:
    residuals = record.residuals

    def _top_table(column: str, label: str) -> pd.DataFrame:
        top = residuals.sort_values(column, ascending=False).head(_TOPN_LIMIT).reset_index(drop=True).copy()
        return pd.DataFrame(
            {
                "分析仪": record.analyzer,
                "气体": record.gas.upper(),
                "排序维度": label,
                "rank": range(1, len(top) + 1),
                "index": top["index"],
                "点位行号": top["PointRow"],
                "点位相位": top["PointPhase"],
                "点位标签": top["PointTag"],
                "Y_true": top["Y_true"],
                "Y_pred_orig": top["Y_pred_orig"],
                "Y_pred_simple": top["Y_pred_simple"],
                "error_orig": top["error_orig"],
                "error_simple": top["error_simple"],
                "pred_diff": top["pred_diff"],
                "abs_error_orig": top["abs_error_orig"],
                "abs_error_simple": top["abs_error_simple"],
                "abs_pred_diff": top["abs_pred_diff"],
            }
        )

    return [
        _top_table("abs_error_orig", "原始误差"),
        _top_table("abs_error_simple", "简化误差"),
        _top_table("abs_pred_diff", "预测差值"),
    ]


def export_ratio_poly_report(
    results: Sequence[SamplingResult],
    *,
    out_dir: Path,
    coeff_cfg: CoefficientsConfig,
    expected_analyzers: Optional[Sequence[str]] = None,
    reference_on_aligned_rows: bool = True,
) -> Optional[Path]:
    summary_frame = build_analyzer_summary_frame(
        results,
        expected_analyzers=expected_analyzers,
        reference_on_aligned_rows=reference_on_aligned_rows,
    )
    return export_ratio_poly_report_from_summary_frame(summary_frame, out_dir=out_dir, coeff_cfg=coeff_cfg)


def export_ratio_poly_report_from_summary_files(
    summary_paths: Sequence[str | Path],
    *,
    out_dir: Path,
    coeff_cfg: CoefficientsConfig,
    expected_analyzers: Optional[Sequence[str]] = None,
    reference_on_aligned_rows: bool = True,
) -> Optional[Path]:
    summary_frame = load_summary_workbook_rows(summary_paths)
    return export_ratio_poly_report_from_summary_frame(
        summary_frame,
        out_dir=out_dir,
        coeff_cfg=coeff_cfg,
    )


def export_ratio_poly_report_from_summary_frame(
    summary_frame: pd.DataFrame,
    *,
    out_dir: Path,
    coeff_cfg: CoefficientsConfig,
) -> Optional[Path]:
    if summary_frame.empty:
        return None

    temperature_key = str(coeff_cfg.report_temperature_key or "Temp")
    pressure_key = str(coeff_cfg.report_pressure_key or "P_fit")
    records = build_ratio_poly_fit_records(summary_frame, coeff_cfg=coeff_cfg)
    if not records:
        return None

    notes = _build_notes_frame(summary_frame, temperature_key=temperature_key, pressure_key=pressure_key)
    quality_bundle = build_quality_analysis_bundle(summary_frame)
    summary_rows = [
        _build_summary_row(record, coeff_cfg=coeff_cfg, temperature_key=temperature_key, pressure_key=pressure_key)
        for record in records
    ]
    simplified_rows = [_build_coefficient_row(record, simplified=True) for record in records]
    original_rows = [_build_coefficient_row(record, simplified=False) for record in records]
    point_tables = [_build_point_table(record, temperature_key=temperature_key, pressure_key=pressure_key) for record in records]
    range_tables = [_build_range_table(record) for record in records]
    topn_tables = [table for record in records for table in _build_topn_tables(record)]
    download_rows = _build_download_plan_rows(records)

    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / str(coeff_cfg.report_output_name or "calibration_coefficients.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        notes.to_excel(writer, sheet_name="说明", index=False)
        quality_bundle.summary.to_excel(writer, sheet_name=_PREFIT_QUALITY_SHEET, index=False)
        quality_bundle.detail.to_excel(writer, sheet_name=_QUALITY_DETAIL_SHEET, index=False)
        quality_bundle.notes.to_excel(writer, sheet_name=_QUALITY_NOTES_SHEET, index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="汇总", index=False)
        pd.DataFrame(simplified_rows).to_excel(writer, sheet_name="简化系数", index=False)
        pd.DataFrame(original_rows).to_excel(writer, sheet_name="原始系数", index=False)
        pd.concat(point_tables, ignore_index=True).to_excel(writer, sheet_name="逐点对账", index=False)
        pd.concat(range_tables, ignore_index=True).to_excel(writer, sheet_name="分区间分析", index=False)
        pd.concat(topn_tables, ignore_index=True).to_excel(writer, sheet_name="误差TopN", index=False)
        pd.DataFrame(download_rows).to_excel(writer, sheet_name="download_plan", index=False)
        summary_frame.to_excel(writer, sheet_name="分析仪汇总", index=False)

    device_eval_path = out_dir / "设备评估.xlsx"
    with pd.ExcelWriter(device_eval_path, engine="openpyxl") as writer:
        quality_bundle.summary.to_excel(writer, sheet_name=_QUALITY_WORKBOOK_SUMMARY_SHEET, index=False)
        quality_bundle.detail.to_excel(writer, sheet_name=_QUALITY_WORKBOOK_DETAIL_SHEET, index=False)
        quality_bundle.notes.to_excel(writer, sheet_name=_QUALITY_WORKBOOK_NOTES_SHEET, index=False)

    _apply_summary_fill(output_path)
    return output_path
