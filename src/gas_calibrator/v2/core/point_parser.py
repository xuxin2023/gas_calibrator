"""
点位文件解析器。

支持 `xlsx/csv/json` 三种格式，并兼容旧版点位列名。
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

from ..domain.pressure_selection import (
    effective_pressure_mode,
    normalize_pressure_mode,
    normalize_pressure_selection_token,
    pressure_target_label,
)
from ..exceptions import DataParseError
from ..utils import as_float, as_int
from .models import CalibrationPoint


@dataclass(frozen=True)
class PointFilter:
    """点位过滤器。"""

    temperature_min: Optional[float] = None
    temperature_max: Optional[float] = None
    co2_ppm_values: list[float] = field(default_factory=list)
    routes: list[str] = field(default_factory=list)
    max_points: Optional[int] = None
    point_indices: list[int] = field(default_factory=list)


@dataclass(frozen=True)
class TemperatureGroup:
    """温度组。"""

    temperature_c: float
    points: list[CalibrationPoint] = field(default_factory=list)


@dataclass(frozen=True)
class LegacyExcelPointLoader:
    """Adapter that reuses the V1 Excel point loader for legacy point tables."""

    missing_pressure_policy: str = "require"
    carry_forward_h2o: bool = False

    def load(self, path: str | Path) -> List[CalibrationPoint]:
        from gas_calibrator.data.points import load_points_from_excel

        legacy_points = load_points_from_excel(
            path,
            missing_pressure_policy=self.missing_pressure_policy,
            carry_forward_h2o=self.carry_forward_h2o,
        )
        return [self._to_v2_point(point) for point in legacy_points]

    @staticmethod
    def looks_like_legacy_sheet(rows: Sequence[Sequence[Any]]) -> bool:
        if len(rows) < 2:
            return False

        first_row = LegacyExcelPointLoader._non_empty_cells(rows[0])
        second_row = LegacyExcelPointLoader._non_empty_cells(rows[1])
        if len(first_row) != 1 or len(second_row) < 2:
            return False

        first_text = first_row[0]
        if "normalized calibration points" in first_text:
            return True

        return (
            "temp" not in first_text
            and "temperature" not in first_text
            and any("temp" in value for value in second_row)
            and any("co2" in value for value in second_row)
        )

    @staticmethod
    def _non_empty_cells(row: Sequence[Any]) -> list[str]:
        values: list[str] = []
        for value in row:
            text = str(value or "").strip().lower()
            if text:
                values.append(text)
        return values

    @staticmethod
    def _to_v2_point(point: Any) -> CalibrationPoint:
        route = "h2o" if bool(getattr(point, "is_h2o_point", False)) else "co2"
        raw_mode = normalize_pressure_mode(
            getattr(point, "_pressure_mode", getattr(point, "pressure_mode", ""))
        )
        raw_token = normalize_pressure_selection_token(
            getattr(point, "_pressure_selection_token", getattr(point, "pressure_selection_token", ""))
        )
        pressure_hpa = point.target_pressure_hpa
        resolved_mode = effective_pressure_mode(
            pressure_hpa=pressure_hpa,
            pressure_mode=raw_mode,
            pressure_selection_token=raw_token,
        )
        if resolved_mode == "ambient_open":
            pressure_hpa = None
        return CalibrationPoint(
            index=int(point.index),
            temperature_c=float(point.temp_chamber_c),
            co2_ppm=point.co2_ppm,
            humidity_pct=point.hgen_rh_pct,
            pressure_hpa=pressure_hpa,
            route=route,
            humidity_generator_temp_c=point.hgen_temp_c,
            dewpoint_c=point.dewpoint_c,
            h2o_mmol=point.h2o_mmol,
            raw_h2o=point.raw_h2o,
            co2_group=point.co2_group,
            cylinder_nominal_ppm=getattr(point, "cylinder_nominal_ppm", None),
            pressure_mode=resolved_mode,
            pressure_target_label=pressure_target_label(
                pressure_hpa=pressure_hpa,
                pressure_mode=resolved_mode,
                pressure_selection_token=raw_token,
                explicit_label=getattr(point, "_pressure_target_label", getattr(point, "pressure_target_label", None)),
            ),
            pressure_selection_token=raw_token,
        )


class PointParser:
    """V2 点位解析器。"""

    def __init__(self, legacy_excel_loader: Optional[LegacyExcelPointLoader] = None) -> None:
        self.legacy_excel_loader = legacy_excel_loader or LegacyExcelPointLoader()

    def parse(
        self,
        path: str | Path,
        point_filter: Optional[PointFilter] = None,
    ) -> List[CalibrationPoint]:
        """
        解析点位文件。

        Args:
            path: 文件路径。
            point_filter: 可选过滤器。

        Returns:
            解析并过滤后的点位列表。
        """
        file_path = Path(path)
        if not file_path.exists():
            raise DataParseError(str(file_path), reason="points file does not exist")

        suffix = file_path.suffix.lower()
        if suffix == ".json":
            points = self._parse_json(file_path)
        elif suffix == ".csv":
            points = self._parse_csv(file_path)
        elif suffix == ".xlsx":
            points = self._parse_xlsx(file_path)
        else:
            raise DataParseError(str(file_path), reason=f"unsupported points file format: {suffix}")

        if point_filter is not None:
            return self.filter(points, point_filter)
        return points

    def filter(
        self,
        points: Sequence[CalibrationPoint],
        point_filter: PointFilter,
    ) -> List[CalibrationPoint]:
        """按过滤器筛选点位。"""
        filtered: list[CalibrationPoint] = []
        route_values = {str(value).strip().lower() for value in point_filter.routes}
        point_indices = set(point_filter.point_indices)
        co2_values = {float(value) for value in point_filter.co2_ppm_values}

        for point in points:
            if point_filter.temperature_min is not None and point.temperature_c < point_filter.temperature_min:
                continue
            if point_filter.temperature_max is not None and point.temperature_c > point_filter.temperature_max:
                continue
            if route_values and point.route.lower() not in route_values:
                continue
            if point_indices and point.index not in point_indices:
                continue
            if co2_values:
                if point.co2_ppm is None or float(point.co2_ppm) not in co2_values:
                    continue
            filtered.append(point)
            if point_filter.max_points is not None and len(filtered) >= point_filter.max_points:
                break

        return filtered

    def group_by_temperature(
        self,
        points: Sequence[CalibrationPoint],
    ) -> List[TemperatureGroup]:
        """按温度分组，保持原始出现顺序。"""
        groups: list[TemperatureGroup] = []
        by_temperature: Dict[float, list[CalibrationPoint]] = {}
        order: list[float] = []

        for point in points:
            key = float(point.temp_chamber_c)
            if key not in by_temperature:
                by_temperature[key] = []
                order.append(key)
            by_temperature[key].append(point)

        for key in order:
            groups.append(TemperatureGroup(temperature_c=key, points=list(by_temperature[key])))
        return groups

    def _parse_json(self, path: Path) -> List[CalibrationPoint]:
        with path.open("r", encoding="utf-8") as fh:
            payload = json.load(fh)

        rows = payload.get("points", payload) if isinstance(payload, dict) else payload
        if not isinstance(rows, list):
            raise DataParseError(str(path), reason="JSON payload must be a list or contain 'points'")
        return [self._row_to_point(index, row) for index, row in enumerate(rows, start=1)]

    def _parse_csv(self, path: Path) -> List[CalibrationPoint]:
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            return [self._row_to_point(index, row) for index, row in enumerate(reader, start=1)]

    def _parse_xlsx(self, path: Path) -> List[CalibrationPoint]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise DataParseError(str(path), reason=f"openpyxl unavailable: {exc}") from exc

        workbook = load_workbook(path, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if not rows:
            return []

        if self.legacy_excel_loader.looks_like_legacy_sheet(rows):
            try:
                return self.legacy_excel_loader.load(path)
            except Exception as exc:
                raise DataParseError(str(path), reason=f"failed to parse legacy points sheet: {exc}") from exc

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        data_rows: list[dict[str, Any]] = []
        for row in rows[1:]:
            payload = {
                headers[column_index]: row[column_index]
                for column_index in range(min(len(headers), len(row)))
                if headers[column_index]
            }
            if payload:
                data_rows.append(payload)

        return [self._row_to_point(index, row) for index, row in enumerate(data_rows, start=1)]

    def _row_to_point(self, index: int, row: Any) -> CalibrationPoint:
        if not isinstance(row, dict):
            raise DataParseError(f"point #{index}", reason="point row must be a mapping")

        normalized = self._normalize_row_keys(row)
        point_index = as_int(
            self._pick_first(normalized, "index", "point_index", "point", "序号"),
            default=index,
            allow_none=False,
        )
        temperature_c = as_float(
            self._pick_first(
                normalized,
                "temperature",
                "temperature_c",
                "temp_chamber_c",
                "temp",
                "t_chamber",
                "t",
                "温度",
            ),
            allow_none=False,
        )
        if temperature_c is None:
            raise DataParseError(f"point #{index}", reason="temperature is required")

        co2_ppm = as_float(self._pick_first(normalized, "co2_ppm", "co2", "CO2"))
        humidity_pct = as_float(
            self._pick_first(
                normalized,
                "humidity",
                "humidity_pct",
                "hgen_rh_pct",
                "humidity_generator_rh_pct",
                "rh",
                "rh_pct",
                "湿度",
            )
        )
        pressure_hpa = as_float(
            self._pick_first(
                normalized,
                "pressure_hpa",
                "target_pressure_hpa",
                "pressure",
                "气压",
            )
        )
        raw_pressure_mode = normalize_pressure_mode(
            self._pick_first(normalized, "pressure_mode", "pressure_state", "pressure_route_mode")
        )
        raw_pressure_token = normalize_pressure_selection_token(
            self._pick_first(
                normalized,
                "pressure_selection_token",
                "pressure_token",
                "selected_pressure_point",
            )
        )
        humidity_generator_temp_c = as_float(
            self._pick_first(
                normalized,
                "humidity_generator_temp_c",
                "hgen_temp_c",
                "humidity_temp_c",
                "humidity_generator_temperature_c",
                "temp_hgen_c",
                "hgen_temp",
            )
        )
        dewpoint_c = as_float(self._pick_first(normalized, "dewpoint_c", "dew_point_c", "dewpoint", "td"))
        h2o_mmol = as_float(self._pick_first(normalized, "h2o_mmol", "h2o_mmol_target", "h2o"))
        raw_h2o = self._pick_first(normalized, "raw_h2o", "h2o_text", "humidity_text")
        co2_group = self._pick_first(normalized, "co2_group", "group", "co2_source_group")
        cylinder_nominal_ppm = as_float(
            self._pick_first(
                normalized,
                "cylinder_nominal_ppm",
                "nominal_ppm",
                "bottle_ppm",
                "cylinder_ppm",
            )
        )
        route = str(self._pick_first(normalized, "route", "路线") or "").strip().lower()

        resolved_pressure_mode = effective_pressure_mode(
            pressure_hpa=pressure_hpa,
            pressure_mode=raw_pressure_mode,
            pressure_selection_token=raw_pressure_token,
        )
        if resolved_pressure_mode == "ambient_open":
            pressure_hpa = None
        resolved_pressure_label = pressure_target_label(
            pressure_hpa=pressure_hpa,
            pressure_mode=resolved_pressure_mode,
            pressure_selection_token=raw_pressure_token,
            explicit_label=self._pick_first(normalized, "pressure_target_label", "pressure_label"),
        )

        if route not in {"h2o", "co2"}:
            has_h2o_payload = (
                humidity_pct is not None
                or humidity_generator_temp_c is not None
                or dewpoint_c is not None
                or h2o_mmol is not None
            )
            route = "h2o" if has_h2o_payload and co2_ppm is None else "co2"

        return CalibrationPoint(
            index=int(point_index),
            temperature_c=float(temperature_c),
            co2_ppm=co2_ppm,
            humidity_pct=humidity_pct,
            pressure_hpa=pressure_hpa,
            route=route,
            humidity_generator_temp_c=humidity_generator_temp_c,
            dewpoint_c=dewpoint_c,
            h2o_mmol=h2o_mmol,
            raw_h2o=None if raw_h2o is None else str(raw_h2o),
            co2_group=None if co2_group is None else str(co2_group).strip() or None,
            cylinder_nominal_ppm=cylinder_nominal_ppm,
            pressure_mode=resolved_pressure_mode,
            pressure_target_label=resolved_pressure_label,
            pressure_selection_token=raw_pressure_token,
        )

    @staticmethod
    def _normalize_row_keys(row: Dict[str, Any]) -> Dict[str, Any]:
        normalized: Dict[str, Any] = {}
        for key, value in row.items():
            text = str(key or "").strip()
            if not text:
                continue
            normalized[text] = value
            normalized[text.lower()] = value
        return normalized

    @staticmethod
    def _pick_first(row: Dict[str, Any], *keys: str) -> Any:
        for key in keys:
            if key in row and row[key] not in ("", None):
                return row[key]
        return None
