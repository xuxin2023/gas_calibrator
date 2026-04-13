from __future__ import annotations

import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from ..h2o_summary_selection import normalize_h2o_summary_selection
from ..coefficients.fit_ratio_poly import fit_ratio_poly_rt_p
from ..coefficients.model_metrics import compute_metrics
from ..coefficients.prediction_analysis import analyze_by_range

_TEMP_RE = re.compile(r"([-+]?\d+(?:\.\d+)?)°C")


@dataclass
class CorrectedFitBundle:
    analyzer: str
    gas: str
    data_scope: str
    selected_frame: pd.DataFrame
    summary_row: Dict[str, Any]
    simplified_row: Dict[str, Any]
    original_row: Dict[str, Any]
    point_table: pd.DataFrame
    range_table: pd.DataFrame
    top_error_orig: pd.DataFrame
    top_error_simple: pd.DataFrame
    top_pred_diff: pd.DataFrame


def _safe_float(value: Any) -> float | None:
    try:
        numeric = float(value)
    except Exception:
        return None
    if math.isfinite(numeric):
        return numeric
    return None


def _phase_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"气路", "co2"}:
        return "co2"
    if text in {"水路", "h2o"}:
        return "h2o"
    return text


def _env_temp_from_row(row: Mapping[str, Any]) -> float | None:
    for key in ("PointTitle", "PointTag"):
        match = _TEMP_RE.search(str(row.get(key) or ""))
        if match:
            return _safe_float(match.group(1))
    return None


def _resolve_h2o_selection(selection: Mapping[str, Any] | None) -> Dict[str, Any]:
    return normalize_h2o_summary_selection(selection)


def _anchor_threshold_for_row(
    row: Mapping[str, Any],
    selection_cfg: Mapping[str, Any],
) -> tuple[float | None, float | None]:
    env_temp = _safe_float(row.get("EnvTempC"))
    tolerance_c = float(selection_cfg.get("temp_tolerance_c", 0.6) or 0.6)
    keyed_limits = dict(selection_cfg.get("co2_zero_ppm_anchor_max_ppm_h2o_dew_by_temp_c") or {})
    for raw_key, raw_value in keyed_limits.items():
        target_temp = _safe_float(raw_key)
        limit = _safe_float(raw_value)
        if target_temp is None or limit is None or env_temp is None:
            continue
        if abs(env_temp - target_temp) <= tolerance_c:
            return limit, target_temp
    return _safe_float(selection_cfg.get("co2_zero_ppm_anchor_max_ppm_h2o_dew_default")), env_temp


def select_corrected_fit_rows_with_diagnostics(
    frame: pd.DataFrame,
    *,
    gas: str,
    temperature_key: str = "Temp",
    selection: Mapping[str, Any] | None = None,
) -> Dict[str, pd.DataFrame]:
    selection_cfg = _resolve_h2o_selection(selection)
    working = frame.copy()
    if "PhaseKey" in working.columns:
        working["PhaseKey"] = working["PhaseKey"].map(_phase_key)
    else:
        working["PhaseKey"] = working.get("PointPhase", "").map(_phase_key)
    working["FitTemp"] = pd.to_numeric(working.get(temperature_key), errors="coerce")
    working["EnvTempC"] = pd.to_numeric(working.get("EnvTempC"), errors="coerce")
    working["ppm_CO2_Tank_num"] = pd.to_numeric(working.get("ppm_CO2_Tank"), errors="coerce")
    working["ppm_H2O_Dew_num"] = pd.to_numeric(working.get("ppm_H2O_Dew"), errors="coerce")
    working["SelectionOrigin"] = ""

    gas_key = str(gas or "").strip().lower()
    if gas_key == "co2":
        selected = working[working["PhaseKey"] == "co2"].copy()
        selected.loc[:, "SelectionOrigin"] = "co2_phase"
        return {
            "selected_frame": selected,
            "h2o_anchor_gate_hits": pd.DataFrame(),
        }

    if gas_key != "h2o":
        raise ValueError(f"Unsupported gas: {gas}")

    phase_mask = (
        (working["PhaseKey"] == "h2o")
        if selection_cfg["include_h2o_phase"]
        else pd.Series(False, index=working.index)
    )
    co2_mask = working["PhaseKey"] == "co2"

    all_co2_temp_mask = pd.Series(False, index=working.index)
    for target in selection_cfg["include_co2_temp_groups_c"]:
        all_co2_temp_mask = all_co2_temp_mask | (
            co2_mask & (working["EnvTempC"] - float(target)).abs().le(float(selection_cfg["temp_tolerance_c"]))
        )

    zero_temp_mask = pd.Series(False, index=working.index)
    if selection_cfg["include_co2_zero_ppm_rows"]:
        zero_ppm_mask = working["ppm_CO2_Tank_num"].sub(float(selection_cfg["co2_zero_ppm_target"])).abs().le(
            float(selection_cfg["co2_zero_ppm_tolerance"])
        )
        for target in selection_cfg["include_co2_zero_ppm_temp_groups_c"]:
            zero_temp_mask = zero_temp_mask | (
                co2_mask
                & zero_ppm_mask
                & (working["EnvTempC"] - float(target)).abs().le(float(selection_cfg["temp_tolerance_c"]))
            )

    zero_anchor_pass_mask = zero_temp_mask.copy()
    anchor_gate_rows: List[Dict[str, Any]] = []
    if bool(selection_cfg.get("co2_zero_ppm_anchor_quality_gate_enabled", True)):
        candidate_rows = working.loc[zero_temp_mask].copy()
        passed_index = set(candidate_rows.index.tolist())
        for idx, row in candidate_rows.iterrows():
            observed_h2o_dew = _safe_float(row.get("ppm_H2O_Dew_num"))
            require_h2o_dew = bool(selection_cfg.get("co2_zero_ppm_anchor_require_h2o_dew", True))
            threshold, matched_temp = _anchor_threshold_for_row(row, selection_cfg)
            reason = ""
            if observed_h2o_dew is None and require_h2o_dew:
                reason = "anchor_h2o_dew_missing"
            elif threshold is not None and observed_h2o_dew is not None and observed_h2o_dew > threshold:
                reason = "anchor_h2o_dew_above_limit"
            if reason:
                passed_index.discard(idx)
                anchor_gate_rows.append(
                    {
                        "Analyzer": row.get("Analyzer"),
                        "PointRow": row.get("PointRow"),
                        "PointPhase": row.get("PointPhase"),
                        "PointTag": row.get("PointTag"),
                        "PointTitle": row.get("PointTitle"),
                        "EnvTempC": _safe_float(row.get("EnvTempC")),
                        "Temp": _safe_float(row.get("FitTemp")),
                        "ppm_CO2_Tank": _safe_float(row.get("ppm_CO2_Tank_num")),
                        "ppm_H2O_Dew": observed_h2o_dew,
                        "SelectionOrigin": "co2_zero_ppm_anchor",
                        "GateReason": reason,
                        "GateThresholdPpmH2ODew": threshold,
                        "GateMatchedTempC": matched_temp,
                        "SourceFile": row.get("SourceFile"),
                        "SourceStamp": row.get("SourceStamp"),
                    }
                )
        zero_anchor_pass_mask = working.index.to_series().isin(passed_index) & zero_temp_mask

    selected_mask = phase_mask | all_co2_temp_mask | zero_anchor_pass_mask
    working.loc[phase_mask, "SelectionOrigin"] = "h2o_phase"
    working.loc[all_co2_temp_mask, "SelectionOrigin"] = "co2_temp_group"
    working.loc[zero_anchor_pass_mask, "SelectionOrigin"] = "co2_zero_ppm_anchor"
    selected = working[selected_mask].copy()
    gate_hits = pd.DataFrame(anchor_gate_rows)
    return {
        "selected_frame": selected,
        "h2o_anchor_gate_hits": gate_hits,
    }


def _describe_h2o_selection(selection: Mapping[str, Any]) -> str:
    parts: List[str] = []
    if bool(selection.get("include_h2o_phase", True)):
        parts.append("全部水路点")
    co2_groups = list(selection.get("include_co2_temp_groups_c") or [])
    if co2_groups:
        group_text = "/".join(f"{float(value):g}" for value in co2_groups)
        parts.append(f"气路 {group_text}°C 全部点")
    if bool(selection.get("include_co2_zero_ppm_rows", True)):
        zero_groups = list(selection.get("include_co2_zero_ppm_temp_groups_c") or [])
        if zero_groups:
            zero_text = "/".join(f"{float(value):g}" for value in zero_groups)
            parts.append(f"气路 {zero_text}°C 的 0ppm 锚点")
    return " + ".join(parts) if parts else "未启用 H2O 选点"


def load_summary_workbook_rows(paths: Sequence[str | Path]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for raw_path in paths:
        path = Path(raw_path)
        workbook = pd.read_excel(path, sheet_name=None)
        stamp_match = re.search(r"(\d{8}_\d{6})", path.name)
        source_stamp = stamp_match.group(1) if stamp_match else ""
        for analyzer, frame in workbook.items():
            for row in frame.to_dict(orient="records"):
                payload = dict(row)
                payload["Analyzer"] = str(analyzer or "").strip() or str(payload.get("Analyzer") or "").strip()
                payload["SourceFile"] = path.name
                payload["SourcePath"] = str(path)
                payload["SourceStamp"] = source_stamp
                payload["PhaseKey"] = _phase_key(payload.get("PointPhase"))
                payload["EnvTempC"] = _env_temp_from_row(payload)
                rows.append(payload)
    return pd.DataFrame(rows)


def select_corrected_fit_rows(
    frame: pd.DataFrame,
    *,
    gas: str,
    temperature_key: str = "Temp",
    selection: Mapping[str, Any] | None = None,
) -> pd.DataFrame:
    return select_corrected_fit_rows_with_diagnostics(
        frame,
        gas=gas,
        temperature_key=temperature_key,
        selection=selection,
    )["selected_frame"].copy()


def _relative_error(errors: pd.Series, truth: pd.Series) -> pd.Series:
    truth_abs = truth.abs()
    rel = pd.Series([math.nan] * len(truth), index=truth.index, dtype=float)
    mask = truth_abs > 1e-12
    rel.loc[mask] = errors.loc[mask] / truth.loc[mask] * 100.0
    return rel


def _score_summary(rmse: float, r2: float, rmse_pct: float) -> tuple[str, str, str, str]:
    if rmse_pct <= 1.0 and r2 >= 0.95:
        return ("建议采用", "误差水平稳定", "采用", "绿色")
    if rmse_pct <= 3.0 and r2 >= 0.90:
        return ("谨慎采用", "误差可接受", "视业务确认", "黄色")
    return ("暂不建议", "误差偏大", "暂不采用", "红色")


def _normalize_data_scope(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return "按水路纠正规则"
    stripped = text.replace("?", "").strip()
    if not stripped:
        return "按水路纠正规则"
    return text


def _range_bins(gas: str) -> List[float]:
    return [0.0, 200.0, 400.0, 800.0, 1200.0, 2000.0] if gas == "co2" else [0.0, 5.0, 10.0, 20.0, 40.0]


def _build_bundle(
    analyzer: str,
    gas: str,
    data_scope: str,
    selected_frame: pd.DataFrame,
    *,
    target_key: str,
    ratio_key: str,
    temperature_key: str,
    pressure_key: str,
    coeff_cfg: Mapping[str, Any],
) -> CorrectedFitBundle:
    fit_frame = selected_frame.copy()
    if pressure_key not in fit_frame.columns and "BAR" in fit_frame.columns and pressure_key == "P_fit":
        fit_frame["P_fit"] = pd.to_numeric(fit_frame["BAR"], errors="coerce")

    result = fit_ratio_poly_rt_p(
        fit_frame.to_dict(orient="records"),
        gas=gas,
        target_key=target_key,
        ratio_keys=(ratio_key,),
        temp_keys=("FitTemp",),
        pressure_keys=(pressure_key,),
        pressure_scale=float(coeff_cfg.get("pressure_scale", 1.0)),
        ratio_degree=int(coeff_cfg.get("ratio_degree", 3)),
        temperature_offset_c=float(coeff_cfg.get("temperature_offset_c", 273.15)),
        add_intercept=bool(coeff_cfg.get("add_intercept", True)),
        simplify_coefficients=bool(coeff_cfg.get("simplify_coefficients", True)),
        simplification_method=str(coeff_cfg.get("simplification_method", "column_norm")),
        target_digits=int(coeff_cfg.get("target_digits", 6)),
        min_samples=int(coeff_cfg.get("min_samples", 0) or 0),
        train_ratio=float(coeff_cfg.get("train_ratio", 0.7)),
        val_ratio=float(coeff_cfg.get("val_ratio", 0.15)),
        random_seed=int(coeff_cfg.get("random_seed", 42)),
        shuffle_dataset=bool(coeff_cfg.get("shuffle_dataset", True)),
        evaluation_bins=_range_bins(gas),
    )

    residuals = pd.DataFrame(result.residuals).copy()
    residuals["index"] = range(len(residuals))
    residuals["Y_true"] = pd.to_numeric(residuals["target"], errors="coerce")
    residuals["Y_pred_orig"] = pd.to_numeric(residuals["prediction_original"], errors="coerce")
    residuals["error_orig"] = pd.to_numeric(residuals["error_original"], errors="coerce")
    residuals["Y_pred_simple"] = pd.to_numeric(residuals["prediction_simplified"], errors="coerce")
    residuals["error_simple"] = pd.to_numeric(residuals["error_simplified"], errors="coerce")
    residuals["rel_error_orig_pct"] = _relative_error(residuals["error_orig"], residuals["Y_true"])
    residuals["rel_error_simple_pct"] = _relative_error(residuals["error_simple"], residuals["Y_true"])
    residuals["pred_diff"] = residuals["Y_pred_simple"] - residuals["Y_pred_orig"]
    residuals["abs_error_orig"] = residuals["error_orig"].abs()
    residuals["abs_error_simple"] = residuals["error_simple"].abs()
    residuals["abs_pred_diff"] = residuals["pred_diff"].abs()

    orig_metrics = compute_metrics(
        residuals["Y_true"].to_numpy(float),
        residuals["Y_pred_orig"].to_numpy(float),
    )
    simple_metrics = compute_metrics(
        residuals["Y_true"].to_numpy(float),
        residuals["Y_pred_simple"].to_numpy(float),
    )
    rmse_change = float(simple_metrics["RMSE"] - orig_metrics["RMSE"])
    rmse_pct = 0.0 if orig_metrics["RMSE"] == 0.0 else rmse_change / orig_metrics["RMSE"] * 100.0
    effect, effect_note, suggestion, suggestion_note = _score_summary(
        float(simple_metrics["RMSE"]),
        float(simple_metrics["R2"]),
        float(abs(rmse_pct)),
    )

    point_table = pd.DataFrame(
        {
            "分析仪": analyzer,
            "气体": gas.upper(),
            "数据范围": data_scope,
            "index": residuals["index"],
            "点位行号": residuals["PointRow"],
            "点位相位": residuals["PointPhase"],
            "点位标签": residuals["PointTag"],
            "点位标题": residuals["PointTitle"],
            "Y_true": residuals["Y_true"],
            "Y_pred_orig": residuals["Y_pred_orig"],
            "error_orig": residuals["error_orig"],
            "rel_error_orig_pct": residuals["rel_error_orig_pct"],
            "Y_pred_simple": residuals["Y_pred_simple"],
            "error_simple": residuals["error_simple"],
            "rel_error_simple_pct": residuals["rel_error_simple_pct"],
            "pred_diff": residuals["pred_diff"],
            "abs_error_orig": residuals["abs_error_orig"],
            "abs_error_simple": residuals["abs_error_simple"],
            "abs_pred_diff": residuals["abs_pred_diff"],
            "R": pd.to_numeric(residuals["R"], errors="coerce"),
            temperature_key: pd.to_numeric(residuals["T_c"], errors="coerce"),
            pressure_key: pd.to_numeric(residuals["P"], errors="coerce"),
        }
    )

    range_table = analyze_by_range(
        point_table["Y_true"].to_numpy(float),
        point_table["error_orig"].to_numpy(float),
        point_table["error_simple"].to_numpy(float),
        bins=_range_bins(gas),
    )
    range_table.insert(0, "数据范围", data_scope)
    range_table.insert(0, "气体", gas.upper())
    range_table.insert(0, "分析仪", analyzer)

    top_common = ["分析仪", "气体", "排序维度", "rank", "index", "点位行号", "点位相位", "点位标签", "Y_true", "Y_pred_orig", "Y_pred_simple", "error_orig", "error_simple", "pred_diff", "abs_error_orig", "abs_error_simple", "abs_pred_diff"]

    def _top_table(column: str, dimension: str) -> pd.DataFrame:
        top = point_table.sort_values(column, ascending=False).head(10).reset_index(drop=True).copy()
        top["rank"] = range(1, len(top) + 1)
        top["排序维度"] = dimension
        top["气体"] = gas.upper()
        top["分析仪"] = analyzer
        return top[top_common]

    feature_terms = {name: result.feature_terms.get(name, "") for name in result.feature_names}
    simplified_row: Dict[str, Any] = {"分析仪": analyzer, "气体": gas.upper(), "数据范围": data_scope}
    original_row: Dict[str, Any] = {"分析仪": analyzer, "气体": gas.upper(), "数据范围": data_scope}
    for name in result.feature_names:
        simplified_row[name] = result.simplified_coefficients.get(name)
        simplified_row[f"{name}_term"] = feature_terms.get(name, "")
        original_row[name] = result.original_coefficients.get(name)
        original_row[f"{name}_term"] = feature_terms.get(name, "")

    summary_row = {
        "分析仪": analyzer,
        "气体": gas.upper(),
        "数据范围": data_scope,
        "总样本数": int(len(fit_frame)),
        "参与拟合样本数": int(result.n),
        "目标列": target_key,
        "比值列(R)": ratio_key,
        "温度列(T)": temperature_key,
        "压力列(P)": pressure_key,
        "原始方程RMSE": float(orig_metrics["RMSE"]),
        "原始方程R2": float(orig_metrics["R2"]),
        "原始方程Bias": float(orig_metrics["Bias"]),
        "原始方程MaxError": float(orig_metrics["MaxError"]),
        "简化方程RMSE": float(simple_metrics["RMSE"]),
        "简化方程R2": float(simple_metrics["R2"]),
        "简化方程Bias": float(simple_metrics["Bias"]),
        "简化方程MaxError": float(simple_metrics["MaxError"]),
        "RMSE变化量": rmse_change,
        "RMSE相对变化(%)": rmse_pct,
        "简化方程MAE": float((result.stats or {}).get("mae_simplified", point_table["abs_error_simple"].mean())),
        "简化方程最大绝对误差": float(point_table["abs_error_simple"].max()),
        "原始与简化预测最大差值": float(point_table["abs_pred_diff"].max()),
        "原始与简化预测平均差值": float(point_table["pred_diff"].mean()),
        "拟合效果评价": effect,
        "拟合效果摘要": effect_note,
        "综合建议": suggestion,
        "建议说明": suggestion_note,
        "模型": result.model,
        "系数简化方法": str(coeff_cfg.get("simplification_method", "column_norm")),
        "系数有效数字": int(coeff_cfg.get("target_digits", 6)),
        "R多项式阶数": int(coeff_cfg.get("ratio_degree", 3)),
    }

    return CorrectedFitBundle(
        analyzer=analyzer,
        gas=gas.upper(),
        data_scope=data_scope,
        selected_frame=fit_frame,
        summary_row=summary_row,
        simplified_row=simplified_row,
        original_row=original_row,
        point_table=point_table,
        range_table=range_table,
        top_error_orig=_top_table("abs_error_orig", "原始误差"),
        top_error_simple=_top_table("abs_error_simple", "简化误差"),
        top_pred_diff=_top_table("abs_pred_diff", "预测差值"),
    )


def build_corrected_water_points_report(
    summary_paths: Sequence[str | Path],
    *,
    output_path: str | Path,
    coeff_cfg: Mapping[str, Any] | None = None,
    temperature_key: str = "Temp",
    data_scope: str = "按水路纠正规则",
) -> Dict[str, Any]:
    cfg = dict(coeff_cfg or {})
    normalized_scope = _normalize_data_scope(data_scope)
    h2o_selection = _resolve_h2o_selection(cfg.get("h2o_summary_selection"))
    frame = load_summary_workbook_rows(summary_paths)
    analyzers = sorted(str(value) for value in frame["Analyzer"].dropna().unique())

    bundles: List[CorrectedFitBundle] = []
    h2o_anchor_gate_frames: List[pd.DataFrame] = []
    for analyzer in analyzers:
        analyzer_frame = frame[frame["Analyzer"] == analyzer].copy()
        for gas, target_key, ratio_key in (
            ("co2", "ppm_CO2_Tank", "R_CO2"),
            ("h2o", "ppm_H2O_Dew", "R_H2O"),
        ):
            selection_result = select_corrected_fit_rows_with_diagnostics(
                analyzer_frame,
                gas=gas,
                temperature_key=temperature_key,
                selection=h2o_selection,
            )
            selected = selection_result["selected_frame"]
            gate_hits = pd.DataFrame(selection_result.get("h2o_anchor_gate_hits", pd.DataFrame()))
            if not gate_hits.empty:
                gate_hits.insert(0, "DataScope", normalized_scope)
                gate_hits.insert(0, "Gas", "H2O")
                h2o_anchor_gate_frames.append(gate_hits)
            if selected.empty:
                continue
            bundles.append(
                _build_bundle(
                    analyzer,
                    gas,
                    normalized_scope,
                    selected,
                    target_key=target_key,
                    ratio_key=ratio_key,
                    temperature_key=temperature_key,
                    pressure_key="P_fit",
                    coeff_cfg={**cfg, "pressure_scale": 1.0},
                )
            )

    h2o_selected_frames: List[pd.DataFrame] = []
    for bundle in bundles:
        if str(bundle.gas or "").strip().lower() != "h2o":
            continue
        selected = bundle.selected_frame.copy()
        if selected.empty:
            continue
        selected.insert(0, "DataScope", bundle.data_scope)
        selected.insert(0, "Gas", bundle.gas)
        selected.insert(0, "Analyzer", bundle.analyzer)
        keep_columns = [
            column
            for column in (
                "Analyzer",
                "Gas",
                "DataScope",
                "SelectionOrigin",
                "PointRow",
                "PointPhase",
                "PointTag",
                "PointTitle",
                "EnvTempC",
                "Temp",
                "ppm_CO2_Tank",
                "ppm_H2O_Dew",
                "R_CO2",
                "R_H2O",
                "BAR",
                "SourceFile",
                "SourceStamp",
            )
            if column in selected.columns
        ]
        h2o_selected_frames.append(selected.loc[:, keep_columns])
    h2o_anchor_gate_df = (
        pd.concat(h2o_anchor_gate_frames, ignore_index=True)
        if h2o_anchor_gate_frames
        else pd.DataFrame()
    )

    summary_df = pd.DataFrame([bundle.summary_row for bundle in bundles])
    simplified_df = pd.DataFrame([bundle.simplified_row for bundle in bundles])
    original_df = pd.DataFrame([bundle.original_row for bundle in bundles])
    point_df = pd.concat([bundle.point_table for bundle in bundles], ignore_index=True)
    range_df = pd.concat([bundle.range_table for bundle in bundles], ignore_index=True)
    topn_df = pd.concat(
        [*(bundle.top_error_orig for bundle in bundles), *(bundle.top_error_simple for bundle in bundles), *(bundle.top_pred_diff for bundle in bundles)],
        ignore_index=True,
    )
    h2o_selected_df = pd.concat(h2o_selected_frames, ignore_index=True) if h2o_selected_frames else pd.DataFrame()

    note_rows = [
        {"说明项": "数据来源", "说明内容": f"{len(summary_paths)} 份分析仪汇总合并，覆盖 {', '.join(Path(path).name for path in summary_paths)}"},
        {"说明项": "统计口径", "说明内容": "CO2 只使用 PointPhase=气路；H2O 使用全部水路点 + 气路 -20/-10/0℃ 全部点 + 10℃ 仅 0ppm"},
        {"说明项": "温度列", "说明内容": f"拟合温度列使用 {temperature_key}"},
        {"说明项": "修正原因", "说明内容": "按历史反推规则修正 H2O 引用气路点范围"},
        {"说明项": "算法", "说明内容": "ratio_poly_rt_p 全量拟合 + 系数回代验证 + 逐点对账分析"},
        {"说明项": "颜色含义", "说明内容": "绿色=建议采用；黄色=谨慎采用；红色=暂不建议"},
    ]
    note_rows[1] = {
        "说明项": "统计口径",
        "说明内容": f"CO2 仅使用 PointPhase=气路；H2O 使用 {_describe_h2o_selection(h2o_selection)}",
    }
    note_rows.append(
        {
            "说明项": "H2O锚点门禁",
            "说明内容": "对 -20/-10/0°C 的 0ppm 气路锚点额外校验 ppm_H2O_Dew；命中后仅从 H2O 拟合剔除，不影响原始点位记录。",
        }
    )
    notes_df = pd.DataFrame(note_rows)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        notes_df.to_excel(writer, sheet_name="说明", index=False)
        summary_df.to_excel(writer, sheet_name="汇总", index=False)
        simplified_df.to_excel(writer, sheet_name="简化系数", index=False)
        original_df.to_excel(writer, sheet_name="原始系数", index=False)
        point_df.to_excel(writer, sheet_name="逐点对账", index=False)
        range_df.to_excel(writer, sheet_name="分区间分析", index=False)
        topn_df.to_excel(writer, sheet_name="误差TopN", index=False)

    if not h2o_selected_df.empty:
        with pd.ExcelWriter(output, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            h2o_selected_df.to_excel(writer, sheet_name="H2O锚点入选", index=False)
    if not h2o_anchor_gate_df.empty:
        with pd.ExcelWriter(output, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            h2o_anchor_gate_df.to_excel(writer, sheet_name="H2O锚点门禁", index=False)

    workbook = load_workbook(output)
    summary_sheet = workbook["汇总"]
    header_cells = {str(cell.value or "").strip(): cell.column for cell in summary_sheet[1]}
    suggestion_col = header_cells.get("综合建议")
    fill_map = {
        "采用": PatternFill(fill_type="solid", fgColor="C6EFCE"),
        "视业务确认": PatternFill(fill_type="solid", fgColor="FFF2CC"),
        "暂不采用": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    }
    if suggestion_col is not None:
        for row_idx in range(2, summary_sheet.max_row + 1):
            suggestion = str(summary_sheet.cell(row_idx, suggestion_col).value or "").strip()
            fill = fill_map.get(suggestion)
            if fill is None:
                continue
            for col_idx in range(1, summary_sheet.max_column + 1):
                summary_sheet.cell(row_idx, col_idx).fill = fill
    workbook.save(output)
    workbook.close()

    return {
        "output_path": output,
        "summary": summary_df,
        "simplified": simplified_df,
        "original": original_df,
        "points": point_df,
        "ranges": range_df,
        "topn": topn_df,
        "notes": notes_df,
        "h2o_selected_rows": h2o_selected_df,
        "h2o_anchor_gate_hits": h2o_anchor_gate_df,
    }
