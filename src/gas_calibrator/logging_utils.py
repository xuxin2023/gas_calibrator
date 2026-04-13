"""Run log persistence helpers."""

from __future__ import annotations

import csv
import os
import re
import tempfile
from datetime import datetime
from pathlib import Path
from statistics import mean, stdev
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .validation.dewpoint_flush_gate import (
    dewpoint_saturation_pressure_hpa as _shared_dewpoint_saturation_pressure_hpa,
    dewpoint_to_h2o_mmol_per_mol as _shared_dewpoint_to_h2o_mmol_per_mol,
)


def _utc_ts() -> str:
    """Return a local timestamp with millisecond precision for log rows."""
    return datetime.now().isoformat(timespec="milliseconds")


def _summary_text(value: Any, limit: int = 160) -> str:
    """Render a compact single-line summary suitable for CSV logs."""
    if value is None:
        return ""
    text = str(value).replace("\r", "\\r").replace("\n", "\\n").strip()
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."


def _safe_suffix(text: str) -> str:
    out = "".join(ch if ch.isalnum() or ch in {"_", "-"} else "_" for ch in str(text or "").strip().lower())
    while "__" in out:
        out = out.replace("__", "_")
    return out.strip("_")


def _save_workbook_atomic(wb: Workbook, path: Path) -> None:
    """Persist a workbook via temp file replacement to avoid transient 0-byte outputs."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(
        prefix=f".{path.stem}_",
        suffix=path.suffix,
        dir=str(path.parent),
    )
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def _save_csv_atomic(path: Path, fieldnames: List[str], rows: List[Dict[str, Any]]) -> None:
    """Persist a CSV via temp file replacement to avoid partial in-place rewrites."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(
        prefix=f".{path.stem}_",
        suffix=path.suffix,
        dir=str(path.parent),
    )
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        with tmp_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=list(fieldnames),
                extrasaction="ignore",
            )
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
            f.flush()
        os.replace(tmp_path, path)
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def _dewpoint_saturation_pressure_hpa(dewpoint_c: float) -> float:
    """Return saturation vapor pressure at dewpoint in hPa."""
    vapor_pressure = _shared_dewpoint_saturation_pressure_hpa(dewpoint_c)
    if vapor_pressure is None:
        raise ValueError("invalid dewpoint")
    return vapor_pressure


def _dewpoint_to_h2o_mmol_per_mol(dewpoint_c: Any, pressure_hpa: Any) -> Optional[float]:
    """Convert dewpoint and total pressure to water amount fraction in mmol/mol."""
    return _shared_dewpoint_to_h2o_mmol_per_mol(dewpoint_c, pressure_hpa)


_FIELD_LABELS = {
    "point_title": "点位标题",
    "run_id": "运行ID",
    "session_id": "会话ID",
    "device_id": "设备ID(追溯)",
    "gas_type": "气体类型",
    "step": "步骤",
    "point_no": "点位编号",
    "target_value": "目标值",
    "measured_value": "测量值",
    "sample_ts": "采样时间",
    "save_ts": "保存时间",
    "sample_due_ts": "采样目标时间",
    "sample_start_ts": "采样开始时间",
    "sample_end_ts": "采样结束时间",
    "window_start_ts": "窗口开始时间",
    "window_end_ts": "窗口结束时间",
    "sample_count": "窗口样本数",
    "stable_flag": "稳态标记",
    "mode_exit_attempted": "已尝试退出校准模式",
    "mode_exit_confirmed": "已确认退出校准模式",
    "rollback_attempted": "已尝试回滚",
    "rollback_confirmed": "已确认回滚",
    "overall_write_status": "写入总体状态",
    "overall_verify_status": "回读总体状态",
    "overall_rollback_status": "回滚总体状态",
    "sample_elapsed_ms": "采样耗时ms",
    "sample_lag_ms": "采样滞后ms",
    "sample_index": "样本序号",
    "point_phase": "流程阶段",
    "point_tag": "点位标签",
    "point_row": "校准点行号",
    "point_is_h2o": "是否水路点",
    "pressure_mode": "压力执行模式",
    "pressure_target_label": "压力目标标签",
    "route": "采样路线",
    "trace_stage": "流程阶段标签",
    "raw": "原始报文",
    "id": "设备ID",
    "mode": "模式",
    "mode2_field_count": "MODE2字段数",
    "frame_has_data": "分析仪有帧",
    "frame_usable": "分析仪可用帧",
    "frame_status": "分析仪帧状态",
    "frame_cache_ts": "分析仪缓存时间",
    "frame_cache_age_ms": "分析仪缓存年龄ms",
    "frame_rx_ts": "分析仪选中帧接收时间",
    "frame_rx_seq": "分析仪选中帧序号",
    "frame_anchor_delta_ms": "分析仪帧距锚点ms",
    "frame_anchor_side": "分析仪帧锚点侧",
    "frame_match_strategy": "分析仪帧匹配策略",
    "frame_stale": "分析仪帧是否陈旧",
    "frame_source": "分析仪帧来源",
    "frame_is_live": "分析仪帧是否实时",
    "co2_ppm": "二氧化碳浓度ppm",
    "h2o_mmol": "水浓度mmol每mol",
    "co2_density": "二氧化碳密度",
    "h2o_density": "水密度",
    "co2_ratio_f": "二氧化碳比值滤波后",
    "co2_ratio_raw": "二氧化碳比值原始值",
    "h2o_ratio_f": "水比值滤波后",
    "h2o_ratio_raw": "水比值原始值",
    "ref_signal": "参考信号",
    "co2_signal": "二氧化碳信号",
    "h2o_signal": "水信号",
    "pressure_kpa": "分析仪压力kPa",
    "pressure_hpa": "压力控制器压力hPa",
    "controller_pressure": "压力控制器压力hPa",
    "pace_sample_ts": "压力控制器采样时间",
    "pace_anchor_delta_ms": "压力控制器距锚点ms",
    "pace_output_state": "压力控制器输出状态",
    "pace_isolation_state": "压力控制器隔离状态",
    "pace_vent_status": "压力控制器通大气状态",
    "pressure_gauge_raw": "数字压力计原始值",
    "pressure_gauge_hpa": "数字压力计压力hPa",
    "gauge_pressure": "数字压力计压力hPa",
    "pressure_gauge_sample_ts": "数字压力计采样时间",
    "pressure_gauge_anchor_delta_ms": "数字压力计距锚点ms",
    "pressure_target_hpa": "目标压力hPa",
    "temp_set_c": "温箱设定温度C",
    "temp_chamber_c": "温箱目标温度C",
    "env_chamber_temp_c": "温度箱环境温度C",
    "env_chamber_rh_pct": "温度箱环境湿度%",
    "chamber_temp_c": "温度箱温度C",
    "chamber_rh_pct": "温度箱湿度%",
    "chamber_sample_ts": "温箱缓存采样时间",
    "chamber_cache_age_ms": "温箱缓存年龄ms",
    "thermometer_temp_c": "数字温度计温度C",
    "thermometer_sample_ts": "数字温度计缓存采样时间",
    "thermometer_cache_age_ms": "数字温度计缓存年龄ms",
    "case_temp_c": "机壳温度C",
    "dewpoint_c": "露点仪露点C",
    "dew_temp_c": "露点仪温度C",
    "dew_rh_pct": "露点仪湿度%",
    "dewpoint_live_c": "露点仪实时露点C",
    "dew_temp_live_c": "露点仪实时温度C",
    "dew_rh_live_pct": "露点仪实时湿度%",
    "dewpoint_live_sample_ts": "露点仪实时采样时间",
    "dewpoint_live_anchor_delta_ms": "露点仪实时距锚点ms",
    "dew_pressure_hpa": "封压前露点压力hPa",
    "fast_group_anchor_ts": "快采组锚点时间",
    "fast_group_start_ts": "快采组开始时间",
    "fast_group_end_ts": "快采组结束时间",
    "fast_group_span_ms": "快采组跨度ms",
    "hgen_sample_ts": "湿度发生器缓存采样时间",
    "hgen_cache_age_ms": "湿度发生器缓存年龄ms",
    "hgen_raw": "湿度发生器原始报文",
    "status": "状态",
    "co2_ppm_target": "目标二氧化碳浓度ppm",
    "h2o_mmol_target": "目标水浓度mmol每mol",
    "hgen_temp_c": "湿度发生器_目标温度(℃)",
    "hgen_rh_pct": "湿度发生器_目标湿度(%RH)",
    "co2": "二氧化碳",
    "h2o": "水",
    "pressure": "压力",
    "co2_mean": "二氧化碳平均值",
    "co2_std": "二氧化碳标准差",
    "h2o_mean": "水平均值",
    "h2o_std": "水标准差",
    "co2_mean_primary_or_first": "二氧化碳平均值(主分析仪或首台可用)",
    "h2o_mean_primary_or_first": "水平均值(主分析仪或首台可用)",
    "analyzer_mean_mode": "分析仪均值语义",
    "co2_fleet": "二氧化碳全分析仪",
    "h2o_fleet": "水全分析仪",
    "co2_fleet_mean": "二氧化碳全分析仪平均值",
    "co2_fleet_std": "二氧化碳全分析仪标准差",
    "h2o_fleet_mean": "水全分析仪平均值",
    "h2o_fleet_std": "水全分析仪标准差",
    "pressure_mean": "压力平均值",
    "controller_pressure_mean": "压力控制器压力hPa_平均值",
    "gauge_pressure_mean": "数字压力计压力hPa_平均值",
    "env_chamber_temp_c_mean": "温度箱环境温度C_平均值",
    "env_chamber_rh_pct_mean": "温度箱环境湿度%_平均值",
    "analyzer_expected_count": "分析仪应到台数",
    "analyzer_with_frame_count": "分析仪有帧台数",
    "analyzer_usable_count": "分析仪可用台数",
    "analyzer_coverage_text": "分析仪覆盖率",
    "analyzer_integrity": "分析仪数据完整性",
    "analyzer_missing_labels": "缺失分析仪",
    "analyzer_unusable_labels": "异常帧分析仪",
    "dewpoint_gate_result": "封压后露点门禁结果",
    "dewpoint_gate_elapsed_s": "封压后露点门禁耗时s",
    "dewpoint_gate_count": "封压后露点门禁样本数",
    "dewpoint_gate_span_c": "封压后露点门禁跨度C",
    "dewpoint_gate_slope_c_per_s": "封压后露点门禁斜率C每s",
    "dewpoint_time_to_gate": "气路放行露点判稳耗时s",
    "dewpoint_tail_span_60s": "气路放行露点尾窗跨度C",
    "dewpoint_tail_slope_60s": "气路放行露点尾窗斜率C每s",
    "dewpoint_rebound_detected": "气路放行露点反弹检测",
    "flush_gate_status": "气路放行门禁结果",
    "flush_gate_reason": "气路放行门禁原因",
    "preseal_dewpoint_c": "封路前露点快照C",
    "preseal_temp_c": "封路前温度快照C",
    "preseal_rh_pct": "封路前湿度快照%",
    "preseal_pressure_hpa": "封路前压力快照hPa",
    "preseal_trigger_overshoot_hpa": "封路触发超调hPa",
    "postseal_expected_dewpoint_c": "封压后理论露点C",
    "postseal_actual_dewpoint_c": "封压后实际露点C",
    "postseal_physical_delta_c": "封压后物理偏差C",
    "postseal_physical_qc_status": "封压后物理一致性结果",
    "postseal_physical_qc_reason": "封压后物理一致性原因",
    "postseal_timeout_policy": "封压后露点超时策略",
    "postseal_timeout_blocked": "封压后露点超时是否阻断",
    "point_quality_timeout_flag": "点位质量_封压后露点超时标记",
    "dewpoint_gate_pass_live_c": "露点门禁放行实时露点C",
    "presample_long_guard_status": "采样前长稳守护结果",
    "presample_long_guard_reason": "采样前长稳守护原因",
    "presample_long_guard_elapsed_s": "采样前长稳守护耗时s",
    "presample_long_guard_span_c": "采样前长稳守护跨度C",
    "presample_long_guard_slope_c_per_s": "采样前长稳守护斜率C每s",
    "presample_long_guard_rise_c": "采样前长稳守护回升C",
    "first_effective_sample_dewpoint_c": "首个有效样本露点C",
    "postgate_to_first_effective_dewpoint_rise_c": "门禁放行到首个有效样本露点回升C",
    "postsample_late_rebound_status": "采样早期晚回潮结果",
    "postsample_late_rebound_reason": "采样早期晚回潮原因",
    "sampling_window_dewpoint_first_c": "采样窗露点首值C",
    "sampling_window_dewpoint_last_c": "采样窗露点末值C",
    "sampling_window_dewpoint_range_c": "采样窗露点跨度C",
    "sampling_window_dewpoint_rise_c": "采样窗露点首末回升C",
    "sampling_window_dewpoint_slope_c_per_s": "采样窗露点斜率C每s",
    "sampling_window_qc_status": "采样窗露点质控结果",
    "sampling_window_qc_reason": "采样窗露点质控原因",
    "pressure_gauge_stale_count": "数字压力计陈旧样本数",
    "pressure_gauge_total_count": "数字压力计总样本数",
    "pressure_gauge_stale_ratio": "数字压力计陈旧样本占比",
    "point_quality_status": "点位质量结果",
    "point_quality_reason": "点位质量原因",
    "point_quality_flags": "点位质量标记",
    "point_quality_blocked": "点位质量是否阻断",
    "preseal_vent_off_begin_to_route_sealed_ms": "关大气开始到封路ms",
    "route_sealed_to_control_prepare_begin_ms": "封路到控压准备开始ms",
    "pressure_in_limits_to_sampling_begin_ms": "压力达标到采样开始ms",
    "first_valid_pace_ms": "首个有效压力控制器样本偏移ms",
    "first_valid_pressure_gauge_ms": "首个有效数字压力计样本偏移ms",
    "first_valid_dewpoint_ms": "首个有效露点样本偏移ms",
    "first_valid_analyzer_ms": "首个有效分析仪样本偏移ms",
    "effective_sample_started_on_row": "首个全有效样本行",
}

_HGEN_FIELD_LABELS = {
    "Td": "露点(℃)",
    "Tc": "当前温度(℃)",
    "Tf": "霜点(℃)",
    "Ts": "设定温度(℃)",
    "Uw": "当前湿度(%RH)",
    "UwA": "设定湿度(%RH)",
    "temp_c": "目标温度(℃)",
    "rh_pct": "目标湿度(%RH)",
    "Ui": "绝对湿度",
    "Fl": "流量(L/min)",
    "Flux": "流量(L/min)",
    "TA": "目标温度(℃)",
    "FA": "目标流量(L/min)",
    "Pc": "当前压力",
    "Ps": "供气压力",
    "PST": "压力稳定计时(s)",
    "TST": "温度稳定计时(s)",
}

_ANALYZER_SAMPLE_FIELDS = [
    "raw",
    "id",
    "mode",
    "mode2_field_count",
    "frame_cache_ts",
    "frame_cache_age_ms",
    "frame_source",
    "frame_is_live",
    "co2_ppm",
    "h2o_mmol",
    "co2_density",
    "h2o_density",
    "co2_ratio_f",
    "co2_ratio_raw",
    "h2o_ratio_f",
    "h2o_ratio_raw",
    "ref_signal",
    "co2_signal",
    "h2o_signal",
    "chamber_temp_c",
    "case_temp_c",
    "pressure_kpa",
    "status",
]

_COMMON_SHEET_FIELDS = [
    "point_title",
    "sample_index",
    "sample_ts",
    "sample_due_ts",
    "sample_start_ts",
    "sample_end_ts",
    "sample_elapsed_ms",
    "sample_lag_ms",
    "point_phase",
    "point_tag",
    "point_row",
    "point_is_h2o",
    "route",
    "trace_stage",
    "pressure_mode",
    "pressure_target_label",
    "co2_ppm_target",
    "h2o_mmol_target",
    "pressure_target_hpa",
    "temp_set_c",
    "temp_chamber_c",
    "hgen_temp_c",
    "hgen_rh_pct",
    "pace_sample_ts",
    "pressure_hpa",
    "pace_output_state",
    "pace_isolation_state",
    "pace_vent_status",
    "pressure_gauge_sample_ts",
    "pressure_gauge_raw",
    "pressure_gauge_hpa",
    "fast_group_anchor_ts",
    "fast_group_start_ts",
    "fast_group_end_ts",
    "fast_group_span_ms",
    "dewpoint_c",
    "dew_temp_c",
    "dew_rh_pct",
    "dewpoint_sample_ts",
    "dewpoint_live_c",
    "dew_temp_live_c",
    "dew_rh_live_pct",
    "dewpoint_live_sample_ts",
    "dew_pressure_hpa",
    "preseal_dewpoint_c",
    "preseal_temp_c",
    "preseal_rh_pct",
    "preseal_pressure_hpa",
    "preseal_trigger_overshoot_hpa",
    "postseal_expected_dewpoint_c",
    "postseal_actual_dewpoint_c",
    "postseal_physical_delta_c",
    "postseal_physical_qc_status",
    "postseal_physical_qc_reason",
    "postseal_timeout_policy",
    "postseal_timeout_blocked",
    "point_quality_timeout_flag",
    "dewpoint_gate_pass_live_c",
    "presample_long_guard_status",
    "presample_long_guard_reason",
    "presample_long_guard_elapsed_s",
    "presample_long_guard_span_c",
    "presample_long_guard_slope_c_per_s",
    "presample_long_guard_rise_c",
    "first_effective_sample_dewpoint_c",
    "postgate_to_first_effective_dewpoint_rise_c",
    "postsample_late_rebound_status",
    "postsample_late_rebound_reason",
    "sampling_window_dewpoint_first_c",
    "sampling_window_dewpoint_last_c",
    "sampling_window_dewpoint_range_c",
    "sampling_window_dewpoint_rise_c",
    "sampling_window_dewpoint_slope_c_per_s",
    "sampling_window_qc_status",
    "sampling_window_qc_reason",
    "pressure_gauge_stale_count",
    "pressure_gauge_total_count",
    "pressure_gauge_stale_ratio",
    "point_quality_status",
    "point_quality_reason",
    "point_quality_flags",
    "point_quality_blocked",
    "env_chamber_temp_c",
    "env_chamber_rh_pct",
    "chamber_sample_ts",
    "chamber_cache_age_ms",
    "chamber_temp_c",
    "chamber_rh_pct",
    "thermometer_sample_ts",
    "thermometer_cache_age_ms",
    "thermometer_temp_c",
    "hgen_sample_ts",
    "hgen_cache_age_ms",
    "analyzer_expected_count",
    "analyzer_with_frame_count",
    "analyzer_usable_count",
    "analyzer_coverage_text",
    "analyzer_integrity",
    "analyzer_missing_labels",
    "analyzer_unusable_labels",
    "dewpoint_gate_result",
    "dewpoint_gate_elapsed_s",
    "dewpoint_gate_count",
    "dewpoint_gate_span_c",
    "dewpoint_gate_slope_c_per_s",
    "dewpoint_time_to_gate",
    "dewpoint_tail_span_60s",
    "dewpoint_tail_slope_60s",
    "dewpoint_rebound_detected",
    "flush_gate_status",
    "flush_gate_reason",
    "preseal_vent_off_begin_to_route_sealed_ms",
    "route_sealed_to_control_prepare_begin_ms",
    "pressure_in_limits_to_sampling_begin_ms",
    "first_valid_pace_ms",
    "first_valid_pressure_gauge_ms",
    "first_valid_dewpoint_ms",
    "first_valid_analyzer_ms",
    "effective_sample_started_on_row",
]

_READABLE_POINT_FIELDS = [
    "point_title",
    "point_row",
    "point_phase",
    "point_tag",
    "pressure_mode",
    "pressure_target_label",
    "temp_chamber_c",
    "co2_ppm_target",
    "hgen_temp_c",
    "hgen_rh_pct",
    "pressure_target_hpa",
    "analyzer_expected_count",
    "analyzer_with_frame_count",
    "analyzer_usable_count",
    "analyzer_coverage_text",
    "analyzer_integrity",
    "analyzer_missing_labels",
    "analyzer_unusable_labels",
    "dewpoint_gate_result",
    "dewpoint_gate_elapsed_s",
    "dewpoint_gate_count",
    "dewpoint_gate_span_c",
    "dewpoint_gate_slope_c_per_s",
    "preseal_dewpoint_c",
    "preseal_temp_c",
    "preseal_rh_pct",
    "preseal_pressure_hpa",
    "preseal_trigger_overshoot_hpa",
    "postseal_expected_dewpoint_c",
    "postseal_actual_dewpoint_c",
    "postseal_physical_delta_c",
    "postseal_physical_qc_status",
    "postseal_physical_qc_reason",
    "postseal_timeout_policy",
    "postseal_timeout_blocked",
    "point_quality_timeout_flag",
    "dewpoint_gate_pass_live_c",
    "presample_long_guard_status",
    "presample_long_guard_reason",
    "presample_long_guard_elapsed_s",
    "presample_long_guard_span_c",
    "presample_long_guard_slope_c_per_s",
    "presample_long_guard_rise_c",
    "first_effective_sample_dewpoint_c",
    "postgate_to_first_effective_dewpoint_rise_c",
    "postsample_late_rebound_status",
    "postsample_late_rebound_reason",
    "sampling_window_dewpoint_first_c",
    "sampling_window_dewpoint_last_c",
    "sampling_window_dewpoint_range_c",
    "sampling_window_dewpoint_rise_c",
    "sampling_window_dewpoint_slope_c_per_s",
    "sampling_window_qc_status",
    "sampling_window_qc_reason",
    "pressure_gauge_stale_count",
    "pressure_gauge_total_count",
    "pressure_gauge_stale_ratio",
    "point_quality_status",
    "point_quality_reason",
    "point_quality_flags",
    "point_quality_blocked",
    "preseal_vent_off_begin_to_route_sealed_ms",
    "route_sealed_to_control_prepare_begin_ms",
    "pressure_in_limits_to_sampling_begin_ms",
    "first_valid_pace_ms",
    "first_valid_pressure_gauge_ms",
    "first_valid_dewpoint_ms",
    "first_valid_analyzer_ms",
    "effective_sample_started_on_row",
]

_READABLE_MEAN_FIELDS = [
    "controller_pressure_mean",
    "gauge_pressure_mean",
    "dewpoint_c_mean",
    "dew_temp_c_mean",
    "dew_rh_pct_mean",
    "env_chamber_temp_c_mean",
    "env_chamber_rh_pct_mean",
    "hgen_Td_mean",
    "hgen_Tc_mean",
    "hgen_Tf_mean",
    "hgen_Ts_mean",
    "hgen_Uw_mean",
    "hgen_Ui_mean",
    "hgen_Fl_mean",
    "hgen_Pc_mean",
    "hgen_Ps_mean",
]

_READABLE_VALID_COUNT_FIELDS = [
    "co2_valid_count",
    "h2o_valid_count",
    "co2_mean_primary_or_first_valid_count",
    "h2o_mean_primary_or_first_valid_count",
    "controller_pressure_valid_count",
    "gauge_pressure_valid_count",
    "dewpoint_c_valid_count",
    "dew_temp_c_valid_count",
    "dew_rh_pct_valid_count",
]

_ANALYZER_EXPORT_FIELDS = [
    "NUM",
    "PointRow",
    "PointPhase",
    "PointTag",
    "PointTitle",
    "TempSet",
    "HgenTempSet",
    "HgenRhSet",
    "Temp",
    "Dew",
    "DewSampleTs",
    "DewPressurePreseal",
    "P",
    "PSample",
    "PpmH2oDewPressureSource",
    "ppm_CO2_Tank",
    "PressureTarget",
    "AnalyzerCoverage",
    "UsableAnalyzers",
    "ExpectedAnalyzers",
    "PointIntegrity",
    "MissingAnalyzers",
    "UnusableAnalyzers",
    "ValidFrames",
    "TotalFrames",
    "FrameStatus",
    "ppm_H2O_Dew",
    "ppm_CO2",
    "ppm_H2O",
    "R_CO2",
    "R_CO2_dev",
    "R_H2O",
    "R_H2O_dev",
    "Raw_REF",
    "Raw_CO2",
    "Raw_H2O",
    "T1",
    "T2",
    "BAR",
]

_READABLE_SKIP_DYNAMIC_MEAN_KEYS = {
    "co2_mean",
    "h2o_mean",
    "co2_mean_primary_or_first",
    "h2o_mean_primary_or_first",
    "pressure_mean",
    "pressure_gauge_hpa_mean",
    "chamber_temp_c_mean",
    "chamber_rh_pct_mean",
}

_ANALYZER_RATIO_FIELDS = {"R_CO2", "R_CO2_dev", "R_H2O", "R_H2O_dev"}
_THERMOMETER_STALE_DELTA_C = 2.0


def _field_label(key: str) -> str:
    text = str(key)

    direct = _FIELD_LABELS.get(text)
    if direct is not None:
        return direct

    if text.endswith("_valid_count"):
        return _field_label(text[:-12]) + "_有效样本数"
    if text.endswith("_mean"):
        return _field_label(text[:-5]) + "_平均值"
    if text.endswith("_std"):
        return _field_label(text[:-4]) + "_标准差"

    if text.startswith("hgen_"):
        suffix = text[5:]
        return f"湿度发生器_{_HGEN_FIELD_LABELS.get(suffix, _field_label(suffix))}"

    ga_match = re.match(r"^ga(\d+)_(.+)$", text)
    if ga_match:
        idx = int(ga_match.group(1))
        return f"气体分析仪{idx}_{_field_label(ga_match.group(2))}"

    return text


def _translate_row(row: Dict[str, Any]) -> Dict[str, Any]:
    return {_field_label(str(key)): value for key, value in row.items()}


def _phase_display(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text == "h2o":
        return "水路"
    if text == "co2":
        return "气路"
    return str(value or "")


def _analyzer_sheet_name(label: str, analyzer_id: Any = None) -> str:
    text = _safe_suffix(label)
    m = re.match(r"^ga0*(\d+)$", text)
    if m:
        base = f"气体分析仪{int(m.group(1))}"
    else:
        base = f"气体分析仪_{label}"
    id_text = str(analyzer_id).strip() if analyzer_id is not None else ""
    id_text = id_text or "UNKNOWN"
    id_text = re.sub(r"[^0-9A-Za-z_-]+", "_", id_text).strip("_") or "UNKNOWN"
    return f"{base}_ID{id_text}"[:31]


def _ordered_keys(keys: List[str], preferred: List[str]) -> List[str]:
    out: List[str] = []
    seen = set()
    for key in preferred + keys:
        text = str(key)
        if text in seen:
            continue
        seen.add(text)
        out.append(text)
    return out


def _dedupe_keys_by_label(keys: List[str]) -> List[str]:
    out: List[str] = []
    seen_labels = set()
    for key in keys:
        label = _field_label(key)
        if label in seen_labels:
            continue
        seen_labels.add(label)
        out.append(key)
    return out


def _merge_sheet_header_keys(existing_labels: List[Any], header_keys: List[str]) -> List[str]:
    """Preserve existing sheet-column order while appending newly discovered fields."""
    keyed_labels = [
        (idx, str(key), _field_label(str(key)))
        for idx, key in enumerate(header_keys)
    ]
    label_to_entries: Dict[str, List[tuple[int, str]]] = {}
    for idx, key, label in keyed_labels:
        label_to_entries.setdefault(label, []).append((idx, key))

    used_indices = set()
    merged_keys: List[str] = []
    blank_idx = 0
    for value in existing_labels:
        label = "" if value is None else str(value)
        entries = label_to_entries.get(label, [])
        match_idx: Optional[int] = None
        match_key: Optional[str] = None
        while entries:
            candidate_idx, candidate_key = entries.pop(0)
            if candidate_idx in used_indices:
                continue
            match_idx = candidate_idx
            match_key = candidate_key
            break
        if match_idx is None or match_key is None:
            blank_idx += 1
            merged_keys.append(f"__sheet_blank_{blank_idx}__")
            continue
        used_indices.add(match_idx)
        merged_keys.append(match_key)

    for idx, key, _label in keyed_labels:
        if idx in used_indices:
            continue
        merged_keys.append(key)
    return merged_keys


def _trim_trailing_empty_labels(labels: List[Any]) -> List[Any]:
    out = list(labels)
    while out and out[-1] in (None, ""):
        out.pop()
    return out


def _select_reference_temp(
    thermometer_temp: Optional[float],
    chamber_temp: Optional[float],
    *,
    stale_delta_c: float = _THERMOMETER_STALE_DELTA_C,
) -> Optional[float]:
    if thermometer_temp is None:
        return chamber_temp
    if chamber_temp is None:
        return thermometer_temp
    if abs(float(thermometer_temp) - float(chamber_temp)) > float(stale_delta_c):
        return chamber_temp
    return thermometer_temp


class RunLogger:
    """Runtime logger for samples, point summaries and device I/O traces."""

    def __init__(self, out_dir: Path, run_id: Optional[str] = None, cfg: Optional[Dict[str, Any]] = None):
        base_dir = Path(out_dir)
        base_dir.mkdir(parents=True, exist_ok=True)

        self.cfg = cfg or {}
        self.run_id = run_id or datetime.now().strftime("run_%Y%m%d_%H%M%S")
        self.run_dir = base_dir / self.run_id
        self.run_dir.mkdir(parents=True, exist_ok=True)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.samples_path = self.run_dir / f"samples_{stamp}.csv"
        self.points_path = self.run_dir / f"points_{stamp}.csv"
        self.points_readable_path = self.run_dir / f"points_readable_{stamp}.csv"
        self.points_readable_book_path = self.run_dir / f"points_readable_{stamp}.xlsx"
        self.io_path = self.run_dir / f"io_{stamp}.csv"
        self.coefficient_write_path = self.run_dir / f"coefficient_writeback_{stamp}.csv"
        self.h2o_analyzer_book_path = self.run_dir / f"h2o_analyzer_sheets_{stamp}.xlsx"
        self.co2_analyzer_book_path = self.run_dir / f"co2_analyzer_sheets_{stamp}.xlsx"
        self.analyzer_summary_csv_path = self.run_dir / f"分析仪汇总_{stamp}.csv"
        self.analyzer_summary_book_path = self.run_dir / f"分析仪汇总_{stamp}.xlsx"
        self.h2o_analyzer_summary_csv_path = self.run_dir / f"分析仪汇总_水路_{stamp}.csv"
        self.co2_analyzer_summary_csv_path = self.run_dir / f"分析仪汇总_气路_{stamp}.csv"
        self.h2o_analyzer_summary_book_path = self.run_dir / f"分析仪汇总_水路_{stamp}.xlsx"
        self.co2_analyzer_summary_book_path = self.run_dir / f"分析仪汇总_气路_{stamp}.xlsx"

        self._samples_file = self.samples_path.open("w", newline="", encoding="utf-8")
        self._points_file = self.points_path.open("w", newline="", encoding="utf-8")
        self._points_readable_file = self.points_readable_path.open("w", newline="", encoding="utf-8")
        self._analyzer_summary_file = self.analyzer_summary_csv_path.open("w", newline="", encoding="utf-8")
        self._io_file = self.io_path.open("w", newline="", encoding="utf-8")
        self._coefficient_write_file = self.coefficient_write_path.open("w", newline="", encoding="utf-8")

        self._samples_writer: Optional[csv.DictWriter] = None
        self._samples_header: List[str] = []
        self._samples_rows: List[Dict[str, Any]] = []
        self._points_writer: Optional[csv.DictWriter] = None
        self._points_header: List[str] = []
        self._points_rows: List[Dict[str, Any]] = []
        self._points_readable_writer: Optional[csv.DictWriter] = None
        self._points_readable_header: List[str] = []
        self._points_readable_rows: List[Dict[str, Any]] = []
        self._coefficient_write_writer: Optional[csv.DictWriter] = None
        self._coefficient_write_header: List[str] = []
        self._coefficient_write_rows: List[Dict[str, Any]] = []
        self._analyzer_summary_writer: Optional[csv.DictWriter] = None
        self._analyzer_summary_rows_by_target: Dict[str, List[Dict[str, Any]]] = {"all": []}
        self._analyzer_summary_header_by_target: Dict[str, List[str]] = {
            "all": ["Analyzer"] + list(_ANALYZER_EXPORT_FIELDS)
        }
        self._analyzer_summary_phase_files: Dict[str, Any] = {}
        self._analyzer_summary_phase_writers: Dict[str, csv.DictWriter] = {}
        self._points_role_logged = False
        self._point_alias_note_logged = False
        self._analyzer_summary_role_logged = False
        self._io_writer = csv.DictWriter(
            self._io_file,
            fieldnames=[
                "timestamp",
                "port",
                "device",
                "direction",
                "command",
                "response",
                "error",
            ],
        )
        self._io_writer.writeheader()
        self._io_file.flush()

    def _include_fleet_stats(self) -> bool:
        workflow_cfg = self.cfg.get("workflow", {}) if isinstance(self.cfg, dict) else {}
        reporting_cfg = workflow_cfg.get("reporting", {}) if isinstance(workflow_cfg, dict) else {}
        return bool(reporting_cfg.get("include_fleet_stats", False))

    def _sanitize_point_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        cleaned = dict(row)
        if not self._include_fleet_stats():
            for key in (
                "co2_fleet_mean",
                "co2_fleet_std",
                "h2o_fleet_mean",
                "h2o_fleet_std",
            ):
                cleaned.pop(key, None)
        return cleaned

    @staticmethod
    def _merge_header(existing: List[str], keys: List[str]) -> List[str]:
        merged = list(existing)
        for key in keys:
            text = str(key)
            if text not in merged:
                merged.append(text)
        return merged

    @staticmethod
    def _open_csv_append_handle(path: Path):
        return path.open("a+", newline="", encoding="utf-8")

    def _rewrite_dynamic_csv(
        self,
        *,
        path: Path,
        file_attr: str,
        writer_attr: str,
        header_attr: str,
        rows: List[Dict[str, Any]],
        target_header: List[str],
    ) -> None:
        old_header = list(getattr(self, header_attr))
        handle = getattr(self, file_attr, None)
        if handle is not None:
            try:
                handle.close()
            except Exception:
                pass

        try:
            _save_csv_atomic(path, target_header, rows)
            new_handle = self._open_csv_append_handle(path)
            new_writer = csv.DictWriter(
                new_handle,
                fieldnames=list(target_header),
                extrasaction="ignore",
            )
        except Exception:
            try:
                reopened = self._open_csv_append_handle(path)
            except Exception:
                reopened = None
            setattr(self, file_attr, reopened)
            setattr(self, writer_attr, None)
            setattr(self, header_attr, old_header)
            raise

        setattr(self, file_attr, new_handle)
        setattr(self, writer_attr, new_writer)
        setattr(self, header_attr, list(target_header))

    def _log_point_export_role_once(self, row: Dict[str, Any]) -> None:
        if not self._points_role_logged:
            self.log_io(
                port="LOG",
                device="run_logger",
                direction="EVENT",
                command="points-export-role",
                response=(
                    "points.csv / points_readable are execution summaries for point coverage, "
                    "environment and reference overview; they are not the formal per-analyzer "
                    "calibration summary table"
                ),
            )
            self._points_role_logged = True
        if not self._point_alias_note_logged and (
            "co2_mean" in row
            or "h2o_mean" in row
            or "co2_mean_primary_or_first" in row
            or "h2o_mean_primary_or_first" in row
        ):
            self.log_io(
                port="LOG",
                device="run_logger",
                direction="EVENT",
                command="points-mean-alias-semantics",
                response=(
                    "point-level co2_mean/h2o_mean aliases are convenience fields based on the "
                    "primary analyzer or first usable analyzer; they are not fleet means and do "
                    "not replace analyzer_summary"
                ),
            )
            self._point_alias_note_logged = True

    def _log_analyzer_summary_role_once(self) -> None:
        if self._analyzer_summary_role_logged:
            return
        self.log_io(
            port="LOG",
            device="run_logger",
            direction="EVENT",
            command="analyzer-summary-role",
            response=(
                "analyzer_summary csv/xlsx are the formal per-analyzer summaries for calibration "
                "fit, validation and manual review"
            ),
        )
        self._analyzer_summary_role_logged = True

    def log_sample(self, row: Dict[str, Any]) -> None:
        row = _translate_row(row)
        self._append_samples_csv_row(row)

    def log_point(self, row: Dict[str, Any]) -> None:
        row = self._sanitize_point_row(row)
        self._log_point_export_role_once(row)
        translated = _translate_row(row)
        self._append_points_csv_row(translated)

        readable = self._build_readable_point_row(row)
        self._append_readable_point_csv(readable)
        try:
            self._append_readable_point_workbook(readable)
        except Exception as exc:
            self.log_io(
                port="LOG",
                device="run_logger",
                direction="WARN",
                command="readable-point-workbook",
                response="csv-only fallback",
                error=exc,
            )

    def _append_samples_csv_row(self, row: Dict[str, Any]) -> None:
        stored = dict(row)
        self._samples_rows.append(stored)
        merged_header = self._merge_header(self._samples_header, list(stored.keys()))
        if self._samples_writer is None:
            self._rewrite_samples_csv()
            return
        if merged_header != self._samples_header:
            self._rewrite_samples_csv(target_header=merged_header)
            return

        try:
            self._samples_writer.writerow(stored)
            self._samples_file.flush()
        except Exception:
            self._samples_writer = None
            self._rewrite_samples_csv()

    def _rewrite_samples_csv(self, target_header: Optional[List[str]] = None) -> None:
        header = list(target_header or self._samples_header or self._merge_header([], [key for row in self._samples_rows for key in row.keys()]))
        self._rewrite_dynamic_csv(
            path=self.samples_path,
            file_attr="_samples_file",
            writer_attr="_samples_writer",
            header_attr="_samples_header",
            rows=self._samples_rows,
            target_header=header,
        )

    def _append_points_csv_row(self, row: Dict[str, Any]) -> None:
        stored = dict(row)
        self._points_rows.append(stored)
        merged_header = self._merge_header(self._points_header, list(stored.keys()))
        if self._points_writer is None:
            self._rewrite_points_csv()
            return
        if merged_header != self._points_header:
            self._rewrite_points_csv(target_header=merged_header)
            return

        try:
            self._points_writer.writerow(stored)
            self._points_file.flush()
        except Exception:
            self._points_writer = None
            self._rewrite_points_csv()

    def _rewrite_points_csv(self, target_header: Optional[List[str]] = None) -> None:
        header = list(target_header or self._points_header or self._merge_header([], [key for row in self._points_rows for key in row.keys()]))
        self._rewrite_dynamic_csv(
            path=self.points_path,
            file_attr="_points_file",
            writer_attr="_points_writer",
            header_attr="_points_header",
            rows=self._points_rows,
            target_header=header,
        )

    def log_coefficient_write(self, row: Dict[str, Any]) -> None:
        stored = dict(row)
        self._coefficient_write_rows.append(stored)
        merged_header = self._merge_header(self._coefficient_write_header, list(stored.keys()))
        if self._coefficient_write_writer is None:
            self._rewrite_coefficient_write_csv(target_header=merged_header)
            return
        if merged_header != self._coefficient_write_header:
            self._rewrite_coefficient_write_csv(target_header=merged_header)
            return

        try:
            self._coefficient_write_writer.writerow(stored)
            self._coefficient_write_file.flush()
        except Exception:
            self._coefficient_write_writer = None
            self._rewrite_coefficient_write_csv(target_header=merged_header)

    def _rewrite_coefficient_write_csv(self, target_header: Optional[List[str]] = None) -> None:
        header = list(
            target_header
            or self._coefficient_write_header
            or self._merge_header([], [key for row in self._coefficient_write_rows for key in row.keys()])
        )
        self._rewrite_dynamic_csv(
            path=self.coefficient_write_path,
            file_attr="_coefficient_write_file",
            writer_attr="_coefficient_write_writer",
            header_attr="_coefficient_write_header",
            rows=self._coefficient_write_rows,
            target_header=header,
        )

    def _append_readable_point_csv(self, row: Dict[str, Any]) -> None:
        stored = dict(row)
        self._points_readable_rows.append(stored)
        merged_header = self._merge_header(self._points_readable_header, list(stored.keys()))
        if self._points_readable_writer is None:
            self._rewrite_readable_point_csv()
            return
        if merged_header != self._points_readable_header:
            self._rewrite_readable_point_csv(target_header=merged_header)
            return

        try:
            self._points_readable_writer.writerow(stored)
            self._points_readable_file.flush()
        except Exception:
            self._points_readable_writer = None
            self._rewrite_readable_point_csv()

    def _rewrite_readable_point_csv(self, target_header: Optional[List[str]] = None) -> None:
        header = list(
            target_header
            or self._points_readable_header
            or self._merge_header([], [key for row in self._points_readable_rows for key in row.keys()])
        )
        self._rewrite_dynamic_csv(
            path=self.points_readable_path,
            file_attr="_points_readable_file",
            writer_attr="_points_readable_writer",
            header_attr="_points_readable_header",
            rows=self._points_readable_rows,
            target_header=header,
        )

    def _build_readable_point_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        readable_row = dict(row)
        if "controller_pressure_mean" not in readable_row and "pressure_mean" in readable_row:
            readable_row["controller_pressure_mean"] = readable_row.get("pressure_mean")
        if "controller_pressure_valid_count" not in readable_row and "pressure_valid_count" in readable_row:
            readable_row["controller_pressure_valid_count"] = readable_row.get("pressure_valid_count")
        if "gauge_pressure_mean" not in readable_row and "pressure_gauge_hpa_mean" in readable_row:
            readable_row["gauge_pressure_mean"] = readable_row.get("pressure_gauge_hpa_mean")
        if "gauge_pressure_valid_count" not in readable_row and "pressure_gauge_hpa_valid_count" in readable_row:
            readable_row["gauge_pressure_valid_count"] = readable_row.get("pressure_gauge_hpa_valid_count")
        if "env_chamber_temp_c_mean" not in readable_row and "chamber_temp_c_mean" in readable_row:
            readable_row["env_chamber_temp_c_mean"] = readable_row.get("chamber_temp_c_mean")
        if "env_chamber_rh_pct_mean" not in readable_row and "chamber_rh_pct_mean" in readable_row:
            readable_row["env_chamber_rh_pct_mean"] = readable_row.get("chamber_rh_pct_mean")

        selected: Dict[str, Any] = {}
        for key in _READABLE_POINT_FIELDS:
            if key in readable_row:
                selected[key] = readable_row.get(key)

        for key in _READABLE_MEAN_FIELDS:
            if key in readable_row:
                selected[key] = readable_row.get(key)

        for key in _READABLE_VALID_COUNT_FIELDS:
            if key in readable_row:
                selected[key] = readable_row.get(key)

        if self._include_fleet_stats():
            for key in (
                "co2_fleet_mean",
                "co2_fleet_std",
                "h2o_fleet_mean",
                "h2o_fleet_std",
            ):
                if key in readable_row:
                    selected[key] = readable_row.get(key)

        dynamic_mean_keys = []
        for key in readable_row.keys():
            text = str(key)
            if not text.endswith("_mean"):
                continue
            if text in selected:
                continue
            if text in _READABLE_SKIP_DYNAMIC_MEAN_KEYS:
                continue
            if re.match(r"^ga\d+_", text):
                continue
            dynamic_mean_keys.append(text)

        ordered_keys = _dedupe_keys_by_label(_ordered_keys(dynamic_mean_keys, _READABLE_MEAN_FIELDS))
        for key in ordered_keys:
            selected[key] = readable_row.get(key)

        dynamic_valid_count_keys = []
        for key in readable_row.keys():
            text = str(key)
            if not text.endswith("_valid_count"):
                continue
            if text in selected:
                continue
            if re.match(r"^ga\d+_", text):
                continue
            dynamic_valid_count_keys.append(text)

        ordered_valid_count_keys = _dedupe_keys_by_label(
            _ordered_keys(dynamic_valid_count_keys, _READABLE_VALID_COUNT_FIELDS)
        )
        for key in ordered_valid_count_keys:
            selected[key] = readable_row.get(key)

        return _translate_row(selected)

    def _append_readable_point_workbook(self, row: Dict[str, Any]) -> None:
        path = self.points_readable_book_path
        if path.exists():
            wb = load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "点位总览"

        header = list(row.keys())
        existing_header = self._worksheet_header(ws)
        if not existing_header:
            merged_header = list(header)
        else:
            merged_header = list(existing_header)
            for key in header:
                if key not in merged_header:
                    merged_header.append(key)

        self._write_worksheet_header(ws, merged_header)
        ws.append([row.get(key) for key in merged_header])
        self._apply_readable_point_row_style(ws, ws.max_row, row)
        self._apply_readable_point_alerts(ws, ws.max_row, merged_header, row)
        try:
            self._format_worksheet(ws)
            _save_workbook_atomic(wb, path)
        finally:
            wb.close()

    @staticmethod
    def _worksheet_header(ws) -> List[str]:
        if ws.max_row < 1:
            return []
        header = [
            ws.cell(row=1, column=i).value
            for i in range(1, ws.max_column + 1)
        ]
        while header and header[-1] in (None, ""):
            header.pop()
        return [str(value) for value in header if value not in (None, "")]

    @staticmethod
    def _write_worksheet_header(ws, header: List[str]) -> None:
        for idx, key in enumerate(header, start=1):
            ws.cell(row=1, column=idx).value = key

    @staticmethod
    def _summary_float(values: List[Any]) -> Optional[float]:
        numeric: List[float] = []
        for value in values:
            if value in (None, ""):
                continue
            try:
                numeric.append(float(value))
            except Exception:
                continue
        if not numeric:
            return None
        return round(mean(numeric), 6)

    @staticmethod
    def _summary_std(values: List[Any]) -> Optional[float]:
        numeric: List[float] = []
        for value in values:
            if value in (None, ""):
                continue
            try:
                numeric.append(float(value))
            except Exception:
                continue
        if len(numeric) < 2:
            return None
        return round(stdev(numeric), 6)

    @staticmethod
    def _summary_first_nonempty(values: List[Any]) -> Any:
        for value in values:
            if value not in (None, ""):
                return value
        return None

    def _summary_dew_h2o_context(self, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
        dewpoint = self._summary_float([row.get("dewpoint_c") for row in rows])
        preseal_pressure = self._summary_float([row.get("dew_pressure_hpa") for row in rows])
        sample_gauge_pressure = self._summary_float([row.get("pressure_gauge_hpa") for row in rows])
        controller_pressure = self._summary_float([row.get("pressure_hpa") for row in rows])
        pressure_source = ""
        pressure_for_ppm = preseal_pressure
        if pressure_for_ppm is not None:
            pressure_source = "dew_pressure_hpa(preseal)"
        elif sample_gauge_pressure is not None:
            pressure_for_ppm = sample_gauge_pressure
            pressure_source = "pressure_gauge_hpa(sample)"
        elif controller_pressure is not None:
            pressure_for_ppm = controller_pressure
            pressure_source = "pressure_hpa(controller_sample)"
        return {
            "ppm_h2o_dew": _dewpoint_to_h2o_mmol_per_mol(dewpoint, pressure_for_ppm),
            "dew_sample_ts": self._summary_first_nonempty([row.get("dewpoint_sample_ts") for row in rows]),
            "dew_pressure_preseal": preseal_pressure,
            "p_sample": sample_gauge_pressure,
            "pressure_source": pressure_source,
        }

    def _summary_reference_on_aligned_rows(self) -> bool:
        workflow_cfg = self.cfg.get("workflow", {}) if isinstance(self.cfg, dict) else {}
        summary_cfg = workflow_cfg.get("summary_alignment", {}) if isinstance(workflow_cfg, dict) else {}
        return bool(summary_cfg.get("reference_on_aligned_rows", True))

    @staticmethod
    def _sample_prefixed_value(row: Dict[str, Any], prefix: str, key: str) -> Any:
        prefixed_key = f"{prefix}_{key}"
        value = row.get(prefixed_key)
        if value in (None, "") and prefix == "ga01":
            value = row.get(key)
        return value

    @staticmethod
    def _sample_prefixed_has_data(row: Dict[str, Any], prefix: str) -> bool:
        value = row.get(f"{prefix}_frame_has_data")
        if value not in (None, ""):
            return bool(value)
        return any(
            RunLogger._sample_prefixed_value(row, prefix, key) not in (None, "")
            for key in ("co2_ppm", "h2o_mmol", "co2_ratio_f", "h2o_ratio_f", "ref_signal")
        )

    @staticmethod
    def _sample_prefixed_usable(row: Dict[str, Any], prefix: str) -> bool:
        value = row.get(f"{prefix}_frame_usable")
        if value not in (None, ""):
            return bool(value)
        return RunLogger._sample_prefixed_has_data(row, prefix)

    def _build_analyzer_summary_row(
        self,
        rows: List[Dict[str, Any]],
        *,
        label: str,
        num: int,
    ) -> Dict[str, Any]:
        prefix = _safe_suffix(label)
        sheet_rows = rows
        first = sheet_rows[0] if sheet_rows else {}
        analyzer_rows = [row for row in sheet_rows if self._sample_prefixed_usable(row, prefix)]
        aligned_rows = analyzer_rows or sheet_rows
        # Keep the summary reference quantities aligned with this analyzer's usable
        # sample set whenever possible. Otherwise ratio/ppm means may be computed from
        # one subset while Dew/P reference terms come from a different subset.
        reference_rows = aligned_rows if self._summary_reference_on_aligned_rows() else sheet_rows

        def _common_mean(key: str, *, rows_override: Optional[List[Dict[str, Any]]] = None) -> Optional[float]:
            source_rows = rows_override if rows_override is not None else sheet_rows
            return self._summary_float([row.get(key) for row in source_rows])

        def _common_mean_any(keys: List[str], *, rows_override: Optional[List[Dict[str, Any]]] = None) -> Optional[float]:
            for key in keys:
                value = _common_mean(key, rows_override=rows_override)
                if value is not None:
                    return value
            return None

        def _analyzer_mean(key: str) -> Optional[float]:
            return self._summary_float(
                [
                    self._sample_prefixed_value(row, prefix, key)
                    for row in sheet_rows
                    if self._sample_prefixed_usable(row, prefix)
                ]
            )

        def _analyzer_std(key: str) -> Optional[float]:
            return self._summary_std(
                [
                    self._sample_prefixed_value(row, prefix, key)
                    for row in sheet_rows
                    if self._sample_prefixed_usable(row, prefix)
                ]
            )

        thermometer_temp = _common_mean("thermometer_temp_c", rows_override=aligned_rows)
        chamber_temp = _common_mean_any(
            ["env_chamber_temp_c", "chamber_temp_c"],
            rows_override=aligned_rows,
        )
        reference_temp = _select_reference_temp(thermometer_temp, chamber_temp)
        dew_context = self._summary_dew_h2o_context(reference_rows)
        total_frames = sum(1 for row in sheet_rows if self._sample_prefixed_has_data(row, prefix))
        valid_frames = sum(1 for row in sheet_rows if self._sample_prefixed_usable(row, prefix))
        if total_frames == 0:
            frame_status = "无帧"
        elif valid_frames == total_frames:
            frame_status = "全部可用"
        elif valid_frames == 0:
            frame_status = "仅异常帧"
        else:
            frame_status = "部分可用"

        self.log_io(
            port="LOG",
            device="run_logger",
            direction="EVENT",
            command="analyzer-summary-align",
            response=(
                f"label={str(label or '').upper()} total={total_frames} valid={valid_frames} "
                f"ref_rows={'aligned' if reference_rows is aligned_rows else 'all'} "
                f"aligned_count={len(aligned_rows)}"
            ),
        )
        if _phase_display(first.get("point_phase")) == "水路":
            self.log_io(
                port="LOG",
                device="run_logger",
                direction="EVENT",
                command="analyzer-summary-h2o-reference",
                response=(
                    f"label={str(label or '').upper()} Dew=preseal_snapshot "
                    f"ppm_H2O_Dew_source={dew_context.get('pressure_source') or 'none'} "
                    f"DewPressurePreseal={dew_context.get('dew_pressure_preseal')} "
                    f"PSample={dew_context.get('p_sample')}"
                ),
            )

        return {
            "NUM": num,
            "PointRow": first.get("point_row"),
            "PointPhase": _phase_display(first.get("point_phase")),
            "PointTag": first.get("point_tag"),
            "PointTitle": first.get("point_title"),
            "PressureMode": first.get("pressure_mode"),
            "PressureTargetLabel": first.get("pressure_target_label"),
            "TempSet": first.get("temp_chamber_c"),
            "HgenTempSet": first.get("hgen_temp_c"),
            "HgenRhSet": first.get("hgen_rh_pct"),
            "Temp": reference_temp,
            "Dew": _common_mean("dewpoint_c", rows_override=reference_rows),
            "DewSampleTs": dew_context.get("dew_sample_ts"),
            "DewPressurePreseal": dew_context.get("dew_pressure_preseal"),
            "P": _common_mean("pressure_gauge_hpa", rows_override=reference_rows),
            "PSample": dew_context.get("p_sample"),
            "PpmH2oDewPressureSource": dew_context.get("pressure_source"),
            "ppm_CO2_Tank": first.get("co2_ppm_target"),
            "PressureTarget": first.get("pressure_target_hpa"),
            "AnalyzerCoverage": first.get("analyzer_coverage_text"),
            "UsableAnalyzers": first.get("analyzer_usable_count"),
            "ExpectedAnalyzers": first.get("analyzer_expected_count"),
            "PointIntegrity": first.get("analyzer_integrity"),
            "MissingAnalyzers": first.get("analyzer_missing_labels"),
            "UnusableAnalyzers": first.get("analyzer_unusable_labels"),
            "ValidFrames": valid_frames,
            "TotalFrames": total_frames,
            "FrameStatus": frame_status,
            "ppm_H2O_Dew": dew_context.get("ppm_h2o_dew"),
            "ppm_CO2": _analyzer_mean("co2_ppm"),
            "ppm_H2O": _analyzer_mean("h2o_mmol"),
            "R_CO2": _analyzer_mean("co2_ratio_f"),
            "R_CO2_dev": _analyzer_std("co2_ratio_f"),
            "R_H2O": _analyzer_mean("h2o_ratio_f"),
            "R_H2O_dev": _analyzer_std("h2o_ratio_f"),
            "Raw_REF": _analyzer_mean("ref_signal"),
            "Raw_CO2": _analyzer_mean("co2_signal"),
            "Raw_H2O": _analyzer_mean("h2o_signal"),
            "T1": _analyzer_mean("chamber_temp_c"),
            "T2": _analyzer_mean("case_temp_c"),
            "BAR": _analyzer_mean("pressure_kpa"),
        }

    def _next_analyzer_summary_num(self, analyzer_label: str) -> int:
        path = self.analyzer_summary_book_path
        sheet_name = analyzer_label.upper()[:31]
        if not path.exists():
            return 1
        wb = load_workbook(path, read_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return 1
            ws = wb[sheet_name]
            return max(ws.max_row, 1)
        finally:
            wb.close()

    @staticmethod
    def _analyzer_summary_phase_key(phase: Any) -> Optional[str]:
        text = str(phase or "").strip().lower()
        if text in {"水路", "h2o"}:
            return "h2o"
        if text in {"气路", "co2"}:
            return "co2"
        return None

    def _analyzer_summary_csv_path_for_key(self, key: str) -> Path:
        if key == "h2o":
            return self.h2o_analyzer_summary_csv_path
        if key == "co2":
            return self.co2_analyzer_summary_csv_path
        return self.analyzer_summary_csv_path

    def _analyzer_summary_book_path_for_key(self, key: str) -> Path:
        if key == "h2o":
            return self.h2o_analyzer_summary_book_path
        if key == "co2":
            return self.co2_analyzer_summary_book_path
        return self.analyzer_summary_book_path

    def _analyzer_summary_writer_for_target(self, target_key: str) -> Optional[csv.DictWriter]:
        if target_key == "all":
            return self._analyzer_summary_writer
        return self._analyzer_summary_phase_writers.get(target_key)

    def _set_analyzer_summary_writer_for_target(self, target_key: str, writer: Optional[csv.DictWriter]) -> None:
        if target_key == "all":
            self._analyzer_summary_writer = writer
        else:
            if writer is None:
                self._analyzer_summary_phase_writers.pop(target_key, None)
            else:
                self._analyzer_summary_phase_writers[target_key] = writer

    def _analyzer_summary_file_for_target(self, target_key: str):
        if target_key == "all":
            return self._analyzer_summary_file
        return self._analyzer_summary_phase_files.get(target_key)

    def _set_analyzer_summary_file_for_target(self, target_key: str, handle: Any) -> None:
        if target_key == "all":
            self._analyzer_summary_file = handle
        else:
            if handle is None:
                self._analyzer_summary_phase_files.pop(target_key, None)
            else:
                self._analyzer_summary_phase_files[target_key] = handle

    def _rewrite_analyzer_summary_csv_target(
        self,
        target_key: str,
        target_header: Optional[List[str]] = None,
    ) -> None:
        old_header = list(self._analyzer_summary_header_by_target.get(target_key, ["Analyzer"] + list(_ANALYZER_EXPORT_FIELDS)))
        old_handle = self._analyzer_summary_file_for_target(target_key)
        if old_handle is not None:
            try:
                old_handle.close()
            except Exception:
                pass

        header = list(
            target_header
            or self._analyzer_summary_header_by_target.get(target_key, [])
            or ["Analyzer"] + list(_ANALYZER_EXPORT_FIELDS)
        )
        rows = self._analyzer_summary_rows_by_target.get(target_key, [])
        path = self._analyzer_summary_csv_path_for_key(target_key)
        try:
            _save_csv_atomic(path, header, rows)
            new_handle = self._open_csv_append_handle(path)
            new_writer = csv.DictWriter(
                new_handle,
                fieldnames=list(header),
                extrasaction="ignore",
            )
        except Exception:
            try:
                reopened = self._open_csv_append_handle(path)
            except Exception:
                reopened = None
            self._set_analyzer_summary_file_for_target(target_key, reopened)
            self._set_analyzer_summary_writer_for_target(target_key, None)
            self._analyzer_summary_header_by_target[target_key] = old_header
            raise

        self._set_analyzer_summary_file_for_target(target_key, new_handle)
        self._set_analyzer_summary_writer_for_target(target_key, new_writer)
        self._analyzer_summary_header_by_target[target_key] = list(header)

    def _append_analyzer_summary_csv_row_to_target(
        self,
        target_key: str,
        analyzer_label: str,
        row: Dict[str, Any],
    ) -> None:
        csv_row = {"Analyzer": analyzer_label.upper()}
        for key, value in row.items():
            if key in _ANALYZER_RATIO_FIELDS and value not in (None, ""):
                try:
                    csv_row[key] = f"{float(value):.6f}"
                    continue
                except Exception:
                    pass
            csv_row[key] = value

        rows = self._analyzer_summary_rows_by_target.setdefault(target_key, [])
        rows.append(dict(csv_row))
        current_header = self._analyzer_summary_header_by_target.setdefault(
            target_key,
            ["Analyzer"] + list(_ANALYZER_EXPORT_FIELDS),
        )
        merged_header = self._merge_header(current_header, list(csv_row.keys()))
        writer = self._analyzer_summary_writer_for_target(target_key)
        handle = self._analyzer_summary_file_for_target(target_key)

        if writer is None:
            self._rewrite_analyzer_summary_csv_target(target_key)
            return
        if merged_header != current_header:
            self._rewrite_analyzer_summary_csv_target(target_key, target_header=merged_header)
            return

        try:
            writer.writerow(csv_row)
            if handle is not None:
                handle.flush()
        except Exception:
            self._set_analyzer_summary_writer_for_target(target_key, None)
            self._rewrite_analyzer_summary_csv_target(target_key)

    def _append_analyzer_summary_csv_row(self, analyzer_label: str, row: Dict[str, Any]) -> None:
        self._append_analyzer_summary_csv_row_to_target("all", analyzer_label, row)
        phase_key = self._analyzer_summary_phase_key(row.get("PointPhase"))
        if phase_key is not None:
            self._append_analyzer_summary_csv_row_to_target(phase_key, analyzer_label, row)

    def _append_analyzer_summary_workbook_row_to_target(self, path: Path, analyzer_label: str, row: Dict[str, Any]) -> None:
        if path.exists():
            wb = load_workbook(path)
        else:
            wb = Workbook()
            if wb.sheetnames == ["Sheet"]:
                wb.remove(wb["Sheet"])

        try:
            sheet_name = analyzer_label.upper()[:31]
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)

            header = _ANALYZER_EXPORT_FIELDS
            existing = [ws.cell(row=1, column=i + 1).value for i in range(max(ws.max_column, 1))]
            is_blank_header = all(value is None for value in existing)
            if is_blank_header:
                ws.append(header)
                if ws.max_row >= 2 and ws["A1"].value is None:
                    ws.delete_rows(1, 1)
            elif existing != header:
                raise RuntimeError(f"Analyzer summary workbook header mismatch: {sheet_name}")

            ws.append([row.get(key) for key in header])
            self._apply_analyzer_summary_row_style(ws, ws.max_row, row)
            self._format_worksheet(ws)
            _save_workbook_atomic(wb, path)
        finally:
            wb.close()

    def _append_analyzer_summary_workbook_row(self, analyzer_label: str, row: Dict[str, Any]) -> None:
        self._append_analyzer_summary_workbook_row_to_target(self.analyzer_summary_book_path, analyzer_label, row)
        phase_key = self._analyzer_summary_phase_key(row.get("PointPhase"))
        if phase_key is not None:
            self._append_analyzer_summary_workbook_row_to_target(
                self._analyzer_summary_book_path_for_key(phase_key),
                analyzer_label,
                row,
            )

    @staticmethod
    def _apply_analyzer_summary_row_style(ws: Any, row_idx: int, row: Dict[str, Any]) -> None:
        phase = str(row.get("PointPhase") or "").strip()
        if phase == "水路":
            fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
        elif phase == "气路":
            fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        else:
            fill = PatternFill(fill_type="solid", fgColor="EDEDED")

        title_col = None
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            if ws.cell(row=1, column=col).value == "PointTitle":
                title_col = col

        if title_col is not None:
            ws.cell(row=row_idx, column=title_col).alignment = Alignment(vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header in _ANALYZER_RATIO_FIELDS:
                ws.cell(row=row_idx, column=col).number_format = "0.000000"

    @staticmethod
    def _apply_readable_point_row_style(ws: Any, row_idx: int, row: Dict[str, Any]) -> None:
        phase = str(row.get("流程阶段") or "").strip().lower()
        if phase == "h2o":
            fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
        elif phase == "co2":
            fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        else:
            fill = PatternFill(fill_type="solid", fgColor="EDEDED")

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    @staticmethod
    def _apply_readable_point_alerts(ws: Any, row_idx: int, header: List[str], row: Dict[str, Any]) -> None:
        alert_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
        label_to_col = {str(name): idx + 1 for idx, name in enumerate(header)}

        def _to_float(value: Any) -> Optional[float]:
            if value in (None, ""):
                return None
            try:
                return float(value)
            except Exception:
                return None

        def _mark_if_outside(actual_labels: Any, target_labels: Any, threshold: float) -> None:
            labels = [actual_labels] if isinstance(actual_labels, str) else list(actual_labels or [])
            actual_label = next((label for label in labels if label in label_to_col), None)
            if actual_label is None:
                return
            target_options = [target_labels] if isinstance(target_labels, str) else list(target_labels or [])
            target_label = next((label for label in target_options if label in row), None)
            if target_label is None:
                return
            actual = _to_float(row.get(actual_label))
            target = _to_float(row.get(target_label))
            col = label_to_col.get(actual_label)
            if actual is None or target is None or col is None:
                return
            if abs(actual - target) > float(threshold):
                ws.cell(row=row_idx, column=col).fill = alert_fill

        def _mark_if_outside_dynamic(actual_labels: Any, target_labels: Any, min_threshold: float, ratio: float) -> None:
            labels = [actual_labels] if isinstance(actual_labels, str) else list(actual_labels or [])
            actual_label = next((label for label in labels if label in label_to_col), None)
            if actual_label is None:
                return
            target_options = [target_labels] if isinstance(target_labels, str) else list(target_labels or [])
            target_label = next((label for label in target_options if label in row), None)
            if target_label is None:
                return
            actual = _to_float(row.get(actual_label))
            target = _to_float(row.get(target_label))
            col = label_to_col.get(actual_label)
            if actual is None or target is None or col is None:
                return
            threshold = max(float(min_threshold), abs(target) * float(ratio))
            if abs(actual - target) > threshold:
                ws.cell(row=row_idx, column=col).fill = alert_fill

        _mark_if_outside_dynamic(
            ["二氧化碳平均值(主分析仪或首台可用)", "二氧化碳平均值"],
            "目标二氧化碳浓度ppm",
            min_threshold=20.0,
            ratio=0.05,
        )
        _mark_if_outside(["压力控制器压力hPa_平均值", "压力平均值"], "目标压力hPa", threshold=5.0)
        _mark_if_outside(["数字压力计压力hPa_平均值"], "目标压力hPa", threshold=5.0)
        _mark_if_outside(["温度箱环境温度C_平均值", "温度箱温度C_平均值"], "温箱目标温度C", threshold=0.5)
        _mark_if_outside(
            "湿度发生器_当前温度(℃)_平均值",
            ["湿度发生器_目标温度(℃)", "湿度发生器目标温度C"],
            threshold=0.5,
        )
        _mark_if_outside(
            "湿度发生器_当前湿度(%RH)_平均值",
            ["湿度发生器_目标湿度(%RH)", "湿度发生器目标湿度%"],
            threshold=3.0,
        )

    @staticmethod
    def _format_worksheet(ws: Any) -> None:
        if ws.max_row < 1 or ws.max_column < 1:
            return

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        sample_rows = min(ws.max_row, 100)
        for col in range(1, ws.max_column + 1):
            values: List[str] = []
            for row in range(1, sample_rows + 1):
                value = ws.cell(row=row, column=col).value
                if value is None:
                    continue
                values.append(str(value))
            max_len = max((len(v) for v in values), default=8)
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max(max_len + 2, 10), 36)

    def log_point_samples(
        self,
        point_row: int,
        rows: List[Dict[str, Any]],
        phase: Optional[str] = None,
        tag: Optional[str] = None,
    ) -> Path:
        """Write one CSV for one calibration point's stable sample rows."""
        point_id = int(point_row)
        suffix_parts = []
        phase_text = _safe_suffix(str(phase or ""))
        if phase_text:
            suffix_parts.append(phase_text)
        tag_text = _safe_suffix(str(tag or ""))
        if tag_text:
            suffix_parts.append(tag_text)
        suffix = "_" + "_".join(suffix_parts) if suffix_parts else ""
        out_path = self.run_dir / f"point_{point_id:04d}{suffix}_samples.csv"
        if not rows:
            out_path.write_text("", encoding="utf-8")
            return out_path

        translated_rows = [_translate_row(row) for row in rows]
        fields: List[str] = []
        seen = set()
        for row in translated_rows:
            for key in row.keys():
                if key in seen:
                    continue
                seen.add(key)
                fields.append(str(key))

        with out_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(translated_rows)
        return out_path

    def log_analyzer_workbook(
        self,
        rows: List[Dict[str, Any]],
        *,
        analyzer_labels: List[str],
        phase: str = "",
        write_summary: bool = True,
    ) -> Path:
        if not rows or not analyzer_labels:
            return self._analyzer_book_path_for_phase(phase)
        self._log_analyzer_summary_role_once()
        if write_summary:
            self.log_analyzer_summary(rows, analyzer_labels=analyzer_labels)

        out_path = self._analyzer_book_path_for_phase(phase)

        if out_path.exists():
            wb = load_workbook(out_path)
        else:
            wb = Workbook()
            if wb.sheetnames == ["Sheet"]:
                wb.remove(wb["Sheet"])

        try:
            row_keys: List[str] = []
            seen_keys = set()
            for row in rows:
                for key in row.keys():
                    text = str(key)
                    if text in seen_keys:
                        continue
                    seen_keys.add(text)
                    row_keys.append(text)

            common_dynamic = [
                key
                for key in row_keys
                if not re.match(r"^ga\d+_", key)
                and key not in _ANALYZER_SAMPLE_FIELDS
            ]
            common_fields = _dedupe_keys_by_label(_ordered_keys(common_dynamic, _COMMON_SHEET_FIELDS))

            for label in analyzer_labels:
                prefix = _safe_suffix(label)
                analyzer_id = None
                id_key = f"{prefix}_id"
                for row in rows:
                    value = row.get(id_key)
                    if value not in (None, ""):
                        analyzer_id = value
                        break
                sheet_name = _analyzer_sheet_name(label, analyzer_id)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(title=sheet_name)

                analyzer_dynamic = [
                    key[len(prefix) + 1 :]
                    for key in row_keys
                    if key.startswith(prefix + "_")
                ]
                analyzer_fields = _dedupe_keys_by_label(_ordered_keys(analyzer_dynamic, _ANALYZER_SAMPLE_FIELDS))

                header_keys = common_fields + analyzer_fields
                header_labels = [_field_label(key) for key in header_keys]
                existing = _trim_trailing_empty_labels(
                    [ws.cell(row=1, column=i + 1).value for i in range(max(ws.max_column, 1))]
                )
                is_blank_header = not existing
                if is_blank_header:
                    ws.append(header_labels)
                    if ws.max_row >= 2 and ws["A1"].value is None:
                        ws.delete_rows(1, 1)
                else:
                    header_keys = _merge_sheet_header_keys(existing, header_keys)
                    header_labels = [_field_label(key) for key in header_keys]
                    if len(existing) < len(header_labels):
                        for offset, label_text in enumerate(header_labels[len(existing) :], start=len(existing) + 1):
                            ws.cell(row=1, column=offset).value = label_text

                for row in rows:
                    sheet_row: Dict[str, Any] = {}
                    for key in common_fields:
                        sheet_row[key] = row.get(key)
                    for key in analyzer_fields:
                        prefixed_key = f"{prefix}_{key}"
                        value = row.get(prefixed_key)
                        if value is None and prefix == "ga01":
                            value = row.get(key)
                        sheet_row[key] = value
                    ws.append([sheet_row.get(key) for key in header_keys])

                self._format_worksheet(ws)

            _save_workbook_atomic(wb, out_path)
        finally:
            wb.close()
        return out_path

    def log_analyzer_summary(
        self,
        rows: List[Dict[str, Any]],
        *,
        analyzer_labels: List[str],
    ) -> Path:
        if not rows:
            return self.analyzer_summary_book_path
        self._log_analyzer_summary_role_once()
        csv_failures: List[str] = []
        build_failures: List[str] = []
        for label in analyzer_labels:
            try:
                summary_num = self._next_analyzer_summary_num(label)
                summary_row = self._build_analyzer_summary_row(rows, label=label, num=summary_num)
            except Exception as exc:
                build_failures.append(str(label))
                self.log_io(
                    port="LOG",
                    device="run_logger",
                    direction="WARN",
                    command="analyzer-summary-build",
                    response=f"label={str(label or '').upper()} skipped",
                    error=exc,
                )
                continue
            try:
                self._append_analyzer_summary_csv_row(label, summary_row)
            except Exception as exc:
                csv_failures.append(str(label))
                self.log_io(
                    port="LOG",
                    device="run_logger",
                    direction="WARN",
                    command="analyzer-summary-csv",
                    response=f"label={str(label or '').upper()} skipped",
                    error=exc,
                )
                continue
            try:
                self._append_analyzer_summary_workbook_row(label, summary_row)
            except Exception as exc:
                self.log_io(
                    port="LOG",
                    device="run_logger",
                    direction="WARN",
                    command="analyzer-summary-workbook",
                    response=f"label={str(label or '').upper()} csv-kept",
                    error=exc,
                )
        if build_failures or csv_failures:
            problems = []
            if build_failures:
                problems.append(f"build={','.join(build_failures)}")
            if csv_failures:
                problems.append(f"csv={','.join(csv_failures)}")
            raise RuntimeError("analyzer summary partial failures: " + "; ".join(problems))
        return self.analyzer_summary_book_path

    def _analyzer_book_path_for_phase(self, phase: str) -> Path:
        text = str(phase or "").strip().lower()
        if text == "co2":
            return self.co2_analyzer_book_path
        return self.h2o_analyzer_book_path

    def log_io(
        self,
        *,
        port: str,
        device: str,
        direction: str,
        command: Any = None,
        response: Any = None,
        error: Any = None,
    ) -> None:
        """Append one IO trace row."""
        limit = 400 if str(direction or "").strip().upper() in {"EVENT", "WARN"} else 160
        self._io_writer.writerow(
            {
                "timestamp": _utc_ts(),
                "port": port,
                "device": device,
                "direction": direction,
                "command": _summary_text(command, limit=limit),
                "response": _summary_text(response, limit=limit),
                "error": _summary_text(error, limit=limit),
            }
        )
        self._io_file.flush()

    @staticmethod
    def _delete_path_if_empty(path: Optional[Path]) -> None:
        if not isinstance(path, Path) or not path.exists():
            return
        try:
            if path.stat().st_size <= 0:
                path.unlink()
        except OSError:
            pass

    def _prune_empty_core_exports(self) -> None:
        if not self._samples_rows:
            self._delete_path_if_empty(self.samples_path)
        if not self._points_rows:
            self._delete_path_if_empty(self.points_path)
        if not self._points_readable_rows:
            self._delete_path_if_empty(self.points_readable_path)
        if not self._coefficient_write_rows:
            self._delete_path_if_empty(self.coefficient_write_path)
        if not list(self._analyzer_summary_rows_by_target.get("all") or []):
            self._delete_path_if_empty(self.analyzer_summary_csv_path)
            self._delete_path_if_empty(self.h2o_analyzer_summary_csv_path)
            self._delete_path_if_empty(self.co2_analyzer_summary_csv_path)

    def close(self) -> None:
        try:
            self._samples_file.close()
        finally:
            try:
                self._points_file.close()
            finally:
                try:
                    self._points_readable_file.close()
                finally:
                    try:
                        self._coefficient_write_file.close()
                    finally:
                        try:
                            self._analyzer_summary_file.close()
                        finally:
                            try:
                                for handle in self._analyzer_summary_phase_files.values():
                                    try:
                                        handle.close()
                                    except Exception:
                                        pass
                            finally:
                                self._io_file.close()
        self._prune_empty_core_exports()
