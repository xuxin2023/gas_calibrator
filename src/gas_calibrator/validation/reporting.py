"""Shared validation report writers.

These helpers are intentionally sidecar-only. They are used by validation tools
that run without changing the V1 production calibration workflow timing.
"""

from __future__ import annotations

import csv
import json
import os
import re
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


_EXCEL_WORKSHEET_TITLE_LIMIT = 31
_INVALID_WORKSHEET_TITLE_RE = re.compile(r"[\x00-\x1f\\*?:/\[\]]")


@dataclass
class ValidationMetadata:
    """Common metadata persisted by sidecar validation tools."""

    tool_name: str
    created_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))
    analyzers: List[str] = field(default_factory=list)
    input_paths: List[str] = field(default_factory=list)
    output_dir: str = ""
    config_path: str = ""
    config_summary: Dict[str, Any] = field(default_factory=dict)
    notes: List[str] = field(default_factory=list)


def _normalize_table_rows(rows: Iterable[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    for row in rows:
        normalized.append(dict(row))
    return normalized


def _table_header(rows: List[Dict[str, Any]]) -> List[str]:
    header: List[str] = []
    for row in rows:
        for key in row.keys():
            if key not in header:
                header.append(str(key))
    return header


def _fs_path(path: Path) -> Path:
    """Return a filesystem path usable for Windows paths beyond MAX_PATH."""

    if os.name != "nt":
        return path
    resolved = path if path.is_absolute() else path.resolve()
    text = str(resolved)
    if text.startswith("\\\\?\\"):
        return resolved
    if text.startswith("\\\\"):
        return Path("\\\\?\\UNC\\" + text.lstrip("\\"))
    return Path("\\\\?\\" + text)


def _write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    _fs_path(path.parent).mkdir(parents=True, exist_ok=True)
    header = _table_header(rows)
    with _fs_path(path).open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=header)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _autosize_sheet(ws) -> None:
    for column_cells in ws.columns:
        width = 10
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(60, len(value) + 2))
        ws.column_dimensions[column_cells[0].column_letter].width = width


def _safe_worksheet_title(value: object, existing_titles: Iterable[str]) -> str:
    base = _INVALID_WORKSHEET_TITLE_RE.sub("_", str(value)).strip() or "sheet"
    existing = {str(title).casefold() for title in existing_titles}
    candidate = base[:_EXCEL_WORKSHEET_TITLE_LIMIT]
    if candidate.casefold() not in existing:
        return candidate
    index = 2
    while True:
        suffix = f"_{index}"
        candidate = f"{base[: _EXCEL_WORKSHEET_TITLE_LIMIT - len(suffix)]}{suffix}"
        if candidate.casefold() not in existing:
            return candidate
        index += 1


def write_validation_report(
    output_dir: str | Path,
    *,
    prefix: str,
    metadata: ValidationMetadata,
    tables: Mapping[str, Iterable[Mapping[str, Any]]],
) -> Dict[str, Path]:
    """Write a validation workbook plus per-table CSV artifacts."""

    root = Path(output_dir).resolve()
    _fs_path(root).mkdir(parents=True, exist_ok=True)

    workbook_path = root / f"{prefix}.xlsx"
    metadata_path = root / f"{prefix}_meta.json"
    workbook = Workbook()
    meta_ws = workbook.active
    meta_ws.title = "meta"
    meta_ws.append(["field", "value"])
    for key, value in asdict(metadata).items():
        if isinstance(value, (dict, list)):
            rendered = json.dumps(value, ensure_ascii=False, indent=2)
        else:
            rendered = value
        meta_ws.append([key, rendered])
    meta_ws.freeze_panes = "A2"
    for cell in meta_ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")
    _autosize_sheet(meta_ws)

    outputs: Dict[str, Path] = {"workbook": workbook_path, "metadata": metadata_path}
    for table_name, raw_rows in tables.items():
        rows = _normalize_table_rows(raw_rows)
        csv_path = root / f"{table_name}.csv"
        _write_csv(csv_path, rows)
        outputs[f"{table_name}_csv"] = csv_path

        ws = workbook.create_sheet(
            title=_safe_worksheet_title(table_name, workbook.sheetnames)
        )
        header = _table_header(rows)
        if header:
            ws.append(header)
            for row in rows:
                ws.append([row.get(key) for key in header])
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = Font(bold=True)
        _autosize_sheet(ws)

    workbook.save(str(_fs_path(workbook_path)))
    _fs_path(metadata_path).write_text(
        json.dumps(asdict(metadata), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    workbook.close()
    return outputs
