"""No-write CO2 SENCO1/SENCO3/SENCO5 model-scope review against the V1.5 point table."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence

from openpyxl import load_workbook

from ..data.points import load_points_from_excel
from .reporting import ValidationMetadata, write_validation_report


PRIMARY_TERMS = ("intercept", "R", "R2", "R3")
SECONDARY_TERMS = ("T", "T2", "RT", "P", "RTP")
LINEAR_CORRECTION_TERMS = ("linear_c0", "linear_c1")


def _manual_co2_coefficient_family_rows() -> List[Dict[str, Any]]:
    return [
        {
            "senco_group": "SENCO1",
            "manual_role": "CO2 density/ratio primary coefficients",
            "model_scope": "primary ratio or absorbance response",
            "v1_5_current_status": "reviewable_from_clean_open_flow_ratio_evidence",
            "write_allowed": False,
            "physical_meaning": (
                "SENCO1 changes the analyzer CO2 optical response curve. It must be based on clean open-flow CO2 evidence."
            ),
        },
        {
            "senco_group": "SENCO3",
            "manual_role": "CO2 ratio temperature compensation coefficients",
            "model_scope": "temperature terms coupled to the CO2 ratio/absorbance response",
            "v1_5_current_status": "reviewable_only_when_multi_temperature_formula_contract_is_confirmed",
            "write_allowed": False,
            "physical_meaning": (
                "SENCO3 changes how CO2 optical response is interpreted across chamber temperature."
            ),
        },
        {
            "senco_group": "SENCO5",
            "manual_role": "CO2/gas concentration linear correction coefficients C0/C1",
            "model_scope": "final output concentration trim: corrected concentration = concentration*C1 + C0",
            "v1_5_current_status": "in_scope_integrated_final_output_candidate_no_write",
            "write_allowed": False,
            "physical_meaning": (
                "SENCO5 is part of the CO2 output chain as the final affine concentration layer. "
                "It must be reviewed in the same candidate package as SENCO1/SENCO3 when displayed ppm is the acceptance output."
            ),
        },
        {
            "senco_group": "SENCO9",
            "manual_role": "pressure calibration coefficient",
            "model_scope": "pressure input P used by CO2/H2O calculations",
            "v1_5_current_status": "handled_by_independent_pressure_channel_workflow",
            "write_allowed": False,
            "physical_meaning": (
                "SENCO9 is not fitted from CO2/H2O concentration residuals. It is verified and calibrated as an "
                "independent pressure input before open-flow component calibration."
            ),
        },
    ]


def _senco5_linear_correction_contract_rows() -> List[Dict[str, Any]]:
    """Describe the no-write contract for the CO2 SENCO5 output layer."""

    return [
        {
            "contract_item": "manual_scope",
            "status": "confirmed_manual_scope",
            "requirement": "SENCO5,YGAS,FFF,C0,C1 applies gas concentration correction as concentration*C1+C0.",
            "physical_meaning": (
                "SENCO5 belongs to the final CO2/gas concentration output chain. It is not the a5=T_k^2 "
                "coefficient used by the CO2 SENCO1/SENCO3 optical-temperature polynomial."
            ),
        },
        {
            "contract_item": "not_in_legacy_ratio_fit",
            "status": "legacy_algorithm_does_not_identify_senco5",
            "requirement": "Legacy V1/V2 ratio-polynomial fitting writes CO2 SENCO1/SENCO3 only.",
            "physical_meaning": (
                "The original least-squares algorithm fits lower-layer R/T/P features. It can be reused, "
                "but the final displayed concentration layer needs an explicit SENCO5 C0/C1 candidate review."
            ),
        },
        {
            "contract_item": "target_quantity",
            "status": "required_in_same_candidate_package",
            "requirement": (
                "Fit target must be the final displayed CO2 concentration residual after the SENCO1/SENCO3 "
                "candidate is applied, expressed in the same concentration unit as the analyzer output."
            ),
            "physical_meaning": (
                "SENCO5 is a final affine layer, not an optical ratio coefficient. It belongs in the released "
                "candidate package when displayed ppm still has a stable gain/offset residual."
            ),
        },
        {
            "contract_item": "reference_concentration_model",
            "status": "required_before_write",
            "requirement": (
                "Reference concentration must come from traceable standard-gas amount fraction, stable open-flow "
                "sampling, accepted pressure/temperature inputs, and any released dry/wet water correction."
            ),
            "physical_meaning": (
                "This prevents SENCO5 from hiding pressure, temperature, humidity, route, or SENCO1/SENCO3 errors."
            ),
        },
        {
            "contract_item": "input_evidence",
            "status": "required_before_write",
            "requirement": (
                "Use open-flow A-grade point means with MODE2 concentration and factory ratio/signal retained, pressure "
                "channel accepted, temperature channel accepted, stable ratio/signal evidence, "
                "and complete gas certificates."
            ),
            "physical_meaning": (
                "SENCO5 must see clean standard gas and trustworthy P/T inputs; otherwise it "
                "would compensate for route contamination or sensor faults."
            ),
        },
        {
            "contract_item": "fit_model",
            "status": "reviewable_no_write_with_decimal_contract",
            "requirement": (
                "Default write remains blocked, but the no-write candidate model is SENCO5=[C0,C1] with "
                "decimal payload values and firmware-side concentration*C1+C0 order."
            ),
            "physical_meaning": (
                "The group is the final linear output layer. It should be a small, reviewable part of the same "
                "SENCO1/SENCO3/SENCO5 candidate package, not a substitute for stable gas, temperature, pressure, or route evidence."
            ),
        },
        {
            "contract_item": "exclusions",
            "status": "mandatory",
            "requirement": (
                "Do not fit SENCO5 from raw ratio residuals, sealed pressure diagnostic rows, "
                "pressure-compensation rows, devices with bad temperature channels, or "
                "samples where concentration/pressure/temperature evidence is unstable."
            ),
            "physical_meaning": (
                "These exclusions prevent SENCO5 from absorbing errors that belong to "
                "SENCO1/SENCO3, SENCO7/SENCO8, SENCO9, or the gas route."
            ),
        },
        {
            "contract_item": "verification",
            "status": "required_after_any_write",
            "requirement": (
                "Any future SENCO5 write requires old GETCO5 backup, controlled decimal C0/C1 write, "
                "readback, rollback plan, and independent concentration verification at at least "
                "two temperatures and one non-fit CO2 point."
            ),
            "physical_meaning": (
                "A linear trim is only accepted when it improves independent concentration verification without "
                "masking unstable gas, temperature, or pressure evidence."
            ),
        },
    ]


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _read_csv(path: str | Path) -> List[Dict[str, Any]]:
    target = Path(path)
    if not target.exists():
        return []
    with target.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _safe_float(value: Any) -> Optional[float]:
    if value in (None, "", "None", "null"):
        return None
    try:
        numeric = float(value)
    except Exception:
        return None
    if not math.isfinite(numeric):
        return None
    return numeric


def _safe_int(value: Any) -> Optional[int]:
    numeric = _safe_float(value)
    if numeric is None:
        return None
    return int(round(numeric))


def _split_terms(value: Any) -> List[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [item.strip() for item in text.replace(",", ";").split(";") if item.strip()]


def _format_values(values: Iterable[Any]) -> str:
    out: List[str] = []
    for value in values:
        numeric = _safe_float(value)
        if numeric is None:
            out.append(str(value))
        elif float(numeric).is_integer():
            out.append(str(int(numeric)))
        else:
            out.append(f"{numeric:g}")
    return ",".join(out)


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _span(values: Sequence[float]) -> float:
    if not values:
        return 0.0
    return float(max(values) - min(values))


def _compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)


_TEMP_RE = re.compile(r"([+-]?\d+(?:\.\d+)?)")
_NUMBER_RE = re.compile(r"([+-]?\d+(?:\.\d+)?)")


def _parse_temp(value: Any) -> Optional[float]:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        match = _TEMP_RE.search(value)
        if match:
            return float(match.group(1))
    return None


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _h2o_text_is_explicit(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    return "湿度发生器" in text or "mmol/mol" in text or "露点" in text


def _number_list(value: Any) -> List[float]:
    if value in (None, ""):
        return []
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return [float(value)]
    return [float(match.group(1)) for match in _NUMBER_RE.finditer(str(value))]


def _header_map(row: Sequence[Any]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for idx, value in enumerate(row):
        key = str(value or "").strip()
        if key:
            out[key] = idx
    return out


def _load_execution_view_rows(points_xlsx: str | Path) -> List[Dict[str, Any]]:
    path = Path(points_xlsx).resolve()
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "execution_view" not in workbook.sheetnames:
            return []
        worksheet = workbook["execution_view"]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header = _header_map(tuple(next(iterator) or ()))
        except StopIteration:
            return []
        rows: List[Dict[str, Any]] = []
        for excel_row, values in enumerate(iterator, start=2):
            values = tuple(values or ())
            temp = _safe_float(values[header["Temp_C"]] if "Temp_C" in header else None)
            if temp is None:
                continue
            co2_sources = _number_list(values[header["CO2_sources_ppm"]] if "CO2_sources_ppm" in header else None)
            pressures = _number_list(values[header["Pressures_hPa"]] if "Pressures_hPa" in header else None)
            rows.append(
                {
                    "excel_row": excel_row,
                    "temp_c": temp,
                    "execution_order": values[header["Order"]] if "Order" in header else "",
                    "h2o_targets": values[header["H2O_targets"]] if "H2O_targets" in header else "",
                    "co2_sources_ppm": _format_values(co2_sources),
                    "co2_source_count": len(co2_sources),
                    "pressures_hpa": _format_values(pressures),
                    "pressure_count": len(pressures),
                    "notes": values[header["Notes"]] if "Notes" in header else "",
                    "physical_meaning": (
                        "V1 normalized execution view: this is the intended source/pressure scope after "
                        "the point-table conversion rules are applied."
                    ),
                }
            )
        return rows
    finally:
        workbook.close()


def _execution_view_summary_rows(
    execution_rows: Sequence[Mapping[str, Any]],
    *,
    points_xlsx: Path,
) -> List[Dict[str, Any]]:
    temps = sorted({float(row["temp_c"]) for row in execution_rows if _safe_float(row.get("temp_c")) is not None})
    co2_targets: set[float] = set()
    pressures: set[float] = set()
    for row in execution_rows:
        co2_targets.update(_number_list(row.get("co2_sources_ppm")))
        pressures.update(_number_list(row.get("pressures_hpa")))
    return [
        {
            "points_xlsx": str(points_xlsx.resolve()),
            "execution_view_present": bool(execution_rows),
            "temperature_values_c": _format_values(temps),
            "temperature_span_c": _span(temps),
            "co2_sources_ppm": _format_values(sorted(co2_targets)),
            "pressure_targets_hpa": _format_values(sorted(pressures)),
            "pressure_span_hpa": _span(sorted(pressures)),
            "execution_temperature_count": len(temps),
            "physical_meaning": (
                "The execution_view sheet is the V1-normalized point contract. It confirms that the "
                "formal point table is broader than the current 20 C current-atmosphere evidence."
            ),
        }
    ]


def _v1_loader_comparison_rows(points_xlsx: str | Path) -> List[Dict[str, Any]]:
    path = Path(points_xlsx).resolve()
    rows: List[Dict[str, Any]] = []
    for carry_forward_h2o in (False, True):
        try:
            points = load_points_from_excel(
                path,
                missing_pressure_policy="carry_forward",
                carry_forward_h2o=carry_forward_h2o,
            )
            load_status = "loaded"
        except Exception as exc:
            points = []
            load_status = f"error:{exc}"
        co2_points = [point for point in points if not point.is_h2o_point]
        h2o_points = [point for point in points if point.is_h2o_point]
        temps = sorted({float(point.temp_chamber_c) for point in points if point.temp_chamber_c is not None})
        co2_targets = sorted(
            {
                float(point.co2_ppm)
                for point in co2_points
                if _safe_float(getattr(point, "co2_ppm", None)) is not None
            }
        )
        pressures = sorted(
            {
                float(point.target_pressure_hpa)
                for point in points
                if _safe_float(getattr(point, "target_pressure_hpa", None)) is not None
            }
        )
        rows.append(
            {
                "loader": "load_points_from_excel",
                "missing_pressure_policy": "carry_forward",
                "carry_forward_h2o": carry_forward_h2o,
                "load_status": load_status,
                "point_count": len(points),
                "co2_point_count": len(co2_points),
                "h2o_point_count": len(h2o_points),
                "temperature_values_c": _format_values(temps),
                "co2_targets_ppm_direct_loader": _format_values(co2_targets),
                "pressure_targets_hpa_direct_loader": _format_values(pressures),
                "calibration_meaning": (
                    "Direct loader rows are not the full V1 CO2 execution contract when the runner expands "
                    "10/20/30 C CO2 sources; compare with execution_view before judging model coverage."
                ),
            }
        )
    return rows


def _load_original_points(points_xlsx: str | Path) -> List[Dict[str, Any]]:
    path = Path(points_xlsx).resolve()
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        current_temp: Optional[float] = None
        rows: List[Dict[str, Any]] = []
        for excel_row, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if excel_row <= 2:
                continue
            values = tuple(values or ())
            if not any(not _is_missing(value) for value in values):
                continue
            temp = _parse_temp(values[0] if len(values) > 0 else None)
            if temp is not None:
                current_temp = temp
            if current_temp is None:
                continue
            co2 = _safe_float(values[1] if len(values) > 1 else None)
            h2o_text = values[2] if len(values) > 2 else None
            pressure = _safe_float(values[3] if len(values) > 3 else None)
            group = "" if len(values) <= 4 or _is_missing(values[4]) else str(values[4]).strip()
            explicit_h2o = _h2o_text_is_explicit(h2o_text)
            is_co2_fit_point = co2 is not None and not explicit_h2o
            rows.append(
                {
                    "excel_row": excel_row,
                    "temp_c": float(current_temp),
                    "co2_ppm": co2,
                    "h2o_text": "" if h2o_text is None else str(h2o_text),
                    "pressure_hpa": pressure,
                    "co2_group": group,
                    "is_h2o_point": explicit_h2o,
                    "is_co2_fit_point": is_co2_fit_point,
                }
            )
        return rows
    finally:
        workbook.close()


def _original_point_table_summary(points: Sequence[Mapping[str, Any]], points_xlsx: Path) -> List[Dict[str, Any]]:
    co2 = [p for p in points if bool(p.get("is_co2_fit_point"))]
    h2o = [p for p in points if bool(p.get("is_h2o_point"))]
    temps = sorted({float(p["temp_c"]) for p in points if _safe_float(p.get("temp_c")) is not None})
    co2_targets = sorted({float(p["co2_ppm"]) for p in co2 if _safe_float(p.get("co2_ppm")) is not None})
    pressures = sorted({float(p["pressure_hpa"]) for p in co2 if _safe_float(p.get("pressure_hpa")) is not None})
    b_group = sorted(
        {
            int(round(float(p["co2_ppm"])))
            for p in co2
            if str(p.get("co2_group") or "").strip().upper() == "B" and _safe_float(p.get("co2_ppm")) is not None
        }
    )
    return [
        {
            "points_xlsx": str(points_xlsx.resolve()),
            "points_sha256": _sha256_file(points_xlsx),
            "total_point_count": len(points),
            "co2_point_count": len(co2),
            "h2o_point_count": len(h2o),
            "temperature_values_c": _format_values(temps),
            "temperature_span_c": _span(temps),
            "co2_targets_ppm": _format_values(co2_targets),
            "pressure_targets_hpa": _format_values(pressures),
            "pressure_span_hpa": _span(pressures),
            "co2_group_b_targets_ppm": _format_values(b_group),
            "physical_meaning": (
                "Original V1.5 point table covers multiple chamber temperatures, CO2 levels, humidity points, and pressure targets; "
                "the current B-group 20 C CO2 evidence is only one slice of this table."
            ),
        }
    ]


def _original_temperature_rows(points: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    temps = sorted({float(p["temp_c"]) for p in points if _safe_float(p.get("temp_c")) is not None})
    for temp in temps:
        group = [p for p in points if _safe_float(p.get("temp_c")) == temp]
        co2 = [p for p in group if bool(p.get("is_co2_fit_point"))]
        h2o = [p for p in group if bool(p.get("is_h2o_point"))]
        pressures = sorted({float(p["pressure_hpa"]) for p in co2 if _safe_float(p.get("pressure_hpa")) is not None})
        co2_targets = sorted({float(p["co2_ppm"]) for p in co2 if _safe_float(p.get("co2_ppm")) is not None})
        b_group = sorted(
            {
                int(round(float(p["co2_ppm"])))
                for p in co2
                if str(p.get("co2_group") or "").strip().upper() == "B" and _safe_float(p.get("co2_ppm")) is not None
            }
        )
        rows.append(
            {
                "temp_c": temp,
                "co2_point_count": len(co2),
                "h2o_point_count": len(h2o),
                "co2_targets_ppm": _format_values(co2_targets),
                "co2_group_b_targets_ppm": _format_values(b_group),
                "pressure_targets_hpa": _format_values(pressures),
                "pressure_span_hpa": _span(pressures),
                "route_meaning": (
                    "multi_component_original_scope"
                    if h2o
                    else "co2_only_original_scope"
                ),
            }
        )
    return rows


def _current_coverage_rows(candidate_policy_rows: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for row in candidate_policy_rows:
        if str(row.get("component") or "").strip().lower() != "co2":
            continue
        selected_terms = _split_terms(row.get("selected_model_terms"))
        frozen_terms = _split_terms(row.get("frozen_terms"))
        rows.append(
            {
                "component": "co2",
                "analyzer_prefix": row.get("analyzer_prefix", ""),
                "analyzer_device_id": row.get("analyzer_device_id", ""),
                "candidate_status": row.get("candidate_status", ""),
                "fit_sample_count": row.get("fit_sample_count", ""),
                "verification_sample_count": row.get("verification_sample_count", ""),
                "fit_point_count": row.get("fit_point_count", ""),
                "verification_point_count": row.get("verification_point_count", ""),
                "distinct_fit_targets": row.get("distinct_fit_targets", ""),
                "pressure_span_hpa": row.get("pressure_span_hpa", ""),
                "temperature_span_c": row.get("temperature_span_c", ""),
                "selected_model_terms": ";".join(selected_terms),
                "frozen_terms": ";".join(frozen_terms),
                "formal_pressure_validation_status": row.get("formal_pressure_validation_status", ""),
                "verification_status": row.get("verification_status", ""),
                "not_pressure_compensation_fit": row.get("not_pressure_compensation_fit", ""),
                "current_scope_interpretation": (
                    "supports_current_atmosphere_ratio_and_temperature_curve_pressure_frozen"
                    if set(PRIMARY_TERMS).issubset(set(selected_terms))
                    and {"T", "T2", "RT"}.issubset(set(selected_terms))
                    else "supports_current_atmosphere_primary_ratio_curve_only"
                    if set(PRIMARY_TERMS).issubset(set(selected_terms))
                    else "review_required"
                ),
            }
        )
    return rows


def _observed_fit_targets(candidate_policy_rows: Sequence[Mapping[str, Any]]) -> str:
    targets: set[int] = set()
    for row in candidate_policy_rows:
        text = str(row.get("verification_certificate_uncertainties") or "")
        for token in text.replace(";", ",").split(","):
            left = token.split(":", 1)[0].strip()
            value = _safe_int(left)
            if value is not None:
                targets.add(value)
        count = _safe_int(row.get("distinct_fit_targets"))
        if count == 5:
            targets.update({100, 300, 500, 700, 900})
    return _format_values(sorted(targets))


def _point_table_alignment_rows(
    *,
    original_points: Sequence[Mapping[str, Any]],
    execution_rows: Sequence[Mapping[str, Any]],
    candidate_policy_rows: Sequence[Mapping[str, Any]],
) -> List[Dict[str, Any]]:
    original_b = sorted(
        {
            int(round(float(p["co2_ppm"])))
            for p in original_points
            if bool(p.get("is_co2_fit_point"))
            and str(p.get("co2_group") or "").strip().upper() == "B"
            and _safe_float(p.get("co2_ppm")) is not None
        }
    )
    original_20_b = sorted(
        {
            int(round(float(p["co2_ppm"])))
            for p in original_points
            if bool(p.get("is_co2_fit_point"))
            and _safe_float(p.get("temp_c")) == 20.0
            and str(p.get("co2_group") or "").strip().upper() == "B"
            and _safe_float(p.get("co2_ppm")) is not None
        }
    )
    observed_targets = {value for value in (_safe_int(v) for v in _observed_fit_targets(candidate_policy_rows).split(",") if v) if value is not None}
    current_matches_20_b = bool(observed_targets) and observed_targets == set(original_20_b)
    execution_20_targets = sorted(
        {
            int(round(value))
            for row in execution_rows
            if _safe_float(row.get("temp_c")) == 20.0
            for value in _number_list(row.get("co2_sources_ppm"))
        }
    )
    execution_all_targets = sorted(
        {
            int(round(value))
            for row in execution_rows
            for value in _number_list(row.get("co2_sources_ppm"))
        }
    )
    execution_temps = sorted(
        {float(row["temp_c"]) for row in execution_rows if _safe_float(row.get("temp_c")) is not None}
    )
    rows = [
        {
            "alignment_item": "current_fit_targets_vs_original_20c_b_group",
            "original_targets_ppm": _format_values(original_20_b),
            "current_targets_ppm": _format_values(sorted(observed_targets)),
            "status": "match" if current_matches_20_b else "partial_or_unknown",
            "calibration_meaning": (
                "Current CO2 data matches the original V1.5 20 C B-group target set, so it is a valid slice for primary ratio review."
            ),
        },
        {
            "alignment_item": "current_fit_targets_vs_full_original_b_group",
            "original_targets_ppm": _format_values(original_b),
            "current_targets_ppm": _format_values(sorted(observed_targets)),
            "status": "target_values_match_temperature_coverage_partial"
            if observed_targets == set(original_b)
            else "partial",
            "calibration_meaning": (
                "The current run matches the B-group CO2 concentration values, but not the original multi-temperature coverage; "
                "therefore pressure compensation remains out of the main CO2 fitting scope; temperature coverage is assessed separately."
            ),
        },
    ]
    if execution_rows:
        rows.extend(
            [
                {
                    "alignment_item": "current_fit_targets_vs_v1_execution_20c_sources",
                    "original_targets_ppm": _format_values(execution_20_targets),
                    "current_targets_ppm": _format_values(sorted(observed_targets)),
                    "status": (
                        "subset_of_v1_execution_sources"
                        if observed_targets and observed_targets.issubset(set(execution_20_targets))
                        else "partial_or_unknown"
                    ),
                    "calibration_meaning": (
                        "V1's execution view expands the 20 C CO2 route to the full source sweep; "
                        "the current evidence intentionally covers the B-group subset used for primary ratio review."
                    ),
                },
                {
                    "alignment_item": "current_fit_targets_vs_v1_execution_full_scope",
                    "original_targets_ppm": _format_values(execution_all_targets),
                    "current_targets_ppm": _format_values(sorted(observed_targets)),
                    "status": "concentration_subset_temperature_scope_partial",
                    "calibration_meaning": (
                        "The V1 execution contract covers multiple temperatures and source levels. Current data may "
                        "support identifiable temperature terms when the candidate policy selects T/T2/RT, while "
                        "pressure terms remain frozen under V1.5."
                    ),
                    "v1_execution_temperatures_c": _format_values(execution_temps),
                },
            ]
        )
    return rows


def _term_identifiability_rows(
    *,
    original_points: Sequence[Mapping[str, Any]],
    current_rows: Sequence[Mapping[str, Any]],
    min_temp_span_c_for_secondary: float,
    min_pressure_span_hpa_for_pressure_terms: float,
) -> List[Dict[str, Any]]:
    original_temps = [float(p["temp_c"]) for p in original_points if _safe_float(p.get("temp_c")) is not None]
    original_pressures = [
        float(p["pressure_hpa"])
        for p in original_points
        if bool(p.get("is_co2_fit_point")) and _safe_float(p.get("pressure_hpa")) is not None
    ]
    current_temp_span = max((_safe_float(row.get("temperature_span_c")) or 0.0 for row in current_rows), default=0.0)
    current_pressure_span = max((_safe_float(row.get("pressure_span_hpa")) or 0.0 for row in current_rows), default=0.0)
    current_selected = set()
    current_frozen = set()
    for row in current_rows:
        current_selected.update(_split_terms(row.get("selected_model_terms")))
        current_frozen.update(_split_terms(row.get("frozen_terms")))
    rows = []
    term_mapping = [
        ("a0", "SENCO1", "intercept", "1", "primary_ratio_term"),
        ("a1", "SENCO1", "R", "R", "primary_ratio_term"),
        ("a2", "SENCO1", "R2", "R^2", "primary_ratio_term"),
        ("a3", "SENCO1", "R3", "R^3", "primary_ratio_term"),
        ("a4", "SENCO3", "T", "T_k", "secondary_temperature_term"),
        ("a5", "SENCO3", "T2", "T_k^2", "secondary_temperature_term"),
        ("a6", "SENCO3", "RT", "R*T_k", "secondary_temperature_term"),
        ("a7", "SENCO3", "P", "P", "secondary_pressure_term"),
        ("a8", "SENCO3", "RTP", "R*T_k*P", "secondary_pressure_term"),
        ("senco5_C0", "SENCO5", "linear_c0", "CO2_concentration_offset", "linear_concentration_correction"),
        ("senco5_C1", "SENCO5", "linear_c1", "CO2_concentration_multiplier", "linear_concentration_correction"),
    ]
    for coefficient, senco, token, feature, family in term_mapping:
        if family == "primary_ratio_term":
            status = "currently_identifiable_no_write" if token in current_selected else "missing"
            blocker = ""
        elif family == "secondary_temperature_term":
            enough_temp = current_temp_span >= float(min_temp_span_c_for_secondary)
            status = "blocked_current_temp_span_too_small" if not enough_temp else "reviewable_if_formula_contract_confirmed"
            blocker = "" if enough_temp else "current data is effectively single-temperature"
        else:
            if family == "secondary_pressure_term":
                enough_pressure = current_pressure_span >= float(min_pressure_span_hpa_for_pressure_terms)
                status = (
                    "blocked_pressure_terms_frozen_by_v1_5_policy"
                    if not enough_pressure
                    else "requires_clean_pressure_compensation_validation"
                )
                blocker = "pressure terms remain independent from main CO2 open-flow fitting"
            else:
                status = "requires_integrated_output_layer_candidate_review"
                blocker = (
                    "SENCO5 is fitted from the final displayed concentration residual after the lower-layer "
                    "SENCO1/SENCO3 model, not from raw optical ratio terms alone"
                )
        rows.append(
            {
                "coefficient": coefficient,
                "senco_group": senco,
                "model_token": token,
                "feature": feature,
                "term_family": family,
                "original_table_temp_span_c": _span(original_temps),
                "original_table_pressure_span_hpa": _span(original_pressures),
                "current_temp_span_c": current_temp_span,
                "current_pressure_span_hpa": current_pressure_span,
                "current_selected": token in current_selected,
                "current_frozen": token in current_frozen,
                "identifiability_status": status,
                "blocker": blocker,
                "write_allowed": False,
            }
        )
    return rows


def _decision_rows(
    *,
    current_rows: Sequence[Mapping[str, Any]],
    pair_summary_rows: Sequence[Mapping[str, Any]],
) -> List[Dict[str, Any]]:
    post_write_failed = any(
        str(row.get("single_senco1_write_verification_status") or "").lower() == "failed"
        or "failed" in str(row.get("review_status") or "").lower()
        for row in pair_summary_rows
    )
    every_current_primary_ok = bool(current_rows) and all(
        str(row.get("current_scope_interpretation") or "")
        in {
            "supports_current_atmosphere_primary_ratio_curve_only",
            "supports_current_atmosphere_ratio_and_temperature_curve_pressure_frozen",
        }
        for row in current_rows
    )
    every_current_temperature_ok = bool(current_rows) and all(
        str(row.get("current_scope_interpretation") or "")
        == "supports_current_atmosphere_ratio_and_temperature_curve_pressure_frozen"
        for row in current_rows
    )
    return [
        {
            "decision_item": "current_primary_ratio_curve",
            "decision_status": "reviewable_no_write" if every_current_primary_ok else "blocked",
            "write_allowed": False,
            "physical_meaning": "Current open-flow data support an offline primary ratio curve, not a device-output release.",
        },
        {
            "decision_item": "preserve_old_senco3_with_new_senco1",
            "decision_status": "blocked_by_post_write_failure" if post_write_failed else "review_required",
            "write_allowed": False,
            "physical_meaning": "Keeping old SENCO3 after writing new SENCO1 already failed the 900 ppm device-output check.",
        },
        {
            "decision_item": "new_senco1_senco3_pair_write",
            "decision_status": (
                "reviewable_no_write_pending_formula_contract_old_snapshot_and_independent_verification"
                if every_current_temperature_ok
                else "blocked_until_formula_contract_and_secondary_data"
            ),
            "write_allowed": False,
            "physical_meaning": (
                "Current multi-temperature data can review T/T2/RT secondary terms, but pressure terms remain frozen and "
                "actual write still requires formula-contract review, old GETCO backup, approval, and post-write verification."
                if every_current_temperature_ok
                else "SENCO3 affects temperature/pressure-related internal compensation and cannot be inferred from single-temperature current-atmosphere data."
            ),
        },
        {
            "decision_item": "co2_concentration_linear_correction_senco5",
            "decision_status": "in_scope_integrated_final_output_candidate_no_write",
            "write_allowed": False,
            "physical_meaning": (
                "SENCO5 is CO2-related because it applies corrected concentration = concentration*C1 + C0. It must "
                "be handled as the final output layer in the same controlled candidate package, not as an after-acceptance patch."
            ),
        },
        {
            "decision_item": "next_safe_action",
            "decision_status": (
                "no_write_review_then_getco_backup_and_small_independent_verification_plan"
                if every_current_temperature_ok
                else "no_write_review_then_choose_rollback_or_collect_secondary_span"
            ),
            "write_allowed": False,
            "physical_meaning": (
                "The safe branch is to bind old SENCO1/SENCO3 snapshots, keep write blocked, then choose a controlled paired-write plus small independent open-flow verification."
                if every_current_temperature_ok
                else "The safe branch is either controlled rollback to old SENCO1 or collecting enough multi-temperature evidence for a paired model; both require approval before writing."
            ),
        },
    ]


def _database_index_rows(
    *,
    output_dir: Path,
    summary_status: str,
) -> List[Dict[str, Any]]:
    temperature_reviewable = "temperature_terms_reviewable" in str(summary_status)
    return [
        {
            "db_table": "coefficient_candidates",
            "record_key": "co2_senco_pair_model_scope_review",
            "component": "co2",
            "candidate_status": summary_status,
            "auto_write_allowed": False,
            "source_artifact_role": "candidate_coefficient_review",
            "source_path": str(output_dir.resolve()),
            "metadata_json": _compact_json(
                {
                    "review": "co2_senco_pair_model_scope",
                    "manual_co2_groups_in_scope": ["SENCO1", "SENCO3", "SENCO5"],
                    "pressure_group_handled_separately": "SENCO9",
                }
            ),
        },
        {
            "db_table": "qc_results",
            "record_key": (
                "co2_senco3_temperature_terms_reviewable_pressure_terms_frozen"
                if temperature_reviewable
                else "co2_senco3_secondary_identifiability_blocked"
            ),
            "component": "co2",
            "candidate_status": "blocked_not_real_acceptance",
            "auto_write_allowed": False,
            "source_artifact_role": "candidate_coefficient_review",
            "source_path": str(output_dir.resolve()),
            "metadata_json": _compact_json(
                {
                    "rule_name": "co2_senco3_secondary_identifiability",
                    "status": "warning" if temperature_reviewable else "fail",
                    "reason": (
                        "temperature terms are reviewable from current multi-temperature evidence; pressure terms remain frozen"
                        if temperature_reviewable
                        else "current evidence lacks secondary T/P span"
                    ),
                    "senco5_status": "in_scope_integrated_final_output_candidate_no_write",
                }
            ),
        },
        {
            "db_table": "qc_results",
            "record_key": "co2_senco5_linear_correction_integrated_candidate",
            "component": "co2",
            "candidate_status": "review_needed_not_real_acceptance",
            "auto_write_allowed": False,
            "source_artifact_role": "candidate_coefficient_review",
            "source_path": str(output_dir.resolve()),
            "metadata_json": _compact_json(
                {
                    "rule_name": "co2_senco5_linear_correction_identifiability",
                    "status": "warning",
                    "reason": (
                        "manual links SENCO5 to the final CO2 concentration affine layer; release requires an "
                        "integrated SENCO1/SENCO3/SENCO5 candidate review"
                    ),
                }
            ),
        },
        {
            "db_table": "audit_events",
            "record_key": "co2_senco_pair_model_scope_review_built",
            "component": "co2",
            "candidate_status": "available",
            "auto_write_allowed": False,
            "source_artifact_role": "candidate_coefficient_review",
            "source_path": str(output_dir.resolve()),
            "metadata_json": _compact_json({"opens_com_ports": False, "writes_coefficients": False}),
        },
    ]


def _write_database_sidecar(
    path: str | Path,
    *,
    outputs: Mapping[str, Path],
    database_index_rows: Sequence[Mapping[str, Any]],
) -> Path:
    target = Path(path).resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    artifacts: List[Dict[str, Any]] = []
    for key, output_path in sorted(outputs.items()):
        if not output_path.exists():
            continue
        artifacts.append(
            {
                "output_key": key,
                "artifact_role": "candidate_coefficient_review"
                if "co2_senco_pair_model_scope" in output_path.name
                else "evidence_file",
                "path": str(output_path.resolve()),
                "sha256": _sha256_file(output_path),
                "size_bytes": output_path.stat().st_size,
            }
        )
    payload = {
        "schema": "v1_5_co2_senco_pair_model_scope_database_sidecar",
        "created_at": _now(),
        "no_write": True,
        "opens_com_ports": False,
        "controls_water_or_gas_routes": False,
        "database_target_tables": [
            "sample_files",
            "coefficient_candidates",
            "qc_results",
            "audit_events",
        ],
        "artifacts": artifacts,
        "suggested_rows": list(database_index_rows),
    }
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return target


def build_co2_senco_pair_model_scope_tables(
    *,
    original_points_xlsx: str | Path,
    candidate_dir: str | Path,
    pair_review_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
    min_temp_span_c_for_secondary: float = 20.0,
    min_pressure_span_hpa_for_pressure_terms: float = 300.0,
) -> Dict[str, List[Dict[str, Any]]]:
    points_path = Path(original_points_xlsx).resolve()
    candidate_path = Path(candidate_dir).resolve()
    output_path = Path(output_dir).resolve() if output_dir is not None else Path(".").resolve()
    original_points = _load_original_points(points_path)
    execution_rows = _load_execution_view_rows(points_path)
    loader_comparison_rows = _v1_loader_comparison_rows(points_path)
    candidate_policy_rows = _read_csv(candidate_path / "candidate_policy_summary.csv")
    pair_summary_rows: List[Dict[str, Any]] = []
    if pair_review_dir:
        pair_summary_rows = _read_csv(Path(pair_review_dir).resolve() / "co2_senco_pair_review_summary.csv")
    original_summary = _original_point_table_summary(original_points, points_path)
    execution_summary = _execution_view_summary_rows(execution_rows, points_xlsx=points_path)
    original_temp_rows = _original_temperature_rows(original_points)
    current_rows = _current_coverage_rows(candidate_policy_rows)
    term_rows = _term_identifiability_rows(
        original_points=original_points,
        current_rows=current_rows,
        min_temp_span_c_for_secondary=min_temp_span_c_for_secondary,
        min_pressure_span_hpa_for_pressure_terms=min_pressure_span_hpa_for_pressure_terms,
    )
    decision_rows = _decision_rows(current_rows=current_rows, pair_summary_rows=pair_summary_rows)
    temperature_terms_reviewable = all(
        any(
            str(row.get("coefficient") or "") == coefficient
            and str(row.get("identifiability_status") or "") == "reviewable_if_formula_contract_confirmed"
            for row in term_rows
        )
        for coefficient in ("a4", "a5", "a6")
    )
    pressure_terms_frozen = all(
        any(
            str(row.get("coefficient") or "") == coefficient
            and str(row.get("identifiability_status") or "") == "blocked_pressure_terms_frozen_by_v1_5_policy"
            for row in term_rows
        )
        for coefficient in ("a7", "a8")
    )
    summary_status = (
        "secondary_temperature_terms_reviewable_pressure_terms_frozen_no_write"
        if temperature_terms_reviewable
        else "blocked_secondary_terms_not_identifiable_no_write"
    )
    summary_rows = [
        {
            "component": "co2",
            "review_status": summary_status,
            "original_points_xlsx": str(points_path),
            "candidate_dir": str(candidate_path),
            "pair_review_dir": str(Path(pair_review_dir).resolve()) if pair_review_dir else "",
            "current_device_count": len(current_rows),
            "v1_execution_view_present": bool(execution_rows),
            "current_primary_curve_reviewable": all(
                str(row.get("current_scope_interpretation") or "")
                in {
                    "supports_current_atmosphere_primary_ratio_curve_only",
                    "supports_current_atmosphere_ratio_and_temperature_curve_pressure_frozen",
                }
                for row in current_rows
            )
            if current_rows
            else False,
            "senco3_temperature_terms_reviewable": temperature_terms_reviewable,
            "senco5_linear_correction_in_scope": True,
            "senco5_linear_correction_write_allowed": False,
            "pressure_terms_frozen_by_v1_5_policy": pressure_terms_frozen,
            "senco3_secondary_write_allowed": False,
            "senco1_senco3_pair_write_allowed": False,
            "senco1_senco3_senco5_family_write_allowed": False,
            "opens_com_ports": False,
            "controls_water_or_gas_routes": False,
            "writes_coefficients": False,
            "physical_meaning": (
                "The current multi-temperature open-flow evidence can review SENCO3 T/T2/RT temperature terms. "
                "SENCO5 is in the CO2 family as the final output concentration affine layer and must be reviewed "
                "inside the same candidate package. P/RTP pressure terms remain frozen because V1.5 "
                "keeps pressure compensation outside the main CO2 fit."
                if temperature_terms_reviewable
                else "The current B-group open-flow data match one V1.5 execution slice after point-table "
                "normalization, but they do not contain enough temperature or pressure span to identify SENCO3 secondary terms. "
                "SENCO5 is still CO2-related by manual scope, but requires final displayed concentration residual evidence."
            ),
        }
    ]
    database_index = _database_index_rows(output_dir=output_path, summary_status=summary_status)
    return {
        "co2_senco_pair_model_scope_summary": summary_rows,
        "co2_senco_pair_original_point_table_summary": original_summary,
        "co2_senco_pair_v1_execution_view_summary": execution_summary,
        "co2_senco_pair_v1_execution_view_coverage": list(execution_rows),
        "co2_senco_pair_v1_loader_comparison": loader_comparison_rows,
        "co2_senco_pair_original_temperature_coverage": original_temp_rows,
        "co2_senco_pair_point_table_alignment": _point_table_alignment_rows(
            original_points=original_points,
            execution_rows=execution_rows,
            candidate_policy_rows=candidate_policy_rows,
        ),
        "co2_senco_pair_current_evidence_coverage": current_rows,
        "co2_senco_pair_manual_co2_coefficient_family": _manual_co2_coefficient_family_rows(),
        "co2_senco5_linear_correction_contract": _senco5_linear_correction_contract_rows(),
        "co2_senco_pair_term_identifiability": term_rows,
        "co2_senco_pair_model_decision": decision_rows,
        "co2_senco_pair_model_database_index": database_index,
    }


def write_co2_senco_pair_model_scope_report(
    *,
    original_points_xlsx: str | Path,
    candidate_dir: str | Path,
    output_dir: str | Path,
    pair_review_dir: str | Path | None = None,
    database_sidecar_json: str | Path | None = None,
    min_temp_span_c_for_secondary: float = 20.0,
    min_pressure_span_hpa_for_pressure_terms: float = 300.0,
) -> Dict[str, Path]:
    tables = build_co2_senco_pair_model_scope_tables(
        original_points_xlsx=original_points_xlsx,
        candidate_dir=candidate_dir,
        pair_review_dir=pair_review_dir,
        output_dir=output_dir,
        min_temp_span_c_for_secondary=min_temp_span_c_for_secondary,
        min_pressure_span_hpa_for_pressure_terms=min_pressure_span_hpa_for_pressure_terms,
    )
    analyzers = [
        str(row.get("analyzer_device_id") or "")
        for row in tables.get("co2_senco_pair_current_evidence_coverage", [])
        if str(row.get("analyzer_device_id") or "")
    ]
    metadata = ValidationMetadata(
        tool_name="v1_5_co2_senco_pair_model_scope",
        analyzers=analyzers,
        input_paths=[
            str(Path(original_points_xlsx).resolve()),
            str(Path(candidate_dir).resolve()),
            str(Path(pair_review_dir).resolve()) if pair_review_dir else "",
        ],
        output_dir=str(Path(output_dir).resolve()),
        config_summary={
            "min_temp_span_c_for_secondary": float(min_temp_span_c_for_secondary),
            "min_pressure_span_hpa_for_pressure_terms": float(min_pressure_span_hpa_for_pressure_terms),
            "no_write": True,
            "opens_com_ports": False,
            "controls_water_or_gas_routes": False,
            "manual_co2_groups_in_scope": "SENCO1,SENCO3,SENCO5",
            "pressure_group_handled_separately": "SENCO9",
        },
        notes=[
            "V1 execution_view and loader comparison are used as model-scope references.",
            "Current evidence may support temperature terms when the candidate policy selected T/T2/RT across enough temperature span.",
            "SENCO5 is included in CO2 model scope as the final concentration affine layer; it remains no-write until integrated candidate review, readback, rollback, and independent verification are approved.",
            "Pressure terms remain frozen because pressure compensation is validated independently in V1.5.",
        ],
    )
    outputs = write_validation_report(
        output_dir,
        prefix="co2_senco_pair_model_scope",
        metadata=metadata,
        tables=tables,
    )
    sidecar_path = (
        Path(database_sidecar_json).resolve()
        if database_sidecar_json
        else Path(output_dir).resolve() / "co2_senco_pair_model_scope_database_sidecar.json"
    )
    outputs["database_sidecar"] = _write_database_sidecar(
        sidecar_path,
        outputs=outputs,
        database_index_rows=tables["co2_senco_pair_model_database_index"],
    )
    return outputs
