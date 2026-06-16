"""Build V1.5 canonical open-flow ambient point plans.

This module is offline-only. It converts the old V1/V1.5 mixed
temperature/gas/water/pressure point table into a formal V1.5 point contract:
pressure is verified before sampling, then CO2/H2O samples are collected in
open flow at the current atmosphere. Legacy pressure targets are retained only
as excluded diagnostic evidence.
"""

from __future__ import annotations

import hashlib
import json
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence

from openpyxl import load_workbook

from ..data.points import CalibrationPoint, load_points_from_excel
from .reporting import ValidationMetadata, write_validation_report


DEFAULT_CO2_FIT_PPM = (0, 100, 200, 300, 400, 500, 600, 700, 800, 900, 1000)
DEFAULT_CO2_VERIFICATION_PPM: tuple[int, ...] = ()
DEFAULT_PURGE_S = 360.0
DEFAULT_SAMPLE_COUNT = 10
PRESSURE_MODE = "ambient_open"
_NUMBER_RE = re.compile(r"([+-]?\d+(?:\.\d+)?)")


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _safe_float(value: Any) -> Optional[float]:
    if value in (None, "", "None", "null"):
        return None
    try:
        numeric = float(value)
    except Exception:
        return None
    if not math.isfinite(numeric):
        return None
    return numeric


def _format_number(value: Any) -> str:
    numeric = _safe_float(value)
    if numeric is None:
        return str(value or "")
    if float(numeric).is_integer():
        return str(int(numeric))
    return f"{numeric:g}"


def _format_values(values: Iterable[Any]) -> str:
    return ",".join(_format_number(value) for value in values)


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _number_list(value: Any) -> List[float]:
    if value in (None, ""):
        return []
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return [float(value)]
    return [float(match.group(1)) for match in _NUMBER_RE.finditer(str(value))]


def _header_map(row: Sequence[Any]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for idx, value in enumerate(row):
        text = str(value or "").strip()
        if text:
            out[text] = idx
    return out


def _load_execution_view(points_xlsx: str | Path) -> List[Dict[str, Any]]:
    path = Path(points_xlsx).resolve()
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "execution_view" not in workbook.sheetnames:
            return []
        worksheet = workbook["execution_view"]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header = _header_map(tuple(next(iterator) or ()))
        except StopIteration:
            return []
        rows: List[Dict[str, Any]] = []
        for excel_row, values in enumerate(iterator, start=2):
            values = tuple(values or ())
            temp = _safe_float(values[header["Temp_C"]] if "Temp_C" in header else None)
            if temp is None:
                continue
            rows.append(
                {
                    "source_excel_row": excel_row,
                    "temp_c": temp,
                    "execution_order": values[header["Order"]] if "Order" in header else "",
                    "h2o_targets": values[header["H2O_targets"]] if "H2O_targets" in header else "",
                    "co2_sources_ppm": _number_list(
                        values[header["CO2_sources_ppm"]] if "CO2_sources_ppm" in header else None
                    ),
                    "legacy_pressure_targets_hpa": _number_list(
                        values[header["Pressures_hPa"]] if "Pressures_hPa" in header else None
                    ),
                    "notes": values[header["Notes"]] if "Notes" in header else "",
                }
            )
        return rows
    finally:
        workbook.close()


def _co2_group_for_ppm(ppm: int) -> str:
    if ppm in {100, 300, 500, 700, 900}:
        return "B"
    return "A"


def _co2_sample_role(ppm: int, fit_ppm: set[int], verification_ppm: set[int]) -> str:
    if ppm in fit_ppm:
        return "fit"
    if ppm in verification_ppm:
        # Formal V1.5 no longer reserves pre-write CO2 holdout points inside the
        # main open-flow run. Real verification is a separate post-write run.
        return "fit"
    return "diagnostic"


def _fallback_co2_rows_from_loader(points: Sequence[CalibrationPoint]) -> List[Dict[str, Any]]:
    grouped: Dict[float, set[int]] = {}
    for point in points:
        if point.is_h2o_point:
            continue
        ppm = _safe_float(point.co2_ppm)
        if ppm is None:
            continue
        grouped.setdefault(float(point.temp_chamber_c), set()).add(int(round(ppm)))
    rows: List[Dict[str, Any]] = []
    for temp in sorted(grouped.keys()):
        rows.append(
            {
                "source_excel_row": "",
                "temp_c": temp,
                "execution_order": "loader_fallback",
                "h2o_targets": "",
                "co2_sources_ppm": sorted(float(value) for value in grouped[temp]),
                "legacy_pressure_targets_hpa": [],
                "notes": "fallback from load_points_from_excel because execution_view is absent",
            }
        )
    return rows


def _build_co2_canonical_rows(
    *,
    execution_rows: Sequence[Mapping[str, Any]],
    fit_ppm: set[int],
    verification_ppm: set[int],
    purge_s: float,
    sample_count: int,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for execution in sorted(execution_rows, key=lambda row: float(row["temp_c"])):
        temp = float(execution["temp_c"])
        pressure_targets = _format_values(execution.get("legacy_pressure_targets_hpa", []))
        for ppm_value in sorted({int(round(value)) for value in execution.get("co2_sources_ppm", [])}):
            sample_role = _co2_sample_role(ppm_value, fit_ppm, verification_ppm)
            rows.append(
                {
                    "point_id": f"co2_T{_format_number(temp)}_{ppm_value}ppm_ambient",
                    "component": "co2",
                    "temp_c": temp,
                    "source_nominal_ppm": ppm_value,
                    "co2_group": _co2_group_for_ppm(ppm_value),
                    "sample_role": sample_role,
                    "fit_eligible": sample_role == "fit",
                    "verification_eligible": False,
                    "zero_gas_required": ppm_value == 0,
                    "standard_role": "zero_air" if ppm_value == 0 else "co2_standard_gas",
                    "certificate_required": True,
                    "pressure_mode": PRESSURE_MODE,
                    "target_pressure_hpa": "",
                    "pressure_reference_required": True,
                    "pressure_channel_precheck_required": True,
                    "legacy_pressure_targets_excluded_hpa": pressure_targets,
                    "purge_s": float(purge_s),
                    "sample_count": int(sample_count),
                    "analyzer_acquisition": "active_stream_1hz",
                    "runner": "run_v1_5_formal_open_flow_sampling",
                    "runner_args": (
                        f"--temp {temp:g} --co2-source-ppm {ppm_value:g} --co2-group {_co2_group_for_ppm(ppm_value)} "
                        f"--purge-s {float(purge_s):g} --sample-count {int(sample_count)} "
                        "--analyzer-acquisition active_stream_1hz"
                    ),
                    "physical_meaning": (
                        "Open-flow CO2 point: standard gas continuously refreshes the analyzer cells; "
                        "pressure is recorded as an input quantity after independent pressure-channel verification, "
                        "not used as a sealed pressure fitting target."
                    ),
                }
            )
    return rows


def _build_h2o_canonical_rows(
    *,
    h2o_points: Sequence[CalibrationPoint],
    purge_s: float,
    sample_count: int,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    seen: set[tuple[float, float, float, Optional[float], Optional[float]]] = set()
    for point in sorted(
        [item for item in h2o_points if item.is_h2o_point],
        key=lambda item: (
            float(item.temp_chamber_c),
            float(item.hgen_temp_c or 0.0),
            float(item.hgen_rh_pct or 0.0),
            float(item.target_pressure_hpa or 0.0),
        ),
    ):
        hgen_temp = _safe_float(point.hgen_temp_c)
        hgen_rh = _safe_float(point.hgen_rh_pct)
        if hgen_temp is None or hgen_rh is None:
            continue
        key = (
            float(point.temp_chamber_c),
            float(hgen_temp),
            float(hgen_rh),
            _safe_float(point.dewpoint_c),
            _safe_float(point.h2o_mmol),
        )
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "point_id": (
                    f"h2o_T{_format_number(point.temp_chamber_c)}_"
                    f"HGEN{_format_number(hgen_temp)}C_{_format_number(hgen_rh)}RH_ambient"
                ),
                "component": "h2o",
                "temp_c": float(point.temp_chamber_c),
                "hgen_temp_c": hgen_temp,
                "hgen_rh_pct": hgen_rh,
                "reference_dewpoint_c": _safe_float(point.dewpoint_c),
                "reference_h2o_mmol": _safe_float(point.h2o_mmol),
                "sample_role": "fit",
                "fit_eligible": True,
                "verification_eligible": False,
                "standard_role": "humidity_generator_dewpoint_reference",
                "certificate_required": True,
                "pressure_mode": PRESSURE_MODE,
                "target_pressure_hpa": "",
                "pressure_reference_required": True,
                "pressure_channel_precheck_required": True,
                "legacy_pressure_targets_excluded_hpa": _format_number(point.target_pressure_hpa),
                "purge_s": float(purge_s),
                "sample_count": int(sample_count),
                "analyzer_acquisition": "active_stream_1hz",
                "runner": "run_v1_5_formal_h2o_open_flow_sampling",
                "runner_args": (
                    f"--temp {float(point.temp_chamber_c):g} --hgen-temp {hgen_temp:g} --hgen-rh {hgen_rh:g} "
                    f"--purge-s {float(purge_s):g} --sample-count {int(sample_count)} "
                    "--analyzer-acquisition active_stream_1hz --h2o-pressure-presample-policy skip"
                ),
                "physical_meaning": (
                    "Open-flow H2O point: humidified gas continuously refreshes the chain and dewpoint is the "
                    "reference quantity; pressure is evidence for compensation/uncertainty, not a sealed target."
                ),
            }
        )
    return rows


def _excluded_pressure_rows(
    *,
    execution_rows: Sequence[Mapping[str, Any]],
    h2o_points: Sequence[CalibrationPoint],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    seen: set[tuple[str, float, float]] = set()
    for execution in execution_rows:
        temp = float(execution["temp_c"])
        for pressure in execution.get("legacy_pressure_targets_hpa", []):
            key = ("co2_execution_view", temp, float(pressure))
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "source": "co2_execution_view",
                    "temp_c": temp,
                    "legacy_pressure_target_hpa": float(pressure),
                    "formal_v1_5_action": "excluded_from_main_calibration",
                    "reason": "sealed_or_dynamic_pressure_targets_are_not_part_of_open_flow_co2_h2o_fit",
                    "allowed_future_use": "engineering_diagnostic_or_post_hoc_pressure_compensation_validation",
                }
            )
    for point in h2o_points:
        pressure = _safe_float(point.target_pressure_hpa)
        if pressure is None:
            continue
        key = ("h2o_loader_point", float(point.temp_chamber_c), float(pressure))
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "source": "h2o_loader_point",
                "temp_c": float(point.temp_chamber_c),
                "legacy_pressure_target_hpa": float(pressure),
                "formal_v1_5_action": "excluded_from_main_calibration",
                "reason": "h2o_open_flow_uses_current_atmosphere_after_pressure_channel_verification",
                "allowed_future_use": "diagnostic_only_unless_group_stability_and_pressure_compensation_protocol_are_met",
            }
        )
    return rows


def _summary_rows(
    *,
    source_points_xlsx: Path,
    co2_rows: Sequence[Mapping[str, Any]],
    h2o_rows: Sequence[Mapping[str, Any]],
    excluded_rows: Sequence[Mapping[str, Any]],
    execution_view_present: bool,
) -> List[Dict[str, Any]]:
    fit_co2 = [row for row in co2_rows if row.get("sample_role") == "fit"]
    verification_co2 = [row for row in co2_rows if row.get("sample_role") == "verification"]
    return [
        {
            "plan_status": "ready_for_no_write_open_flow_execution",
            "source_points_xlsx": str(source_points_xlsx.resolve()),
            "source_points_sha256": _sha256_file(source_points_xlsx),
            "execution_view_present": execution_view_present,
            "co2_point_count": len(co2_rows),
            "co2_fit_point_count": len(fit_co2),
            "co2_verification_point_count": len(verification_co2),
            "h2o_point_count": len(h2o_rows),
            "excluded_legacy_pressure_target_count": len(excluded_rows),
            "pressure_channel_ordering": (
                "PRECHECK -> PRESSURE_CHANNEL_QUICK_CHECK -> pressure calibration if needed -> "
                "pressure verification pass -> CO2/H2O open-flow sampling"
            ),
            "opens_com_ports": False,
            "controls_water_or_gas_routes": False,
            "writes_coefficients": False,
            "physical_meaning": (
                "This canonical point plan removes legacy sealed-pressure targets from the formal main fit. "
                "Pressure remains a traceable input that must be verified before the open-flow gas/water samples."
            ),
        }
    ]


def build_open_flow_canonical_point_tables(
    *,
    source_points_xlsx: str | Path,
    co2_fit_ppm: Sequence[int] = DEFAULT_CO2_FIT_PPM,
    co2_verification_ppm: Sequence[int] = DEFAULT_CO2_VERIFICATION_PPM,
    purge_s: float = DEFAULT_PURGE_S,
    sample_count: int = DEFAULT_SAMPLE_COUNT,
) -> Dict[str, List[Dict[str, Any]]]:
    source_path = Path(source_points_xlsx).resolve()
    loader_points = load_points_from_excel(
        source_path,
        missing_pressure_policy="carry_forward",
        carry_forward_h2o=False,
    )
    execution_rows = _load_execution_view(source_path)
    if not execution_rows:
        execution_rows = _fallback_co2_rows_from_loader(loader_points)
    h2o_points = [point for point in loader_points if point.is_h2o_point]
    fit_set = {int(value) for value in co2_fit_ppm}
    verification_set = {int(value) for value in co2_verification_ppm}
    co2_rows = _build_co2_canonical_rows(
        execution_rows=execution_rows,
        fit_ppm=fit_set,
        verification_ppm=verification_set,
        purge_s=float(purge_s),
        sample_count=int(sample_count),
    )
    h2o_rows = _build_h2o_canonical_rows(
        h2o_points=h2o_points,
        purge_s=float(purge_s),
        sample_count=int(sample_count),
    )
    excluded_rows = _excluded_pressure_rows(execution_rows=execution_rows, h2o_points=h2o_points)
    summary = _summary_rows(
        source_points_xlsx=source_path,
        co2_rows=co2_rows,
        h2o_rows=h2o_rows,
        excluded_rows=excluded_rows,
        execution_view_present=bool(_load_execution_view(source_path)),
    )
    return {
        "v1_5_open_flow_canonical_summary": summary,
        "co2_open_flow_multitemp_ambient": co2_rows,
        "h2o_open_flow_multitemp_ambient": h2o_rows,
        "legacy_pressure_targets_excluded": excluded_rows,
        "co2_runner_queue": [row for row in co2_rows if row.get("sample_role") in {"fit", "verification"}],
        "h2o_runner_queue": h2o_rows,
    }


def _write_manifest(path: Path, *, outputs: Mapping[str, Path], summary_row: Mapping[str, Any]) -> Path:
    payload = {
        "schema": "v1_5_open_flow_canonical_point_plan_v1",
        "created_at": _now(),
        "plan_status": summary_row.get("plan_status"),
        "pressure_model": "pressure_channel_verified_before_sampling_then_ambient_open_evidence_only",
        "opens_com_ports": False,
        "controls_water_or_gas_routes": False,
        "writes_coefficients": False,
        "artifacts": {
            key: {
                "path": str(value.resolve()),
                "sha256": _sha256_file(value),
                "size_bytes": value.stat().st_size,
            }
            for key, value in outputs.items()
            if value.exists()
        },
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return path


def write_open_flow_canonical_point_plan(
    *,
    source_points_xlsx: str | Path,
    output_dir: str | Path,
    co2_fit_ppm: Sequence[int] = DEFAULT_CO2_FIT_PPM,
    co2_verification_ppm: Sequence[int] = DEFAULT_CO2_VERIFICATION_PPM,
    purge_s: float = DEFAULT_PURGE_S,
    sample_count: int = DEFAULT_SAMPLE_COUNT,
) -> Dict[str, Path]:
    tables = build_open_flow_canonical_point_tables(
        source_points_xlsx=source_points_xlsx,
        co2_fit_ppm=co2_fit_ppm,
        co2_verification_ppm=co2_verification_ppm,
        purge_s=purge_s,
        sample_count=sample_count,
    )
    output_path = Path(output_dir).resolve()
    metadata = ValidationMetadata(
        tool_name="v1_5_open_flow_canonical_point_plan",
        input_paths=[str(Path(source_points_xlsx).resolve())],
        output_dir=str(output_path),
        config_summary={
            "co2_fit_ppm": list(co2_fit_ppm),
            "co2_verification_ppm": list(co2_verification_ppm),
            "purge_s": float(purge_s),
            "sample_count": int(sample_count),
            "pressure_mode": PRESSURE_MODE,
        },
        notes=[
            "Pressure channel must be verified before formal CO2/H2O open-flow sampling.",
            "Legacy sealed/dynamic pressure targets are excluded from the formal main fit.",
            "Zero gas is included as a formal CO2 fit source when its certificate is available.",
        ],
    )
    outputs = write_validation_report(
        output_path,
        prefix="v1_5_open_flow_canonical_point_plan",
        metadata=metadata,
        tables=tables,
    )
    outputs["manifest"] = _write_manifest(
        output_path / "v1_5_open_flow_canonical_point_plan_manifest.json",
        outputs=outputs,
        summary_row=tables["v1_5_open_flow_canonical_summary"][0],
    )
    return outputs
