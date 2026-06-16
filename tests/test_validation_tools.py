import csv
import json
import os
from pathlib import Path
import subprocess
import sys

from gas_calibrator.tools import (
    run_prevalidation_no_sources,
    run_v1_merged_calibration_sidecar,
    validate_dry_collect,
    validate_offline_run,
    validate_pressure_only,
    verify_coefficient_roundtrip,
)
from gas_calibrator.validation.common import analyze_sample_rows
from openpyxl import Workbook


def _write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _base_cfg(output_dir: Path) -> dict:
    return {
        "paths": {
            "output_dir": str(output_dir),
            "points_excel": "points.xlsx",
        },
        "devices": {
            "gas_analyzer": {"enabled": False, "active_send": False, "ftd_hz": 10, "average_co2": 1, "average_h2o": 1},
            "gas_analyzers": [
                {"enabled": True, "name": "GA01", "port": "COM1", "baud": 115200, "device_id": "010", "active_send": False, "ftd_hz": 10, "average_co2": 1, "average_h2o": 1},
            ],
            "pressure_controller": {"enabled": False, "port": "COM2", "baud": 9600, "in_limits_pct": 0.02, "in_limits_time_s": 10},
            "pressure_gauge": {"enabled": False, "port": "COM3", "baud": 9600, "dest_id": "001"},
            "humidity_generator": {"enabled": False, "port": "COM4", "baud": 9600},
            "dewpoint_meter": {"enabled": False, "port": "COM5", "baud": 9600, "station": 1},
            "temperature_chamber": {"enabled": False, "port": "COM6", "baud": 9600, "addr": 1},
            "thermometer": {"enabled": False, "port": "COM7", "baud": 9600},
            "relay": {"enabled": False, "port": "COM8", "baud": 9600, "addr": 1},
            "relay_8": {"enabled": False, "port": "COM9", "baud": 9600, "addr": 1},
        },
        "workflow": {
            "sampling": {
                "count": 3,
                "stable_count": 3,
                "interval_s": 0.0,
                "quality": {"enabled": False},
            }
        },
        "coefficients": {
            "model": "ratio_poly_rt_p",
            "summary_columns": {
                "co2": {"target": "ppm_CO2_Tank", "ratio": "R_CO2", "temperature": "T1", "pressure": "BAR", "pressure_scale": 1.0},
                "h2o": {"target": "ppm_H2O_Dew", "ratio": "R_H2O", "temperature": "T1", "pressure": "BAR", "pressure_scale": 1.0},
            },
            "ratio_poly_fit": {"pressure_source_preference": "reference_first"},
            "fit_h2o": True,
            "save_residuals": False,
            "min_samples": 0,
            "enabled": False,
        },
    }


def _write_csv(path: Path, rows: list[dict]) -> None:
    header = []
    for row in rows:
        for key in row.keys():
            if key not in header:
                header.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=header)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _offline_samples() -> list[dict]:
    rows = []
    for point_row, ratio, target, pressure in [
        (1, 1.00, 400.0, 1000.0),
        (2, 1.10, 500.0, 1005.0),
        (3, 1.20, 600.0, 1010.0),
    ]:
        for idx in range(3):
            rows.append(
                {
                    "point_row": point_row,
                    "point_phase": "co2",
                    "point_tag": f"demo_{point_row}",
                    "point_title": f"demo_{point_row}",
                    "pressure_target_hpa": pressure,
                    "co2_ppm_target": target,
                    "thermometer_temp_c": 20.0 + idx * 0.1,
                    "dewpoint_c": 1.0 + idx * 0.1,
                    "dew_pressure_hpa": pressure,
                    "pressure_gauge_hpa": pressure,
                    "ga01_device_id": "010",
                    "ga01_mode2_field_count": 16,
                    "ga01_status": "OK",
                    "ga01_co2_ppm": target + 0.5,
                    "ga01_h2o_mmol": 1.0,
                    "ga01_co2_ratio_f": ratio + idx * 0.001,
                    "ga01_h2o_ratio_f": 0.2 + idx * 0.001,
                    "ga01_pressure_kpa": 101.0 + idx * 0.01,
                }
            )
    return rows


class _FakeGasAnalyzer:
    def __init__(self, *, pressure_kpa: float = 100.0, device_id: str = "010"):
        self.calls = []
        self._counter = 0
        self.pressure_kpa = float(pressure_kpa)
        self.device_id = str(device_id)

    def set_mode(self, mode):
        self.calls.append(("mode", mode))
        return True

    def set_mode_with_ack(self, mode, require_ack=True):
        self.calls.append(("mode", mode, require_ack))
        return True

    def set_comm_way(self, active):
        self.calls.append(("active", active))
        return True

    def set_comm_way_with_ack(self, active, require_ack=True):
        self.calls.append(("active", active, require_ack))
        return True

    def set_average_filter(self, window_n):
        self.calls.append(("avg_filter", window_n))
        return True

    def set_average_filter_with_ack(self, window_n, require_ack=True):
        self.calls.append(("avg_filter", window_n, require_ack))
        return True

    def set_active_freq(self, hz):
        self.calls.append(("ftd", hz))
        return True

    def set_active_freq_with_ack(self, hz, require_ack=True):
        self.calls.append(("ftd", hz, require_ack))
        return True

    def read_latest_data(self, *args, **kwargs):
        self._counter += 1
        return "YGAS,010,0400.0,01.0,0003.0,00.5,1.0000,1.0001,0.2000,0.2001,04000,05000,02000,020.00,021.00,101.00,OK"

    def _drain_stream_lines(self, *args, **kwargs):
        self.calls.append(("drain_stream", args, kwargs))
        return [
            "YGAS,010,0400.0,01.0,0003.0,00.5,1.0000,1.0001,0.2000,0.2001,04000,05000,02000,020.00,021.00,101.00,OK"
        ]

    def parse_line_mode2(self, _line):
        self._counter += 1
        return {
            "mode2_field_count": 17,
            "status": "OK",
            "co2_ppm": 400.0 + self._counter * 0.1,
            "h2o_mmol": 1.0 + self._counter * 0.01,
            "co2_ratio_f": 1.0 + self._counter * 0.001,
            "h2o_ratio_f": 0.2 + self._counter * 0.001,
            "pressure_kpa": self.pressure_kpa + self._counter * 0.001,
            "chamber_temp_c": 20.0,
            "case_temp_c": 21.0,
            "ref_signal": 4000.0,
            "co2_signal": 5000.0,
            "h2o_signal": 2000.0,
            "id": self.device_id,
            "mode": 2,
        }

    def close(self):
        return None


class _FakePressureGauge:
    def read_pressure(self):
        return 1000.5

    def close(self):
        return None


class _FakeControlledPressureGauge:
    def __init__(self, pace):
        self.pace = pace

    def read_pressure(self, **kwargs):
        return float(self.pace.pressure_hpa)

    def read_pressure_fast(self, **kwargs):
        return float(self.pace.pressure_hpa)

    def close(self):
        return None


class _FakePace:
    def __init__(self):
        self.calls = []
        self.output_state = 0
        self.isolation_state = 1
        self.vent_status = 0
        self.hold_active = False
        self.pressure_hpa = 1000.6
        self.atmosphere_hpa = 1000.6
        self.setpoint_hpa = 1000.6
        self.source_range = "2.00bara"

    def stop_atmosphere_hold(self):
        self.calls.append(("stop_atmosphere_hold",))
        self.hold_active = False
        return True

    def set_units_hpa(self):
        self.calls.append(("set_units_hpa",))
        return True

    def set_range(self, range_name):
        self.calls.append(("set_range", range_name))
        self.source_range = str(range_name)
        return True

    def set_in_limits(self, pct, time_s):
        self.calls.append(("set_in_limits", pct, time_s))
        return True

    def set_output_mode_active(self):
        self.calls.append(("set_output_mode_active",))
        return True

    def set_slew_mode_linear(self):
        self.calls.append(("set_slew_mode_linear",))
        return True

    def set_slew_mode_max(self):
        self.calls.append(("set_slew_mode_max",))
        return True

    def set_slew_rate(self, value_hpa_per_s):
        self.calls.append(("set_slew_rate", value_hpa_per_s))
        return True

    def set_overshoot_allowed(self, enabled):
        self.calls.append(("set_overshoot_allowed", enabled))
        return True

    def set_output(self, enabled):
        self.calls.append(("set_output", enabled))
        self.output_state = 1 if enabled else 0
        return True

    def vent(self, enabled=True):
        self.calls.append(("vent", enabled))
        self.vent_status = 1 if enabled else 0
        return True

    def wait_for_vent_idle(self, **kwargs):
        self.calls.append(("wait_for_vent_idle", dict(kwargs)))
        self.vent_status = 0
        return 0

    def set_isolation_open(self, is_open):
        self.calls.append(("set_isolation_open", is_open))
        self.isolation_state = 1 if is_open else 0
        return True

    def exit_atmosphere_mode(self, **kwargs):
        self.calls.append(("exit_atmosphere_mode", dict(kwargs)))
        self.hold_active = False
        self.output_state = 0
        self.vent_status = 0
        self.isolation_state = 1
        return 0

    def set_setpoint(self, value_hpa):
        self.calls.append(("set_setpoint", value_hpa))
        self.setpoint_hpa = float(value_hpa)
        if "barg" in self.source_range.lower():
            self.pressure_hpa = float(self.atmosphere_hpa) + float(value_hpa)
        else:
            self.pressure_hpa = float(value_hpa)
        return True

    def enable_control_output(self, **kwargs):
        self.calls.append(("enable_control_output", dict(kwargs)))
        self.output_state = 1
        self.isolation_state = 1
        self.vent_status = 0
        return True

    def enter_atmosphere_mode(self, **kwargs):
        self.calls.append(("enter_atmosphere_mode", dict(kwargs)))
        self.output_state = 0
        self.isolation_state = 1
        self.vent_status = 1
        self.hold_active = bool(kwargs.get("hold_open"))

    def is_atmosphere_hold_active(self):
        return self.hold_active

    def get_output_state(self):
        return self.output_state

    def get_isolation_state(self):
        return self.isolation_state

    def get_vent_status(self):
        return self.vent_status

    def read_pressure(self):
        return self.pressure_hpa

    def read_gauge_pressure(self):
        return float(self.pressure_hpa) - float(self.atmosphere_hpa)

    def set_vent_after_valve_open(self, open_after_vent):
        self.calls.append(("set_vent_after_valve_open", open_after_vent))
        return True

    def get_vent_after_valve_open(self):
        self.calls.append(("get_vent_after_valve_open",))
        return False

    def vent_status_allows_control(self, status):
        return int(status) in {0, 2, 3, 4}

    def close(self):
        return None


class _FakeRoundtripAnalyzer:
    def __init__(self, *args, **kwargs):
        self.groups = {
            1: {"C0": 1.0, "C1": 2.0},
            2: {"C0": 3.0, "C1": 4.0},
            3: {"C0": 5.0, "C1": 6.0},
            4: {"C0": 7.0, "C1": 8.0},
        }
        self.calls = []

    def open(self):
        return None

    def close(self):
        return None

    def set_mode(self, mode):
        self.calls.append(("mode", mode))
        return True

    def set_senco(self, index, *coefficients):
        self.calls.append(("set_senco", index, list(coefficients)))
        self.groups[int(index)] = {f"C{i}": float(value) for i, value in enumerate(coefficients)}
        return True

    def read_coefficient_group(self, index, **kwargs):
        return dict(self.groups[int(index)])


def _build_points_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="温度")
    ws.cell(row=2, column=1, value="温度")
    rows = [
        (-20.0, 0.0, None, 1100.0, None),
        (0.0, None, "0℃（湿度发生器） 50%（湿度发生器） -9.16℃（露点温度） 3.0233mmol/mol（气体分析仪）", 1100.0, None),
        (20.0, None, "20℃（湿度发生器） 70%（湿度发生器） 14.36℃（露点温度） 16.3715mmol/mol（气体分析仪）", 500.0, None),
        (30.0, 200.0, None, 1000.0, None),
        (30.0, 500.0, None, 800.0, "B"),
        (30.0, 800.0, None, 700.0, None),
        (40.0, 1000.0, None, 500.0, None),
    ]
    excel_row = 3
    for temp, co2, h2o, pressure, group in rows:
        ws.cell(row=excel_row, column=1, value=temp)
        ws.cell(row=excel_row, column=2, value=co2)
        ws.cell(row=excel_row, column=3, value=h2o)
        ws.cell(row=excel_row, column=4, value=pressure)
        ws.cell(row=excel_row, column=5, value=group)
        excel_row += 1
    wb.save(path)
    wb.close()


def test_validate_offline_run_generates_expected_tables(tmp_path: Path) -> None:
    run_dir = tmp_path / "run_demo"
    run_dir.mkdir()
    cfg = _base_cfg(tmp_path / "logs")
    (run_dir / "runtime_config_snapshot.json").write_text(json.dumps(cfg, ensure_ascii=False), encoding="utf-8")
    _write_csv(run_dir / "samples_demo.csv", _offline_samples())

    assert validate_offline_run.main(["--run-dir", str(run_dir), "--output-dir", str(tmp_path / "out")]) == 0

    out_dir = tmp_path / "out"
    assert (out_dir / "frame_quality_summary.csv").exists()
    assert (out_dir / "pressure_source_check.csv").exists()
    assert (out_dir / "fit_input_overview.csv").exists()


def test_validate_offline_run_marks_single_point_as_fit_skipped(tmp_path: Path) -> None:
    run_dir = tmp_path / "single_point_run"
    run_dir.mkdir()
    cfg = _base_cfg(tmp_path / "logs")
    (run_dir / "runtime_config_snapshot.json").write_text(json.dumps(cfg, ensure_ascii=False), encoding="utf-8")
    rows = [dict(row) for row in _offline_samples() if row["point_row"] == 1]
    for row in rows:
        row.pop("pressure_target_hpa", None)
    _write_csv(run_dir / "samples_single_point.csv", rows)

    assert (
        validate_offline_run.main(
            ["--run-dir", str(run_dir), "--gas", "co2", "--mode", "current", "--output-dir", str(tmp_path / "out")]
        )
        == 0
    )

    with (tmp_path / "out" / "fit_input_overview.csv").open("r", encoding="utf-8-sig", newline="") as handle:
        fit_rows = list(csv.DictReader(handle))
    with (tmp_path / "out" / "conclusion_summary.csv").open("r", encoding="utf-8-sig", newline="") as handle:
        conclusion_rows = list(csv.DictReader(handle))

    assert fit_rows[0]["status"] == "fit_skipped_insufficient_points"
    assert conclusion_rows[0]["risk_level"] == "warn"
    assert conclusion_rows[0]["fit_error_count"] == "0"
    assert conclusion_rows[0]["fit_skipped_count"] == "1"


def test_analyze_sample_rows_keeps_missing_prefixed_analyzer_frames_unusable(tmp_path: Path) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["gas_analyzers"] = [
        {"enabled": True, "name": "GA01", "device_id": "010"},
        {"enabled": True, "name": "GA02", "device_id": "011"},
    ]
    rows = []
    for idx in range(3):
        rows.append(
            {
                "point_row": 1,
                "point_phase": "co2",
                "point_tag": "open_flow_900ppm",
                "co2_ppm_target": 897.04,
                "thermometer_temp_c": 24.5,
                "pressure_gauge_hpa": 1011.8,
                "ga01_device_id": "010",
                "ga01_frame_has_data": True,
                "ga01_frame_usable": True,
                "ga01_raw": (
                    f"YGAS,010,{838 + idx:.3f},01.0,0003.0,00.5,1.2850,1.2851,"
                    "0.2,0.2,04000,05000,02000,020.00,021.00,101.00"
                ),
                "ga01_id": "010",
                "ga01_mode": 2,
                "ga01_mode2_field_count": 16,
                "ga01_mode2_contract_status": "pass",
                "ga01_mode2_contract_reason": "ok",
                "ga01_mode2_qc_status": "pass",
                "ga01_mode2_qc_reason": "可用",
                "ga01_co2_ppm": 838 + idx,
                "ga01_h2o_mmol": 1.0,
                "ga01_co2_ratio_f": 1.285,
                "ga01_h2o_ratio_f": 0.2,
                "ga01_pressure_kpa": 101.0,
                "ga02_frame_has_data": False,
                "ga02_frame_usable": False,
                "ga02_mode2_contract_status": "missing",
                "ga02_mode2_contract_reason": "no_frame",
                "ga02_mode2_qc_status": "missing",
                "ga02_mode2_qc_reason": "no_frame",
            }
        )

    tables = analyze_sample_rows(rows, cfg=cfg, gas="co2", modes=("current",))
    frame_by_analyzer = {row["Analyzer"]: row for row in tables["frame_quality_summary"]}

    assert frame_by_analyzer["GA01"]["TotalFrames"] == 3
    assert frame_by_analyzer["GA01"]["ValidFrames"] == 3
    assert frame_by_analyzer["GA02"]["TotalFrames"] == 0
    assert frame_by_analyzer["GA02"]["ValidFrames"] == 0


def test_merged_sidecar_merge_keeps_gas_and_water_separate() -> None:
    older = Path("D:/old")
    newer = Path("D:/new")
    point_rows = {
        str(older): [
            {"流程阶段": "co2", "温箱目标温度C": 20, "目标二氧化碳浓度ppm": 400, "目标压力hPa": 1000, "来源": "older-gas"},
            {"流程阶段": "h2o", "温箱目标温度C": 20, "湿度发生器目标温度C": 20, "湿度发生器目标湿度%": 70, "目标压力hPa": 500, "来源": "older-water"},
        ],
        str(newer): [
            {"流程阶段": "co2", "温箱目标温度C": 20, "目标二氧化碳浓度ppm": 400, "目标压力hPa": 1000, "来源": "newer-gas"},
        ],
    }

    merged_rows, selected_sources = run_v1_merged_calibration_sidecar._merge_point_rows(
        [older, newer],
        point_rows,
        allowed_gas_ppm=run_v1_merged_calibration_sidecar.DEFAULT_GAS_PPM,
    )

    assert len(merged_rows) == 2
    assert any(row["来源"] == "newer-gas" for row in merged_rows)
    assert any(row["来源"] == "older-water" for row in merged_rows)
    water_key = run_v1_merged_calibration_sidecar._point_identity_from_row(
        {"流程阶段": "h2o", "温箱目标温度C": 20, "湿度发生器目标温度C": 20, "湿度发生器目标湿度%": 70, "目标压力hPa": 500}
    )
    assert selected_sources[water_key]["source_run"] == str(older)
    assert selected_sources[water_key]["phase"] == "h2o"


def test_merged_sidecar_builds_verify_subset_points_workbook(tmp_path: Path) -> None:
    source = tmp_path / "points.xlsx"
    target = tmp_path / "verify_points.xlsx"
    _build_points_workbook(source)

    info = run_v1_merged_calibration_sidecar._build_verify_points_workbook(source, target)

    assert target.exists()
    assert info["point_count"] == 5
    points = run_v1_merged_calibration_sidecar._build_verify_point_rows_from_workbook(target)
    assert len(points) == 5
    assert sum(1 for row in points if row["流程阶段"] == "co2") == 3
    assert sum(1 for row in points if row["流程阶段"] == "h2o") == 2


def test_merged_sidecar_defaults_to_standalone_non_write_mode() -> None:
    args = run_v1_merged_calibration_sidecar._parse_args(["--run-dir", "D:/completed_run"])

    assert args.write_temperature is False
    assert args.write_gas is False
    assert args.run_verify is False


def test_merged_sidecar_non_write_path_survives_without_sqlalchemy() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    src_root = repo_root / "src"
    script = """
import builtins
import json
import tempfile
from pathlib import Path

from openpyxl import Workbook

real_import = builtins.__import__

def fake_import(name, globals=None, locals=None, fromlist=(), level=0):
    if name == "sqlalchemy" or name.startswith("sqlalchemy."):
        raise ModuleNotFoundError("No module named 'sqlalchemy'")
    return real_import(name, globals, locals, fromlist, level)

builtins.__import__ = fake_import

try:
    import gas_calibrator.tools.run_v1_merged_calibration_sidecar as sidecar

    tmp = Path(tempfile.mkdtemp(prefix="merged_sidecar_no_sqlalchemy_"))
    cfg_path = tmp / "cfg.json"
    points_path = tmp / "points.xlsx"
    points_book = Workbook()
    points_book.save(points_path)
    points_book.close()
    cfg_path.write_text(
        json.dumps(
            {
                "paths": {"output_dir": str(tmp / "logs"), "points_excel": str(points_path)},
                "devices": {
                    "gas_analyzer": {"enabled": False},
                    "gas_analyzers": [{"enabled": False, "name": "GA01", "port": "COM1", "baud": 115200, "device_id": "010"}],
                },
                "workflow": {"sampling": {"quality": {"enabled": False}}},
                "coefficients": {
                    "enabled": False,
                    "model": "ratio_poly_rt_p",
                    "summary_columns": {
                        "co2": {"target": "ppm_CO2_Tank", "ratio": "R_CO2", "temperature": "Temp", "pressure": "BAR"},
                        "h2o": {"target": "ppm_H2O_Dew", "ratio": "R_H2O", "temperature": "Temp", "pressure": "BAR"},
                    },
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    run_dir = tmp / "run_demo"
    run_dir.mkdir()

    point_rows = [
        {"流程阶段": "co2", "温箱目标温度C": 20, "目标二氧化碳浓度ppm": 200, "目标压力hPa": 1000, "点位标题": "co2-20"},
        {"流程阶段": "h2o", "温箱目标温度C": 20, "湿度发生器目标温度C": 20, "湿度发生器目标湿度%": 70, "目标压力hPa": 500, "点位标题": "h2o-20"},
    ]
    summary_rows = [
        {"Analyzer": "GA01", "PointPhase": "co2", "TempSet": 20, "ppm_CO2_Tank": 200, "PressureTarget": 1000, "ppm_CO2": 205, "PointRow": 1},
        {"Analyzer": "GA01", "PointPhase": "h2o", "TempSet": 20, "HgenTempSet": 20, "HgenRhSet": 70, "PressureTarget": 500, "ppm_H2O_Dew": 16.0, "ppm_H2O": 15.5, "PointRow": 2},
    ]
    temp_rows = [
        {"analyzer_id": "GA01", "temp_setpoint_c": 20, "ref_temp_c": 20, "analyzer_cell_temp_raw_c": 20.02, "analyzer_shell_temp_raw_c": 20.03},
        {"analyzer_id": "GA01", "temp_setpoint_c": 30, "ref_temp_c": 30, "analyzer_cell_temp_raw_c": 30.01, "analyzer_shell_temp_raw_c": 30.02},
    ]

    sidecar._load_merge_inputs = lambda run_dirs: {
        "summary_rows_by_run": {str(run_dir): summary_rows},
        "point_rows_by_run": {str(run_dir): point_rows},
        "temperature_rows_by_run": {str(run_dir): temp_rows},
    }

    def fake_export_ratio_poly(summary_frame, *, out_dir, coeff_cfg):
        out_dir.mkdir(parents=True, exist_ok=True)
        report = out_dir / "calibration_coefficients.xlsx"
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "汇总"
        ws1.append(["分析仪", "气体", "Constant"])
        ws1.append(["GA01", "CO2", 1.23])
        ws2 = wb.create_sheet("download_plan")
        ws2.append(["Analyzer", "Gas", "PrimaryCommand", "SecondaryCommand", "ModeEnterCommand", "ModeExitCommand"])
        ws2.append(["GA01", "CO2", "SENCO1,YGAS,FFF,1.0,2.0,3.0,4.0", "", "", ""])
        wb.save(report)
        wb.close()
        return report

    sidecar.export_ratio_poly_report_from_summary_frame = fake_export_ratio_poly

    rc = sidecar.main(
        [
            "--config",
            str(cfg_path),
            "--run-dir",
            str(run_dir),
            "--output-dir",
            str(tmp / "out"),
        ]
    )
    assert rc == 0
    assert (tmp / "out" / "merge_manifest.json").exists()
    assert list((tmp / "out").glob("校准汇总与验证结论_*.xlsx"))
finally:
    builtins.__import__ = real_import
"""
    env = os.environ.copy()
    existing_pythonpath = env.get("PYTHONPATH", "")
    env["PYTHONPATH"] = str(src_root) if not existing_pythonpath else os.pathsep.join([str(src_root), existing_pythonpath])
    result = subprocess.run(
        [sys.executable, "-c", script],
        cwd=repo_root,
        env=env,
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, result.stderr or result.stdout


def test_merged_sidecar_postprocess_runtime_skips_sqlalchemy_backed_steps_when_dependency_missing() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    src_root = repo_root / "src"
    script = """
import builtins
import tempfile
from pathlib import Path

real_import = builtins.__import__

def fake_import(name, globals=None, locals=None, fromlist=(), level=0):
    if name == "sqlalchemy" or name.startswith("sqlalchemy."):
        raise ModuleNotFoundError("No module named 'sqlalchemy'")
    return real_import(name, globals, locals, fromlist, level)

builtins.__import__ = fake_import

try:
    import gas_calibrator.v2.adapters.v1_postprocess_runner as runner

    tmp = Path(tempfile.mkdtemp(prefix="postprocess_no_sqlalchemy_"))
    run_dir = tmp / "run_demo"
    run_dir.mkdir()

    database_step = runner._database_import_step(
        run_dir=run_dir,
        artifact_dir=tmp,
        config_path=None,
        dsn=None,
        stage="raw",
    )
    assert database_step["status"] == "skipped"
    assert database_step["dependency"] == "sqlalchemy"
    assert "sqlalchemy" in database_step["reason"].lower()

    analytics_step = runner._analytics_step(
        run_dir=run_dir,
        target_dir=tmp,
        run_id="demo",
        config_path=None,
        dsn=None,
        import_db=False,
        raw_database_step={"status": "skipped"},
        enrich_database_step={"status": "skipped"},
        run_analytics=True,
        skip_analytics=False,
    )
    assert analytics_step["status"] == "skipped"
    assert analytics_step["dependency"] == "sqlalchemy"
    assert "sqlalchemy" in analytics_step["reason"].lower()

    measurement_step = runner._measurement_analytics_step(
        run_dir=run_dir,
        target_dir=tmp,
        run_id="demo",
        config_path=None,
        dsn=None,
        import_db=False,
        raw_database_step={"status": "skipped"},
        enrich_database_step={"status": "skipped"},
        run_measurement_analytics=True,
        skip_measurement_analytics=False,
    )
    assert measurement_step["status"] == "skipped"
    assert measurement_step["dependency"] == "sqlalchemy"
    assert "sqlalchemy" in measurement_step["reason"].lower()
finally:
    builtins.__import__ = real_import
"""
    env = os.environ.copy()
    existing_pythonpath = env.get("PYTHONPATH", "")
    env["PYTHONPATH"] = str(src_root) if not existing_pythonpath else os.pathsep.join([str(src_root), existing_pythonpath])
    result = subprocess.run(
        [sys.executable, "-c", script],
        cwd=repo_root,
        env=env,
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, result.stderr or result.stdout


def test_validate_dry_collect_runs_without_aux_devices(monkeypatch, tmp_path: Path) -> None:
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, _base_cfg(tmp_path / "logs"))
    fake = _FakeGasAnalyzer()

    monkeypatch.setattr(validate_dry_collect, "_build_devices", lambda cfg, io_logger=None: {"gas_analyzer": fake, "gas_analyzer_01": fake})
    monkeypatch.setattr(validate_dry_collect, "_close_devices", lambda devices: None)

    assert validate_dry_collect.main(["--config", str(cfg_path), "--output-dir", str(tmp_path / "out"), "--count", "2", "--interval-s", "0"]) == 0
    assert next((tmp_path / "out").glob("dry_collect_*")).is_dir()


def test_validate_dry_collect_include_temperature_enables_temperature_evidence_devices(tmp_path: Path) -> None:
    cfg = _base_cfg(tmp_path / "logs")

    runtime_cfg = validate_dry_collect._prepare_runtime_cfg(
        cfg,
        include_pressure=False,
        include_temperature=True,
    )

    devices = runtime_cfg["devices"]
    assert devices["temperature_chamber"]["enabled"] is True
    assert devices["thermometer"]["enabled"] is True
    assert devices["relay"]["enabled"] is False
    assert devices["relay_8"]["enabled"] is False
    assert devices["humidity_generator"]["enabled"] is False
    assert devices["dewpoint_meter"]["enabled"] is False
    assert devices["pressure_controller"]["enabled"] is False
    assert devices["pressure_gauge"]["enabled"] is False


def test_validate_dry_collect_read_first_only_keeps_analyzer_startup_commands_off(monkeypatch, tmp_path: Path) -> None:
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, _base_cfg(tmp_path / "logs"))
    fake = _FakeGasAnalyzer()
    captured_cfg = {}

    def _fake_build_devices(cfg, io_logger=None):
        captured_cfg.update(cfg)
        return {"gas_analyzer": fake, "gas_analyzer_01": fake}

    monkeypatch.setattr(validate_dry_collect, "_build_devices", _fake_build_devices)
    monkeypatch.setattr(validate_dry_collect, "_close_devices", lambda devices: None)

    assert (
        validate_dry_collect.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--count",
                "2",
                "--interval-s",
                "0",
                "--read-first-only-analyzers",
            ]
        )
        == 0
    )

    init_cfg = captured_cfg["workflow"]["analyzer_mode2_init"]
    assert init_cfg["read_first_before_config"] is True
    assert init_cfg["sniff_stream_before_config"] is True
    assert init_cfg["skip_config_when_read_first_ready"] is True
    assert init_cfg["write_config_on_read_first_fail"] is False
    assert not any(call[0] in {"mode", "active", "ftd", "avg_filter"} for call in fake.calls)


def test_validate_dry_collect_writes_temperature_evidence_from_io(tmp_path: Path) -> None:
    io_path = tmp_path / "io.csv"
    _write_csv(
        io_path,
        [
            {
                "timestamp": "2026-06-07T10:00:00",
                "port": "COM19",
                "device": "temperature_chamber",
                "direction": "RX",
                "command": "",
                "response": "temp_c=23.9",
                "error": "",
            },
            {
                "timestamp": "2026-06-07T10:00:01",
                "port": "COM19",
                "device": "temperature_chamber",
                "direction": "RX",
                "command": "",
                "response": "rh_pct=50.0",
                "error": "",
            },
            {
                "timestamp": "2026-06-07T10:00:02",
                "port": "COM18",
                "device": "thermometer",
                "direction": "RX",
                "command": "",
                "response": "+024.50C",
                "error": "",
            },
            {
                "timestamp": "2026-06-07T10:00:03",
                "port": "COM23",
                "device": "pressure_controller",
                "direction": "RX",
                "command": "",
                "response": "1000.0",
                "error": "",
            },
        ],
    )

    outputs = validate_dry_collect._write_temperature_evidence_from_io(tmp_path, io_path)

    rows = list(csv.DictReader(outputs["csv"].open("r", encoding="utf-8-sig")))
    assert [row["metric"] for row in rows] == [
        "temperature_chamber_temp_c",
        "temperature_chamber_rh_pct",
        "digital_thermometer_temp_c",
    ]
    summary = json.loads(outputs["summary"].read_text(encoding="utf-8"))
    assert summary["total_temperature_io_rows"] == 3
    assert {row["metric"]: row["last"] for row in summary["summary"]}["digital_thermometer_temp_c"] == 24.5


def test_merged_sidecar_main_writes_summary_workbook(monkeypatch, tmp_path: Path) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    points_path = tmp_path / "points.xlsx"
    _build_points_workbook(points_path)
    cfg["paths"]["points_excel"] = str(points_path)
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, cfg)
    run_dir = tmp_path / "run_demo"
    run_dir.mkdir()

    point_rows = [
        {"流程阶段": "co2", "温箱目标温度C": 20, "目标二氧化碳浓度ppm": 200, "目标压力hPa": 1000, "点位标题": "co2-20"},
        {"流程阶段": "h2o", "温箱目标温度C": 20, "湿度发生器目标温度C": 20, "湿度发生器目标湿度%": 70, "目标压力hPa": 500, "点位标题": "h2o-20"},
    ]
    summary_rows = [
        {"Analyzer": "GA01", "PointPhase": "co2", "TempSet": 20, "ppm_CO2_Tank": 200, "PressureTarget": 1000, "ppm_CO2": 205, "PointRow": 1},
        {"Analyzer": "GA01", "PointPhase": "h2o", "TempSet": 20, "HgenTempSet": 20, "HgenRhSet": 70, "PressureTarget": 500, "ppm_H2O_Dew": 16.0, "ppm_H2O": 15.5, "PointRow": 2},
    ]
    temp_rows = [
        {"analyzer_id": "GA01", "temp_setpoint_c": 20, "ref_temp_c": 20, "analyzer_cell_temp_raw_c": 20.02, "analyzer_shell_temp_raw_c": 20.03},
        {"analyzer_id": "GA01", "temp_setpoint_c": 30, "ref_temp_c": 30, "analyzer_cell_temp_raw_c": 30.01, "analyzer_shell_temp_raw_c": 30.02},
    ]

    monkeypatch.setattr(
        run_v1_merged_calibration_sidecar,
        "_load_merge_inputs",
        lambda run_dirs: {
            "summary_rows_by_run": {str(run_dir): summary_rows},
            "point_rows_by_run": {str(run_dir): point_rows},
            "temperature_rows_by_run": {str(run_dir): temp_rows},
        },
    )

    def _fake_export_ratio_poly(summary_frame, *, out_dir, coeff_cfg):
        out_dir.mkdir(parents=True, exist_ok=True)
        report = out_dir / "calibration_coefficients.xlsx"
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "汇总"
        ws1.append(["分析仪", "气体", "Constant"])
        ws1.append(["GA01", "CO2", 1.23])
        ws2 = wb.create_sheet("download_plan")
        ws2.append(["Analyzer", "Gas", "PrimaryCommand", "SecondaryCommand", "ModeEnterCommand", "ModeExitCommand"])
        ws2.append(["GA01", "CO2", "SENCO1,YGAS,FFF,1.0,2.0,3.0,4.0", "SENCO3,YGAS,FFF,5.0,6.0,7.0,8.0", "MODE,YGAS,FFF,2", "MODE,YGAS,FFF,1"])
        wb.save(report)
        wb.close()
        return report

    monkeypatch.setattr(run_v1_merged_calibration_sidecar, "export_ratio_poly_report_from_summary_frame", _fake_export_ratio_poly)

    assert (
        run_v1_merged_calibration_sidecar.main(
            [
                "--config",
                str(cfg_path),
                "--run-dir",
                str(run_dir),
                "--output-dir",
                str(tmp_path / "out"),
            ]
        )
        == 0
    )

    assert (tmp_path / "out" / "merge_manifest.json").exists()
    assert any((tmp_path / "out").glob("校准汇总与验证结论_*.xlsx"))


def test_validate_pressure_only_exports_pressure_checks(monkeypatch, tmp_path: Path) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["pressure_gauge"]["enabled"] = True
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, cfg)
    fake = _FakeGasAnalyzer()

    monkeypatch.setattr(
        validate_pressure_only,
        "_build_devices",
        lambda cfg, io_logger=None: {"gas_analyzer": fake, "gas_analyzer_01": fake, "pressure_gauge": _FakePressureGauge()},
    )
    monkeypatch.setattr(validate_pressure_only, "_close_devices", lambda devices: None)

    assert validate_pressure_only.main(["--config", str(cfg_path), "--output-dir", str(tmp_path / "out"), "--pressure-points", "ambient,900", "--count", "2", "--interval-s", "0", "--no-prompt"]) == 0
    run_dir = next((tmp_path / "out").glob("pressure_only_*"))
    assert (run_dir / "pressure_source_check.csv").exists()


def test_validate_pressure_only_runtime_defaults_extend_pressure_fast_signal_wait(tmp_path: Path) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg.setdefault("workflow", {}).setdefault("sampling", {})["pre_sample_freshness_timeout_s"] = 1.0
    cfg["workflow"]["sampling"]["pre_sample_signal_max_age_s"] = 0.35
    cfg["workflow"]["analyzer_mode2_init"] = {
        "read_first_before_config": False,
        "sniff_stream_before_config": False,
        "write_config_on_read_first_fail": True,
        "send_active_freq": True,
        "read_first_attempts": 2,
        "ready_consecutive_frames": 1,
        "read_first_retry_delay_s": 0.01,
    }

    runtime_cfg = validate_pressure_only._prepare_runtime_cfg(cfg)
    validate_pressure_only._apply_pressure_only_sampling_runtime_defaults(
        runtime_cfg,
        pre_sample_freshness_timeout_s=3.0,
        pre_sample_signal_max_age_s=1.0,
    )

    sampling_cfg = runtime_cfg["workflow"]["sampling"]
    assert sampling_cfg["pre_sample_freshness_timeout_s"] == 3.0
    assert sampling_cfg["pre_sample_signal_max_age_s"] == 1.0
    assert runtime_cfg["devices"]["relay"]["enabled"] is False
    assert runtime_cfg["devices"]["humidity_generator"]["enabled"] is False
    init_cfg = runtime_cfg["workflow"]["analyzer_mode2_init"]
    assert init_cfg["read_first_before_config"] is True
    assert init_cfg["sniff_stream_before_config"] is True
    assert init_cfg["write_config_on_read_first_fail"] is False
    assert init_cfg["send_active_freq"] is False
    assert init_cfg["read_first_attempts"] >= 10
    assert init_cfg["read_first_retry_delay_s"] >= 0.2
    assert runtime_cfg["metadata"]["pressure_only_analyzer_startup_policy"] == (
        "read_first_no_startup_config_writes"
    )


def test_validate_pressure_only_runtime_keeps_route_relays_for_controlled_points(tmp_path: Path) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["relay"]["enabled"] = True
    cfg["devices"]["relay_8"]["enabled"] = True

    runtime_cfg = validate_pressure_only._prepare_runtime_cfg(
        cfg,
        enable_route_relays_for_control=True,
    )

    assert runtime_cfg["devices"]["relay"]["enabled"] is True
    assert runtime_cfg["devices"]["relay_8"]["enabled"] is True
    assert runtime_cfg["devices"]["humidity_generator"]["enabled"] is False
    assert runtime_cfg["devices"]["dewpoint_meter"]["enabled"] is False
    assert runtime_cfg["metadata"]["pressure_only_route_relays_enabled_for_control"] is True


def test_validate_pressure_only_can_explicitly_configure_1hz_active_upload(tmp_path: Path) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["gas_analyzers"] = [
        {"name": "ga01", "active_send": True, "ftd_hz": 10},
        {"name": "ga02", "active_send": False, "ftd_hz": 10},
    ]
    cfg["workflow"]["analyzer_mode2_init"] = {
        "read_first_before_config": True,
        "sniff_stream_before_config": True,
        "write_config_on_read_first_fail": False,
        "send_active_freq": False,
    }

    runtime_cfg = validate_pressure_only._prepare_runtime_cfg(cfg, analyzer_active_upload_hz=1)

    assert runtime_cfg["devices"]["gas_analyzer"]["active_send"] is True
    assert runtime_cfg["devices"]["gas_analyzer"]["ftd_hz"] == 1
    assert [item["ftd_hz"] for item in runtime_cfg["devices"]["gas_analyzers"]] == [1, 1]
    assert [item["active_send"] for item in runtime_cfg["devices"]["gas_analyzers"]] == [True, True]
    init_cfg = runtime_cfg["workflow"]["analyzer_mode2_init"]
    assert init_cfg["read_first_before_config"] is False
    assert init_cfg["sniff_stream_before_config"] is False
    assert init_cfg["write_config_on_read_first_fail"] is True
    assert init_cfg["send_active_freq"] is True
    assert runtime_cfg["metadata"]["pressure_only_analyzer_startup_policy"] == (
        "controlled_active_upload_1hz_startup_config"
    )
    assert runtime_cfg["metadata"]["ftd_write_enabled"] is True


def test_validate_pressure_only_cli_defaults_to_1hz_active_upload() -> None:
    args = validate_pressure_only._parse_args(["--config", "cfg.json"])

    assert args.analyzer_active_upload_hz == 1
    assert args.no_analyzer_active_upload_config is False


def test_validate_pressure_only_verifies_continuous_atmosphere_hold(monkeypatch, tmp_path: Path) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["pressure_controller"]["enabled"] = True
    cfg["devices"]["pressure_gauge"]["enabled"] = True
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, cfg)
    fake_analyzer = _FakeGasAnalyzer()
    fake_pace = _FakePace()

    monkeypatch.setattr(
        validate_pressure_only,
        "_build_devices",
        lambda cfg, io_logger=None: {
            "gas_analyzer": fake_analyzer,
            "gas_analyzer_01": fake_analyzer,
            "pressure_gauge": _FakePressureGauge(),
            "pace": fake_pace,
        },
    )
    monkeypatch.setattr(validate_pressure_only, "_close_devices", lambda devices: None)

    assert (
        validate_pressure_only.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--pressure-points",
                "ambient",
                "--count",
                "3",
                "--interval-s",
                "0",
                "--no-prompt",
                "--require-continuous-atmosphere-hold",
            ]
        )
        == 0
    )
    run_dir = next((tmp_path / "out").glob("pressure_only_*"))
    quick_check = next(run_dir.glob("pressure_channel_quick_check_*.csv"))
    rows = list(csv.DictReader(quick_check.open("r", encoding="utf-8-sig", newline="")))
    assert rows
    assert {row["pressure_channel_row_status"] for row in rows} == {"paired"}
    assert {row["pressure_atmosphere_hold_status"] for row in rows} == {"verified"}
    assert {row["pressure_atmosphere_hold_active"] for row in rows} == {"True"}
    samples = next(run_dir.glob("samples_*.csv"))
    sample_rows = list(csv.DictReader(samples.open("r", encoding="utf-8-sig", newline="")))
    assert sample_rows
    assert {row["pressure_atmosphere_hold_status"] for row in sample_rows} == {"verified"}
    assert {row["pressure_atmosphere_hold_active"] for row in sample_rows} == {"True"}
    assert any(call[0] == "enter_atmosphere_mode" for call in fake_pace.calls)


def test_validate_pressure_only_controls_non_ambient_pressure_no_write(monkeypatch, tmp_path: Path) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["pressure_controller"]["enabled"] = True
    cfg["devices"]["pressure_gauge"]["enabled"] = True
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, cfg)
    fake_analyzer = _FakeGasAnalyzer()
    fake_pace = _FakePace()
    fake_gauge = _FakeControlledPressureGauge(fake_pace)

    monkeypatch.setattr(
        validate_pressure_only,
        "_build_devices",
        lambda cfg, io_logger=None: {
            "gas_analyzer": fake_analyzer,
            "gas_analyzer_01": fake_analyzer,
            "pressure_gauge": fake_gauge,
            "pace": fake_pace,
        },
    )
    monkeypatch.setattr(validate_pressure_only, "_close_devices", lambda devices: None)

    assert (
        validate_pressure_only.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--run-id",
                "controlled_pressure",
                "--pressure-points",
                "ambient,900",
                "--count",
                "2",
                "--interval-s",
                "0",
                "--no-prompt",
                "--require-continuous-atmosphere-hold",
                "--control-pressure-points",
                "--pressure-control-stable-s",
                "0",
                "--pressure-control-timeout-s",
                "2",
                "--pressure-control-atmosphere-release-wait-s",
                "0",
                "--pressure-control-post-stable-wait-s",
                "0",
                "--pressure-control-analyzer-stream-flush-s",
                "0.01",
            ]
        )
        == 0
    )

    run_dir = tmp_path / "out" / "controlled_pressure"
    sample_rows = list(csv.DictReader(next(run_dir.glob("samples_*.csv")).open("r", encoding="utf-8-sig", newline="")))
    controlled_rows = [row for row in sample_rows if row.get("pressure_control_enabled") == "True"]
    assert controlled_rows
    assert {row["pressure_control_status"] for row in controlled_rows} == {"verified"}
    assert {row["pressure_control_target_hpa"] for row in controlled_rows} == {"900.0"}
    assert {row["pressure_control_atmosphere_release_status"] for row in controlled_rows} == {"verified"}
    assert {row["pressure_control_atmosphere_release_vent_after_valve_open"] for row in controlled_rows} == {""}
    assert {row["pressure_control_analyzer_stream_flush_status"] for row in controlled_rows} == {"done"}
    assert {row["pressure_control_controls_water_or_gas_routes"] for row in controlled_rows} == {"False"}
    assert {row["pressure_control_route_sealed_before_setpoint"] for row in controlled_rows} == {"False"}
    assert {row["pressure_control_route_seal_reason"] for row in controlled_rows} == {"external_or_manual_closed_volume"}
    assert {row["pressure_control_writes_senco"] for row in controlled_rows} == {"False"}
    assert {row["pressure_control_writes_device_id"] for row in controlled_rows} == {"False"}
    quick_check_rows = list(
        csv.DictReader(next(run_dir.glob("pressure_channel_quick_check_*.csv")).open("r", encoding="utf-8-sig", newline=""))
    )
    assert {row["pressure_mode"] for row in quick_check_rows if row["com22_pressure_hpa"] == "900.0"} == {
        "pace_no_write_controlled"
    }
    assert not any(call[0] == "set_vent_after_valve_open" for call in fake_pace.calls)
    assert any(call[0] == "drain_stream" for call in fake_analyzer.calls)
    assert any(call[0] == "set_range" and call[1] == "2.00bara" for call in fake_pace.calls)
    assert any(
        call[0] == "set_setpoint" and abs(float(call[1]) - 900.0) < 1e-9
        for call in fake_pace.calls
    )
    assert any(call[0] == "enable_control_output" for call in fake_pace.calls)
    assert ("set_slew_mode_max",) in fake_pace.calls
    assert ("set_overshoot_allowed", True) in fake_pace.calls
    setpoint_index = next(idx for idx, call in enumerate(fake_pace.calls) if call[0] == "set_setpoint")
    output_index = next(idx for idx, call in enumerate(fake_pace.calls) if call[0] == "enable_control_output")
    assert setpoint_index < output_index
    assert any(call[0] == "set_range" and call[1] == "2.00bara" for call in fake_pace.calls)


def test_validate_pressure_only_defaults_to_absolute_range_without_configured_range_enable(
    monkeypatch,
    tmp_path: Path,
) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["pressure_controller"]["enabled"] = True
    cfg["devices"]["pressure_gauge"]["enabled"] = True
    cfg.setdefault("workflow", {}).setdefault("pressure", {})["control_range"] = "2.00bara"
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, cfg)
    fake_analyzer = _FakeGasAnalyzer()
    fake_pace = _FakePace()
    fake_gauge = _FakeControlledPressureGauge(fake_pace)

    monkeypatch.setattr(
        validate_pressure_only,
        "_build_devices",
        lambda cfg, io_logger=None: {
            "gas_analyzer": fake_analyzer,
            "gas_analyzer_01": fake_analyzer,
            "pressure_gauge": fake_gauge,
            "pace": fake_pace,
        },
    )
    monkeypatch.setattr(validate_pressure_only, "_close_devices", lambda devices: None)

    assert (
        validate_pressure_only.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--run-id",
                "controlled_pressure_no_range_default",
                "--pressure-points",
                "900",
                "--count",
                "1",
                "--interval-s",
                "0",
                "--no-prompt",
                "--control-pressure-points",
                "--pressure-control-stable-s",
                "0",
                "--pressure-control-timeout-s",
                "2",
                "--pressure-control-atmosphere-release-wait-s",
                "0",
                "--pressure-control-post-stable-wait-s",
                "0",
                "--pressure-control-analyzer-stream-flush-s",
                "0.01",
            ]
        )
        == 0
    )

    assert ("set_range", "2.00bara") in fake_pace.calls
    assert ("set_range", "1.00barg") not in fake_pace.calls
    units_index = fake_pace.calls.index(("set_units_hpa",))
    enter_hold_index = next(idx for idx, call in enumerate(fake_pace.calls) if call[0] == "enter_atmosphere_mode")
    slew_index = fake_pace.calls.index(("set_slew_mode_max",))
    over_index = fake_pace.calls.index(("set_overshoot_allowed", True))
    setpoint_index = next(idx for idx, call in enumerate(fake_pace.calls) if call[0] == "set_setpoint")
    assert enter_hold_index < units_index < slew_index < over_index < setpoint_index
    setpoint_call = next(call for call in fake_pace.calls if call[0] == "set_setpoint")
    assert abs(float(setpoint_call[1]) - 900.0) < 1e-9


def test_validate_pressure_only_reuses_closed_volume_between_control_points(
    monkeypatch,
    tmp_path: Path,
) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["pressure_controller"]["enabled"] = True
    cfg["devices"]["pressure_gauge"]["enabled"] = True
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, cfg)
    fake_analyzer = _FakeGasAnalyzer()
    fake_pace = _FakePace()
    fake_gauge = _FakeControlledPressureGauge(fake_pace)

    monkeypatch.setattr(
        validate_pressure_only,
        "_build_devices",
        lambda cfg, io_logger=None: {
            "gas_analyzer": fake_analyzer,
            "gas_analyzer_01": fake_analyzer,
            "pressure_gauge": fake_gauge,
            "pace": fake_pace,
        },
    )
    monkeypatch.setattr(validate_pressure_only, "_close_devices", lambda devices: None)

    assert (
        validate_pressure_only.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--run-id",
                "controlled_pressure_closed_volume_reuse",
                "--pressure-points",
                "1100,500",
                "--count",
                "1",
                "--interval-s",
                "0",
                "--no-prompt",
                "--control-pressure-points",
                "--pressure-control-setpoint-mode",
                "absolute",
                "--pressure-control-stable-s",
                "0",
                "--pressure-control-timeout-s",
                "2",
                "--pressure-control-atmosphere-release-wait-s",
                "0",
                "--pressure-control-post-stable-wait-s",
                "0",
                "--pressure-control-analyzer-stream-flush-s",
                "0.01",
            ]
        )
        == 0
    )

    run_dir = tmp_path / "out" / "controlled_pressure_closed_volume_reuse"
    sample_rows = list(csv.DictReader(next(run_dir.glob("samples_*.csv")).open("r", encoding="utf-8-sig", newline="")))
    controlled_rows = [row for row in sample_rows if row.get("pressure_control_enabled") == "True"]
    assert [row["pressure_control_target_hpa"] for row in controlled_rows] == ["1100.0", "500.0"]
    assert [call for call in fake_pace.calls if call[0] == "set_setpoint"] == [
        ("set_setpoint", 1100.0),
        ("set_setpoint", 500.0),
    ]
    assert sum(1 for call in fake_pace.calls if call[0] == "exit_atmosphere_mode") == 1
    assert controlled_rows[0]["pressure_control_atmosphere_release_reason"] != (
        "reused_closed_pressure_volume_between_control_points"
    )
    assert controlled_rows[1]["pressure_control_atmosphere_release_reason"] == (
        "reused_closed_pressure_volume_between_control_points"
    )


def test_validate_pressure_only_auto_control_uses_absolute_range_above_atmosphere(
    monkeypatch,
    tmp_path: Path,
) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["pressure_controller"]["enabled"] = True
    cfg["devices"]["pressure_gauge"]["enabled"] = True
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, cfg)
    fake_analyzer = _FakeGasAnalyzer()
    fake_pace = _FakePace()
    fake_gauge = _FakeControlledPressureGauge(fake_pace)

    monkeypatch.setattr(
        validate_pressure_only,
        "_build_devices",
        lambda cfg, io_logger=None: {
            "gas_analyzer": fake_analyzer,
            "gas_analyzer_01": fake_analyzer,
            "pressure_gauge": fake_gauge,
            "pace": fake_pace,
        },
    )
    monkeypatch.setattr(validate_pressure_only, "_close_devices", lambda devices: None)

    assert (
        validate_pressure_only.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--run-id",
                "controlled_pressure_auto_absolute",
                "--pressure-points",
                "1100",
                "--count",
                "1",
                "--interval-s",
                "0",
                "--no-prompt",
                "--control-pressure-points",
                "--pressure-control-setpoint-mode",
                "auto",
                "--pressure-control-stable-s",
                "0",
                "--pressure-control-timeout-s",
                "2",
                "--pressure-control-atmosphere-release-wait-s",
                "0",
                "--pressure-control-post-stable-wait-s",
                "0",
                "--pressure-control-analyzer-stream-flush-s",
                "0.01",
            ]
        )
        == 0
    )

    assert ("set_range", "2.00bara") in fake_pace.calls
    assert any(call[0] == "set_setpoint" and abs(float(call[1]) - 1100.0) < 1e-9 for call in fake_pace.calls)


def test_validate_pressure_only_auto_control_uses_com22_reference_for_below_atmosphere(
    monkeypatch,
    tmp_path: Path,
) -> None:
    class ReferenceGauge:
        def read_pressure(self, **kwargs):
            return 1000.6

        def read_pressure_fast(self, **kwargs):
            return float(fake_pace.pressure_hpa)

    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["pressure_controller"]["enabled"] = True
    cfg["devices"]["pressure_gauge"]["enabled"] = True
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, cfg)
    fake_analyzer = _FakeGasAnalyzer()
    fake_pace = _FakePace()
    fake_pace.source_range = "1.00barg"
    fake_pace.pressure_hpa = 0.0
    fake_gauge = ReferenceGauge()

    monkeypatch.setattr(
        validate_pressure_only,
        "_build_devices",
        lambda cfg, io_logger=None: {
            "gas_analyzer": fake_analyzer,
            "gas_analyzer_01": fake_analyzer,
            "pressure_gauge": fake_gauge,
            "pace": fake_pace,
        },
    )
    monkeypatch.setattr(validate_pressure_only, "_close_devices", lambda devices: None)

    assert (
        validate_pressure_only.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--run-id",
                "controlled_pressure_auto_com22_reference",
                "--pressure-points",
                "900",
                "--count",
                "1",
                "--interval-s",
                "0",
                "--no-prompt",
                "--control-pressure-points",
                "--pressure-control-setpoint-mode",
                "auto",
                "--pressure-control-stable-s",
                "0",
                "--pressure-control-timeout-s",
                "2",
                "--pressure-control-atmosphere-release-wait-s",
                "0",
                "--pressure-control-post-stable-wait-s",
                "0",
                "--pressure-control-analyzer-stream-flush-s",
                "0.01",
            ]
        )
        == 0
    )

    assert ("set_range", "1.00barg") in fake_pace.calls
    assert any(call[0] == "set_setpoint" and abs(float(call[1]) + 100.6) < 1e-9 for call in fake_pace.calls)


def test_validate_pressure_only_uses_configured_control_range_when_explicitly_enabled(
    monkeypatch,
    tmp_path: Path,
) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["pressure_controller"]["enabled"] = True
    cfg["devices"]["pressure_gauge"]["enabled"] = True
    cfg.setdefault("workflow", {}).setdefault("pressure", {})["control_range"] = "2.00bara"
    cfg["workflow"]["pressure"]["control_range_enabled"] = True
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, cfg)
    fake_analyzer = _FakeGasAnalyzer()
    fake_pace = _FakePace()
    fake_gauge = _FakeControlledPressureGauge(fake_pace)

    monkeypatch.setattr(
        validate_pressure_only,
        "_build_devices",
        lambda cfg, io_logger=None: {
            "gas_analyzer": fake_analyzer,
            "gas_analyzer_01": fake_analyzer,
            "pressure_gauge": fake_gauge,
            "pace": fake_pace,
        },
    )
    monkeypatch.setattr(validate_pressure_only, "_close_devices", lambda devices: None)

    assert (
        validate_pressure_only.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--run-id",
                "controlled_pressure_explicit_range",
                "--pressure-points",
                "900",
                "--count",
                "1",
                "--interval-s",
                "0",
                "--no-prompt",
                "--control-pressure-points",
                "--pressure-control-setpoint-mode",
                "absolute",
                "--pressure-control-stable-s",
                "0",
                "--pressure-control-timeout-s",
                "2",
                "--pressure-control-atmosphere-release-wait-s",
                "0",
                "--pressure-control-post-stable-wait-s",
                "0",
                "--pressure-control-analyzer-stream-flush-s",
                "0.01",
            ]
        )
        == 0
    )

    assert ("set_range", "2.00bara") in fake_pace.calls


def test_controlled_pressure_accepts_stable_com22_tail_when_pace_reaches_target(monkeypatch) -> None:
    class SequenceGauge:
        def __init__(self):
            self.values = iter([1098.60, 1098.61, 1098.62, 1098.63, 1098.64, 1098.65])

        def read_pressure_fast(self, **kwargs):
            return next(self.values)

    fake_pace = _FakePace()
    fake_pace.pressure_hpa = 1100.02
    fake_pace.setpoint_hpa = 1100.0

    class Runner:
        devices = {"pace": fake_pace, "pressure_gauge": SequenceGauge()}
        cfg = {"workflow": {"pressure": {"control_setpoint_mode": "absolute"}}}

        def _set_pressure_controller_vent(self, enabled, *, reason):
            fake_pace.vent(enabled)
            return True

    runner = Runner()
    now = [0.0]

    monkeypatch.setattr(validate_pressure_only.time, "monotonic", lambda: now[0])

    def fake_sleep(seconds):
        now[0] += 4.4

    monkeypatch.setattr(validate_pressure_only.time, "sleep", fake_sleep)

    fields = validate_pressure_only._wait_for_controlled_pressure_point(
        runner,
        target_hpa=1100.0,
        tolerance_hpa=1.0,
        stable_s=8.0,
        timeout_s=30.0,
        poll_s=0.5,
        slew_mode="max",
        slew_hpa_per_s=10.0,
        atmosphere_release_wait_s=0.0,
        post_stable_wait_s=0.0,
        analyzer_stream_flush_s=0.0,
    )

    assert fields["pressure_control_status"] == "verified"
    assert fields["pressure_control_reason"] == "pace_internal_pressure_in_tolerance_fast_anchor"
    assert fields["pressure_control_sampling_anchor_policy"] == "first_verified_pressure_window"
    assert fields["pressure_control_reference_hpa"] == 1098.62
    assert abs(float(fields["pressure_control_pace_hpa"]) - 1100.0) <= 1.0
    assert fields["pressure_control_stable_s"] >= 8.0


def test_controlled_pressure_accepts_actual_stable_anchor_near_nominal(monkeypatch) -> None:
    class SequenceGauge:
        def __init__(self):
            self.values = iter([903.80, 903.82, 903.85, 903.86])

        def read_pressure_fast(self, **kwargs):
            return next(self.values)

    class NearNominalPace(_FakePace):
        def __init__(self):
            super().__init__()
            self.values = iter([903.20, 903.22, 903.25, 903.26])

        def set_setpoint(self, value_hpa):
            self.calls.append(("set_setpoint", value_hpa))
            self.setpoint_hpa = float(value_hpa)
            return True

        def read_pressure(self):
            self.pressure_hpa = float(next(self.values))
            return self.pressure_hpa

    fake_pace = NearNominalPace()

    class Runner:
        devices = {"pace": fake_pace, "pressure_gauge": SequenceGauge()}
        cfg = {"workflow": {"pressure": {"control_setpoint_mode": "absolute"}}}

        def _set_pressure_controller_vent(self, enabled, *, reason):
            fake_pace.vent(enabled)
            return True

    runner = Runner()
    now = [0.0]
    monkeypatch.setattr(validate_pressure_only.time, "monotonic", lambda: now[0])

    def fake_sleep(seconds):
        now[0] += 2.0

    monkeypatch.setattr(validate_pressure_only.time, "sleep", fake_sleep)

    fields = validate_pressure_only._wait_for_controlled_pressure_point(
        runner,
        target_hpa=900.0,
        tolerance_hpa=1.0,
        stable_s=4.0,
        timeout_s=20.0,
        poll_s=0.5,
        slew_mode="max",
        slew_hpa_per_s=10.0,
        atmosphere_release_wait_s=0.0,
        post_stable_wait_s=0.0,
        analyzer_stream_flush_s=0.0,
        actual_anchor_near_nominal_hpa=5.0,
        actual_anchor_stability_hpa=1.0,
    )

    assert fields["pressure_control_status"] == "verified"
    assert fields["pressure_control_reason"] == (
        "pace_internal_pressure_actual_pressure_stable_near_nominal_fast_anchor"
    )
    assert fields["pressure_control_actual_anchor_policy"] == "actual_reference_pressure"
    assert fields["pressure_control_actual_anchor_reference_hpa"] == 903.85
    assert fields["pressure_control_actual_anchor_control_hpa"] == 903.25
    assert fields["pressure_control_actual_offset_from_nominal_hpa"] == 3.25
    assert fields["pressure_control_sampling_anchor_policy"] == "first_verified_pressure_window"


def test_controlled_pressure_rejects_actual_anchor_far_from_nominal(monkeypatch) -> None:
    class StableGauge:
        def read_pressure_fast(self, **kwargs):
            return 930.0

    class FarNominalPace(_FakePace):
        def set_setpoint(self, value_hpa):
            self.calls.append(("set_setpoint", value_hpa))
            self.setpoint_hpa = float(value_hpa)
            self.pressure_hpa = 930.0
            return True

    fake_pace = FarNominalPace()

    class Runner:
        devices = {"pace": fake_pace, "pressure_gauge": StableGauge()}
        cfg = {"workflow": {"pressure": {"control_setpoint_mode": "absolute"}}}

        def _set_pressure_controller_vent(self, enabled, *, reason):
            fake_pace.vent(enabled)
            return True

    runner = Runner()
    now = [0.0]
    monkeypatch.setattr(validate_pressure_only.time, "monotonic", lambda: now[0])

    def fake_sleep(seconds):
        now[0] += 2.0

    monkeypatch.setattr(validate_pressure_only.time, "sleep", fake_sleep)

    try:
        validate_pressure_only._wait_for_controlled_pressure_point(
            runner,
            target_hpa=900.0,
            tolerance_hpa=1.0,
            stable_s=4.0,
            timeout_s=8.0,
            poll_s=0.5,
            slew_mode="max",
            slew_hpa_per_s=10.0,
            atmosphere_release_wait_s=0.0,
            post_stable_wait_s=0.0,
            analyzer_stream_flush_s=0.0,
            actual_anchor_near_nominal_hpa=5.0,
            actual_anchor_stability_hpa=1.0,
        )
    except RuntimeError as exc:
        message = str(exc)
    else:
        raise AssertionError("Far-from-nominal pressure was incorrectly accepted")

    assert "reference_side_branch_far_from_nominal" in message


def test_controlled_pressure_post_settle_uses_independent_deadline(monkeypatch) -> None:
    class StableGauge:
        def read_pressure_fast(self, **kwargs):
            return 1099.82

    fake_pace = _FakePace()
    fake_pace.pressure_hpa = 1100.08
    fake_pace.setpoint_hpa = 1100.0
    fake_pace.output_state = 1

    class Runner:
        devices = {"pace": fake_pace, "pressure_gauge": StableGauge()}
        cfg = {}

    runner = Runner()
    now = [100.0]
    monkeypatch.setattr(validate_pressure_only.time, "monotonic", lambda: now[0])

    def fake_sleep(seconds):
        now[0] += float(seconds)

    monkeypatch.setattr(validate_pressure_only.time, "sleep", fake_sleep)

    fields = validate_pressure_only._finalize_pressure_control_after_candidate(
        runner,
        pace=fake_pace,
        target_hpa=1100.0,
        tolerance_hpa=1.0,
        stable_s=2.0,
        deadline=99.0,
        poll_s=0.5,
        start=90.0,
        post_stable_wait_s=0.0,
        analyzer_stream_flush_s=0.0,
        release_fields={},
    )

    assert fields is not None
    assert fields["pressure_control_status"] == "verified"
    assert fields["pressure_control_reason"] == "pace_internal_pressure_in_tolerance_after_settle"


def test_controlled_pressure_blocks_when_pace_internal_pressure_does_not_build(monkeypatch) -> None:
    class SequenceGauge:
        def read_pressure_fast(self, **kwargs):
            return 1098.62

    class NoBuildPace(_FakePace):
        def set_setpoint(self, value_hpa):
            self.calls.append(("set_setpoint", value_hpa))
            self.setpoint_hpa = float(value_hpa)
            return True

    fake_pace = NoBuildPace()
    fake_pace.pressure_hpa = 1000.6
    fake_pace.setpoint_hpa = 1000.6

    class Runner:
        devices = {"pace": fake_pace, "pressure_gauge": SequenceGauge()}
        cfg = {}

        def _set_pressure_controller_vent(self, enabled, *, reason):
            fake_pace.vent(enabled)
            return True

    runner = Runner()
    now = [0.0]

    monkeypatch.setattr(validate_pressure_only.time, "monotonic", lambda: now[0])

    def fake_sleep(seconds):
        now[0] += 4.4

    monkeypatch.setattr(validate_pressure_only.time, "sleep", fake_sleep)

    try:
        validate_pressure_only._wait_for_controlled_pressure_point(
            runner,
            target_hpa=1100.0,
            tolerance_hpa=1.0,
            stable_s=8.0,
            timeout_s=20.0,
            poll_s=0.5,
            slew_mode="linear",
            slew_hpa_per_s=10.0,
            atmosphere_release_wait_s=0.0,
            post_stable_wait_s=0.0,
            analyzer_stream_flush_s=0.0,
        )
    except RuntimeError as exc:
        message = str(exc)
    else:
        raise AssertionError("PACE internal pressure did not reach target but control was accepted")

    assert "pace_internal_pressure_far_from_target" in message or (
        "controller_output_on_but_pressure_not_building" in message
    )


def test_pressure_control_wait_trace_records_reference_and_pace() -> None:
    rows = []

    class Runner:
        def _append_pressure_trace_row(self, **kwargs):
            rows.append(kwargs)

    validate_pressure_only._append_pressure_control_wait_trace(
        Runner(),
        target_hpa=1100.0,
        reference_hpa=1004.2,
        pace_hpa=1004.1,
        error_hpa=-95.8,
        stable_for_s=0.0,
        allowed_error_hpa=1.0,
        elapsed_s=12.5,
        reason="reference_pressure_far_from_nominal",
    )

    assert len(rows) == 1
    row = rows[0]
    assert row["trace_stage"] == "pressure_control_wait_poll"
    assert row["pressure_target_hpa"] == 1100.0
    assert row["pressure_gauge_hpa"] == 1004.2
    assert row["pace_pressure_hpa"] == 1004.1
    assert row["trigger_reason"] == "reference_pressure_far_from_nominal"
    assert row["extra_fields"]["pressure_delta_to_target_hpa"] == -95.8
    assert row["extra_fields"]["pressure_in_limit"] is False
    assert "stable_for_s=0.000" in row["extra_fields"]["pressure_stable_evidence"]


def test_validate_pressure_only_closes_aux_vent_after_valve_only_when_enabled(monkeypatch, tmp_path: Path) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["pressure_controller"]["enabled"] = True
    cfg["devices"]["pressure_gauge"]["enabled"] = True
    cfg.setdefault("workflow", {}).setdefault("pressure", {})["vent_after_valve_open"] = True
    cfg["workflow"]["pressure"]["control_range"] = "2.00bara"
    cfg["workflow"]["pressure"]["control_range_enabled"] = True
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, cfg)
    fake_analyzer = _FakeGasAnalyzer()
    fake_pace = _FakePace()
    fake_gauge = _FakeControlledPressureGauge(fake_pace)

    monkeypatch.setattr(
        validate_pressure_only,
        "_build_devices",
        lambda cfg, io_logger=None: {
            "gas_analyzer": fake_analyzer,
            "gas_analyzer_01": fake_analyzer,
            "pressure_gauge": fake_gauge,
            "pace": fake_pace,
        },
    )
    monkeypatch.setattr(validate_pressure_only, "_close_devices", lambda devices: None)

    assert (
        validate_pressure_only.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--run-id",
                "controlled_pressure_aux",
                "--pressure-points",
                "900",
                "--count",
                "1",
                "--interval-s",
                "0",
                "--no-prompt",
                "--control-pressure-points",
                "--pressure-control-setpoint-mode",
                "absolute",
                "--pressure-control-stable-s",
                "0",
                "--pressure-control-timeout-s",
                "2",
                "--pressure-control-atmosphere-release-wait-s",
                "0",
                "--pressure-control-post-stable-wait-s",
                "0",
                "--pressure-control-analyzer-stream-flush-s",
                "0.01",
            ]
        )
        == 0
    )

    assert any(call[0] == "set_vent_after_valve_open" and call[1] is False for call in fake_pace.calls)


def test_validate_pressure_only_exports_all_active_analyzer_pressure_channels(monkeypatch, tmp_path: Path) -> None:
    cfg = _base_cfg(tmp_path / "logs")
    cfg["devices"]["pressure_controller"]["enabled"] = True
    cfg["devices"]["pressure_gauge"]["enabled"] = True
    cfg["devices"]["gas_analyzers"] = [
        {"enabled": True, "name": "ga01", "port": "COM1", "baud": 115200, "device_id": "010", "active_send": True, "ftd_hz": 10, "average_co2": 1, "average_h2o": 1},
        {"enabled": True, "name": "ga02", "port": "COM2", "baud": 115200, "device_id": "011", "active_send": True, "ftd_hz": 10, "average_co2": 1, "average_h2o": 1},
    ]
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, cfg)
    reference_path = tmp_path / "pressure_reference.json"
    _write_json(
        reference_path,
        {
            "device_id": "COM22-DPG-001",
            "certificate_id": "P-CERT-001",
            "certificate_uncertainty": 0.15,
            "valid_until": "2027-01-01",
            "certificate_hash": "pressure-cert-hash",
        },
    )
    fake_ga01 = _FakeGasAnalyzer(pressure_kpa=100.0, device_id="010")
    fake_ga02 = _FakeGasAnalyzer(pressure_kpa=101.0, device_id="011")
    fake_pace = _FakePace()

    monkeypatch.setattr(
        validate_pressure_only,
        "_build_devices",
        lambda cfg, io_logger=None: {
            "gas_analyzer": fake_ga01,
            "gas_analyzer_01": fake_ga01,
            "gas_analyzer_02": fake_ga02,
            "pressure_gauge": _FakePressureGauge(),
            "pace": fake_pace,
        },
    )
    monkeypatch.setattr(validate_pressure_only, "_close_devices", lambda devices: None)

    assert (
        validate_pressure_only.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--run-id",
                "multi_pressure",
                "--pressure-reference-json",
                str(reference_path),
                "--pressure-points",
                "ambient",
                "--count",
                "3",
                "--interval-s",
                "0",
                "--no-prompt",
                "--require-continuous-atmosphere-hold",
            ]
        )
        == 0
    )

    run_dir = tmp_path / "out" / "multi_pressure"
    summary_path = run_dir / "pressure_channel_multi_analyzer_summary.csv"
    rows = list(csv.DictReader(summary_path.open("r", encoding="utf-8-sig", newline="")))
    by_prefix = {row["analyzer_prefix"]: row for row in rows}

    assert set(by_prefix) == {"ga01", "ga02"}
    assert by_prefix["ga01"]["analyzer_device_id"] == "010"
    assert by_prefix["ga02"]["analyzer_device_id"] == "011"
    assert "acquisition channel" in by_prefix["ga01"]["identity_note"]
    assert by_prefix["ga01"]["status"] == "pass"
    assert by_prefix["ga01"]["allowed_for_co2_h2o_formal_work"] == "True"
    assert by_prefix["ga02"]["status"] == "fail"
    assert by_prefix["ga02"]["allowed_for_co2_h2o_formal_work"] == "False"
    assert (run_dir / "pressure_channel_validation" / "pressure_validation_summary.csv").exists()
    assert (run_dir / "pressure_channel_validation_ga02" / "pressure_validation_summary.csv").exists()
    assert (run_dir / "pressure_channel_validation_all" / "pressure_validation_summary.csv").exists()
    assert next(run_dir.glob("pressure_channel_quick_check_multi_pressure.csv")).exists()
    assert next(run_dir.glob("pressure_channel_quick_check_multi_pressure_ga02.csv")).exists()
    all_quick_check = next(run_dir.glob("pressure_channel_quick_check_multi_pressure_all.csv"))
    all_rows = list(csv.DictReader(all_quick_check.open("r", encoding="utf-8-sig", newline="")))
    assert {row["analyzer_prefix"] for row in all_rows} == {"ga01", "ga02"}
    sample_rows = list(csv.DictReader(next(run_dir.glob("samples_*.csv")).open("r", encoding="utf-8-sig", newline="")))
    assert {row["analyzer_sampling_worker_mode"] for row in sample_rows} == {"per_device"}
    assert {row["analyzer_sampling_active_labels"] for row in sample_rows} == {"ga01,ga02"}


def test_verify_coefficient_roundtrip_with_same_value_write(monkeypatch, tmp_path: Path) -> None:
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, _base_cfg(tmp_path / "logs"))
    monkeypatch.setattr(verify_coefficient_roundtrip, "GasAnalyzer", _FakeRoundtripAnalyzer)

    assert verify_coefficient_roundtrip.main(["--config", str(cfg_path), "--analyzer", "GA01", "--write-back-same", "--output-dir", str(tmp_path / "out")]) == 0
    assert any((tmp_path / "out").glob("coefficient_roundtrip_*.xlsx"))


def test_validation_tool_import_smoke() -> None:
    import gas_calibrator.tools.run_prevalidation_no_sources  # noqa: F401
    import gas_calibrator.tools.run_headless  # noqa: F401
    import gas_calibrator.tools.verify_short_run  # noqa: F401
    import gas_calibrator.ui.app  # noqa: F401


def test_run_prevalidation_no_sources_parses_flags() -> None:
    args = run_prevalidation_no_sources._parse_args(
        [
            "--skip-offline",
            "--include-pressure",
            "--include-roundtrip",
            "--allow-write-back-same",
            "--analyzer",
            "GA01",
            "--fail-fast",
        ]
    )

    assert args.skip_offline is True
    assert args.include_pressure is True
    assert args.include_roundtrip is True
    assert args.allow_write_back_same is True
    assert args.analyzer == "GA01"
    assert args.fail_fast is True


def test_run_prevalidation_no_sources_roundtrip_defaults_to_readonly(monkeypatch, tmp_path: Path) -> None:
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, _base_cfg(tmp_path / "logs"))
    captured = {}

    def _roundtrip(argv):
        captured["argv"] = list(argv)
        return 0

    monkeypatch.setattr(run_prevalidation_no_sources.verify_coefficient_roundtrip, "main", _roundtrip)

    assert (
        run_prevalidation_no_sources.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--skip-offline",
                "--skip-dry-collect",
                "--include-roundtrip",
            ]
        )
        == 0
    )
    assert "--write-back-same" not in captured["argv"]
    assert (tmp_path / "out" / "summary.json").exists()


def test_run_prevalidation_no_sources_continues_after_failed_step(monkeypatch, tmp_path: Path) -> None:
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, _base_cfg(tmp_path / "logs"))
    offline_run_dir = tmp_path / "run_demo"
    offline_run_dir.mkdir()
    calls: list[str] = []

    def _offline(_argv):
        calls.append("offline")
        return 0

    def _dry(_argv):
        calls.append("dry_collect")
        return 1

    def _roundtrip(_argv):
        calls.append("roundtrip")
        return 0

    monkeypatch.setattr(run_prevalidation_no_sources.validate_offline_run, "main", _offline)
    monkeypatch.setattr(run_prevalidation_no_sources.validate_dry_collect, "main", _dry)
    monkeypatch.setattr(run_prevalidation_no_sources.verify_coefficient_roundtrip, "main", _roundtrip)

    assert (
        run_prevalidation_no_sources.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--offline-run-dir",
                str(offline_run_dir),
                "--include-roundtrip",
            ]
        )
        == 1
    )
    assert calls == ["offline", "dry_collect", "roundtrip"]

    summary = json.loads((tmp_path / "out" / "summary.json").read_text(encoding="utf-8"))
    assert [step["name"] for step in summary["steps"]] == ["offline", "dry_collect", "roundtrip"]
    assert [step["status"] for step in summary["steps"]] == ["PASS", "FAIL", "PASS"]


def test_run_prevalidation_no_sources_writes_summary_files(monkeypatch, tmp_path: Path) -> None:
    cfg_path = tmp_path / "cfg.json"
    _write_json(cfg_path, _base_cfg(tmp_path / "logs"))

    def _dry(argv):
        out_dir = Path(argv[argv.index("--output-dir") + 1])
        out_dir.mkdir(parents=True, exist_ok=True)
        (out_dir / "dry_collect_report.txt").write_text("ok", encoding="utf-8")
        return 0

    monkeypatch.setattr(run_prevalidation_no_sources.validate_dry_collect, "main", _dry)

    assert (
        run_prevalidation_no_sources.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--skip-offline",
            ]
        )
        == 0
    )
    summary_json = tmp_path / "out" / "summary.json"
    summary_md = tmp_path / "out" / "summary.md"
    assert summary_json.exists()
    assert summary_md.exists()
    assert "dry_collect" in summary_json.read_text(encoding="utf-8")
    md_text = summary_md.read_text(encoding="utf-8")
    assert "dry_collect" in md_text
    assert "frame_quality_summary" in md_text


def test_run_prevalidation_no_sources_old_config_compatible(tmp_path: Path) -> None:
    cfg_path = tmp_path / "cfg_min.json"
    _write_json(
        cfg_path,
        {
            "paths": {"output_dir": str(tmp_path / "logs")},
            "devices": {
                "gas_analyzers": [{"enabled": True, "name": "GA01", "device_id": "010"}],
                "gas_analyzer": {"enabled": False},
            },
        },
    )

    assert (
        run_prevalidation_no_sources.main(
            [
                "--config",
                str(cfg_path),
                "--output-dir",
                str(tmp_path / "out"),
                "--skip-offline",
                "--skip-dry-collect",
            ]
        )
        == 0
    )
    assert (tmp_path / "out" / "summary.json").exists()
