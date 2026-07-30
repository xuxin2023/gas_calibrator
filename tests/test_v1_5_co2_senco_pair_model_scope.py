import csv
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from gas_calibrator.storage.v1_5_evidence.bundle import (
    _artifact_role,
    _build_sidecar_candidates,
    _build_sidecar_qc_rows,
    _load_database_sidecar_rows,
)
from gas_calibrator.tools.export_v1_5_co2_senco_pair_model_scope import main as cli_main
from gas_calibrator.validation.co2_senco_pair_model_scope import (
    build_co2_senco_pair_model_scope_tables,
    write_co2_senco_pair_model_scope_report,
)
from gas_calibrator.validation.reporting import _safe_worksheet_title


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields: list[str] = []
    for row in rows:
        for key in row:
            if key not in fields:
                fields.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def test_safe_worksheet_title_is_bounded_unique_and_excel_compatible():
    first = _safe_worksheet_title("A" * 40, [])
    second = _safe_worksheet_title("a" * 40, [first])
    sanitized = _safe_worksheet_title("bad/name?with*invalid[chars]:", [])

    assert first == "A" * 31
    assert second.endswith("_2")
    assert len(second) == 31
    assert first.casefold() != second.casefold()
    assert sanitized == "bad_name_with_invalid_chars__"


def _write_points(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "points"
    ws.append(["Normalized calibration points", None, None, None, None])
    ws.append(["Temp_C", "CO2_ppm", "H2O_text", "Pressure_hPa", "CO2_group"])
    ws.append([-20, 0, None, 1100, None])
    ws.append([-20, 400, None, 800, None])
    ws.append([-20, 1000, None, 500, None])
    for ppm, pressure, group in [
        (0, 1100, None),
        (100, 1000, "B"),
        (200, 1000, None),
        (300, 900, "B"),
        (400, 900, None),
        (500, 800, "B"),
        (600, 800, None),
        (700, 700, "B"),
        (800, 700, None),
        (900, 600, "B"),
        (1000, 600, None),
    ]:
        ws.append([20, ppm, None, pressure, group])
    ws.append([40, 0, None, 1100, None])
    ws.append([40, 400, None, 900, None])
    ws.append([40, 1000, None, 500, None])
    ev = wb.create_sheet("execution_view")
    ev.append(["Temp_C", "Order", "H2O_targets", "CO2_sources_ppm", "Pressures_hPa", "Notes"])
    ev.append([-20, "CO2 only", None, "0, 400, 1000", "1100, 800, 550", "sub-zero fixed CO2 only"])
    ev.append([20, "H2O then CO2", "20C / 30,50,70%RH", "0, 100, 200, 300, 400, 500, 600, 700, 800, 900, 1000", "1100, 1000, 900, 800, 700, 600, 550", "200ppm can be skipped by skip_co2_ppm"])
    ev.append([40, "H2O then CO2", "30C / 50,70%RH", "0, 400, 1000", "1100, 900, 550", "use real points only"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _make_inputs(tmp_path: Path) -> tuple[Path, Path, Path]:
    points_xlsx = tmp_path / "points_original.xlsx"
    candidate_dir = tmp_path / "candidate_no_write"
    pair_review_dir = tmp_path / "pair_review"
    _write_points(points_xlsx)
    _write_csv(
        candidate_dir / "candidate_policy_summary.csv",
        [
            {
                "component": "co2",
                "analyzer_prefix": "ga01",
                "analyzer_device_id": "023",
                "candidate_status": "verification_passed",
                "fit_sample_count": 100,
                "verification_sample_count": 30,
                "fit_point_count": 5,
                "verification_point_count": 1,
                "distinct_fit_targets": 5,
                "pressure_span_hpa": 1.5,
                "temperature_span_c": 0.04,
                "selected_model_terms": "intercept;R;R2;R3",
                "frozen_terms": "P;RP;RTP;T;T2;RT",
                "verification_status": "pass",
                "formal_pressure_validation_status": "pass",
                "not_pressure_compensation_fit": "True",
            }
        ],
    )
    _write_csv(
        pair_review_dir / "co2_senco_pair_review_summary.csv",
        [
            {
                "review_status": "blocked_single_senco1_write_failed_pair_review_required",
                "single_senco1_write_verification_status": "failed",
            }
        ],
    )
    return points_xlsx, candidate_dir, pair_review_dir


def _make_multitemp_inputs(tmp_path: Path) -> tuple[Path, Path, Path]:
    points_xlsx, candidate_dir, pair_review_dir = _make_inputs(tmp_path)
    _write_csv(
        candidate_dir / "candidate_policy_summary.csv",
        [
            {
                "component": "co2",
                "analyzer_prefix": "ga01",
                "analyzer_device_id": "023",
                "candidate_status": "fit_ready_requires_verification",
                "fit_sample_count": 450,
                "verification_sample_count": 0,
                "fit_point_count": 45,
                "verification_point_count": 0,
                "distinct_fit_targets": 11,
                "pressure_span_hpa": 13.0,
                "temperature_span_c": 60.0,
                "selected_model_terms": "intercept;R;R2;R3;T;T2;RT",
                "frozen_terms": "P;RP;RTP",
                "verification_status": "missing",
                "formal_pressure_validation_status": "pass",
                "not_pressure_compensation_fit": "True",
            }
        ],
    )
    return points_xlsx, candidate_dir, pair_review_dir


def test_co2_senco_pair_model_scope_marks_primary_slice_and_blocks_secondary(tmp_path):
    points_xlsx, candidate_dir, pair_review_dir = _make_inputs(tmp_path)

    tables = build_co2_senco_pair_model_scope_tables(
        original_points_xlsx=points_xlsx,
        candidate_dir=candidate_dir,
        pair_review_dir=pair_review_dir,
        output_dir=tmp_path / "out",
    )

    summary = tables["co2_senco_pair_model_scope_summary"][0]
    assert summary["review_status"] == "blocked_secondary_terms_not_identifiable_no_write"
    assert summary["senco1_senco3_pair_write_allowed"] is False
    assert summary["senco5_linear_correction_in_scope"] is True
    assert summary["senco5_linear_correction_write_allowed"] is False
    assert summary["v1_execution_view_present"] is True
    family = {row["senco_group"]: row for row in tables["co2_senco_pair_manual_co2_coefficient_family"]}
    assert family["SENCO1"]["v1_5_current_status"] == "reviewable_from_clean_open_flow_ratio_evidence"
    assert family["SENCO5"]["v1_5_current_status"] == "in_scope_integrated_final_output_candidate_no_write"
    assert family["SENCO9"]["v1_5_current_status"] == "handled_by_independent_pressure_channel_workflow"
    senco5_contract = {
        row["contract_item"]: row for row in tables["co2_senco5_linear_correction_contract"]
    }
    assert senco5_contract["manual_scope"]["status"] == "confirmed_manual_scope"
    assert senco5_contract["not_in_legacy_ratio_fit"]["status"] == "legacy_algorithm_does_not_identify_senco5"
    assert senco5_contract["target_quantity"]["status"] == "required_in_same_candidate_package"
    assert senco5_contract["fit_model"]["status"] == "reviewable_no_write_with_decimal_contract"
    assert "raw ratio residuals" in senco5_contract["exclusions"]["requirement"]
    execution_summary = tables["co2_senco_pair_v1_execution_view_summary"][0]
    assert execution_summary["execution_view_present"] is True
    assert "100" in execution_summary["co2_sources_ppm"]
    loader_rows = tables["co2_senco_pair_v1_loader_comparison"]
    assert {row["carry_forward_h2o"] for row in loader_rows} == {False, True}
    alignment = {
        row["alignment_item"]: row for row in tables["co2_senco_pair_point_table_alignment"]
    }
    assert alignment["current_fit_targets_vs_original_20c_b_group"]["status"] == "match"
    assert alignment["current_fit_targets_vs_v1_execution_20c_sources"]["status"] == "subset_of_v1_execution_sources"
    terms = {row["coefficient"]: row for row in tables["co2_senco_pair_term_identifiability"]}
    assert terms["a0"]["identifiability_status"] == "currently_identifiable_no_write"
    assert terms["a4"]["identifiability_status"] == "blocked_current_temp_span_too_small"
    assert terms["a7"]["identifiability_status"] == "blocked_pressure_terms_frozen_by_v1_5_policy"
    assert terms["senco5_C0"]["senco_group"] == "SENCO5"
    assert terms["senco5_C0"]["identifiability_status"] == "requires_integrated_output_layer_candidate_review"
    assert terms["senco5_C1"]["write_allowed"] is False
    decisions = {row["decision_item"]: row for row in tables["co2_senco_pair_model_decision"]}
    assert decisions["preserve_old_senco3_with_new_senco1"]["decision_status"] == "blocked_by_post_write_failure"
    assert (
        decisions["co2_concentration_linear_correction_senco5"]["decision_status"]
        == "in_scope_integrated_final_output_candidate_no_write"
    )


def test_co2_senco_pair_model_scope_marks_multitemp_temperature_terms_reviewable(tmp_path):
    points_xlsx, candidate_dir, pair_review_dir = _make_multitemp_inputs(tmp_path)

    tables = build_co2_senco_pair_model_scope_tables(
        original_points_xlsx=points_xlsx,
        candidate_dir=candidate_dir,
        pair_review_dir=pair_review_dir,
        output_dir=tmp_path / "out",
    )

    summary = tables["co2_senco_pair_model_scope_summary"][0]
    assert summary["review_status"] == "secondary_temperature_terms_reviewable_pressure_terms_frozen_no_write"
    assert summary["senco3_temperature_terms_reviewable"] is True
    assert summary["senco5_linear_correction_in_scope"] is True
    assert summary["pressure_terms_frozen_by_v1_5_policy"] is True
    terms = {row["coefficient"]: row for row in tables["co2_senco_pair_term_identifiability"]}
    assert terms["a4"]["identifiability_status"] == "reviewable_if_formula_contract_confirmed"
    assert terms["a5"]["identifiability_status"] == "reviewable_if_formula_contract_confirmed"
    assert terms["a6"]["identifiability_status"] == "reviewable_if_formula_contract_confirmed"
    assert terms["a7"]["identifiability_status"] == "blocked_pressure_terms_frozen_by_v1_5_policy"
    assert terms["senco5_C0"]["identifiability_status"] == "requires_integrated_output_layer_candidate_review"
    decisions = {row["decision_item"]: row for row in tables["co2_senco_pair_model_decision"]}
    assert (
        decisions["new_senco1_senco3_pair_write"]["decision_status"]
        == "reviewable_no_write_pending_formula_contract_old_snapshot_and_independent_verification"
    )


def test_co2_senco_pair_model_scope_writer_and_database_sidecar(tmp_path):
    points_xlsx, candidate_dir, pair_review_dir = _make_inputs(tmp_path)
    output_dir = tmp_path / "out"

    outputs = write_co2_senco_pair_model_scope_report(
        original_points_xlsx=points_xlsx,
        candidate_dir=candidate_dir,
        pair_review_dir=pair_review_dir,
        output_dir=output_dir,
    )

    assert outputs["workbook"].exists()
    workbook = load_workbook(outputs["workbook"], read_only=True)
    try:
        sheet_names = workbook.sheetnames
        assert all(len(name) <= 31 for name in sheet_names)
        assert len({name.casefold() for name in sheet_names}) == len(sheet_names)
        execution_sheets = [
            name
            for name in sheet_names
            if name.startswith("co2_senco_pair_v1_execution")
        ]
        assert len(execution_sheets) == 2
        assert any(name.endswith("_2") for name in execution_sheets)
    finally:
        workbook.close()
    assert outputs["co2_senco_pair_term_identifiability_csv"].exists()
    sidecar = json.loads(outputs["database_sidecar"].read_text(encoding="utf-8"))
    assert sidecar["no_write"] is True
    assert "coefficient_candidates" in sidecar["database_target_tables"]
    sidecar_rows = _load_database_sidecar_rows(
        [{"id": "artifact-1", "artifact_role": "candidate_coefficient_review", "path": str(outputs["database_sidecar"])}]
    )
    candidates = _build_sidecar_candidates(run_db_id="run-db", sidecar_rows=sidecar_rows)
    qc_rows = _build_sidecar_qc_rows(run_db_id="run-db", sidecar_rows=sidecar_rows)
    assert candidates[0]["candidate_status"] == "blocked_secondary_terms_not_identifiable_no_write"
    assert qc_rows[0]["rule_name"] == "co2_senco3_secondary_identifiability"
    assert {row["rule_name"] for row in qc_rows} >= {
        "co2_senco3_secondary_identifiability",
        "co2_senco5_linear_correction_identifiability",
    }


def test_co2_senco_pair_model_scope_cli_and_artifact_role(tmp_path):
    points_xlsx, candidate_dir, pair_review_dir = _make_inputs(tmp_path)
    output_dir = tmp_path / "out"

    rc = cli_main(
        [
            "--original-points-xlsx",
            str(points_xlsx),
            "--candidate-dir",
            str(candidate_dir),
            "--pair-review-dir",
            str(pair_review_dir),
            "--output-dir",
            str(output_dir),
        ]
    )

    assert rc == 0
    artifact = output_dir / "co2_senco_pair_model_scope_summary.csv"
    assert artifact.exists()
    assert (
        _artifact_role(artifact, plan_path=None, pressure_reference_path=None)
        == "candidate_coefficient_review"
    )
