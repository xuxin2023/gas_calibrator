import csv
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from gas_calibrator.v2.core import CalibrationPhase, CalibrationStatus
from gas_calibrator.v2.core.data_writer import SUMMARY_SCHEMA_VERSION, DataWriter
from gas_calibrator.v2.core.models import CalibrationPoint, SamplingResult


def _sample_results() -> list[SamplingResult]:
    point = CalibrationPoint(
        index=1,
        temperature_c=25.0,
        co2_ppm=400.0,
        pressure_hpa=1000.0,
        route="co2",
    )
    return [
        SamplingResult(
            point=point,
            analyzer_id="ga01",
            timestamp=datetime(2026, 3, 17, 12, 0, 0, tzinfo=timezone.utc),
            h2o_signal=10.5,
            co2_signal=401.2,
            temperature_c=25.0,
            pressure_hpa=1000.0,
            dew_point_c=5.0,
        ),
        SamplingResult(
            point=point,
            analyzer_id="ga02",
            timestamp=datetime(2026, 3, 17, 12, 0, 1, tzinfo=timezone.utc),
            h2o_signal=10.7,
            co2_signal=401.6,
            temperature_c=25.0,
            pressure_hpa=1000.2,
            dew_point_c=5.1,
        ),
    ]


def test_data_writer_creates_run_directory(tmp_path: Path) -> None:
    writer = DataWriter(str(tmp_path), "run_test")

    assert writer.run_dir.exists()
    assert writer.run_dir.is_dir()


def test_data_writer_writes_samples_csv(tmp_path: Path) -> None:
    writer = DataWriter(str(tmp_path), "run_test")

    path = writer.write_samples(_sample_results())

    csv_path = Path(path)
    assert csv_path.exists()
    with csv_path.open("r", encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))
    assert len(rows) == 2
    assert rows[0]["point_index"] == "1"
    assert rows[0]["route"] == "co2"
    assert rows[0]["analyzer_id"] == "ga01"
    assert rows[0]["co2_signal"] == "401.2"


def test_data_writer_writes_samples_excel(tmp_path: Path) -> None:
    writer = DataWriter(str(tmp_path), "run_test")

    path = writer.write_samples_excel(_sample_results())

    workbook = load_workbook(path)
    sheet = workbook.active
    assert sheet.title == "samples"
    assert sheet.max_row == 3
    assert sheet.cell(row=2, column=2).value == 1
    assert sheet.cell(row=2, column=6).value == 400
    assert sheet.cell(row=2, column=9).value == "ga01"
    workbook.close()


def test_data_writer_writes_summary_json(tmp_path: Path) -> None:
    writer = DataWriter(str(tmp_path), "run_test")
    status = CalibrationStatus(
        phase=CalibrationPhase.COMPLETED,
        total_points=2,
        completed_points=2,
        progress=1.0,
        message="done",
        elapsed_s=12.5,
    )

    path = writer.write_summary(
        status,
        {"sample_count": 2},
        started_at="2026-03-17T12:00:00",
        ended_at="2026-03-17T12:00:13",
        warnings=1,
        errors=0,
        software_version="test-version",
    )

    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    assert payload["schema_version"] == SUMMARY_SCHEMA_VERSION
    assert payload["run_id"] == "run_test"
    assert payload["software_version"] == "test-version"
    assert payload["software_build_id"] == "test-version"
    assert payload["started_at"] == "2026-03-17T12:00:00"
    assert payload["ended_at"] == "2026-03-17T12:00:13"
    assert payload["points_total"] == 2
    assert payload["points_completed"] == 2
    assert payload["total_points"] == 2
    assert payload["completed_points"] == 2
    assert payload["warnings"] == 1
    assert payload["errors"] == 0
    assert payload["progress"] == 1.0
    assert payload["status"]["phase"] == "completed"
    assert payload["status"]["total_points"] == 2
    assert payload["status"]["completed_points"] == 2
    assert payload["status"]["progress"] == 1.0
    assert payload["evidence_source"] == "diagnostic"
    assert payload["not_real_acceptance_evidence"] is True
    assert payload["acceptance_level"] == "diagnostic"
    assert payload["promotion_state"] == "dry_run_only"
    assert payload["stats"]["sample_count"] == 2


def test_data_writer_writes_simulation_boundary_fields_when_requested(tmp_path: Path) -> None:
    writer = DataWriter(str(tmp_path), "run_test")
    status = CalibrationStatus(
        phase=CalibrationPhase.COMPLETED,
        total_points=1,
        completed_points=1,
        progress=1.0,
        message="done",
        elapsed_s=1.0,
    )

    path = writer.write_summary(
        status,
        {"sample_count": 1},
        simulation_mode=True,
    )

    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    assert payload["evidence_source"] == "simulated_protocol"
    assert payload["not_real_acceptance_evidence"] is True
    assert payload["acceptance_level"] == "offline_regression"
    assert payload["promotion_state"] == "dry_run_only"


def test_data_writer_promotes_config_safety_to_summary_top_level(tmp_path: Path) -> None:
    writer = DataWriter(str(tmp_path), "run_test")
    status = CalibrationStatus(
        phase=CalibrationPhase.COMPLETED,
        total_points=1,
        completed_points=1,
        progress=1.0,
        message="done",
        elapsed_s=1.0,
    )

    config_safety = {
        "classification": "simulation_real_port_inventory_risk",
        "execution_gate": {"status": "blocked"},
    }
    config_safety_review = {
        "status": "blocked",
        "summary": "Step 2 默认工作流已拦截当前配置。",
        "warnings": ["检测到非仿真设备端口。"],
        "execution_gate": {"status": "blocked"},
    }

    path = writer.write_summary(
        status,
        {
            "sample_count": 1,
            "config_safety": config_safety,
            "config_safety_review": config_safety_review,
        },
        simulation_mode=True,
    )

    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    assert payload["config_safety"]["classification"] == "simulation_real_port_inventory_risk"
    assert payload["config_safety"]["execution_gate"]["status"] == "blocked"
    assert payload["config_safety_review"]["status"] == "blocked"
    assert payload["config_safety_review"]["execution_gate"]["status"] == "blocked"
    assert payload["stats"]["config_safety"]["classification"] == "simulation_real_port_inventory_risk"
    assert payload["stats"]["config_safety_review"]["status"] == "blocked"


def test_data_writer_writes_points_readable_csv(tmp_path: Path) -> None:
    writer = DataWriter(str(tmp_path), "run_test")

    path = writer.write_points_readable(
        [
            {
                "point_index": 1,
                "point_phase": "co2",
                "point_tag": "co2_1",
                "execution_status": "usable",
                "thermometer_temp_c_mean": 25.2,
                "thermometer_reference_status": "healthy",
                "pressure_gauge_hpa_mean": 998.4,
                "pressure_reference_status": "healthy",
                "reference_quality": "healthy",
                "AnalyzerCoverage": "1/1",
            }
        ]
    )

    csv_path = Path(path)
    with csv_path.open("r", encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))
    assert csv_path.exists()
    assert rows[0]["point_tag"] == "co2_1"
    assert rows[0]["thermometer_temp_c_mean"] == "25.2"
    assert rows[0]["pressure_reference_status"] == "healthy"
    assert rows[0]["reference_quality"] == "healthy"
