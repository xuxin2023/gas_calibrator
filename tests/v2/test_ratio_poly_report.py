from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from gas_calibrator.v2.config import CoefficientsConfig
from gas_calibrator.v2.core.models import CalibrationPoint, SamplingResult
from gas_calibrator.v2.export.ratio_poly_report import (
    _format_senco_float,
    build_analyzer_summary_frame,
    build_quality_analysis_bundle,
    export_ratio_poly_report,
    select_corrected_fit_rows,
)


def _sample(
    *,
    point_index: int,
    route: str,
    analyzer_id: str,
    temperature_c: float,
    co2_target: float | None,
    humidity_pct: float | None,
    pressure_hpa: float,
    co2_ratio: float,
    h2o_ratio: float,
    dew_point_c: float,
    thermometer_temp_c: float | None = None,
    pressure_gauge_hpa: float | None = None,
    frame_usable: bool = True,
    sample_index: int = 1,
) -> SamplingResult:
    point = CalibrationPoint(
        index=point_index,
        temperature_c=temperature_c,
        co2_ppm=co2_target,
        humidity_pct=humidity_pct,
        pressure_hpa=pressure_hpa,
        route=route,
    )
    return SamplingResult(
        point=point,
        analyzer_id=analyzer_id,
        timestamp=datetime(2026, 3, 19, 12, 0, 0, tzinfo=timezone.utc),
        co2_ppm=co2_target,
        h2o_mmol=5.0 + point_index,
        co2_signal=1000.0 + point_index,
        h2o_signal=500.0 + point_index,
        co2_ratio_f=co2_ratio,
        h2o_ratio_f=h2o_ratio,
        ref_signal=2000.0 + point_index,
        temperature_c=temperature_c,
        pressure_hpa=pressure_hpa,
        pressure_gauge_hpa=pressure_gauge_hpa,
        thermometer_temp_c=thermometer_temp_c,
        dew_point_c=dew_point_c,
        analyzer_pressure_kpa=pressure_hpa / 10.0,
        analyzer_chamber_temp_c=temperature_c + 0.2,
        case_temp_c=temperature_c + 0.8,
        frame_usable=frame_usable,
        frame_has_data=True,
        frame_status="ok" if frame_usable else "unusable",
        sample_index=sample_index,
    )


def _fit_cfg() -> CoefficientsConfig:
    return CoefficientsConfig.from_dict(
        {
            "enabled": True,
            "auto_fit": True,
            "model": "ratio_poly_rt_p",
            "summary_columns": {
                "co2": {"target": "ppm_CO2_Tank", "ratio": "R_CO2", "temperature": "Temp", "pressure": "BAR"},
                "h2o": {"target": "ppm_H2O_Dew", "ratio": "R_H2O", "temperature": "Temp", "pressure": "BAR"},
            },
        }
    )


def test_build_analyzer_summary_frame_groups_results() -> None:
    samples = [
        _sample(
            point_index=1,
            route="co2",
            analyzer_id="ga01",
            temperature_c=20.0,
            co2_target=400.0,
            humidity_pct=None,
            pressure_hpa=1000.0,
            co2_ratio=1.2,
            h2o_ratio=0.2,
            dew_point_c=2.0,
        ),
        _sample(
            point_index=1,
            route="co2",
            analyzer_id="ga01",
            temperature_c=20.0,
            co2_target=400.0,
            humidity_pct=None,
            pressure_hpa=1000.0,
            co2_ratio=1.4,
            h2o_ratio=0.3,
            dew_point_c=2.5,
        ),
    ]

    frame = build_analyzer_summary_frame(samples)

    assert len(frame) == 1
    row = frame.iloc[0]
    assert row["Analyzer"] == "GA01"
    assert row["PointPhase"] == "气路"
    assert abs(float(row["R_CO2"]) - 1.3) < 1e-9
    assert row["ppm_H2O_Dew"] is not None
    assert abs(float(row["BAR"]) - 100.0) < 1e-9


def test_build_analyzer_summary_frame_uses_aligned_reference_rows_and_fleet_fields() -> None:
    usable = _sample(
        point_index=1,
        route="co2",
        analyzer_id="ga01",
        temperature_c=20.0,
        co2_target=400.0,
        humidity_pct=None,
        pressure_hpa=1000.0,
        co2_ratio=1.2,
        h2o_ratio=0.2,
        dew_point_c=2.0,
        thermometer_temp_c=25.0,
        pressure_gauge_hpa=998.0,
        sample_index=1,
    )
    unusable = _sample(
        point_index=1,
        route="co2",
        analyzer_id="ga01",
        temperature_c=20.0,
        co2_target=400.0,
        humidity_pct=None,
        pressure_hpa=950.0,
        co2_ratio=1.3,
        h2o_ratio=0.2,
        dew_point_c=8.0,
        thermometer_temp_c=35.0,
        pressure_gauge_hpa=930.0,
        frame_usable=False,
        sample_index=2,
    )

    aligned = build_analyzer_summary_frame(
        [usable, unusable],
        expected_analyzers=["GA01", "GA02"],
        reference_on_aligned_rows=True,
    )
    unaligned = build_analyzer_summary_frame(
        [usable, unusable],
        expected_analyzers=["GA01", "GA02"],
        reference_on_aligned_rows=False,
    )

    aligned_row = aligned.iloc[0]
    unaligned_row = unaligned.iloc[0]

    assert aligned_row["AnalyzerCoverage"] == "1/2"
    assert int(aligned_row["UsableAnalyzers"]) == 1
    assert int(aligned_row["ExpectedAnalyzers"]) == 2
    assert aligned_row["MissingAnalyzers"] == "GA02"
    assert bool(aligned_row["ReferenceAlignedRows"]) is True
    assert float(aligned_row["ReferenceThermometerTempC"]) == 25.0
    assert float(aligned_row["ReferencePressureHpa"]) == 998.0
    assert float(aligned_row["Temp"]) == 25.0
    assert float(aligned_row["P"]) == 998.0
    assert float(unaligned_row["ReferenceThermometerTempC"]) == 25.0
    assert float(unaligned_row["Temp"]) == 25.0
    assert float(unaligned_row["ReferencePressureHpa"]) == 964.0


def test_select_corrected_fit_rows_keeps_only_configured_dry_gas_points() -> None:
    frame = build_analyzer_summary_frame(
        [
            _sample(
                point_index=1,
                route="h2o",
                analyzer_id="ga01",
                temperature_c=20.0,
                co2_target=None,
                humidity_pct=30.0,
                pressure_hpa=1000.0,
                co2_ratio=1.0,
                h2o_ratio=0.2,
                dew_point_c=1.0,
            ),
            _sample(
                point_index=2,
                route="co2",
                analyzer_id="ga01",
                temperature_c=-20.0,
                co2_target=400.0,
                humidity_pct=None,
                pressure_hpa=1000.0,
                co2_ratio=1.1,
                h2o_ratio=0.1,
                dew_point_c=-10.0,
            ),
            _sample(
                point_index=3,
                route="co2",
                analyzer_id="ga01",
                temperature_c=10.0,
                co2_target=0.0,
                humidity_pct=None,
                pressure_hpa=1000.0,
                co2_ratio=1.2,
                h2o_ratio=0.1,
                dew_point_c=-9.0,
            ),
            _sample(
                point_index=4,
                route="co2",
                analyzer_id="ga01",
                temperature_c=10.0,
                co2_target=400.0,
                humidity_pct=None,
                pressure_hpa=1000.0,
                co2_ratio=1.3,
                h2o_ratio=0.1,
                dew_point_c=-8.0,
            ),
        ]
    )

    selected = select_corrected_fit_rows(
        frame,
        gas="h2o",
        selection=CoefficientsConfig().h2o_summary_selection,
        temperature_key="Temp",
    )

    assert set(selected["PointRow"]) == {1, 2, 3}


def test_build_quality_analysis_bundle_falls_back_to_bar_when_p_is_blank() -> None:
    samples: list[SamplingResult] = []
    point_index = 1
    for temp, pressure in ((20.0, 900.0), (30.0, 1000.0)):
        for target in (0.0, 400.0, 800.0, 1200.0):
            samples.append(
                _sample(
                    point_index=point_index,
                    route="co2",
                    analyzer_id="ga01",
                    temperature_c=temp,
                    co2_target=target,
                    humidity_pct=None,
                    pressure_hpa=pressure,
                    co2_ratio=1.48 - 0.0002 * target + 0.0003 * temp + 0.00005 * pressure,
                    h2o_ratio=0.18 + 0.0002 * point_index,
                    dew_point_c=-5.0 + point_index * 0.2,
                )
            )
            point_index += 1

    for idx in range(point_index, point_index + 8):
        temp = 10.0 + float(idx % 4)
        pressure = 850.0 + float((idx % 3) * 40.0)
        dew_point = -6.0 + idx * 0.4
        samples.append(
            _sample(
                point_index=idx,
                route="h2o",
                analyzer_id="ga01",
                temperature_c=temp,
                co2_target=None,
                humidity_pct=30.0 + idx,
                pressure_hpa=pressure,
                co2_ratio=1.0 + 0.00015 * idx,
                h2o_ratio=0.45 + 0.012 * idx + 0.002 * temp + 0.00015 * pressure,
                dew_point_c=dew_point,
            )
        )

    frame = build_analyzer_summary_frame(samples)
    frame["P"] = pd.NA

    quality = build_quality_analysis_bundle(frame)

    assert not quality.summary.empty
    assert quality.summary["RISE ppm"].notna().all()
    assert quality.summary["Bias ppm"].notna().all()
    assert quality.summary["R\u00b2\u8d28\u91cf"].notna().all()
    assert not quality.summary["\u7ed3\u8bba"].eq("\u6570\u636e\u7f3a\u5931").all()


def test_export_ratio_poly_report_creates_workbook_and_download_plan(tmp_path: Path) -> None:
    samples: list[SamplingResult] = []
    point_index = 1
    for temp, pressure in ((20.0, 900.0), (30.0, 1000.0)):
        for target in [0.0, 200.0, 400.0, 600.0, 800.0, 1000.0, 1200.0, 1400.0, 1600.0]:
            samples.append(
                _sample(
                    point_index=point_index,
                    route="co2",
                    analyzer_id="ga01",
                    temperature_c=temp,
                    co2_target=target,
                    humidity_pct=None,
                    pressure_hpa=pressure,
                    co2_ratio=1.52 - 0.00024 * target + 0.00045 * temp + 0.00008 * pressure,
                    h2o_ratio=0.2 + 0.0001 * point_index,
                    dew_point_c=1.0 + point_index * 0.2,
                )
            )
            point_index += 1

    for idx in range(point_index, point_index + 9):
        temp = 10.0 + float(idx % 4)
        pressure = 850.0 + float((idx % 3) * 40.0)
        dew_point = -5.0 + idx * 0.5
        samples.append(
            _sample(
                point_index=idx,
                route="h2o",
                analyzer_id="ga01",
                temperature_c=temp,
                co2_target=None,
                humidity_pct=30.0 + idx,
                pressure_hpa=pressure,
                co2_ratio=1.0 + 0.0002 * idx,
                h2o_ratio=0.5 + 0.015 * idx + 0.002 * temp + 0.0002 * pressure,
                dew_point_c=dew_point,
            )
        )

    output = export_ratio_poly_report(samples, out_dir=tmp_path, coeff_cfg=_fit_cfg())

    assert output is not None
    assert output.exists()

    quality = pd.read_excel(output, sheet_name="拟合前基础质量检查", dtype={"ID": str})
    assert {"设备", "ID", "RISE ppm", "Bias ppm", "R²质量", "结论"}.issubset(quality.columns)
    assert quality.loc[0, "设备"] == "GA01"
    assert pd.notna(quality.loc[0, "RISE ppm"])
    assert float(quality.loc[0, "RISE ppm"]).is_integer()
    assert float(quality.loc[0, "Bias ppm"]).is_integer()
    assert round(float(quality.loc[0, "R²质量"]), 2) == float(quality.loc[0, "R²质量"])
    detail = pd.read_excel(output, sheet_name="数据质量分析_分通道")
    assert {
        "设备",
        "气路RISE ppm",
        "气路Bias ppm",
        "气路R²",
        "水路RISE ppm(等效)",
        "水路Bias ppm(等效)",
        "水路R²",
        "RISE ppm",
        "Bias ppm",
        "R²质量",
    }.issubset(detail.columns)
    quality_notes = pd.read_excel(output, sheet_name="数据质量分析_说明")
    assert "说明" in quality_notes.columns
    assert quality_notes["说明"].fillna("").str.contains("共同点位").any()

    summary = pd.read_excel(output, sheet_name="汇总")
    assert {
        "总样本数",
        "参与拟合样本数",
        "原始方程Bias",
        "原始方程MaxError",
        "简化方程Bias",
        "简化方程MaxError",
        "RMSE变化量",
        "RMSE相对变化(%)",
        "简化方程MAE",
        "简化方程最大绝对误差",
        "原始与简化预测最大差值",
        "原始与简化预测平均差值",
        "拟合效果评价",
        "拟合效果摘要",
        "综合建议",
        "建议说明",
    }.issubset(summary.columns)

    topn = pd.read_excel(output, sheet_name="误差TopN")
    assert {
        "分析仪",
        "气体",
        "排序维度",
        "rank",
        "点位行号",
        "点位相位",
        "点位标签",
        "abs_error_orig",
        "abs_error_simple",
        "abs_pred_diff",
    }.issubset(topn.columns)
    assert set(topn["排序维度"]) == {"原始误差", "简化误差", "预测差值"}

    plan = pd.read_excel(output, sheet_name="download_plan")
    assert {"Analyzer", "Gas", "PrimaryCommand", "SecondaryCommand"}.issubset(plan.columns)
    assert set(plan["Gas"]) == {"CO2", "H2O"}
    assert plan.loc[plan["Gas"] == "CO2", "PrimaryCommand"].item().startswith("SENCO1,YGAS,FFF,")
    assert plan.loc[plan["Gas"] == "CO2", "SecondaryCommand"].item().startswith("SENCO3,YGAS,FFF,")
    assert plan.loc[plan["Gas"] == "H2O", "PrimaryCommand"].item().startswith("SENCO2,YGAS,FFF,")
    assert plan.loc[plan["Gas"] == "H2O", "SecondaryCommand"].item().startswith("SENCO4,YGAS,FFF,")
    assert "e+" not in plan.loc[plan["Gas"] == "CO2", "PrimaryCommand"].item()
    assert plan.loc[plan["Gas"] == "CO2", "PrimaryCommand"].item().endswith(",0.00000e00,0.00000e00")

    device_eval = tmp_path / "设备评估.xlsx"
    assert device_eval.exists()
    device_book = pd.ExcelFile(device_eval)
    assert device_book.sheet_names == ["汇总", "分通道明细", "说明"]
    eval_frame = pd.read_excel(device_eval, sheet_name="汇总", dtype={"ID": str})
    assert list(eval_frame.columns) == ["设备", "ID", "RISE ppm", "Bias ppm", "R²质量", "结论"]
    eval_detail = pd.read_excel(device_eval, sheet_name="分通道明细")
    assert {"设备", "气路RISE ppm", "水路RISE ppm(等效)", "R²质量"}.issubset(eval_detail.columns)

    workbook = load_workbook(output)
    try:
        simplified_sheet = workbook["简化系数"]
        assert simplified_sheet["D2"].number_format == "0.00000E00"
        assert simplified_sheet["F2"].number_format == "0.00000E00"
        original_sheet = workbook["原始系数"]
        assert original_sheet["D2"].number_format == "0.00000E00"
        assert original_sheet["F2"].number_format == "0.00000E00"
    finally:
        workbook.close()


def test_export_ratio_poly_report_skips_h2o_for_pure_gas_input(tmp_path: Path) -> None:
    samples: list[SamplingResult] = []
    point_index = 1
    for temp, pressure in ((-20.0, 1100.0), (0.0, 900.0), (20.0, 700.0)):
        for target in (0.0, 200.0, 400.0, 600.0, 800.0, 1000.0):
            samples.append(
                _sample(
                    point_index=point_index,
                    route="co2",
                    analyzer_id="ga01",
                    temperature_c=temp,
                    co2_target=target,
                    humidity_pct=None,
                    pressure_hpa=pressure,
                    co2_ratio=1.45 - 0.00018 * target + 0.00035 * temp + 0.00005 * pressure,
                    h2o_ratio=0.18 + 0.0002 * point_index,
                    dew_point_c=-15.0 + point_index * 0.3,
                )
            )
            point_index += 1

    output = export_ratio_poly_report(samples, out_dir=tmp_path, coeff_cfg=_fit_cfg())

    assert output is not None
    plan = pd.read_excel(output, sheet_name="download_plan")
    assert set(plan["Gas"]) == {"CO2"}
    assert plan["PrimaryCommand"].str.startswith("SENCO1,YGAS,FFF,").all()
    assert plan["SecondaryCommand"].str.startswith("SENCO3,YGAS,FFF,").all()

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        summary_sheet = workbook[workbook.sheetnames[4]]
        summary_gases = {
            row[1]
            for row in summary_sheet.iter_rows(min_row=2, values_only=True)
            if row and row[1] is not None
        }
        assert summary_gases == {"CO2"}

        simplified_sheet = workbook[workbook.sheetnames[5]]
        simplified_gases = {
            row[1]
            for row in simplified_sheet.iter_rows(min_row=2, values_only=True)
            if row and row[1] is not None
        }
        assert simplified_gases == {"CO2"}
    finally:
        workbook.close()
    return

    summary = pd.read_excel(output, sheet_name="姹囨€?")
    assert set(summary["姘斾綋"]) == {"CO2"}

    simplified = pd.read_excel(output, sheet_name="绠€鍖栫郴鏁?")
    assert set(simplified["姘斾綋"]) == {"CO2"}


def test_format_senco_float_matches_manual_style() -> None:
    assert _format_senco_float(65916.6) == "6.59166e04"
    assert _format_senco_float(-106614.0) == "-1.06614e05"
    assert _format_senco_float(1.0) == "1.00000e00"
    assert _format_senco_float(-0.000123456) == "-1.23456e-04"
    assert _format_senco_float(0.0) == "0.00000e00"
