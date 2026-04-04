import csv
import math
from pathlib import Path

import gas_calibrator.logging_utils as logging_utils_module
import pytest
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from gas_calibrator.logging_utils import RunLogger, _dewpoint_to_h2o_mmol_per_mol


def test_run_logger_creates_run_dir_and_io_csv(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_sample({"a": 1, "b": "x"})
    logger.log_point({"point": 1, "mean": 12.3})
    point_csv = logger.log_point_samples(1, [{"k": 1, "v": "a"}, {"k": 2, "v": "b"}])
    point_h2o_csv = logger.log_point_samples(1, [{"k": 3, "v": "c"}], phase="h2o")
    logger.log_io(
        port="COM1",
        device="test_dev",
        direction="TX",
        command="CMD,1",
        response="",
        error="",
    )
    logger.close()

    assert logger.run_dir.exists()
    assert logger.samples_path.exists()
    assert logger.points_path.exists()
    assert logger.points_readable_path.exists()
    assert logger.points_readable_book_path.exists()
    assert logger.io_path.exists()
    assert point_csv.exists()
    assert point_h2o_csv.exists()

    with logger.io_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    tx_rows = [row for row in rows if row["direction"] == "TX"]
    assert len(tx_rows) == 1
    row = tx_rows[0]
    assert row["port"] == "COM1"
    assert row["device"] == "test_dev"
    assert row["direction"] == "TX"
    assert row["command"] == "CMD,1"

    with point_csv.open("r", encoding="utf-8", newline="") as f:
        point_rows = list(csv.DictReader(f))
    assert len(point_rows) == 2
    assert point_rows[0]["k"] == "1"
    assert point_h2o_csv.name == "point_0001_h2o_samples.csv"

    with logger.samples_path.open("r", encoding="utf-8", newline="") as f:
        header = next(csv.reader(f))
    assert header[0] == "a"

    with logger.samples_path.open("r", encoding="utf-8", newline="") as f:
        sample_rows = list(csv.DictReader(f))
    assert len(sample_rows) == 1
    assert sample_rows[0]["a"] == "1"


def test_run_logger_prunes_empty_core_exports_on_close(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_io(port="LOG", device="test", direction="EVENT", command="noop", response="ok")
    logger.close()

    assert not logger.samples_path.exists()
    assert not logger.points_path.exists()
    assert not logger.points_readable_path.exists()
    assert not logger.analyzer_summary_csv_path.exists()
    assert logger.io_path.exists()


def test_run_logger_saves_workbooks_atomically(monkeypatch, tmp_path: Path) -> None:
    replace_calls = []
    real_replace = logging_utils_module.os.replace

    def _spy_replace(src, dst):
        replace_calls.append((Path(src), Path(dst)))
        return real_replace(src, dst)

    monkeypatch.setattr(logging_utils_module.os, "replace", _spy_replace)

    logger = RunLogger(tmp_path)
    logger.log_point({"point": 1, "mean": 12.3})
    logger.log_analyzer_workbook(
        [
            {
                "point_title": "20°C环境，400ppm，1000hPa",
                "sample_ts": "2026-03-14T10:00:00.000",
                "point_phase": "co2",
                "point_row": 1,
                "co2_ppm_target": 400.0,
                "temp_chamber_c": 20.0,
                "pressure_target_hpa": 1000.0,
                "pressure_gauge_hpa": 1000.0,
                "ga01_id": "001",
                "ga01_co2_ppm": 401.0,
                "ga01_h2o_mmol": 18.4,
                "ga01_co2_ratio_f": 0.1230,
                "ga01_h2o_ratio_f": 0.4560,
                "ga01_ref_signal": 111.0,
                "ga01_co2_signal": 222.0,
                "ga01_h2o_signal": 333.0,
                "ga01_chamber_temp_c": 19.91,
                "ga01_case_temp_c": 20.55,
                "ga01_pressure_kpa": 101.20,
            }
        ],
        analyzer_labels=["ga01"],
        phase="co2",
    )
    logger.close()

    replaced_targets = {dst.name for _src, dst in replace_calls}
    assert logger.points_readable_book_path.name in replaced_targets
    assert logger.co2_analyzer_book_path.name in replaced_targets
    assert logger.analyzer_summary_book_path.name in replaced_targets


def test_run_logger_log_sample_expands_header_without_dropping_rows(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_sample({"sample_ts": "2026-03-25T10:00:00.000", "co2_ppm": 400.0})
    logger.log_sample(
        {
            "sample_ts": "2026-03-25T10:00:01.000",
            "co2_ppm": 401.0,
            "dew_pressure_hpa": 999.8,
        }
    )
    logger.close()

    with logger.samples_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 2
    assert rows[0]["采样时间"] == "2026-03-25T10:00:00.000"
    assert rows[1]["采样时间"] == "2026-03-25T10:00:01.000"
    assert rows[0]["封压前露点压力hPa"] == ""
    assert rows[1]["封压前露点压力hPa"] == "999.8"


def test_run_logger_log_sample_recovers_after_first_rewrite_failure(monkeypatch, tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    real_save = logging_utils_module._save_csv_atomic
    failed = {"done": False}

    def _fail_once(path, fieldnames, rows):
        if Path(path) == logger.samples_path and not failed["done"]:
            failed["done"] = True
            raise RuntimeError("sample-rewrite-boom")
        return real_save(path, fieldnames, rows)

    monkeypatch.setattr(logging_utils_module, "_save_csv_atomic", _fail_once)

    with pytest.raises(RuntimeError, match="sample-rewrite-boom"):
        logger.log_sample({"sample_ts": "2026-03-25T10:00:00.000", "co2_ppm": 400.0})
    logger.log_sample(
        {
            "sample_ts": "2026-03-25T10:00:01.000",
            "co2_ppm": 401.0,
            "dew_pressure_hpa": 999.8,
        }
    )
    logger.close()

    with logger.samples_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 2
    assert rows[0]["采样时间"] == "2026-03-25T10:00:00.000"
    assert rows[1]["封压前露点压力hPa"] == "999.8"


def test_run_logger_log_point_expands_raw_points_header_without_dropping_rows(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_point(
        {
            "point_title": "点1",
            "point_row": 1,
            "point_phase": "co2",
            "controller_pressure_mean": 1000.0,
        }
    )
    logger.log_point(
        {
            "point_title": "点2",
            "point_row": 2,
            "point_phase": "h2o",
            "controller_pressure_mean": 1100.0,
            "env_chamber_temp_c_mean": 20.1,
        }
    )
    logger.close()

    with logger.points_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 2
    assert rows[0]["点位标题"] == "点1"
    assert rows[1]["点位标题"] == "点2"
    assert rows[0]["温度箱环境温度C_平均值"] == ""
    assert rows[1]["温度箱环境温度C_平均值"] == "20.1"


def test_run_logger_log_point_recovers_after_first_rewrite_failure(monkeypatch, tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    real_save = logging_utils_module._save_csv_atomic
    failed = {"done": False}

    def _fail_once(path, fieldnames, rows):
        if Path(path) == logger.points_path and not failed["done"]:
            failed["done"] = True
            raise RuntimeError("point-rewrite-boom")
        return real_save(path, fieldnames, rows)

    monkeypatch.setattr(logging_utils_module, "_save_csv_atomic", _fail_once)

    with pytest.raises(RuntimeError, match="point-rewrite-boom"):
        logger.log_point(
            {
                "point_title": "点1",
                "point_row": 1,
                "point_phase": "co2",
                "controller_pressure_mean": 1000.0,
            }
        )
    logger.log_point(
        {
            "point_title": "点2",
            "point_row": 2,
            "point_phase": "h2o",
            "controller_pressure_mean": 1100.0,
            "env_chamber_temp_c_mean": 20.1,
        }
    )
    logger.close()

    with logger.points_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 2
    assert rows[0]["点位标题"] == "点1"
    assert rows[1]["温度箱环境温度C_平均值"] == "20.1"


def test_readable_points_csv_recovers_after_first_rewrite_failure(monkeypatch, tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    real_save = logging_utils_module._save_csv_atomic
    failed = {"done": False}

    def _fail_once(path, fieldnames, rows):
        if Path(path) == logger.points_readable_path and not failed["done"]:
            failed["done"] = True
            raise RuntimeError("readable-rewrite-boom")
        return real_save(path, fieldnames, rows)

    monkeypatch.setattr(logging_utils_module, "_save_csv_atomic", _fail_once)

    with pytest.raises(RuntimeError, match="readable-rewrite-boom"):
        logger._append_readable_point_csv({"点位标题": "点1", "流程阶段": "气路"})
    logger._append_readable_point_csv({"点位标题": "点2", "流程阶段": "水路", "目标压力hPa": 1100.0})
    logger.close()

    with logger.points_readable_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 2
    assert rows[0]["点位标题"] == "点1"
    assert rows[1]["目标压力hPa"] == "1100.0"


def test_readable_points_include_analyzer_integrity_fields(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_point(
        {
            "point_title": "点1",
            "point_row": 1,
            "point_phase": "co2",
            "point_tag": "demo",
            "temp_chamber_c": 20.0,
            "co2_ppm_target": 400.0,
            "pressure_target_hpa": 1000.0,
            "analyzer_expected_count": 8,
            "analyzer_with_frame_count": 4,
            "analyzer_usable_count": 3,
            "analyzer_coverage_text": "3/8",
            "analyzer_integrity": "部分缺失且含异常帧",
            "analyzer_missing_labels": "GA03,GA04,GA05,GA06",
            "analyzer_unusable_labels": "GA08",
            "co2_mean": 401.0,
            "pressure_mean": 1000.0,
        }
    )
    logger.close()

    with logger.points_readable_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 1
    row = rows[0]
    assert row["分析仪覆盖率"] == "3/8"
    assert row["分析仪数据完整性"] == "部分缺失且含异常帧"
    assert row["缺失分析仪"] == "GA03,GA04,GA05,GA06"
    assert row["异常帧分析仪"] == "GA08"


def test_run_logger_writes_readable_points_csv_with_means_only(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_point(
        {
            "point_title": "20°C环境，二氧化碳400ppm，气压1000hPa",
            "point_row": 1,
            "point_phase": "co2",
            "point_tag": "co2_groupa_400ppm_1000hpa",
            "temp_chamber_c": 20.0,
            "co2_ppm_target": 400.0,
            "hgen_temp_c": 20.0,
            "hgen_rh_pct": 30.0,
            "pressure_target_hpa": 1000.0,
            "co2_mean": 401.0,
            "co2_std": 0.5,
            "pressure_mean": 1000.1,
            "pressure_gauge_hpa_mean": 1000.2,
            "dewpoint_c_mean": 1.2,
            "hgen_Td_mean": 0.8,
            "ga01_co2_ppm_mean": 401.0,
        }
    )
    logger.close()

    with logger.points_readable_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 1
    row = rows[0]
    assert row["点位标题"] == "20°C环境，二氧化碳400ppm，气压1000hPa"
    assert row["压力控制器压力hPa_平均值"] == "1000.1"
    assert row["数字压力计压力hPa_平均值"] == "1000.2"
    assert row["露点仪露点C_平均值"] == "1.2"
    assert row["湿度发生器_露点(℃)_平均值"] == "0.8"
    assert "二氧化碳平均值(主分析仪或首台可用)" not in row
    assert "分析仪均值语义" not in row
    assert "二氧化碳标准差" not in row
    assert "气体分析仪1_二氧化碳浓度ppm_平均值" not in row
    assert "二氧化碳平均值" not in row

    with logger.points_readable_path.open("r", encoding="utf-8", newline="") as f:
        header = next(csv.reader(f))
    assert header[:9] == [
        "点位标题",
        "校准点行号",
        "流程阶段",
        "点位标签",
        "温箱目标温度C",
        "目标二氧化碳浓度ppm",
        "湿度发生器_目标温度(℃)",
        "湿度发生器_目标湿度(%RH)",
        "目标压力hPa",
    ]
    assert header[9:13] == [
        "压力控制器压力hPa_平均值",
        "数字压力计压力hPa_平均值",
        "露点仪露点C_平均值",
        "湿度发生器_露点(℃)_平均值",
    ]

    wb = load_workbook(logger.points_readable_book_path)
    try:
        ws = wb["点位总览"]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref == ws.dimensions
        assert ws["A1"].font.bold is True
        book_header = [ws.cell(row=1, column=i).value for i in range(1, len(header) + 1)]
        assert book_header == header
        assert ws["A2"].value == "20°C环境，二氧化碳400ppm，气压1000hPa"
        assert ws["A2"].fill.fgColor.rgb in {"00FFF2CC", "FFF2CC"}
    finally:
        wb.close()


def test_readable_points_workbook_highlights_large_deviation_cells(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_point(
        {
            "point_title": "20°C环境，二氧化碳400ppm，气压1000hPa",
            "point_row": 1,
            "point_phase": "co2",
            "point_tag": "co2_groupa_400ppm_1000hpa",
            "temp_chamber_c": 20.0,
            "co2_ppm_target": 400.0,
            "hgen_temp_c": 20.0,
            "hgen_rh_pct": 30.0,
            "pressure_target_hpa": 1000.0,
            "co2_mean": 460.0,
            "pressure_mean": 1008.0,
            "pressure_gauge_hpa_mean": 1007.0,
            "chamber_temp_c_mean": 20.8,
            "hgen_Tc_mean": 20.7,
            "hgen_Uw_mean": 34.5,
        }
    )
    logger.close()

    wb = load_workbook(logger.points_readable_book_path)
    try:
        ws = wb["点位总览"]
        header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        label_to_col = {name: idx + 1 for idx, name in enumerate(header)}
        red_fills = {"00F4CCCC", "FFF4CCCC"}
        yellow_fills = {"00FFF2CC", "FFF2CC"}

        assert "二氧化碳平均值(主分析仪或首台可用)" not in label_to_col
        assert ws.cell(row=2, column=label_to_col["压力控制器压力hPa_平均值"]).fill.fgColor.rgb in red_fills
        assert ws.cell(row=2, column=label_to_col["数字压力计压力hPa_平均值"]).fill.fgColor.rgb in red_fills
        assert ws.cell(row=2, column=label_to_col["温度箱环境温度C_平均值"]).fill.fgColor.rgb in red_fills
        assert ws.cell(row=2, column=label_to_col["湿度发生器_当前温度(℃)_平均值"]).fill.fgColor.rgb in red_fills
        assert ws.cell(row=2, column=label_to_col["湿度发生器_当前湿度(%RH)_平均值"]).fill.fgColor.rgb in red_fills
        assert ws.cell(row=2, column=label_to_col["点位标题"]).fill.fgColor.rgb in yellow_fills
    finally:
        wb.close()


def test_readable_points_workbook_expands_header_without_aborting(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_point(
        {
            "point_title": "20°C环境，二氧化碳400ppm，气压1000hPa",
            "point_row": 1,
            "point_phase": "co2",
            "point_tag": "co2_groupa_400ppm_1000hpa",
            "temp_chamber_c": 20.0,
            "co2_ppm_target": 400.0,
            "pressure_target_hpa": 1000.0,
            "co2_mean": 401.0,
            "pressure_mean": 1000.1,
        }
    )
    logger.log_point(
        {
            "point_title": "0°C环境，湿度发生器0°C/50%RH，气压1100hPa",
            "point_row": 9,
            "point_phase": "h2o",
            "point_tag": "h2o_0c_50rh_1100hpa",
            "temp_chamber_c": 0.0,
            "hgen_temp_c": 0.0,
            "hgen_rh_pct": 50.0,
            "pressure_target_hpa": 1100.0,
            "pressure_mean": 1100.0,
            "dewpoint_c_mean": -8.5,
            "hgen_Uw_mean": 48.2,
            "chamber_temp_c_mean": 0.1,
        }
    )
    logger.close()

    wb = load_workbook(logger.points_readable_book_path)
    try:
        ws = wb["点位总览"]
        header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        assert "湿度发生器_当前湿度(%RH)_平均值" in header
        assert "温度箱环境温度C_平均值" in header
        header_map = {name: idx + 1 for idx, name in enumerate(header)}
        assert ws.max_row == 3
        assert ws.cell(row=3, column=header_map["点位标题"]).value == "0°C环境，湿度发生器0°C/50%RH，气压1100hPa"
        assert ws.cell(row=3, column=header_map["湿度发生器_当前湿度(%RH)_平均值"]).value == 48.2
        assert ws.cell(row=3, column=header_map["温度箱环境温度C_平均值"]).value == 0.1
    finally:
        wb.close()


def test_build_analyzer_summary_row_aligns_reference_rows_by_default(tmp_path: Path) -> None:
    logger = RunLogger(
        tmp_path,
        cfg={"workflow": {"summary_alignment": {"reference_on_aligned_rows": True}}},
    )
    rows = []
    for index in range(10):
        usable = index < 7
        dewpoint = 0.0 if usable else 20.0
        pressure = 1000.0 if usable else 500.0
        rows.append(
            {
                "point_row": 1,
                "point_phase": "co2",
                "point_title": "demo",
                "pressure_target_hpa": 1000.0,
                "co2_ppm_target": 400.0,
                "ga01_frame_has_data": True,
                "ga01_frame_usable": usable,
                "ga01_co2_ppm": 400.0 + index,
                "ga01_h2o_mmol": 1.0 + index * 0.01,
                "ga01_co2_ratio_f": 1.0 + index * 0.001,
                "ga01_h2o_ratio_f": 0.2 + index * 0.001,
                "ga01_pressure_kpa": 101.3,
                "dewpoint_c": dewpoint,
                "dew_pressure_hpa": pressure,
                "pressure_gauge_hpa": pressure,
            }
        )

    summary = logger._build_analyzer_summary_row(rows, label="ga01", num=1)
    logger.close()

    assert summary["ValidFrames"] == 7
    assert summary["TotalFrames"] == 10
    assert summary["Dew"] == 0.0
    assert summary["P"] == 1000.0
    assert math.isclose(
        float(summary["ppm_H2O_Dew"]),
        float(_dewpoint_to_h2o_mmol_per_mol(0.0, 1000.0)),
        rel_tol=0.0,
        abs_tol=1e-6,
    )


def test_build_analyzer_summary_row_exports_h2o_reference_semantics_fields(tmp_path: Path) -> None:
    logger = RunLogger(
        tmp_path,
        cfg={"workflow": {"summary_alignment": {"reference_on_aligned_rows": True}}},
    )
    rows = [
        {
            "point_row": 1,
            "point_phase": "h2o",
            "point_title": "demo",
            "pressure_target_hpa": 1000.0,
            "co2_ppm_target": 0.0,
            "ga01_frame_has_data": True,
            "ga01_frame_usable": True,
            "ga01_co2_ppm": 401.0,
            "ga01_h2o_mmol": 1.1,
            "ga01_co2_ratio_f": 1.01,
            "ga01_h2o_ratio_f": 0.21,
            "ga01_pressure_kpa": 101.3,
            "dewpoint_c": 1.0,
            "dewpoint_sample_ts": "2026-03-25T10:00:00.000",
            "dew_pressure_hpa": 999.0,
            "pressure_gauge_hpa": 1002.0,
        }
    ]

    summary = logger._build_analyzer_summary_row(rows, label="ga01", num=1)
    logger.close()

    assert summary["DewSampleTs"] == "2026-03-25T10:00:00.000"
    assert summary["DewPressurePreseal"] == 999.0
    assert summary["PSample"] == 1002.0
    assert summary["PpmH2oDewPressureSource"] == "dew_pressure_hpa(preseal)"
    assert math.isclose(
        float(summary["ppm_H2O_Dew"]),
        float(_dewpoint_to_h2o_mmol_per_mol(1.0, 999.0)),
        rel_tol=0.0,
        abs_tol=1e-6,
    )


def test_analyzer_summary_export_includes_h2o_reference_semantics_columns(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_analyzer_workbook(
        [
            {
                "point_title": "H2O demo",
                "sample_ts": "2026-03-25T10:00:00.000",
                "point_phase": "h2o",
                "point_row": 1,
                "co2_ppm_target": 0.0,
                "temp_chamber_c": 20.0,
                "pressure_target_hpa": 1000.0,
                "pressure_gauge_hpa": 1002.0,
                "dewpoint_c": 1.0,
                "dewpoint_sample_ts": "2026-03-25T10:00:00.000",
                "dew_pressure_hpa": 999.0,
                "ga01_id": "001",
                "ga01_co2_ppm": 401.0,
                "ga01_h2o_mmol": 1.1,
                "ga01_co2_ratio_f": 1.01,
                "ga01_h2o_ratio_f": 0.21,
                "ga01_ref_signal": 111.0,
                "ga01_co2_signal": 222.0,
                "ga01_h2o_signal": 333.0,
                "ga01_chamber_temp_c": 19.91,
                "ga01_case_temp_c": 20.55,
                "ga01_pressure_kpa": 101.20,
                "ga01_frame_has_data": True,
                "ga01_frame_usable": True,
            }
        ],
        analyzer_labels=["ga01"],
        phase="h2o",
    )
    logger.close()

    with logger.analyzer_summary_csv_path.open("r", encoding="utf-8", newline="") as f:
        header = next(csv.reader(f))

    assert "DewSampleTs" in header
    assert "DewPressurePreseal" in header
    assert "PSample" in header
    assert "PpmH2oDewPressureSource" in header


def test_build_analyzer_summary_row_missing_new_cfg_still_uses_safe_default(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    rows = [
        {
            "point_row": 1,
            "point_phase": "co2",
            "point_title": "demo",
            "pressure_target_hpa": 1000.0,
            "co2_ppm_target": 400.0,
            "ga01_frame_has_data": True,
            "ga01_frame_usable": True,
            "ga01_co2_ppm": 401.0,
            "ga01_h2o_mmol": 1.1,
            "ga01_co2_ratio_f": 1.01,
            "ga01_h2o_ratio_f": 0.21,
            "ga01_pressure_kpa": 101.3,
            "dewpoint_c": 1.0,
            "dew_pressure_hpa": 999.0,
            "pressure_gauge_hpa": 999.0,
        }
    ]

    summary = logger._build_analyzer_summary_row(rows, label="ga01", num=1)
    logger.close()

    assert summary["Dew"] == 1.0
    assert summary["P"] == 999.0
    assert summary["FrameStatus"] == "全部可用"


def test_points_exports_exclude_fleet_stats_by_default(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_point(
        {
            "point_title": "点1",
            "point_row": 1,
            "point_phase": "co2",
            "pressure_target_hpa": 1000.0,
            "controller_pressure_mean": 1000.0,
            "co2_fleet_mean": 401.0,
            "co2_fleet_std": 1.2,
            "h2o_fleet_mean": 1.1,
            "h2o_fleet_std": 0.1,
        }
    )
    logger.close()

    with logger.points_path.open("r", encoding="utf-8", newline="") as f:
        header = next(csv.reader(f))
    assert "二氧化碳全分析仪平均值" not in header
    assert "二氧化碳全分析仪标准差" not in header

    with logger.points_readable_path.open("r", encoding="utf-8", newline="") as f:
        readable_header = next(csv.reader(f))
    assert "二氧化碳全分析仪平均值" not in readable_header
    assert "二氧化碳全分析仪标准差" not in readable_header

    wb = load_workbook(logger.points_readable_book_path)
    try:
        ws = wb["点位总览"]
        book_header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        assert "二氧化碳全分析仪平均值" not in book_header
        assert "二氧化碳全分析仪标准差" not in book_header
    finally:
        wb.close()


def test_points_exports_include_fleet_stats_when_enabled(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path, cfg={"workflow": {"reporting": {"include_fleet_stats": True}}})
    logger.log_point(
        {
            "point_title": "点1",
            "point_row": 1,
            "point_phase": "co2",
            "pressure_target_hpa": 1000.0,
            "controller_pressure_mean": 1000.0,
            "co2_fleet_mean": 401.0,
            "co2_fleet_std": 1.2,
            "h2o_fleet_mean": 1.1,
            "h2o_fleet_std": 0.1,
        }
    )
    logger.close()

    with logger.points_path.open("r", encoding="utf-8", newline="") as f:
        header = next(csv.reader(f))
    assert "二氧化碳全分析仪平均值" in header
    assert "二氧化碳全分析仪标准差" in header

    with logger.points_readable_path.open("r", encoding="utf-8", newline="") as f:
        readable_header = next(csv.reader(f))
    with logger.points_readable_path.open("r", encoding="utf-8", newline="") as f:
        readable_rows = list(csv.DictReader(f))
    assert "二氧化碳全分析仪平均值" in readable_header
    assert "二氧化碳全分析仪标准差" in readable_header
    assert readable_rows[0]["二氧化碳全分析仪平均值"] == "401.0"
    assert readable_rows[0]["二氧化碳全分析仪标准差"] == "1.2"


def test_points_export_logs_role_and_alias_semantics_once(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_point(
        {
            "point_title": "点1",
            "point_row": 1,
            "point_phase": "co2",
            "co2_mean": 401.0,
            "pressure_target_hpa": 1000.0,
        }
    )
    logger.log_point(
        {
            "point_title": "点2",
            "point_row": 2,
            "point_phase": "co2",
            "co2_mean": 402.0,
            "pressure_target_hpa": 1000.0,
        }
    )
    logger.close()

    with logger.io_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    commands = [row["command"] for row in rows]
    assert commands.count("points-export-role") == 1
    assert commands.count("points-mean-alias-semantics") == 1


def test_log_analyzer_workbook_preserves_summary_when_workbook_save_fails(monkeypatch, tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    rows = [
        {
            "point_title": "20°C环境，400ppm，1000hPa",
            "sample_ts": "2026-03-14T10:00:00.000",
            "point_phase": "co2",
            "point_row": 1,
            "co2_ppm_target": 400.0,
            "temp_chamber_c": 20.0,
            "pressure_target_hpa": 1000.0,
            "pressure_gauge_hpa": 1000.0,
            "ga01_id": "001",
            "ga01_co2_ppm": 401.0,
            "ga01_h2o_mmol": 18.4,
            "ga01_co2_ratio_f": 0.1230,
            "ga01_h2o_ratio_f": 0.4560,
            "ga01_ref_signal": 111.0,
            "ga01_co2_signal": 222.0,
            "ga01_h2o_signal": 333.0,
            "ga01_chamber_temp_c": 19.91,
            "ga01_case_temp_c": 20.55,
            "ga01_pressure_kpa": 101.20,
            "ga01_frame_has_data": True,
            "ga01_frame_usable": True,
        }
    ]

    real_save = logging_utils_module._save_workbook_atomic

    def _fail_only_analyzer_book(wb, path):
        if Path(path) == logger.co2_analyzer_book_path:
            raise RuntimeError("workbook-save-boom")
        return real_save(wb, path)

    monkeypatch.setattr(logging_utils_module, "_save_workbook_atomic", _fail_only_analyzer_book)

    with pytest.raises(RuntimeError, match="workbook-save-boom"):
        logger.log_analyzer_workbook(rows, analyzer_labels=["ga01"], phase="co2", write_summary=True)
    logger.close()

    assert logger.analyzer_summary_csv_path.exists()
    with logger.analyzer_summary_csv_path.open("r", encoding="utf-8", newline="") as f:
        summary_rows = list(csv.DictReader(f))
    assert len(summary_rows) == 1
    assert summary_rows[0]["Analyzer"] == "GA01"


def test_log_analyzer_summary_csv_recovers_after_first_write_failure(monkeypatch, tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)

    def _rows_for(label: str, analyzer_id: str, co2_ppm: float) -> list[dict[str, object]]:
        prefix = label.lower()
        return [
            {
                "point_title": "20°C环境，400ppm，1000hPa",
                "sample_ts": "2026-03-14T10:00:00.000",
                "point_phase": "co2",
                "point_row": 1,
                "co2_ppm_target": 400.0,
                "temp_chamber_c": 20.0,
                "pressure_target_hpa": 1000.0,
                "pressure_gauge_hpa": 1000.0,
                f"{prefix}_id": analyzer_id,
                f"{prefix}_co2_ppm": co2_ppm,
                f"{prefix}_h2o_mmol": 18.4,
                f"{prefix}_co2_ratio_f": 0.1230,
                f"{prefix}_h2o_ratio_f": 0.4560,
                f"{prefix}_pressure_kpa": 101.20,
                f"{prefix}_frame_has_data": True,
                f"{prefix}_frame_usable": True,
            }
        ]

    real_save = logging_utils_module._save_csv_atomic
    failed = {"done": False}

    def _fail_once(path, fieldnames, rows):
        if Path(path) == logger.analyzer_summary_csv_path and not failed["done"]:
            failed["done"] = True
            raise RuntimeError("summary-csv-boom")
        return real_save(path, fieldnames, rows)

    monkeypatch.setattr(logging_utils_module, "_save_csv_atomic", _fail_once)

    with pytest.raises(RuntimeError, match="analyzer summary partial failures"):
        logger.log_analyzer_summary(_rows_for("ga01", "001", 401.0), analyzer_labels=["ga01"])

    logger.log_analyzer_summary(_rows_for("ga02", "002", 402.0), analyzer_labels=["ga02"])
    logger.close()

    with logger.analyzer_summary_csv_path.open("r", encoding="utf-8", newline="") as f:
        summary_rows = list(csv.DictReader(f))

    assert [row["Analyzer"] for row in summary_rows] == ["GA01", "GA02"]


def test_log_analyzer_summary_continues_after_one_analyzer_workbook_failure(monkeypatch, tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    rows = [
        {
            "point_title": "20°C环境，400ppm，1000hPa",
            "sample_ts": "2026-03-14T10:00:00.000",
            "point_phase": "co2",
            "point_row": 1,
            "co2_ppm_target": 400.0,
            "temp_chamber_c": 20.0,
            "pressure_target_hpa": 1000.0,
            "pressure_gauge_hpa": 1000.0,
            "ga01_id": "001",
            "ga01_co2_ppm": 401.0,
            "ga01_h2o_mmol": 18.4,
            "ga01_co2_ratio_f": 0.1230,
            "ga01_h2o_ratio_f": 0.4560,
            "ga01_pressure_kpa": 101.20,
            "ga01_frame_has_data": True,
            "ga01_frame_usable": True,
            "ga02_id": "002",
            "ga02_co2_ppm": 402.0,
            "ga02_h2o_mmol": 18.5,
            "ga02_co2_ratio_f": 0.1240,
            "ga02_h2o_ratio_f": 0.4570,
            "ga02_pressure_kpa": 101.10,
            "ga02_frame_has_data": True,
            "ga02_frame_usable": True,
            "ga03_id": "003",
            "ga03_co2_ppm": 403.0,
            "ga03_h2o_mmol": 18.6,
            "ga03_co2_ratio_f": 0.1250,
            "ga03_h2o_ratio_f": 0.4580,
            "ga03_pressure_kpa": 101.00,
            "ga03_frame_has_data": True,
            "ga03_frame_usable": True,
        }
    ]

    real_append = logger._append_analyzer_summary_workbook_row

    def _fail_only_second(label, row):
        if str(label).lower() == "ga02":
            raise RuntimeError("summary-workbook-boom")
        return real_append(label, row)

    monkeypatch.setattr(logger, "_append_analyzer_summary_workbook_row", _fail_only_second)

    logger.log_analyzer_summary(rows, analyzer_labels=["ga01", "ga02", "ga03"])
    logger.close()

    with logger.analyzer_summary_csv_path.open("r", encoding="utf-8", newline="") as f:
        summary_rows = list(csv.DictReader(f))

    analyzers = [row["Analyzer"] for row in summary_rows]
    assert analyzers == ["GA01", "GA02", "GA03"]


def test_log_analyzer_summary_csv_append_failure_recovers_with_later_analyzer(monkeypatch, tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)

    def _rows_for(label: str, analyzer_id: str, co2_ppm: float) -> list[dict[str, object]]:
        prefix = label.lower()
        return [
            {
                "point_title": "20°C环境，400ppm，1000hPa",
                "sample_ts": "2026-03-14T10:00:00.000",
                "point_phase": "co2",
                "point_row": 1,
                "co2_ppm_target": 400.0,
                "temp_chamber_c": 20.0,
                "pressure_target_hpa": 1000.0,
                "pressure_gauge_hpa": 1000.0,
                f"{prefix}_id": analyzer_id,
                f"{prefix}_co2_ppm": co2_ppm,
                f"{prefix}_h2o_mmol": 18.4,
                f"{prefix}_co2_ratio_f": 0.1230,
                f"{prefix}_h2o_ratio_f": 0.4560,
                f"{prefix}_pressure_kpa": 101.20,
                f"{prefix}_frame_has_data": True,
                f"{prefix}_frame_usable": True,
            }
        ]

    logger.log_analyzer_summary(_rows_for("ga01", "001", 401.0), analyzer_labels=["ga01"])

    real_writerow = logger._analyzer_summary_writer.writerow
    failed = {"done": False}

    def _fail_once(row):
        if row.get("Analyzer") == "GA02" and not failed["done"]:
            failed["done"] = True
            raise RuntimeError("summary-append-boom")
        return real_writerow(row)

    monkeypatch.setattr(logger._analyzer_summary_writer, "writerow", _fail_once)

    rows = _rows_for("ga02", "002", 402.0) + _rows_for("ga03", "003", 403.0)
    logger.log_analyzer_summary(rows, analyzer_labels=["ga02", "ga03"])
    logger.close()

    with logger.analyzer_summary_csv_path.open("r", encoding="utf-8", newline="") as f:
        summary_rows = list(csv.DictReader(f))

    assert [row["Analyzer"] for row in summary_rows] == ["GA01", "GA02", "GA03"]


def test_append_readable_point_workbook_closes_workbook_when_save_fails(monkeypatch, tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    wb = logging_utils_module.Workbook()
    real_close = wb.close
    closed = {"called": False}

    def _close():
        closed["called"] = True
        return real_close()

    monkeypatch.setattr(wb, "close", _close)
    monkeypatch.setattr(logging_utils_module, "Workbook", lambda: wb)
    monkeypatch.setattr(
        logging_utils_module,
        "_save_workbook_atomic",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("readable-book-boom")),
    )

    with pytest.raises(RuntimeError, match="readable-book-boom"):
        logger._append_readable_point_workbook({"点位标题": "点1", "流程阶段": "气路"})
    logger.close()

    assert closed["called"] is True


def test_readable_points_csv_expands_header_without_duplicate_rows(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_point(
        {
            "point_title": "20°C环境，二氧化碳400ppm，气压1000hPa",
            "point_row": 1,
            "point_phase": "co2",
            "point_tag": "co2_groupa_400ppm_1000hpa",
            "temp_chamber_c": 20.0,
            "co2_ppm_target": 400.0,
            "pressure_target_hpa": 1000.0,
            "co2_mean": 401.0,
            "pressure_mean": 1000.1,
        }
    )
    logger.log_point(
        {
            "point_title": "0°C环境，湿度发生器0°C/50%RH，气压1100hPa",
            "point_row": 9,
            "point_phase": "h2o",
            "point_tag": "h2o_0c_50rh_1100hpa",
            "temp_chamber_c": 0.0,
            "hgen_temp_c": 0.0,
            "hgen_rh_pct": 50.0,
            "pressure_target_hpa": 1100.0,
            "pressure_mean": 1100.0,
            "dewpoint_c_mean": -8.5,
            "hgen_Uw_mean": 48.2,
            "chamber_temp_c_mean": 0.1,
        }
    )
    logger.close()

    with logger.points_readable_path.open("r", encoding="utf-8", newline="") as f:
        raw_rows = list(csv.reader(f))

    assert len(raw_rows) == 3
    header = raw_rows[0]
    assert "湿度发生器_当前湿度(%RH)_平均值" in header
    assert "温度箱环境温度C_平均值" in header

    with logger.points_readable_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 2
    assert rows[0]["点位标题"] == "20°C环境，二氧化碳400ppm，气压1000hPa"
    assert rows[0]["湿度发生器_当前湿度(%RH)_平均值"] == ""
    assert rows[1]["点位标题"] == "0°C环境，湿度发生器0°C/50%RH，气压1100hPa"
    assert rows[1]["湿度发生器_当前湿度(%RH)_平均值"] == "48.2"
    assert rows[1]["温度箱环境温度C_平均值"] == "0.1"


def test_run_logger_writes_analyzer_workbook_with_shared_sheets(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    rows_1 = [
        {
            "point_title": "20°C环境，湿度发生器20°C/30%RH，气压1100hPa",
            "sample_ts": "2026-03-07T16:00:00.000",
            "point_phase": "h2o",
            "point_row": 21,
            "pressure_hpa": 1100.1,
            "dewpoint_c": 1.2,
            "hgen_Fl": 2.3,
            "hgen_Flux": 2.3,
            "ga01_id": "001",
            "ga01_h2o_ratio_f": 0.801,
            "ga01_co2_ratio_f": 0.002,
            "ga02_id": "002",
            "ga02_h2o_ratio_f": 0.780,
            "ga02_co2_ratio_f": 0.003,
        },
        {
            "point_title": "20°C环境，湿度发生器20°C/30%RH，气压1100hPa",
            "sample_ts": "2026-03-07T16:00:10.000",
            "point_phase": "h2o",
            "point_row": 21,
            "pressure_hpa": 1100.2,
            "dewpoint_c": 1.2,
            "hgen_Fl": 2.4,
            "hgen_Flux": 2.4,
            "ga01_id": "001",
            "ga01_h2o_ratio_f": 0.802,
            "ga01_co2_ratio_f": 0.002,
            "ga02_id": "002",
            "ga02_h2o_ratio_f": 0.781,
            "ga02_co2_ratio_f": 0.003,
        },
    ]
    rows_2 = [
        {
            "point_title": "20°C环境，湿度发生器20°C/30%RH，气压1000hPa",
            "sample_ts": "2026-03-07T16:02:00.000",
            "point_phase": "h2o",
            "point_row": 22,
            "pressure_hpa": 1000.1,
            "dewpoint_c": 1.3,
            "hgen_Fl": 2.5,
            "hgen_Flux": 2.5,
            "ga01_id": "001",
            "ga01_h2o_ratio_f": 0.803,
            "ga02_id": "002",
            "ga02_h2o_ratio_f": 0.782,
        }
    ]

    book_path = logger.log_analyzer_workbook(rows_1, analyzer_labels=["ga01", "ga02"], phase="h2o")
    logger.log_analyzer_workbook(rows_2, analyzer_labels=["ga01", "ga02"], phase="h2o")
    logger.close()

    assert book_path == logger.h2o_analyzer_book_path
    assert book_path.exists()

    wb = load_workbook(book_path)
    try:
        assert wb.sheetnames == ["气体分析仪1_ID001", "气体分析仪2_ID002"]
        ws1 = wb["气体分析仪1_ID001"]
        ws2 = wb["气体分析仪2_ID002"]
        assert ws1["A1"].value == "点位标题"
        assert ws2["A1"].value == "点位标题"
        assert ws1.max_row == 4
        assert ws2.max_row == 4
        assert ws1["A2"].value == "20°C环境，湿度发生器20°C/30%RH，气压1100hPa"
        assert ws1["A4"].value == "20°C环境，湿度发生器20°C/30%RH，气压1000hPa"
        headers_1 = [ws1.cell(row=1, column=i).value for i in range(1, ws1.max_column + 1)]
        headers_2 = [ws2.cell(row=1, column=i).value for i in range(1, ws2.max_column + 1)]
        assert "水比值滤波后" in headers_1
        assert "水比值滤波后" in headers_2
        assert "露点仪露点C" in headers_1
        assert headers_1.count("湿度发生器_流量(L/min)") == 1
        assert ws1.freeze_panes == "A2"
        assert ws1.auto_filter.ref == ws1.dimensions
        assert ws1["A1"].font.bold is True
        assert ws1.column_dimensions["A"].width >= 10
    finally:
        wb.close()


def test_run_logger_writes_co2_workbook_separately(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    rows = [
        {
            "point_title": "20°C环境，二氧化碳200ppm，气压1000hPa",
            "sample_ts": "2026-03-07T16:03:00.000",
            "point_phase": "co2",
            "point_row": 31,
            "pressure_hpa": 1000.0,
            "ga01_id": "001",
            "ga01_co2_ratio_f": 0.123,
        }
    ]

    book_path = logger.log_analyzer_workbook(rows, analyzer_labels=["ga01"], phase="co2")
    logger.close()

    assert book_path == logger.co2_analyzer_book_path
    assert book_path.exists()
    assert not logger.h2o_analyzer_book_path.exists()


def test_run_logger_extends_analyzer_workbook_header_for_late_mode2_extra_fields(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    rows_1 = [
        {
            "point_title": "30°C环境，二氧化碳1000ppm，气压900hPa",
            "sample_ts": "2026-03-21T23:54:57.047",
            "point_phase": "co2",
            "point_tag": "co2_groupa_1000ppm_900hpa",
            "point_row": 47,
            "co2_ppm_target": 1000.0,
            "temp_chamber_c": 30.0,
            "pressure_target_hpa": 900.0,
            "pressure_gauge_hpa": 899.551,
            "ga06_id": "005",
            "ga06_mode": 2,
            "ga06_mode2_field_count": 16,
            "ga06_co2_ppm": 1.242,
            "ga06_h2o_mmol": 0.701,
            "ga06_co2_ratio_f": 1.2413,
            "ga06_h2o_ratio_f": 0.7008,
            "ga06_pressure_kpa": 91.94,
        }
    ]
    rows_2 = [
        {
            "point_title": "30°C环境，二氧化碳1000ppm，气压900hPa",
            "sample_ts": "2026-03-21T23:55:07.047",
            "point_phase": "co2",
            "point_tag": "co2_groupa_1000ppm_900hpa",
            "point_row": 47,
            "co2_ppm_target": 1000.0,
            "temp_chamber_c": 30.0,
            "pressure_target_hpa": 900.0,
            "pressure_gauge_hpa": 899.551,
            "ga06_id": "005",
            "ga06_mode": 2,
            "ga06_mode2_field_count": 20,
            "ga06_co2_ppm": 1.243,
            "ga06_h2o_mmol": 0.702,
            "ga06_co2_ratio_f": 1.2423,
            "ga06_h2o_ratio_f": 0.7018,
            "ga06_pressure_kpa": 91.95,
            "ga06_mode2_extra_01": "A",
            "ga06_mode2_extra_02": "B",
            "ga06_mode2_extra_03": "C",
            "ga06_mode2_extra_04": "D",
        }
    ]
    rows_3 = [
        {
            "point_title": "30°C环境，二氧化碳1000ppm，气压800hPa",
            "sample_ts": "2026-03-21T23:58:53.531",
            "point_phase": "co2",
            "point_tag": "co2_groupa_1000ppm_800hpa",
            "point_row": 50,
            "co2_ppm_target": 1000.0,
            "temp_chamber_c": 30.0,
            "pressure_target_hpa": 800.0,
            "pressure_gauge_hpa": 799.658,
            "ga06_id": "005",
            "ga06_mode": 2,
            "ga06_mode2_field_count": 16,
            "ga06_co2_ppm": 1.259,
            "ga06_h2o_mmol": 0.701,
            "ga06_co2_ratio_f": 1.2581,
            "ga06_h2o_ratio_f": 0.7007,
            "ga06_pressure_kpa": 81.94,
        }
    ]

    logger.log_analyzer_workbook(rows_1, analyzer_labels=["ga06"], phase="co2")
    logger.log_analyzer_workbook(rows_2, analyzer_labels=["ga06"], phase="co2")
    logger.log_analyzer_workbook(rows_3, analyzer_labels=["ga06"], phase="co2")
    logger.close()

    wb = load_workbook(logger.co2_analyzer_book_path)
    try:
        ws = wb["气体分析仪6_ID005"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        extra_cols = [headers.index(f"mode2_extra_{idx:02d}") + 1 for idx in range(1, 5)]
        assert ws.max_row == 4
        assert headers[-4:] == ["mode2_extra_01", "mode2_extra_02", "mode2_extra_03", "mode2_extra_04"]
        assert [ws.cell(row=2, column=col).value for col in extra_cols] == [None, None, None, None]
        assert [ws.cell(row=3, column=col).value for col in extra_cols] == ["A", "B", "C", "D"]
        assert [ws.cell(row=4, column=col).value for col in extra_cols] == [None, None, None, None]
    finally:
        wb.close()


def test_run_logger_preserves_analyzer_workbook_header_when_common_dynamic_key_order_changes(
    tmp_path: Path,
) -> None:
    logger = RunLogger(tmp_path)
    rows_1 = [
        {
            "point_title": "点1",
            "sample_ts": "2026-04-04T23:25:55.597",
            "point_phase": "co2",
            "point_tag": "co2_groupa_400ppm_ambient",
            "point_row": 4,
            "pace_anchor_delta_ms": 304.05,
            "pace_sample_ts": "2026-04-04T23:25:55.293",
            "pressure_hpa": 1004.4226074,
            "pressure_gauge_anchor_delta_ms": 11814.237,
            "pressure_gauge_error": "fast_signal_stale",
            "dewpoint_live_anchor_delta_ms": 111.617,
            "dewpoint_live_sample_ts": "2026-04-04T23:25:55.486",
            "dewpoint_live_c": -38.01,
            "dew_temp_live_c": 35.51,
            "dew_rh_live_pct": 0.28,
            "ga01_id": "015",
            "ga01_co2_ratio_f": 1.2661,
            "pressure_error": "",
        }
    ]
    rows_2 = [
        {
            "point_title": "点2",
            "sample_ts": "2026-04-04T23:32:22.328",
            "point_phase": "co2",
            "point_tag": "co2_groupa_1000ppm_ambient",
            "point_row": 5,
            "pace_anchor_delta_ms": 1024.197,
            "pressure_error": "fast_signal_stale",
            "pressure_gauge_anchor_delta_ms": 12810.108,
            "pressure_gauge_error": "fast_signal_stale",
            "dewpoint_live_anchor_delta_ms": 411.399,
            "dewpoint_live_sample_ts": "2026-04-04T23:32:21.917",
            "dewpoint_live_c": -38.24,
            "dew_temp_live_c": 35.62,
            "dew_rh_live_pct": 0.27,
            "ga01_id": "015",
            "ga01_co2_ratio_f": 1.1425,
        }
    ]

    logger.log_analyzer_workbook(rows_1, analyzer_labels=["ga01"], phase="co2")
    logger.log_analyzer_workbook(rows_2, analyzer_labels=["ga01"], phase="co2")
    logger.close()

    wb = load_workbook(logger.co2_analyzer_book_path)
    try:
        ws = wb["气体分析仪1_ID015"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        pressure_error_col = headers.index("pressure_error") + 1
        pace_sample_ts_col = headers.index("压力控制器采样时间") + 1
        assert ws.max_row == 3
        assert pressure_error_col > pace_sample_ts_col
        assert ws.cell(row=2, column=pressure_error_col).value in (None, "")
        assert ws.cell(row=3, column=pressure_error_col).value == "fast_signal_stale"
    finally:
        wb.close()


def test_run_logger_writes_analyzer_summary_csv_and_workbook(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    expected_ppm_h2o_dew = _dewpoint_to_h2o_mmol_per_mol(1.3, 1000.3)
    rows = [
        {
            "point_title": "20°C环境，湿度发生器20°C/50%RH，气压1000hPa",
            "sample_ts": "2026-03-13T10:00:00.000",
            "point_phase": "h2o",
            "point_row": 31,
            "co2_ppm_target": 400.0,
            "h2o_mmol_target": 18.5,
            "temp_chamber_c": 20.0,
            "hgen_temp_c": 20.0,
            "hgen_rh_pct": 50.0,
            "pressure_target_hpa": 1000.0,
            "chamber_temp_c": 20.10,
            "dewpoint_c": 1.30,
            "dew_pressure_hpa": 1000.3,
            "pressure_gauge_hpa": 1000.2,
            "ga01_id": "001",
            "ga01_co2_ppm": 401.0,
            "ga01_h2o_mmol": 18.4,
            "ga01_co2_ratio_f": 0.1230,
            "ga01_h2o_ratio_f": 0.4560,
            "ga01_ref_signal": 111.0,
            "ga01_co2_signal": 222.0,
            "ga01_h2o_signal": 333.0,
            "ga01_chamber_temp_c": 19.91,
            "ga01_case_temp_c": 20.55,
            "ga01_pressure_kpa": 101.20,
        },
        {
            "point_title": "20°C环境，湿度发生器20°C/50%RH，气压1000hPa",
            "sample_ts": "2026-03-13T10:00:10.000",
            "point_phase": "h2o",
            "point_row": 31,
            "co2_ppm_target": 400.0,
            "h2o_mmol_target": 18.5,
            "temp_chamber_c": 20.0,
            "hgen_temp_c": 20.0,
            "hgen_rh_pct": 50.0,
            "pressure_target_hpa": 1000.0,
            "chamber_temp_c": 20.30,
            "dewpoint_c": 1.30,
            "dew_pressure_hpa": 1000.3,
            "pressure_gauge_hpa": 1000.4,
            "ga01_id": "001",
            "ga01_co2_ppm": 403.0,
            "ga01_h2o_mmol": 18.6,
            "ga01_co2_ratio_f": 0.1270,
            "ga01_h2o_ratio_f": 0.4600,
            "ga01_ref_signal": 113.0,
            "ga01_co2_signal": 224.0,
            "ga01_h2o_signal": 335.0,
            "ga01_chamber_temp_c": 19.95,
            "ga01_case_temp_c": 20.65,
            "ga01_pressure_kpa": 101.40,
        },
    ]

    logger.log_analyzer_workbook(rows, analyzer_labels=["ga01"], phase="h2o")
    logger.close()

    assert logger.analyzer_summary_csv_path.exists()
    assert logger.analyzer_summary_book_path.exists()

    with logger.analyzer_summary_csv_path.open("r", encoding="utf-8", newline="") as f:
        csv_rows = list(csv.DictReader(f))

    assert len(csv_rows) == 1
    csv_row = csv_rows[0]
    assert csv_row["Analyzer"] == "GA01"
    assert csv_row["NUM"] == "1"
    assert csv_row["PointRow"] == "31"
    assert csv_row["PointPhase"] == "水路"
    assert csv_row["PointTitle"] == "20°C环境，湿度发生器20°C/50%RH，气压1000hPa"
    assert csv_row["TempSet"] == "20.0"
    assert csv_row["HgenTempSet"] == "20.0"
    assert csv_row["HgenRhSet"] == "50.0"
    assert csv_row["Temp"] == "20.2"
    assert csv_row["Dew"] == "1.3"
    assert csv_row["AnalyzerCoverage"] == ""
    assert csv_row["PointIntegrity"] == ""
    assert csv_row["ValidFrames"] == "2"
    assert csv_row["TotalFrames"] == "2"
    assert csv_row["FrameStatus"] == "全部可用"
    assert csv_row["DewSampleTs"] == ""
    assert csv_row["DewPressurePreseal"] == "1000.3"
    assert csv_row["P"] == "1000.3"
    assert csv_row["PSample"] == "1000.3"
    assert csv_row["PpmH2oDewPressureSource"] == "dew_pressure_hpa(preseal)"
    assert csv_row["ppm_CO2_Tank"] == "400.0"
    assert csv_row["PressureTarget"] == "1000.0"
    assert math.isclose(float(csv_row["ppm_H2O_Dew"]), float(expected_ppm_h2o_dew), rel_tol=0.0, abs_tol=1e-6)
    assert csv_row["ppm_CO2"] == "402.0"
    assert csv_row["ppm_H2O"] == "18.5"
    assert csv_row["R_CO2"] == "0.125000"
    assert csv_row["R_H2O"] == "0.458000"
    assert csv_row["R_CO2_dev"] == "0.002828"
    assert csv_row["R_H2O_dev"] == "0.002828"
    assert csv_row["Raw_REF"] == "112.0"
    assert csv_row["Raw_CO2"] == "223.0"
    assert csv_row["Raw_H2O"] == "334.0"
    assert csv_row["T1"] == "19.93"
    assert csv_row["T2"] == "20.6"
    assert csv_row["BAR"] == "101.3"

    wb = load_workbook(logger.analyzer_summary_book_path)
    try:
        ws = wb["GA01"]
        header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        assert header == [
            "NUM",
            "PointRow",
            "PointPhase",
            "PointTag",
            "PointTitle",
            "TempSet",
            "HgenTempSet",
            "HgenRhSet",
            "Temp",
            "Dew",
            "DewSampleTs",
            "DewPressurePreseal",
            "P",
            "PSample",
            "PpmH2oDewPressureSource",
            "ppm_CO2_Tank",
            "PressureTarget",
            "AnalyzerCoverage",
            "UsableAnalyzers",
            "ExpectedAnalyzers",
            "PointIntegrity",
            "MissingAnalyzers",
            "UnusableAnalyzers",
            "ValidFrames",
            "TotalFrames",
            "FrameStatus",
            "ppm_H2O_Dew",
            "ppm_CO2",
            "ppm_H2O",
            "R_CO2",
            "R_CO2_dev",
            "R_H2O",
            "R_H2O_dev",
            "Raw_REF",
            "Raw_CO2",
            "Raw_H2O",
            "T1",
            "T2",
            "BAR",
        ]
        assert ws["A2"].value == 1
        assert ws["B2"].value == 31
        assert ws["C2"].value == "水路"
        assert ws["E2"].value == "20°C环境，湿度发生器20°C/50%RH，气压1000hPa"
        assert ws["F2"].value == 20.0
        assert ws["G2"].value == 20.0
        assert ws["H2"].value == 50.0
        assert ws["I2"].value == 20.2
        assert ws["P2"].value == 400.0
        assert ws["T2"].value is None
        assert ws["X2"].value == 2
        assert ws["Y2"].value == 2
        assert ws["Z2"].value == "全部可用"
        assert math.isclose(float(ws["AA2"].value), float(expected_ppm_h2o_dew), rel_tol=0.0, abs_tol=1e-6)
        assert ws["AD2"].value == 0.125
        assert ws["AD2"].number_format == "0.000000"
        assert ws["AE2"].value is not None
        assert ws["AE2"].number_format == "0.000000"
        assert ws["A2"].fill.fgColor.rgb in {"00DDEBF7", "DDEBF7"}
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref == ws.dimensions
    finally:
        wb.close()


def test_dewpoint_to_h2o_mmol_per_mol_uses_pressure_and_negative_dewpoint() -> None:
    warm = _dewpoint_to_h2o_mmol_per_mol(1.3, 1000.3)
    cold = _dewpoint_to_h2o_mmol_per_mol(-10.0, 900.0)

    assert warm is not None
    assert cold is not None
    assert math.isclose(warm, 6.711995, rel_tol=0.0, abs_tol=1e-6)
    assert math.isclose(cold, 2.888299, rel_tol=0.0, abs_tol=1e-6)


def test_analyzer_summary_prefers_digital_thermometer_temperature(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    rows = [
        {
            "point_title": "20°C环境，二氧化碳400ppm，气压1000hPa",
            "point_row": 3,
            "point_phase": "co2",
            "point_tag": "co2_400_1000",
            "co2_ppm_target": 400.0,
            "temp_chamber_c": 20.0,
            "pressure_target_hpa": 1000.0,
            "chamber_temp_c": 20.30,
            "thermometer_temp_c": 20.62,
            "pressure_gauge_hpa": 1000.4,
            "ga01_id": "001",
            "ga01_co2_ppm": 401.0,
            "ga01_h2o_mmol": 18.4,
            "ga01_co2_ratio_f": 0.1230,
            "ga01_h2o_ratio_f": 0.4560,
            "ga01_ref_signal": 111.0,
            "ga01_co2_signal": 222.0,
            "ga01_h2o_signal": 333.0,
            "ga01_chamber_temp_c": 19.91,
            "ga01_case_temp_c": 20.55,
            "ga01_pressure_kpa": 101.20,
        },
        {
            "point_title": "20°C环境，二氧化碳400ppm，气压1000hPa",
            "point_row": 3,
            "point_phase": "co2",
            "point_tag": "co2_400_1000",
            "co2_ppm_target": 400.0,
            "temp_chamber_c": 20.0,
            "pressure_target_hpa": 1000.0,
            "chamber_temp_c": 20.10,
            "thermometer_temp_c": 20.58,
            "pressure_gauge_hpa": 1000.2,
            "ga01_id": "001",
            "ga01_co2_ppm": 402.0,
            "ga01_h2o_mmol": 18.5,
            "ga01_co2_ratio_f": 0.1270,
            "ga01_h2o_ratio_f": 0.4600,
            "ga01_ref_signal": 113.0,
            "ga01_co2_signal": 224.0,
            "ga01_h2o_signal": 335.0,
            "ga01_chamber_temp_c": 19.95,
            "ga01_case_temp_c": 20.65,
            "ga01_pressure_kpa": 101.40,
        },
    ]

    logger.log_analyzer_workbook(rows, analyzer_labels=["ga01"], phase="co2")
    logger.close()

    with logger.analyzer_summary_csv_path.open("r", encoding="utf-8", newline="") as f:
        csv_rows = list(csv.DictReader(f))

    assert len(csv_rows) == 1
    csv_row = csv_rows[0]
    assert csv_row["Temp"] == "20.6"
    assert csv_row["T1"] == "19.93"


def test_analyzer_summary_falls_back_to_chamber_temperature_when_thermometer_is_stale(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    rows = [
        {
            "point_title": "0°C环境，二氧化碳0ppm，气压1100hPa",
            "point_row": 9,
            "point_phase": "co2",
            "point_tag": "co2_0_1100",
            "co2_ppm_target": 0.0,
            "temp_chamber_c": 0.0,
            "pressure_target_hpa": 1100.0,
            "chamber_temp_c": 0.02,
            "thermometer_temp_c": 18.09,
            "pressure_gauge_hpa": 1099.8,
            "ga01_id": "001",
            "ga01_co2_ppm": 0.2,
            "ga01_h2o_mmol": 0.3,
            "ga01_co2_ratio_f": 1.2000,
            "ga01_h2o_ratio_f": 0.8000,
            "ga01_ref_signal": 100.0,
            "ga01_co2_signal": 200.0,
            "ga01_h2o_signal": 300.0,
            "ga01_chamber_temp_c": 0.31,
            "ga01_case_temp_c": 0.52,
            "ga01_pressure_kpa": 108.10,
        },
        {
            "point_title": "0°C环境，二氧化碳0ppm，气压1100hPa",
            "point_row": 9,
            "point_phase": "co2",
            "point_tag": "co2_0_1100",
            "co2_ppm_target": 0.0,
            "temp_chamber_c": 0.0,
            "pressure_target_hpa": 1100.0,
            "chamber_temp_c": -0.01,
            "thermometer_temp_c": 18.09,
            "pressure_gauge_hpa": 1099.7,
            "ga01_id": "001",
            "ga01_co2_ppm": 0.1,
            "ga01_h2o_mmol": 0.4,
            "ga01_co2_ratio_f": 1.2100,
            "ga01_h2o_ratio_f": 0.8100,
            "ga01_ref_signal": 101.0,
            "ga01_co2_signal": 201.0,
            "ga01_h2o_signal": 301.0,
            "ga01_chamber_temp_c": 0.33,
            "ga01_case_temp_c": 0.55,
            "ga01_pressure_kpa": 108.12,
        },
    ]

    logger.log_analyzer_workbook(rows, analyzer_labels=["ga01"], phase="co2")
    logger.close()

    with logger.analyzer_summary_csv_path.open("r", encoding="utf-8", newline="") as f:
        csv_rows = list(csv.DictReader(f))

    assert len(csv_rows) == 1
    csv_row = csv_rows[0]
    assert csv_row["Temp"] == "0.005"
    assert csv_row["T1"] == "0.32"


def test_analyzer_summary_temp_uses_same_usable_samples_as_analyzer_values(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    rows = [
        {
            "point_title": "20°C环境，二氧化碳400ppm，气压1000hPa",
            "point_row": 3,
            "point_phase": "co2",
            "point_tag": "co2_400_1000",
            "co2_ppm_target": 400.0,
            "temp_chamber_c": 20.0,
            "pressure_target_hpa": 1000.0,
            "chamber_temp_c": 20.00,
            "thermometer_temp_c": 20.00,
            "pressure_gauge_hpa": 1000.4,
            "ga01_frame_usable": True,
            "ga01_id": "001",
            "ga01_co2_ppm": 401.0,
            "ga01_h2o_mmol": 18.4,
            "ga01_co2_ratio_f": 0.1230,
            "ga01_h2o_ratio_f": 0.4560,
            "ga01_ref_signal": 111.0,
            "ga01_co2_signal": 222.0,
            "ga01_h2o_signal": 333.0,
            "ga01_chamber_temp_c": 19.91,
            "ga01_case_temp_c": 20.55,
            "ga01_pressure_kpa": 101.20,
        },
        {
            "point_title": "20°C环境，二氧化碳400ppm，气压1000hPa",
            "point_row": 3,
            "point_phase": "co2",
            "point_tag": "co2_400_1000",
            "co2_ppm_target": 400.0,
            "temp_chamber_c": 20.0,
            "pressure_target_hpa": 1000.0,
            "chamber_temp_c": 30.00,
            "thermometer_temp_c": 30.00,
            "pressure_gauge_hpa": 1000.2,
            "ga01_frame_usable": False,
            "ga01_id": "001",
            "ga01_co2_ppm": 999.0,
            "ga01_h2o_mmol": 99.9,
            "ga01_co2_ratio_f": 9.9990,
            "ga01_h2o_ratio_f": 9.9990,
            "ga01_ref_signal": 113.0,
            "ga01_co2_signal": 224.0,
            "ga01_h2o_signal": 335.0,
            "ga01_chamber_temp_c": 29.95,
            "ga01_case_temp_c": 30.65,
            "ga01_pressure_kpa": 101.40,
        },
    ]

    logger.log_analyzer_workbook(rows, analyzer_labels=["ga01"], phase="co2")
    logger.close()

    with logger.analyzer_summary_csv_path.open("r", encoding="utf-8", newline="") as f:
        csv_rows = list(csv.DictReader(f))

    assert len(csv_rows) == 1
    csv_row = csv_rows[0]
    assert csv_row["Temp"] == "20.0"
    assert csv_row["ppm_CO2"] == "401.0"


def test_analyzer_summary_marks_partial_integrity_and_ignores_unusable_frames(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    rows = [
        {
            "point_title": "点1",
            "sample_ts": "2026-03-14T10:00:00.000",
            "point_phase": "co2",
            "point_tag": "demo",
            "point_row": 1,
            "co2_ppm_target": 400.0,
            "pressure_target_hpa": 1000.0,
            "pressure_gauge_hpa": 1000.0,
            "analyzer_expected_count": 2,
            "analyzer_with_frame_count": 2,
            "analyzer_usable_count": 1,
            "analyzer_coverage_text": "1/2",
            "analyzer_integrity": "含异常帧",
            "analyzer_missing_labels": "",
            "analyzer_unusable_labels": "GA02",
            "ga01_id": "001",
            "ga01_frame_has_data": True,
            "ga01_frame_usable": True,
            "ga01_co2_ppm": 401.0,
            "ga01_h2o_mmol": 2.0,
            "ga01_co2_ratio_f": 0.123,
            "ga02_id": "002",
            "ga02_frame_has_data": True,
            "ga02_frame_usable": False,
            "ga02_co2_ppm": 3000.0,
            "ga02_h2o_mmol": 72.0,
            "ga02_co2_ratio_f": 9.999,
        }
    ]

    logger.log_analyzer_workbook(rows, analyzer_labels=["ga01", "ga02"], phase="co2")
    logger.close()

    with logger.analyzer_summary_csv_path.open("r", encoding="utf-8", newline="") as f:
        csv_rows = list(csv.DictReader(f))

    ga01 = next(row for row in csv_rows if row["Analyzer"] == "GA01")
    ga02 = next(row for row in csv_rows if row["Analyzer"] == "GA02")
    assert ga01["AnalyzerCoverage"] == "1/2"
    assert ga01["PointIntegrity"] == "含异常帧"
    assert ga01["ValidFrames"] == "1"
    assert ga01["TotalFrames"] == "1"
    assert ga01["FrameStatus"] == "全部可用"
    assert ga02["AnalyzerCoverage"] == "1/2"
    assert ga02["PointIntegrity"] == "含异常帧"
    assert ga02["ValidFrames"] == "0"
    assert ga02["TotalFrames"] == "1"
    assert ga02["FrameStatus"] == "仅异常帧"
    assert ga02["ppm_CO2"] == ""


def test_analyzer_summary_rows_follow_execution_order(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    rows_1 = [
        {
            "point_title": "点1",
            "point_phase": "co2",
            "point_tag": "row1",
            "point_row": 1,
            "co2_ppm_target": 0.0,
            "dewpoint_c": 1.0,
            "dew_pressure_hpa": 1000.0,
            "pressure_gauge_hpa": 1000.0,
            "ga01_id": "001",
            "ga01_co2_ratio_f": 0.1,
            "ga01_h2o_ratio_f": 0.2,
        }
    ]
    rows_2 = [
        {
            "point_title": "点2",
            "point_phase": "co2",
            "point_tag": "row2",
            "point_row": 2,
            "co2_ppm_target": 400.0,
            "dewpoint_c": 2.0,
            "dew_pressure_hpa": 1000.0,
            "pressure_gauge_hpa": 1000.0,
            "ga01_id": "001",
            "ga01_co2_ratio_f": 0.3,
            "ga01_h2o_ratio_f": 0.4,
        }
    ]

    logger.log_analyzer_workbook(rows_1, analyzer_labels=["ga01"], phase="co2")
    logger.log_analyzer_workbook(rows_2, analyzer_labels=["ga01"], phase="co2")
    logger.close()

    with logger.analyzer_summary_csv_path.open("r", encoding="utf-8", newline="") as f:
        csv_rows = list(csv.DictReader(f))

    assert [row["PointTitle"] for row in csv_rows] == ["点1", "点2"]
    assert [row["PointRow"] for row in csv_rows] == ["1", "2"]
    assert [row["PointPhase"] for row in csv_rows] == ["气路", "气路"]


def test_log_point_readable_exports_include_valid_counts(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_point(
        {
            "point_title": "点1",
            "point_row": 1,
            "point_phase": "co2",
            "point_tag": "row1",
            "pressure_target_hpa": 1100.0,
            "controller_pressure_mean": 1100.0,
            "controller_pressure_valid_count": 3,
            "gauge_pressure_mean": 1099.3,
            "gauge_pressure_valid_count": 1,
            "dewpoint_c_mean": -16.0,
            "dewpoint_c_valid_count": 2,
            "dew_temp_c_mean": 17.1,
            "dew_temp_c_valid_count": 2,
            "dew_rh_pct_mean": 7.6,
            "dew_rh_pct_valid_count": 2,
            "co2_mean_primary_or_first": 600.0,
            "co2_mean_primary_or_first_valid_count": 10,
        }
    )
    logger.close()

    with logger.points_readable_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 1
    row = rows[0]
    assert row["压力控制器压力hPa_有效样本数"] == "3"
    assert row["数字压力计压力hPa_有效样本数"] == "1"
    assert row["露点仪露点C_有效样本数"] == "2"
    assert row["露点仪温度C_有效样本数"] == "2"
    assert row["露点仪湿度%_有效样本数"] == "2"
    assert row["二氧化碳平均值(主分析仪或首台可用)_有效样本数"] == "10"


def test_log_analyzer_exports_skip_empty_analyzer_labels(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    rows = [{"point_title": "点1", "point_phase": "co2", "point_tag": "row1", "point_row": 1}]

    summary_path = logger.log_analyzer_summary(rows, analyzer_labels=[])
    workbook_path = logger.log_analyzer_workbook(rows, analyzer_labels=[], phase="co2", write_summary=False)
    logger.close()

    assert summary_path == logger.analyzer_summary_book_path
    assert workbook_path == logger.co2_analyzer_book_path
    assert not summary_path.exists()
    assert not workbook_path.exists()


def test_readable_points_include_pressure_timing_fields(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_point(
        {
            "point_title": "点1",
            "point_row": 1,
            "point_phase": "co2",
            "point_tag": "co2_row1",
            "pressure_target_hpa": 500.0,
            "preseal_vent_off_begin_to_route_sealed_ms": 2300.0,
            "route_sealed_to_control_prepare_begin_ms": 180.0,
            "pressure_in_limits_to_sampling_begin_ms": 10050.0,
        }
    )
    logger.close()

    with logger.points_readable_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 1
    row = rows[0]
    assert row["关大气开始到封路ms"] == "2300.0"
    assert row["封路到控压准备开始ms"] == "180.0"
    assert row["压力达标到采样开始ms"] == "10050.0"


def test_readable_points_include_co2_quality_guard_fields(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_point(
        {
            "point_title": "点1",
            "point_row": 1,
            "point_phase": "co2",
            "point_tag": "co2_row1",
            "preseal_dewpoint_c": -18.0,
            "preseal_temp_c": 20.0,
            "preseal_rh_pct": 5.0,
            "preseal_pressure_hpa": 1140.0,
            "preseal_trigger_overshoot_hpa": 29.4,
            "postseal_expected_dewpoint_c": -24.1,
            "postseal_actual_dewpoint_c": -23.8,
            "postseal_physical_delta_c": 0.3,
            "postseal_physical_qc_status": "fail",
            "postseal_physical_qc_reason": "abs_delta_c=0.300>max_abs_delta_c=0.200;policy=warn",
            "postseal_timeout_policy": "warn",
            "postseal_timeout_blocked": False,
            "point_quality_timeout_flag": True,
            "dewpoint_gate_pass_live_c": -24.2,
            "presample_long_guard_status": "warn",
            "presample_long_guard_reason": "timeout_elapsed_s=20.000;rise_c=0.180>max_rise_c=0.120;policy=warn",
            "presample_long_guard_elapsed_s": 20.0,
            "presample_long_guard_span_c": 0.22,
            "presample_long_guard_slope_c_per_s": 0.03,
            "presample_long_guard_rise_c": 0.18,
            "first_effective_sample_dewpoint_c": -23.7,
            "postgate_to_first_effective_dewpoint_rise_c": 0.5,
            "postsample_late_rebound_status": "warn",
            "postsample_late_rebound_reason": "rise_c=0.500>max_rise_c=0.120;policy=warn",
            "sampling_window_dewpoint_first_c": -24.1,
            "sampling_window_dewpoint_last_c": -23.5,
            "sampling_window_dewpoint_range_c": 0.6,
            "sampling_window_dewpoint_rise_c": 0.6,
            "sampling_window_dewpoint_slope_c_per_s": 0.066667,
            "sampling_window_qc_status": "warn",
            "sampling_window_qc_reason": "range_c=0.600>max_range_c=0.200;policy=warn",
            "pressure_gauge_stale_count": 10,
            "pressure_gauge_total_count": 10,
            "pressure_gauge_stale_ratio": 1.0,
            "point_quality_status": "fail",
            "point_quality_reason": "postseal_timeout(policy=warn);pressure_gauge_stale_ratio=1.000>reject_max=0.500",
            "point_quality_flags": "postseal_timeout,pressure_gauge_stale_ratio",
            "point_quality_blocked": True,
        }
    )
    logger.close()

    with logger.points_readable_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 1
    row = rows[0]
    assert row["封路前露点快照C"] == "-18.0"
    assert row["封路前温度快照C"] == "20.0"
    assert row["封路前湿度快照%"] == "5.0"
    assert row["封路前压力快照hPa"] == "1140.0"
    assert row["封路触发超调hPa"] == "29.4"
    assert row["封压后理论露点C"] == "-24.1"
    assert row["封压后实际露点C"] == "-23.8"
    assert row["封压后物理偏差C"] == "0.3"
    assert row["封压后物理一致性结果"] == "fail"
    assert row["封压后露点超时策略"] == "warn"
    assert row["点位质量_封压后露点超时标记"] == "True"
    assert row["露点门禁放行实时露点C"] == "-24.2"
    assert row["采样前长稳守护结果"] == "warn"
    assert row["采样前长稳守护耗时s"] == "20.0"
    assert row["采样前长稳守护回升C"] == "0.18"
    assert row["首个有效样本露点C"] == "-23.7"
    assert row["门禁放行到首个有效样本露点回升C"] == "0.5"
    assert row["采样早期晚回潮结果"] == "warn"
    assert row["采样窗露点首值C"] == "-24.1"
    assert row["采样窗露点末值C"] == "-23.5"
    assert row["采样窗露点跨度C"] == "0.6"
    assert row["采样窗露点质控结果"] == "warn"
    assert row["数字压力计陈旧样本数"] == "10"
    assert row["数字压力计总样本数"] == "10"
    assert row["数字压力计陈旧样本占比"] == "1.0"
    assert row["点位质量结果"] == "fail"
    assert row["点位质量标记"] == "postseal_timeout,pressure_gauge_stale_ratio"


def test_log_sample_expands_header_for_point_quality_fields(tmp_path: Path) -> None:
    logger = RunLogger(tmp_path)
    logger.log_sample(
        {
            "sample_ts": "2026-04-04T10:00:00.000",
            "point_quality_status": "warn",
            "point_quality_reason": "postseal_timeout(policy=warn)",
            "point_quality_flags": "postseal_timeout",
            "point_quality_blocked": False,
            "postseal_timeout_policy": "warn",
            "dewpoint_gate_pass_live_c": -24.1,
            "presample_long_guard_status": "warn",
            "presample_long_guard_reason": "timeout_elapsed_s=20.000;policy=warn",
            "presample_long_guard_elapsed_s": 20.0,
            "presample_long_guard_span_c": 0.21,
            "presample_long_guard_slope_c_per_s": 0.025,
            "presample_long_guard_rise_c": 0.16,
            "first_effective_sample_dewpoint_c": -23.7,
            "postgate_to_first_effective_dewpoint_rise_c": 0.4,
            "postsample_late_rebound_status": "warn",
            "postsample_late_rebound_reason": "rise_c=0.400>max_rise_c=0.120;policy=warn",
            "sampling_window_dewpoint_first_c": -24.1,
            "sampling_window_dewpoint_last_c": -23.5,
            "sampling_window_dewpoint_range_c": 0.6,
            "sampling_window_dewpoint_rise_c": 0.6,
            "sampling_window_dewpoint_slope_c_per_s": 0.066667,
            "sampling_window_qc_status": "warn",
            "sampling_window_qc_reason": "range_c=0.600>max_range_c=0.200;policy=warn",
        }
    )
    logger.close()

    with logger.samples_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    assert len(rows) == 1
    row = rows[0]
    assert row["点位质量结果"] == "warn"
    assert row["点位质量原因"] == "postseal_timeout(policy=warn)"
    assert row["点位质量标记"] == "postseal_timeout"
    assert row["封压后露点超时策略"] == "warn"
    assert row["露点门禁放行实时露点C"] == "-24.1"
    assert row["采样前长稳守护结果"] == "warn"
    assert row["采样前长稳守护原因"] == "timeout_elapsed_s=20.000;policy=warn"
    assert row["首个有效样本露点C"] == "-23.7"
    assert row["门禁放行到首个有效样本露点回升C"] == "0.4"
    assert row["采样早期晚回潮结果"] == "warn"
    assert row["采样窗露点质控结果"] == "warn"
