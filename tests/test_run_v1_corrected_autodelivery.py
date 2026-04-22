from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from gas_calibrator.tools import run_v1_corrected_autodelivery as module


def test_extract_run_device_ids_uses_most_common_value(tmp_path: Path) -> None:
    run_dir = tmp_path / "run_1"
    run_dir.mkdir()
    frame = pd.DataFrame(
        [
            {"气体分析仪1_设备ID": "7", "气体分析仪2_设备ID": "021"},
            {"气体分析仪1_设备ID": "007", "气体分析仪2_设备ID": "021"},
            {"气体分析仪1_设备ID": "007", "气体分析仪2_设备ID": "21"},
        ]
    )
    frame.to_csv(run_dir / "samples_20260407.csv", index=False, encoding="utf-8-sig")

    mapping = module.extract_run_device_ids(run_dir)

    assert mapping == {"GA01": "007", "GA02": "021"}


def test_build_corrected_download_plan_rows_maps_groups_by_gas() -> None:
    simplified = pd.DataFrame(
        [
            {"分析仪": "GA01", "气体": "CO2", **{f"a{i}": float(i + 1) for i in range(9)}},
            {"分析仪": "GA01", "气体": "H2O", **{f"a{i}": float((i + 1) * 10) for i in range(9)}},
        ]
    )

    rows = module.build_corrected_download_plan_rows(simplified, actual_device_ids={"GA01": "086"})

    assert len(rows) == 2
    co2 = next(row for row in rows if row["Gas"] == "CO2")
    h2o = next(row for row in rows if row["Gas"] == "H2O")
    assert co2["ActualDeviceId"] == "086"
    assert h2o["ActualDeviceId"] == "086"
    assert co2["PrimarySENCO"] == "1"
    assert co2["SecondarySENCO"] == "3"
    assert co2["PrimaryCommand"].startswith("SENCO1,YGAS,FFF,1.00000e00,2.00000e00,3.00000e00,4.00000e00")
    assert co2["SecondaryCommand"].startswith("SENCO3,YGAS,FFF,5.00000e00,6.00000e00,7.00000e00,8.00000e00,9.00000e00")
    assert h2o["PrimarySENCO"] == "2"
    assert h2o["SecondarySENCO"] == "4"


def test_build_corrected_download_plan_rows_fills_missing_a7_a8_with_zero() -> None:
    simplified = pd.DataFrame(
        [
            {"分析仪": "GA01", "气体": "CO2", **{f"a{i}": float(i + 1) for i in range(7)}},
        ]
    )

    rows = module.build_corrected_download_plan_rows(simplified, actual_device_ids={"GA01": "086"})

    assert len(rows) == 1
    row = rows[0]
    assert row["a7"] == 0.0
    assert row["a8"] == 0.0
    assert row["SecondaryC3"] == "0.00000e00"
    assert row["SecondaryC4"] == "0.00000e00"
    assert row["PrimaryCommand"].startswith("SENCO1,YGAS,FFF,1.00000e00,2.00000e00,3.00000e00,4.00000e00")
    assert row["SecondaryCommand"].endswith("5.00000e00,6.00000e00,7.00000e00,0.00000e00,0.00000e00,0.00000e00")


def test_compute_pressure_offset_rows_uses_ambient_pressure_gauge_samples(tmp_path: Path) -> None:
    run_dir = tmp_path / "run_2"
    run_dir.mkdir()
    frame = pd.DataFrame(
        [
            {
                "压力执行模式": "ambient_open",
                "数字压力计压力hPa": 1019.5,
                "气体分析仪1_设备ID": "005",
                "气体分析仪1_分析仪压力kPa": 103.0,
                "气体分析仪1_分析仪可用帧": True,
            },
            {
                "压力执行模式": "ambient_open",
                "数字压力计压力hPa": 1019.7,
                "气体分析仪1_设备ID": "005",
                "气体分析仪1_分析仪压力kPa": 103.1,
                "气体分析仪1_分析仪可用帧": True,
            },
            {
                "压力执行模式": "sealed_controlled",
                "数字压力计压力hPa": 500.0,
                "气体分析仪1_设备ID": "005",
                "气体分析仪1_分析仪压力kPa": 50.0,
                "气体分析仪1_分析仪可用帧": True,
            },
        ]
    )
    frame.to_csv(run_dir / "samples_20260407.csv", index=False, encoding="utf-8-sig")

    rows = module.compute_pressure_offset_rows(run_dir)

    assert len(rows) == 1
    row = rows[0]
    assert row["Analyzer"] == "GA01"
    assert row["DeviceId"] == "005"
    assert row["Samples"] == 2
    assert abs(float(row["OffsetA_kPa"]) + 1.09) < 1e-9
    assert str(row["Command"]).startswith("SENCO9,YGAS,FFF,-1.09000e00,1.00000e00")


def test_compute_pressure_offset_rows_marks_large_gauge_controller_gap_as_not_recommended(tmp_path: Path) -> None:
    run_dir = tmp_path / "run_2_backpressure"
    run_dir.mkdir()
    device_id_col = "气体分析仪1_设备ID"
    pressure_col = "气体分析仪1_分析仪压力kPa"
    usable_col = "气体分析仪1_分析仪可用帧"
    frame = pd.DataFrame(
        [
            {
                "PressureMode": "ambient_open",
                "pressure_gauge_hpa": 1019.0 + idx * 0.1,
                "pressure_controller_hpa": 1014.0 + idx * 0.1,
                device_id_col: "005",
                pressure_col: 103.0 + idx * 0.01,
                usable_col: True,
            }
            for idx in range(5)
        ]
    )
    frame.to_csv(run_dir / "samples_20260407.csv", index=False, encoding="utf-8-sig")

    rows = module.compute_pressure_offset_rows(run_dir, fallback_to_controller=True)

    assert len(rows) == 1
    row = rows[0]
    assert row["GaugeControllerOverlapSamples"] == 5
    assert abs(float(row["GaugeControllerMeanAbsDiff_hPa"]) - 5.0) < 1e-9
    assert abs(float(row["GaugeControllerMaxAbsDiff_hPa"]) - 5.0) < 1e-9
    assert row["PressureWriteRecommended"] is False
    assert row["PressureWriteReason"] == "ambient_open_backpressure_too_large"


def test_load_startup_pressure_calibration_rows_uses_latest_summary(tmp_path: Path) -> None:
    run_dir = tmp_path / "run_3"
    older = run_dir / "startup_pressure_sensor_calibration_20260407_120000"
    newer = run_dir / "startup_pressure_sensor_calibration_20260407_130000"
    older.mkdir(parents=True)
    newer.mkdir(parents=True)
    pd.DataFrame(
        [
            {"Analyzer": "GA01", "DeviceId": "005", "Samples": 4, "OffsetA_kPa": -1.0, "WriteApplied": True, "ReadbackOk": True, "Status": "ok"},
        ]
    ).to_csv(older / "summary.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(
        [
            {"Analyzer": "GA01", "DeviceId": "005", "Samples": 5, "OffsetA_kPa": -2.5, "WriteApplied": True, "ReadbackOk": True, "Status": "ok"},
        ]
    ).to_csv(newer / "summary.csv", index=False, encoding="utf-8-sig")

    rows = module.load_startup_pressure_calibration_rows(run_dir)

    assert len(rows) == 1
    row = rows[0]
    assert row["Analyzer"] == "GA01"
    assert row["DeviceId"] == "005"
    assert row["Samples"] == 5
    assert abs(float(row["OffsetA_kPa"]) + 2.5) < 1e-9
    assert row["ReferenceSource"] == "startup_pressure_sensor_calibration"
    assert row["PressureWriteRecommended"] is True
    assert row["PressureWriteReason"] == ""
    assert str(row["Command"]).startswith("SENCO9,YGAS,FFF,-2.50000e00,1.00000e00")
    assert str(row["SourceSummary"]).endswith("startup_pressure_sensor_calibration_20260407_130000\\summary.csv")


def test_load_startup_pressure_calibration_rows_marks_unstable_reference_as_not_recommended(tmp_path: Path) -> None:
    run_dir = tmp_path / "run_3_unstable"
    target = run_dir / "startup_pressure_sensor_calibration_20260407_130000"
    target.mkdir(parents=True)
    pd.DataFrame(
        [
            {"Analyzer": "GA01", "DeviceId": "005", "Samples": 5, "OffsetA_kPa": -2.5, "WriteApplied": True, "ReadbackOk": True, "Status": "ok"},
        ]
    ).to_csv(target / "summary.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(
        [
            {"Analyzer": "GA01", "DeviceId": "005", "ReferenceHpa": 999.0, "ReferenceSource": "pressure_gauge", "AnalyzerPressureKPa": 99.9, "OffsetA_kPa": 0.0, "FrameOk": True},
            {"Analyzer": "GA01", "DeviceId": "005", "ReferenceHpa": 1002.2, "ReferenceSource": "pressure_gauge", "AnalyzerPressureKPa": 99.9, "OffsetA_kPa": 0.0, "FrameOk": True},
            {"Analyzer": "GA01", "DeviceId": "005", "ReferenceHpa": 1001.5, "ReferenceSource": "pressure_gauge", "AnalyzerPressureKPa": 99.9, "OffsetA_kPa": 0.0, "FrameOk": True},
        ]
    ).to_csv(target / "detail.csv", index=False, encoding="utf-8-sig")

    rows = module.load_startup_pressure_calibration_rows(run_dir)

    assert len(rows) == 1
    row = rows[0]
    assert row["StartupDetailSamples"] == 3
    assert row["StartupPressureGaugeSamples"] == 3
    assert abs(float(row["StartupReferenceSpanHpa"]) - 3.2) < 1e-9
    assert row["PressureWriteRecommended"] is False
    assert row["PressureWriteReason"] == "startup_pressure_reference_unstable"


def test_build_corrected_delivery_prefers_startup_pressure_rows(tmp_path: Path, monkeypatch) -> None:
    run_dir = tmp_path / "run_4"
    out_dir = tmp_path / "out_4"
    run_dir.mkdir()
    out_dir.mkdir()
    report_kwargs: dict[str, object] = {}
    appended_sheets: list[str] = []

    pd.DataFrame(
        [
            {
                "analyzer_id": "GA01",
                "analyzer_device_id": "005",
                "snapshot_time": "2026-04-12T10:00:00",
                "route_type": "co2",
                "ref_temp_c": 0.0,
                "cell_temp_raw_c": 60.0,
                "shell_temp_raw_c": -40.0,
                "cell_temp_span_c": 0.0,
                "shell_temp_span_c": 0.0,
                "valid_for_cell_fit": False,
                "valid_for_shell_fit": False,
                "cell_fit_gate_reason": "hard_bad_value",
                "shell_fit_gate_reason": "hard_bad_value",
            }
        ]
    ).to_csv(run_dir / "temperature_calibration_observations.csv", index=False, encoding="utf-8-sig")

    monkeypatch.setattr(module, "_filter_no_500_summary_paths", lambda *_args, **_kwargs: ([run_dir / "gas.csv", run_dir / "water.csv"], []))
    monkeypatch.setattr(
        module,
        "_append_dataframe_sheet",
        lambda _path, sheet_name, _frame: appended_sheets.append(str(sheet_name)),
    )
    def _fake_report(*_args, **_kwargs):
        report_kwargs.update(_kwargs)
        return {
            "summary": pd.DataFrame([{"鍒嗘瀽浠?": "GA01", "姘斾綋": "CO2"}]),
            "simplified": pd.DataFrame([{"鍒嗘瀽浠?": "GA01", "姘斾綋": "CO2", **{f"a{i}": float(i + 1) for i in range(9)}}]),
            "h2o_selected_rows": pd.DataFrame(
                [
                    {
                        "Analyzer": "GA01",
                        "PointRow": 3,
                        "PointTag": "co2_m20_0",
                        "SelectionOrigin": "co2_zero_ppm_anchor",
                        "EnvTempC": -20.0,
                    }
                ]
            ),
            "h2o_anchor_gate_hits": pd.DataFrame(
                [
                    {
                        "Analyzer": "GA01",
                        "PointRow": 12,
                        "PointTag": "co2_0_0",
                        "GateReason": "anchor_h2o_dew_above_limit",
                    }
                ]
            ),
        }

    monkeypatch.setattr(module, "build_corrected_water_points_report", _fake_report)
    monkeypatch.setattr(module, "extract_run_device_ids", lambda *_args, **_kwargs: {"GA01": "005"})
    monkeypatch.setattr(module, "load_temperature_coefficient_rows", lambda *_args, **_kwargs: [{"analyzer_id": "GA01", "senco_channel": "SENCO7", "A": 1, "B": 2, "C": 3, "D": 4}])
    monkeypatch.setattr(module, "load_startup_pressure_calibration_rows", lambda *_args, **_kwargs: [{"Analyzer": "GA01", "DeviceId": "005", "OffsetA_kPa": -2.5, "Command": "SENCO9,YGAS,FFF,-2.50000e00,1.00000e00,0.00000e00,0.00000e00"}])
    compute_calls = {"count": 0}

    def _fake_compute(*_args, **_kwargs):
        compute_calls["count"] += 1
        return [{"Analyzer": "GA01", "DeviceId": "005", "OffsetA_kPa": -9.9}]

    monkeypatch.setattr(module, "compute_pressure_offset_rows", _fake_compute)

    result = module.build_corrected_delivery(
        run_dir=run_dir,
        output_dir=out_dir,
        coeff_cfg={"h2o_summary_selection": {"include_co2_temp_groups_c": [], "include_co2_zero_ppm_temp_groups_c": [-20.0, -10.0, 0.0]}},
        pressure_row_source="startup_calibration",
    )

    assert compute_calls["count"] == 0
    assert result["pressure_rows"] == [{"Analyzer": "GA01", "DeviceId": "005", "OffsetA_kPa": -2.5, "Command": "SENCO9,YGAS,FFF,-2.50000e00,1.00000e00,0.00000e00,0.00000e00"}]
    assert report_kwargs["coeff_cfg"]["h2o_summary_selection"] == {"include_co2_temp_groups_c": [], "include_co2_zero_ppm_temp_groups_c": [-20.0, -10.0, 0.0]}
    assert report_kwargs["coeff_cfg"]["original_selected_pressure_points"] is None
    assert report_kwargs["coeff_cfg"]["selected_pressure_points_source"] == "runtime_snapshot"
    assert "H2O锚点入选" in appended_sheets
    assert "H2O锚点门禁" in appended_sheets
    assert "推荐运行结构提示" in appended_sheets
    assert "温补异常快照" in appended_sheets
    assert result["h2o_selected_rows"] == [{"Analyzer": "GA01", "PointRow": 3, "PointTag": "co2_m20_0", "SelectionOrigin": "co2_zero_ppm_anchor", "EnvTempC": -20.0, "ActualDeviceId": "005"}]
    assert result["h2o_anchor_gate_hits"] == [{"Analyzer": "GA01", "PointRow": 12, "PointTag": "co2_0_0", "GateReason": "anchor_h2o_dew_above_limit", "ActualDeviceId": "005"}]
    assert result["temperature_gate_hits"][0]["analyzer_id"] == "GA01"
    assert any(item["CheckCode"] == "pressure_structure" for item in result["run_structure_hints"])


def test_build_corrected_delivery_passes_selected_pressure_points_to_report(tmp_path: Path, monkeypatch) -> None:
    run_dir = tmp_path / "run_ambient_only"
    out_dir = tmp_path / "out_ambient_only"
    run_dir.mkdir()
    out_dir.mkdir()
    (run_dir / "runtime_config_snapshot.json").write_text(
        json.dumps({"workflow": {"selected_pressure_points": ["ambient"]}}),
        encoding="utf-8",
    )
    report_kwargs: dict[str, object] = {}

    monkeypatch.setattr(module, "_filter_no_500_summary_paths", lambda *_args, **_kwargs: ([run_dir / "summary.xlsx"], []))

    def _fake_report(*_args, **kwargs):
        report_kwargs.update(kwargs)
        return {
            "summary": pd.DataFrame([{"分析仪": "GA01", "气体": "CO2"}]),
            "simplified": pd.DataFrame([{"分析仪": "GA01", "气体": "CO2", **{f'a{i}': float(i + 1) for i in range(9)}}]),
            "original": pd.DataFrame([{"分析仪": "GA01", "气体": "CO2"}]),
            "points": pd.DataFrame([{"分析仪": "GA01", "气体": "CO2"}]),
            "ranges": pd.DataFrame([{"分析仪": "GA01", "气体": "CO2"}]),
            "topn": pd.DataFrame([{"分析仪": "GA01", "气体": "CO2"}]),
            "h2o_selected_rows": pd.DataFrame(),
            "h2o_anchor_gate_hits": pd.DataFrame(),
        }

    monkeypatch.setattr(module, "build_corrected_water_points_report", _fake_report)
    monkeypatch.setattr(module, "extract_run_device_ids", lambda *_args, **_kwargs: {"GA01": "005"})
    monkeypatch.setattr(module, "load_temperature_coefficient_rows", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(module, "load_startup_pressure_calibration_rows", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(module, "_append_dataframe_sheet", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(module, "_annotate_workbook_with_actual_device_ids", lambda *_args, **_kwargs: None)

    result = module.build_corrected_delivery(
        run_dir=run_dir,
        output_dir=out_dir,
        coeff_cfg={"target_digits": 6},
        pressure_row_source="startup_calibration",
    )

    assert report_kwargs["coeff_cfg"]["target_digits"] == 6
    assert report_kwargs["coeff_cfg"]["selected_pressure_points"] == ["ambient"]
    assert report_kwargs["coeff_cfg"]["original_selected_pressure_points"] == ["ambient"]
    assert report_kwargs["coeff_cfg"]["selected_pressure_points_source"] == "runtime_snapshot"
    assert report_kwargs["gas_temperature_keys"] == {}
    assert result["pressure_rows"] == []


def test_write_coefficients_to_live_devices_can_skip_pressure_rows(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(
        module,
        "scan_live_targets",
        lambda *_args, **_kwargs: [
            {
                "Analyzer": "GA01",
                "Port": "COM35",
                "Baudrate": 115200,
                "Timeout": 0.6,
                "ConfiguredDeviceId": "005",
                "LiveDeviceId": "005",
                "ActiveSend": True,
                "FtdHz": 10,
                "AverageFilter": 49,
            }
        ],
    )

    writes: list[int] = []

    class _FakeGasAnalyzer:
        def __init__(self, *args, **kwargs) -> None:
            self.values = {1: [1.0] * 6, 7: [7.0] * 4, 8: [8.0] * 4}

        def open(self) -> None:
            return None

        def close(self) -> None:
            return None

        def set_comm_way_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_mode_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_active_freq_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_average_filter_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_senco(self, group: int, coeffs) -> bool:
            writes.append(int(group))
            self.values[int(group)] = [float(value) for value in coeffs]
            return True

        def read_coefficient_group(self, group: int):
            values = self.values[int(group)]
            return {f"C{idx}": float(value) for idx, value in enumerate(values)}

    monkeypatch.setattr(module, "GasAnalyzer", _FakeGasAnalyzer)

    result = module.write_coefficients_to_live_devices(
        cfg={},
        output_dir=tmp_path / "write_out",
        download_plan_rows=[{"Analyzer": "GA01", "PrimaryCommand": "SENCO1,YGAS,FFF,1,1,1,1,0,0", "SecondaryCommand": ""}],
        temperature_rows=[{"analyzer_id": "GA01", "senco_channel": "SENCO7", "A": 1, "B": 2, "C": 3, "D": 4}],
        pressure_rows=[{"Analyzer": "GA01", "OffsetA_kPa": -2.5}],
        actual_device_ids={"GA01": "005"},
        write_pressure_rows=False,
    )

    assert 9 not in writes
    assert result["summary_rows"][0]["MatchedGroups"] == 2


def test_write_coefficients_to_live_devices_skips_unrecommended_pressure_rows(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(
        module,
        "scan_live_targets",
        lambda *_args, **_kwargs: [
            {
                "Analyzer": "GA01",
                "Port": "COM35",
                "Baudrate": 115200,
                "Timeout": 0.6,
                "ConfiguredDeviceId": "005",
                "LiveDeviceId": "005",
                "ActiveSend": True,
                "FtdHz": 10,
                "AverageFilter": 49,
            }
        ],
    )

    writes: list[int] = []

    class _FakeGasAnalyzer:
        def __init__(self, *args, **kwargs) -> None:
            return None

        def open(self) -> None:
            return None

        def close(self) -> None:
            return None

        def set_comm_way_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_mode_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_active_freq_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_average_filter_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_senco(self, group: int, coeffs) -> bool:
            writes.append(int(group))
            return True

    monkeypatch.setattr(module, "GasAnalyzer", _FakeGasAnalyzer)

    result = module.write_coefficients_to_live_devices(
        cfg={},
        output_dir=tmp_path / "write_out_skip_pressure",
        download_plan_rows=[],
        temperature_rows=[],
        pressure_rows=[
            {
                "Analyzer": "GA01",
                "OffsetA_kPa": -2.5,
                "PressureWriteRecommended": False,
                "PressureWriteReason": "ambient_open_backpressure_too_large",
            }
        ],
        actual_device_ids={"GA01": "005"},
        write_pressure_rows=True,
    )

    assert writes == []
    assert result["summary_rows"][0]["ExpectedGroups"] == 0
    assert result["detail_rows"][0]["Group"] == 9
    assert result["detail_rows"][0]["Error"] == "SKIPPED_PRESSURE_WRITE:ambient_open_backpressure_too_large"


def test_write_coefficients_to_live_devices_retries_transient_readback_failure(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(
        module,
        "scan_live_targets",
        lambda *_args, **_kwargs: [
            {
                "Analyzer": "GA01",
                "Port": "COM35",
                "Baudrate": 115200,
                "Timeout": 0.6,
                "ConfiguredDeviceId": "005",
                "LiveDeviceId": "005",
                "ActiveSend": True,
                "FtdHz": 10,
                "AverageFilter": 49,
            }
        ],
    )

    class _FakeGasAnalyzer:
        def __init__(self, *args, **kwargs) -> None:
            self.values = {9: [-2.5, 1.0, 0.0, 0.0]}
            self.read_attempts = {9: 0}

        def open(self) -> None:
            return None

        def close(self) -> None:
            return None

        def set_comm_way_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_mode_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_active_freq_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_average_filter_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_senco(self, group: int, coeffs) -> bool:
            self.values[int(group)] = [float(value) for value in coeffs]
            return True

        def read_coefficient_group(self, group: int):
            group = int(group)
            self.read_attempts[group] = self.read_attempts.get(group, 0) + 1
            if group == 9 and self.read_attempts[group] == 1:
                raise RuntimeError("GETCO9 read failed: NO_RESPONSE")
            values = self.values[group]
            return {f"C{idx}": float(value) for idx, value in enumerate(values)}

    monkeypatch.setattr(module, "GasAnalyzer", _FakeGasAnalyzer)

    result = module.write_coefficients_to_live_devices(
        cfg={},
        output_dir=tmp_path / "write_out_retry",
        download_plan_rows=[],
        temperature_rows=[],
        pressure_rows=[{"Analyzer": "GA01", "OffsetA_kPa": -2.5}],
        actual_device_ids={"GA01": "005"},
        write_pressure_rows=True,
    )

    assert result["summary_rows"][0]["MatchedGroups"] == 1
    assert result["detail_rows"][0]["Group"] == 9
    assert result["detail_rows"][0]["ReadbackOk"] is True


def test_write_coefficients_to_live_devices_requires_explicit_c0_readback_source(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(
        module,
        "scan_live_targets",
        lambda *_args, **_kwargs: [
            {
                "Analyzer": "GA01",
                "Port": "COM35",
                "Baudrate": 115200,
                "Timeout": 0.6,
                "ConfiguredDeviceId": "005",
                "LiveDeviceId": "005",
                "ActiveSend": True,
                "FtdHz": 10,
                "AverageFilter": 49,
            }
        ],
    )

    class _FakeGasAnalyzer:
        def __init__(self, *args, **kwargs) -> None:
            self.before_values = {1: [10.0, 20.0, 30.0, 40.0, 0.0, 0.0]}
            self.values = {1: [10.0, 20.0, 30.0, 40.0, 0.0, 0.0]}

        def open(self) -> None:
            return None

        def close(self) -> None:
            return None

        def set_comm_way_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_mode_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_active_freq_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_average_filter_with_ack(self, *args, **kwargs) -> bool:
            return True

        def set_senco(self, group: int, coeffs) -> bool:
            self.values[int(group)] = [float(value) for value in coeffs]
            return True

        def read_coefficient_group_capture(self, group: int):
            group = int(group)
            values = self.values[group]
            coefficients = {f"C{idx}": float(value) for idx, value in enumerate(values)}
            if values != self.before_values[group]:
                return {
                    "source": "parsed_from_ambiguous_line",
                    "coefficients": coefficients,
                    "source_line": "YGAS,005,0782.713,00.000,0.99,0.99,031.94,104.24,0001,2769 <C0:1,C1:2,C2:3,C3:4,C4:0,C5:0>",
                    "source_line_has_explicit_c0": False,
                    "raw_transcript_lines": [
                        "YGAS,005,0782.713,00.000,0.99,0.99,031.94,104.24,0001,2769 <C0:1,C1:2,C2:3,C3:4,C4:0,C5:0>"
                    ],
                    "attempt_transcripts": [
                        {
                            "attempt": 1,
                            "lines": [
                                "YGAS,005,0782.713,00.000,0.99,0.99,031.94,104.24,0001,2769 <C0:1,C1:2,C2:3,C3:4,C4:0,C5:0>"
                            ],
                        }
                    ],
                    "command": "GETCO,YGAS,005,1\r\n",
                    "target_id": "005",
                    "error": "AMBIGUOUS_COEFFICIENT_LINE",
                }
            return {
                "source": "parsed_from_explicit_c0_line",
                "coefficients": coefficients,
                "source_line": "<C0:10,C1:20,C2:30,C3:40,C4:0,C5:0>",
                "source_line_has_explicit_c0": True,
                "raw_transcript_lines": ["<C0:10,C1:20,C2:30,C3:40,C4:0,C5:0>"],
                "attempt_transcripts": [{"attempt": 1, "lines": ["<C0:10,C1:20,C2:30,C3:40,C4:0,C5:0>"]}],
                "command": "GETCO,YGAS,005,1\r\n",
                "target_id": "005",
                "error": "",
            }

    monkeypatch.setattr(module, "GasAnalyzer", _FakeGasAnalyzer)

    result = module.write_coefficients_to_live_devices(
        cfg={},
        output_dir=tmp_path / "write_out_explicit",
        download_plan_rows=[{"Analyzer": "GA01", "PrimaryCommand": "SENCO1,YGAS,FFF,1,2,3,4,0,0", "SecondaryCommand": ""}],
        temperature_rows=[],
        pressure_rows=[],
        actual_device_ids={"GA01": "005"},
        write_pressure_rows=False,
    )

    row = result["detail_rows"][0]
    assert row["ReadbackOk"] is False
    assert row["ReadbackVerified"] is False
    assert row["ReadbackTruthSource"] == "ambiguous"
    assert row["ReadbackSource"] == "parsed_from_ambiguous_line"
    assert row["ReadbackSourceHasExplicitC0"] is False
    assert "READBACK_SOURCE_UNTRUSTED:parsed_from_ambiguous_line" in row["Error"]
    truth_groups_csv = Path(result["writeback_truth_groups_path"]).read_text(encoding="utf-8-sig")
    truth_summary = json.loads(Path(result["writeback_truth_summary_path"]).read_text(encoding="utf-8"))
    assert ",1,ambiguous,False," in truth_groups_csv
    assert truth_summary["verified_group_count"] == 0
    assert truth_summary["truth_source_counts"]["ambiguous"] == 1


def test_write_coefficients_to_live_devices_writes_transcript_and_truth_artifacts(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(
        module,
        "scan_live_targets",
        lambda *_args, **_kwargs: [
            {
                "Analyzer": "GA01",
                "Port": "COM35",
                "Baudrate": 115200,
                "Timeout": 0.6,
                "ConfiguredDeviceId": "005",
                "LiveDeviceId": "005",
                "ActiveSend": True,
                "FtdHz": 10,
                "AverageFilter": 49,
            }
        ],
    )

    class _FakeGasAnalyzer:
        def __init__(self, *args, **kwargs) -> None:
            self.port = str(args[0] if args else "COM35")
            self.device_id = "005"
            self.io_logger = kwargs.get("io_logger")
            self.values = {
                1: [10.0, 20.0, 30.0, 40.0, 0.0, 0.0],
                7: [7.0, 8.0, 9.0, 10.0],
            }

        def _log(self, *, direction: str, command: str = "", response: str = "", error: str = "") -> None:
            if self.io_logger is None:
                return
            self.io_logger.log_io(
                port=self.port,
                device=self.device_id,
                direction=direction,
                command=command or None,
                response=response or None,
                error=error or None,
            )

        def open(self) -> None:
            return None

        def close(self) -> None:
            return None

        def set_comm_way_with_ack(self, active: bool, **_kwargs) -> bool:
            command = f"SETCOMWAY,YGAS,FFF,{1 if active else 0}\r\n"
            self._log(direction="tx", command=command)
            self._log(direction="rx", response="YGAS,005,T")
            return True

        def set_mode_with_ack(self, mode: int, **_kwargs) -> bool:
            command = f"MODE,YGAS,FFF,{int(mode)}\r\n"
            self._log(direction="tx", command=command)
            self._log(direction="rx", response="YGAS,005,T")
            return True

        def set_active_freq_with_ack(self, hz: int, **_kwargs) -> bool:
            command = f"FTD,YGAS,FFF,{int(hz)}\r\n"
            self._log(direction="tx", command=command)
            self._log(direction="rx", response="YGAS,005,T")
            return True

        def set_average_filter_with_ack(self, value: int, **_kwargs) -> bool:
            command = f"AVERAGE_FILTER,YGAS,FFF,{int(value)}\r\n"
            self._log(direction="tx", command=command)
            self._log(direction="rx", response="YGAS,005,T")
            return True

        def set_senco(self, group: int, coeffs) -> bool:
            values = [float(value) for value in coeffs]
            self.values[int(group)] = values
            payload = ",".join(str(value) for value in values)
            self._log(direction="tx", command=f"SENCO{int(group)},YGAS,FFF,{payload}\r\n")
            self._log(direction="rx", response="YGAS,005,T")
            return True

        def read_coefficient_group_capture(self, group: int):
            values = [float(value) for value in self.values[int(group)]]
            coeff_tokens = ",".join(f"C{idx}:{value:g}" for idx, value in enumerate(values))
            source_line = f"<{coeff_tokens}>"
            command = f"GETCO,YGAS,005,{int(group)}\r\n"
            self._log(direction="tx", command=command)
            self._log(direction="rx", response=source_line)
            return {
                "source": "parsed_from_explicit_c0_line",
                "coefficients": {f"C{idx}": float(value) for idx, value in enumerate(values)},
                "source_line": source_line,
                "source_line_has_explicit_c0": True,
                "raw_transcript_lines": [source_line],
                "attempt_transcripts": [{"attempt": 1, "lines": [source_line]}],
                "command": command,
                "target_id": "005",
                "error": "",
            }

    monkeypatch.setattr(module, "GasAnalyzer", _FakeGasAnalyzer)

    result = module.write_coefficients_to_live_devices(
        cfg={},
        output_dir=tmp_path / "write_out_truth",
        download_plan_rows=[{"Analyzer": "GA01", "PrimaryCommand": "SENCO1,YGAS,FFF,1,2,3,4,0,0", "SecondaryCommand": ""}],
        temperature_rows=[{"analyzer_id": "GA01", "senco_channel": "SENCO7", "A": 1, "B": 2, "C": 3, "D": 4}],
        pressure_rows=[],
        actual_device_ids={"GA01": "005"},
        write_pressure_rows=False,
    )

    row = result["detail_rows"][0]
    transcript_log = Path(result["writeback_raw_transcript_path"]).read_text(encoding="utf-8")
    truth_summary = json.loads(Path(result["writeback_truth_summary_path"]).read_text(encoding="utf-8"))
    truth_groups = Path(result["writeback_truth_groups_path"]).read_text(encoding="utf-8-sig")

    assert row["ReadbackOk"] is True
    assert row["ReadbackVerified"] is True
    assert row["ReadbackTruthSource"] == "explicit_c0"
    assert "SENCO1,YGAS,FFF" in transcript_log
    assert "GETCO,YGAS,005,1" in transcript_log
    assert "restore_comm_way" in transcript_log
    assert truth_summary["all_groups_verified"] is True
    assert truth_summary["truth_source_counts"]["explicit_c0"] == 2
    assert ",1,explicit_c0,True," in truth_groups
    assert ",7,explicit_c0,True," in truth_groups


def test_annotate_workbook_with_actual_device_ids_inserts_column(tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["Analyzer", "Metric"])
    ws.append(["GA01", 1.23])
    ws.append(["GA02", 4.56])
    wb.save(workbook_path)
    wb.close()

    module._annotate_workbook_with_actual_device_ids(workbook_path, {"GA01": "086", "GA02": "008"})

    wb = load_workbook(workbook_path)
    try:
        ws = wb["summary"]
        assert ws.cell(1, 2).value == "ActualDeviceId"
        assert ws.cell(2, 2).value == "086"
        assert ws.cell(3, 2).value == "008"
    finally:
        wb.close()


def test_run_from_cli_runs_short_verify_after_successful_write(tmp_path: Path, monkeypatch) -> None:
    run_dir = tmp_path / "run_ok"
    run_dir.mkdir()
    cfg_path = run_dir / "runtime_config_snapshot.json"
    cfg_path.write_text("{}", encoding="utf-8")
    captured: dict[str, object] = {}

    monkeypatch.setattr(
        module,
        "build_corrected_delivery",
        lambda **_kwargs: {
            "report_path": str(run_dir / "calibration_coefficients.xlsx"),
            "output_dir": str(run_dir / "corrected"),
            "filtered_summary_paths": [],
            "filter_stats": [],
            "actual_device_ids": {"GA01": "086"},
            "download_plan_rows": [],
            "temperature_rows": [],
            "pressure_rows": [],
            "pressure_row_source": "startup_calibration",
        },
    )
    monkeypatch.setattr(module, "load_config", lambda _path: {"_base_dir": str(tmp_path)})
    monkeypatch.setattr(
        module,
        "write_coefficients_to_live_devices",
        lambda **_kwargs: {
            "scan_rows": [],
            "summary_rows": [{"Analyzer": "GA01", "Status": "ok"}],
            "detail_rows": [],
        },
    )

    from gas_calibrator.tools import verify_short_run as verify_module

    monkeypatch.setattr(
        verify_module,
        "run_short_verification",
        lambda **kwargs: captured.update(kwargs) or {"ok": True, "run_dir": str(tmp_path / "short_verify_run")},
    )
    monkeypatch.setattr(module, "_append_dataframe_sheet", lambda *_args, **_kwargs: None)

    result = module.run_from_cli(
        run_dir=str(run_dir),
        config_path=str(cfg_path),
        output_dir=str(tmp_path / "out"),
        write_devices=True,
        verify_short_run_cfg={
            "enabled": True,
            "temp_c": 20.0,
            "skip_co2_ppm": [500],
            "enable_connect_check": False,
            "points_excel": "configs/points_tiny_short_run_20c_even500.xlsx",
        },
    )

    assert captured["actual_device_ids"] == {"GA01": "086"}
    assert captured["temp_c"] == 20.0
    assert captured["skip_co2_ppm"] == [500]
    assert str(captured["points_excel_override"]).endswith("configs\\points_tiny_short_run_20c_even500.xlsx")
    assert result["short_verify_outputs"]["ok"] is True


def test_run_from_cli_skips_short_verify_when_writeback_incomplete(tmp_path: Path, monkeypatch) -> None:
    run_dir = tmp_path / "run_partial"
    run_dir.mkdir()
    cfg_path = run_dir / "runtime_config_snapshot.json"
    cfg_path.write_text("{}", encoding="utf-8")

    monkeypatch.setattr(
        module,
        "build_corrected_delivery",
        lambda **_kwargs: {
            "report_path": str(run_dir / "calibration_coefficients.xlsx"),
            "output_dir": str(run_dir / "corrected"),
            "filtered_summary_paths": [],
            "filter_stats": [],
            "actual_device_ids": {"GA01": "086"},
            "download_plan_rows": [],
            "temperature_rows": [],
            "pressure_rows": [],
            "pressure_row_source": "startup_calibration",
        },
    )
    monkeypatch.setattr(module, "load_config", lambda _path: {"_base_dir": str(tmp_path)})
    monkeypatch.setattr(
        module,
        "write_coefficients_to_live_devices",
        lambda **_kwargs: {
            "scan_rows": [],
            "summary_rows": [{"Analyzer": "GA01", "Status": "partial"}],
            "detail_rows": [],
        },
    )
    monkeypatch.setattr(module, "_append_dataframe_sheet", lambda *_args, **_kwargs: None)

    result = module.run_from_cli(
        run_dir=str(run_dir),
        config_path=str(cfg_path),
        output_dir=str(tmp_path / "out"),
        write_devices=True,
        verify_short_run_cfg={"enabled": True},
    )

    assert result["short_verify_outputs"]["skipped"] is True
    assert result["short_verify_outputs"]["reason"] == "writeback_incomplete"


def test_build_corrected_delivery_applies_postrun_pressure_points_override_and_temperature_keys(
    tmp_path: Path,
    monkeypatch,
) -> None:
    run_dir = tmp_path / "run_override"
    out_dir = tmp_path / "out_override"
    run_dir.mkdir()
    out_dir.mkdir()
    (run_dir / "runtime_config_snapshot.json").write_text(
        json.dumps({"workflow": {"selected_pressure_points": ["ambient", 500]}}),
        encoding="utf-8",
    )
    report_kwargs: dict[str, object] = {}

    monkeypatch.setattr(module, "_filter_no_500_summary_paths", lambda *_args, **_kwargs: ([run_dir / "summary.xlsx"], []))

    def _fake_report(*_args, **kwargs):
        report_kwargs.update(kwargs)
        return {
            "summary": pd.DataFrame([{"Analyzer": "GA01", "Gas": "CO2"}]),
            "simplified": pd.DataFrame([{"Analyzer": "GA01", "Gas": "CO2", **{f"a{i}": float(i + 1) for i in range(9)}}]),
            "original": pd.DataFrame([{"Analyzer": "GA01", "Gas": "CO2", **{f"a{i}": float(i + 101) for i in range(9)}}]),
            "points": pd.DataFrame([{"Analyzer": "GA01", "Gas": "CO2"}]),
            "ranges": pd.DataFrame([{"Analyzer": "GA01", "Gas": "CO2"}]),
            "topn": pd.DataFrame([{"Analyzer": "GA01", "Gas": "CO2"}]),
            "h2o_selected_rows": pd.DataFrame(),
            "h2o_anchor_gate_hits": pd.DataFrame(),
        }

    monkeypatch.setattr(module, "build_corrected_water_points_report", _fake_report)
    monkeypatch.setattr(module, "extract_run_device_ids", lambda *_args, **_kwargs: {"GA01": "005"})
    monkeypatch.setattr(module, "load_temperature_coefficient_rows", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(module, "load_startup_pressure_calibration_rows", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(module, "_append_dataframe_sheet", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(module, "_annotate_workbook_with_actual_device_ids", lambda *_args, **_kwargs: None)

    result = module.build_corrected_delivery(
        run_dir=run_dir,
        output_dir=out_dir,
        coeff_cfg={
            "postrun_selected_pressure_points_override": ["ambient"],
            "summary_columns": {
                "co2": {"temperature": "thermometer_temp_c"},
                "h2o": {"temperature": "dew_temp_c"},
            },
        },
        pressure_row_source="startup_calibration",
    )

    assert report_kwargs["coeff_cfg"]["selected_pressure_points"] == ["ambient"]
    assert report_kwargs["coeff_cfg"]["original_selected_pressure_points"] == ["ambient", 500]
    assert report_kwargs["coeff_cfg"]["selected_pressure_points_source"] == "postrun_override"
    assert report_kwargs["gas_temperature_keys"] == {"co2": "thermometer_temp_c", "h2o": "dew_temp_c"}
    assert result["pressure_points_summary"] == {
        "original": ["ambient", 500],
        "effective": ["ambient"],
        "source": "postrun_override",
    }


def test_build_corrected_delivery_without_override_keeps_runtime_snapshot_pressure_points(
    tmp_path: Path,
    monkeypatch,
) -> None:
    run_dir = tmp_path / "run_mixed"
    out_dir = tmp_path / "out_mixed"
    run_dir.mkdir()
    out_dir.mkdir()
    (run_dir / "runtime_config_snapshot.json").write_text(
        json.dumps({"workflow": {"selected_pressure_points": ["ambient", 500]}}),
        encoding="utf-8",
    )
    report_kwargs: dict[str, object] = {}

    monkeypatch.setattr(module, "_filter_no_500_summary_paths", lambda *_args, **_kwargs: ([run_dir / "summary.xlsx"], []))
    monkeypatch.setattr(
        module,
        "build_corrected_water_points_report",
        lambda *_args, **kwargs: report_kwargs.update(kwargs)
        or {
            "summary": pd.DataFrame([{"Analyzer": "GA01", "Gas": "CO2"}]),
            "simplified": pd.DataFrame([{"Analyzer": "GA01", "Gas": "CO2", **{f"a{i}": float(i + 1) for i in range(9)}}]),
            "original": pd.DataFrame([{"Analyzer": "GA01", "Gas": "CO2"}]),
            "points": pd.DataFrame([{"Analyzer": "GA01", "Gas": "CO2"}]),
            "ranges": pd.DataFrame([{"Analyzer": "GA01", "Gas": "CO2"}]),
            "topn": pd.DataFrame([{"Analyzer": "GA01", "Gas": "CO2"}]),
            "h2o_selected_rows": pd.DataFrame(),
            "h2o_anchor_gate_hits": pd.DataFrame(),
        },
    )
    monkeypatch.setattr(module, "extract_run_device_ids", lambda *_args, **_kwargs: {"GA01": "005"})
    monkeypatch.setattr(module, "load_temperature_coefficient_rows", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(module, "load_startup_pressure_calibration_rows", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(module, "_append_dataframe_sheet", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(module, "_annotate_workbook_with_actual_device_ids", lambda *_args, **_kwargs: None)

    result = module.build_corrected_delivery(
        run_dir=run_dir,
        output_dir=out_dir,
        coeff_cfg={},
        pressure_row_source="startup_calibration",
    )

    assert report_kwargs["coeff_cfg"]["selected_pressure_points"] == ["ambient", 500]
    assert report_kwargs["coeff_cfg"]["original_selected_pressure_points"] == ["ambient", 500]
    assert report_kwargs["coeff_cfg"]["selected_pressure_points_source"] == "runtime_snapshot"
    assert result["pressure_points_summary"] == {
        "original": ["ambient", 500],
        "effective": ["ambient", 500],
        "source": "runtime_snapshot",
    }


def test_build_corrected_download_plan_rows_falls_back_to_original_when_simplified_degrades() -> None:
    simplified = pd.DataFrame(
        [
            {"Analyzer": "GA01", "Gas": "CO2", **{f"a{i}": float(i + 1) for i in range(9)}},
        ]
    )
    original = pd.DataFrame(
        [
            {"Analyzer": "GA01", "Gas": "CO2", **{f"a{i}": float(i + 101) for i in range(9)}},
        ]
    )
    summary = pd.DataFrame(
        [
            {
                "Analyzer": "GA01",
                "Gas": "CO2",
                "original_rmse": 10.0,
                "simplified_rmse": 15.0,
                "max_prediction_diff_between_original_and_simplified": 6.0,
            }
        ]
    )

    rows = module.build_corrected_download_plan_rows(
        simplified,
        actual_device_ids={"GA01": "086"},
        original_frame=original,
        summary_frame=summary,
        corrected_cfg={},
    )

    assert len(rows) == 1
    row = rows[0]
    assert row["CoefficientSource"] == "original_fallback"
    assert row["FallbackReason"] == "simplified_rmse_ratio_exceeded;simplified_prediction_diff_exceeded"
    assert row["a0"] == 101.0
    assert row["ActualDeviceId"] == "086"


def test_build_corrected_download_plan_rows_keeps_simplified_when_guard_not_triggered() -> None:
    simplified = pd.DataFrame(
        [
            {"Analyzer": "GA01", "Gas": "H2O", **{f"a{i}": float(i + 1) for i in range(7)}},
        ]
    )
    original = pd.DataFrame(
        [
            {"Analyzer": "GA01", "Gas": "H2O", **{f"a{i}": float(i + 101) for i in range(7)}},
        ]
    )
    summary = pd.DataFrame(
        [
            {
                "Analyzer": "GA01",
                "Gas": "H2O",
                "original_rmse": 1.0,
                "simplified_rmse": 1.1,
                "max_prediction_diff_between_original_and_simplified": 0.1,
            }
        ]
    )

    rows = module.build_corrected_download_plan_rows(
        simplified,
        actual_device_ids={"GA01": "086"},
        original_frame=original,
        summary_frame=summary,
        corrected_cfg={},
    )

    assert len(rows) == 1
    row = rows[0]
    assert row["CoefficientSource"] == "simplified"
    assert row["FallbackReason"] == ""
    assert row["a0"] == 1.0
    assert row["a7"] == 0.0
    assert row["a8"] == 0.0


def test_run_from_cli_writes_separated_fit_and_writeback_summaries(tmp_path: Path, monkeypatch) -> None:
    run_dir = tmp_path / "run_summary"
    run_dir.mkdir()
    cfg_path = run_dir / "runtime_config_snapshot.json"
    cfg_path.write_text("{}", encoding="utf-8")
    out_dir = tmp_path / "out_summary"

    monkeypatch.setattr(
        module,
        "build_corrected_delivery",
        lambda **_kwargs: {
            "report_path": str(run_dir / "calibration_coefficients.xlsx"),
            "output_dir": str(out_dir),
            "filtered_summary_paths": [],
            "filter_stats": [{"source": "summary.xlsx", "original_rows": 10, "removed_rows": 2, "kept_rows": 8}],
            "actual_device_ids": {"GA01": "086"},
            "download_plan_rows": [{"Analyzer": "GA01", "Gas": "CO2", "CoefficientSource": "original_fallback", "FallbackReason": "simplified_rmse_ratio_exceeded"}],
            "temperature_rows": [],
            "pressure_rows": [],
            "pressure_row_source": "startup_calibration",
            "pressure_points_summary": {"original": ["ambient", 500], "effective": ["ambient"], "source": "postrun_override"},
            "fit_quality_summary": [
                {
                    "Analyzer": "GA01",
                    "Gas": "CO2",
                    "TemperatureColumnUsed": "thermometer_temp_c",
                    "ModelFeaturePolicy": "ambient_only_fallback",
                    "FitInputQuality": "fail",
                    "FitInputWarning": "ratio_span_too_small",
                    "DeliveryRecommendationCode": "forbid_download",
                    "OverallSuggestion": "暂不建议",
                }
            ],
            "coefficient_source_summary": [
                {
                    "Analyzer": "GA01",
                    "ActualDeviceId": "086",
                    "Gas": "CO2",
                    "CoefficientSource": "original_fallback",
                    "FallbackReason": "simplified_rmse_ratio_exceeded",
                }
            ],
            "run_structure_hints": [],
        },
    )
    monkeypatch.setattr(module, "load_config", lambda _path: {"_base_dir": str(tmp_path)})
    monkeypatch.setattr(
        module,
        "write_coefficients_to_live_devices",
        lambda **_kwargs: {
            "scan_rows": [],
            "summary_rows": [{"Analyzer": "GA01", "TargetDeviceId": "086", "LiveDeviceId": "086", "Status": "partial", "MatchedGroups": 1, "ExpectedGroups": 2}],
            "detail_rows": [{"Analyzer": "GA01", "Group": "1", "ReadbackOk": False, "Error": "READBACK_MISMATCH"}],
        },
    )
    monkeypatch.setattr(module, "_append_dataframe_sheet", lambda *_args, **_kwargs: None)

    result = module.run_from_cli(
        run_dir=str(run_dir),
        config_path=str(cfg_path),
        output_dir=str(out_dir),
        write_devices=True,
    )

    summary_json = json.loads((out_dir / "autodelivery_summary.json").read_text(encoding="utf-8"))
    summary_md = (out_dir / "summary.md").read_text(encoding="utf-8")

    assert summary_json["fit_quality_summary"][0]["FitInputQuality"] == "fail"
    assert summary_json["coefficient_source_summary"][0]["CoefficientSource"] == "original_fallback"
    assert summary_json["device_write_verify_summary"][0]["Status"] == "partial"
    assert summary_json["corrected_fit_quality"] == "fail"
    assert summary_json["device_write_verify_quality"] == "partial"
    assert summary_json["runtime_parity_quality"] == "not_audited"
    assert summary_json["final_write_ready"] is False
    assert "## fit_quality_summary" in summary_md
    assert "## coefficient_source_summary" in summary_md
    assert "## runtime_parity_summary" in summary_md
    assert "## device_write_verify_summary" in summary_md
    assert "## write_readiness_summary" in summary_md
    assert result["device_write_verify_summary"][0]["FailureReasons"] == "READBACK_MISMATCH"


def test_run_from_cli_uses_runtime_parity_pass_for_final_write_ready(tmp_path: Path, monkeypatch) -> None:
    run_dir = tmp_path / "run_parity_pass"
    run_dir.mkdir()
    cfg_path = run_dir / "runtime_config_snapshot.json"
    cfg_path.write_text("{}", encoding="utf-8")
    out_dir = tmp_path / "out_parity_pass"
    parity_path = tmp_path / "runtime_parity_summary.json"
    parity_path.write_text(
        json.dumps(
            {
                "parity_verdict": "parity_pass",
                "runtime_parity_quality": "pass",
                "legacy_stream_only": False,
                "candidate_rows": [
                    {"candidate_name": "ratio_f_plus_temperature", "candidate_status": "tested", "rmse": 0.1}
                ],
                "best_candidate": {"candidate_name": "ratio_f_plus_temperature"},
            }
        ),
        encoding="utf-8",
    )
    appended_sheets: list[str] = []

    monkeypatch.setattr(
        module,
        "build_corrected_delivery",
        lambda **_kwargs: {
            "report_path": str(run_dir / "calibration_coefficients.xlsx"),
            "output_dir": str(out_dir),
            "filtered_summary_paths": [],
            "filter_stats": [],
            "actual_device_ids": {"GA01": "086"},
            "download_plan_rows": [{"Analyzer": "GA01", "Gas": "CO2", "CoefficientSource": "simplified", "FallbackReason": ""}],
            "temperature_rows": [],
            "pressure_rows": [],
            "pressure_row_source": "startup_calibration",
            "pressure_points_summary": {"original": ["ambient"], "effective": ["ambient"], "source": "runtime_snapshot"},
            "fit_quality_summary": [
                {
                    "Analyzer": "GA01",
                    "Gas": "CO2",
                    "TemperatureColumnUsed": "thermometer_temp_c",
                    "ModelFeaturePolicy": "ambient_only_fallback",
                    "FitInputQuality": "ok",
                    "FitInputWarning": "",
                    "DeliveryRecommendationCode": "ok",
                    "OverallSuggestion": "建议下发",
                }
            ],
            "coefficient_source_summary": [
                {
                    "Analyzer": "GA01",
                    "ActualDeviceId": "086",
                    "Gas": "CO2",
                    "CoefficientSource": "simplified",
                    "FallbackReason": "",
                }
            ],
            "run_structure_hints": [],
        },
    )
    monkeypatch.setattr(module, "load_config", lambda _path: {"_base_dir": str(tmp_path)})
    monkeypatch.setattr(
        module,
        "write_coefficients_to_live_devices",
        lambda **_kwargs: {
            "scan_rows": [],
            "summary_rows": [{"Analyzer": "GA01", "TargetDeviceId": "086", "LiveDeviceId": "086", "Status": "ok", "MatchedGroups": 2, "ExpectedGroups": 2}],
            "detail_rows": [],
        },
    )
    monkeypatch.setattr(module, "_append_dataframe_sheet", lambda _path, sheet_name, _frame: appended_sheets.append(str(sheet_name)))

    result = module.run_from_cli(
        run_dir=str(run_dir),
        config_path=str(cfg_path),
        output_dir=str(out_dir),
        write_devices=True,
        runtime_parity_summary_path=str(parity_path),
    )

    summary_json = json.loads((out_dir / "autodelivery_summary.json").read_text(encoding="utf-8"))
    summary_md = (out_dir / "summary.md").read_text(encoding="utf-8")

    assert summary_json["runtime_parity_quality"] == "pass"
    assert summary_json["device_write_verify_quality"] == "pass"
    assert summary_json["corrected_fit_quality"] == "pass"
    assert summary_json["final_write_ready"] is True
    assert result["readiness_code"] == "all_gates_passed"
    assert "runtime_parity_summary" in appended_sheets
    assert "write_readiness_summary" in appended_sheets
    assert "- final_write_ready: True" in summary_md


def test_run_from_cli_uses_runtime_probe_summary_for_legacy_stream_gate(tmp_path: Path, monkeypatch) -> None:
    run_dir = tmp_path / "run_probe_legacy"
    run_dir.mkdir()
    cfg_path = run_dir / "runtime_config_snapshot.json"
    cfg_path.write_text("{}", encoding="utf-8")
    out_dir = tmp_path / "out_probe_legacy"
    probe_path = tmp_path / "baseline_stream_summary.json"
    probe_path.write_text(
        json.dumps(
            {
                "probe_type": "baseline_ygas_stream",
                "stream_formats_seen": ["legacy"],
                "visible_runtime_inputs_available": [
                    "target_available",
                    "legacy_signal_available",
                    "temperature_available",
                ],
                "visible_runtime_inputs_missing": [
                    "ratio_f_available",
                    "ratio_raw_available",
                    "signal_available",
                    "ref_signal_available",
                ],
                "conclusion_hint": "legacy runtime stream does not expose ratio or chamber/case inputs needed for parity",
            }
        ),
        encoding="utf-8",
    )

    monkeypatch.setattr(
        module,
        "build_corrected_delivery",
        lambda **_kwargs: {
            "report_path": str(run_dir / "calibration_coefficients.xlsx"),
            "output_dir": str(out_dir),
            "filtered_summary_paths": [],
            "filter_stats": [],
            "actual_device_ids": {"GA03": "079"},
            "download_plan_rows": [{"Analyzer": "GA03", "Gas": "CO2", "CoefficientSource": "simplified", "FallbackReason": ""}],
            "temperature_rows": [],
            "pressure_rows": [],
            "pressure_row_source": "startup_calibration",
            "pressure_points_summary": {"original": ["ambient"], "effective": ["ambient"], "source": "runtime_snapshot"},
            "fit_quality_summary": [
                {
                    "Analyzer": "GA03",
                    "Gas": "CO2",
                    "TemperatureColumnUsed": "Temp",
                    "ModelFeaturePolicy": "explicit_config",
                    "FitInputQuality": "ok",
                    "FitInputWarning": "",
                    "DeliveryRecommendationCode": "ok",
                    "OverallSuggestion": "candidate_only",
                }
            ],
            "coefficient_source_summary": [
                {
                    "Analyzer": "GA03",
                    "ActualDeviceId": "079",
                    "Gas": "CO2",
                    "CoefficientSource": "simplified",
                    "FallbackReason": "",
                }
            ],
            "run_structure_hints": [],
        },
    )
    monkeypatch.setattr(module, "load_config", lambda _path: {"_base_dir": str(tmp_path)})
    monkeypatch.setattr(module, "_append_dataframe_sheet", lambda *_args, **_kwargs: None)

    result = module.run_from_cli(
        run_dir=str(run_dir),
        config_path=str(cfg_path),
        output_dir=str(out_dir),
        write_devices=False,
        runtime_parity_summary_path=str(probe_path),
    )

    summary_json = json.loads((out_dir / "autodelivery_summary.json").read_text(encoding="utf-8"))

    assert summary_json["corrected_fit_quality"] == "pass"
    assert summary_json["device_write_verify_quality"] == "not_requested"
    assert summary_json["runtime_parity_quality"] == "parity_inconclusive_missing_runtime_inputs"
    assert summary_json["final_write_ready"] is False
    assert summary_json["readiness_reason"] == "legacy_stream_insufficient_for_runtime_parity"
    assert result["runtime_parity_summary"][0]["LegacyStreamOnly"] is True
    assert result["runtime_parity_summary"][0]["ParityVerdict"] == "parity_inconclusive_missing_runtime_inputs"
